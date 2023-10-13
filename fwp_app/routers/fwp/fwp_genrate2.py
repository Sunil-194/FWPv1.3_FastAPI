from fpdf import FPDF,HTMLMixin
from os.path import join
from PIL import ImageColor
import sys
import os, glob
import json
import requests
import datetime as dt
import openpyxl
import matplotlib.pyplot as plt
import pandas as pd
import seaborn as sns
import numpy as np
import matplotlib.ticker as tick
from matplotlib import font_manager as fm
import matplotlib.font_manager as font_manager
import math
import matplotlib.pyplot as plt
from matplotlib.pyplot import gca
import locale 
import PyPDF2
import math 
import traceback
from PyPDF2 import PdfWriter
import boto3




class InvalidUserID(Exception):
    
    def __init__(self, userID, message="Mobile_Number"):
        self.userID = userID
        self.message = message
        super().__init__(self.message)
# API Setup

def api_call(json_data,save_path):          
    final_pdf_name = json_data['meta']['user_uuid']
    money_sign_pdf(json_data,final_pdf_name,save_path)

 
#//*---PDF INDEX NUMBER SETUP-----*//
your_fin_prof_idx = 0 
your_1_view_idx = 0 
your_fin_analysis_idx = 0 
your_fw_plan_idx = 0 
fin_feat_product_list = 0
best_practices_idx = 0 

#//*-----Index Text of Page--**////
def index_text(pdf,col):
    #//*---Page Index Number----*//
    pdf.set_xy(px2MM(1870), px2MM(1018))  
    pdf.set_font('LeagueSpartan-Regular', size=px2pts(30))
    pdf.set_text_color(*hex2RGB(col))
    pdf.cell(px2MM(20), px2MM(42),str(pdf.page_no()),align='R')


 

#//*----------setting of Pdf Pages---*//

# 1e3 == 1000 where 'e' is exponential power of 10 (10 raise to power 3 = 1* 10^3)
#//*--Setting 0 to 0.0K
def format_cash(amount):
    negative_flag = False
    if amount < 0:
        amount = abs(amount)
        negative_flag = True
        
    def truncate_float(number, places):
        return round(int(number * (10 ** places)) / 10 ** places,1)
    
    if amount < 1e3:
        x = str(truncate_float((amount / 1e3), 2))
        if not negative_flag:
            return x + "K"
        else:
            return '-'+x + "K"

    if 1e3 <= amount < 1e5:
        x = str(truncate_float((amount / 1e5) * 100, 2))
        if not negative_flag:
            return x + "K"
        else:
            return '-'+x + "K"

    if 1e5 <= amount < 1e7:
        x = str(truncate_float((amount / 1e7) * 100, 2))
        if not negative_flag:
            return x + "L"
        else:
            return '-'+x + "L"

    if amount >= 1e7:
        x = str(truncate_float(amount / 1e7, 2))
        if not negative_flag:
            return x + "Cr"
        else:
            return '-'+x + "Cr"
        
#//*--Setting 0 to 0.0L    
def format_cash2(amount):
    negative_flag = False
    if amount < 0:
        amount = abs(amount)
        negative_flag = True
        
    def truncate_float(number, places):
        return round(int(number * (10 ** places)) / 10 ** places,1)
    
    if amount <= 1e1:
        x = str(truncate_float((amount / 1e5), 2))
        if not negative_flag:
            return x + "L"
        else:
            return '-'+x + "L"

    if 1e1 < amount <= 1e3:
        x = str(truncate_float((amount / 1e3), 2))
        if not negative_flag:
            return x + "K"
        else:
            return '-'+x + "K"
        # return amount

    if 1e3 <= amount < 1e5:
        x = str(truncate_float((amount / 1e5) * 100, 2))
        if not negative_flag:
            return x + "K"
        else:
            return '-'+x + "K"

    if 1e5 <= amount < 1e7:
        x = str(truncate_float((amount / 1e7) * 100, 2))
        if not negative_flag:
            return x + "L"
        else:
            return '-'+x + "L"

    if amount >= 1e7:
        x = str(truncate_float(amount / 1e7, 2))
        if not negative_flag:
            return x + "Cr"
        else:
            return '-'+x + "Cr"
  
#//*---If val=300 print 300 (insted of 0.3K)    
def format_cash3(amount):
    negative_flag = False
    if amount < 0:
        amount = abs(amount)
        negative_flag = True
        
    def truncate_float(number, places):
        return round(int(number * (10 ** places)) / 10 ** places,1)

    if amount < 1e3:
        x = str(truncate_float((amount), 2))
        if not negative_flag:
            return x
        else:
            return '-'+x
        # return amount

    if 1e3 <= amount < 1e5:
        x = str(truncate_float((amount / 1e5) * 100, 2))
        if not negative_flag:
            return x + "K"
        else:
            return '-'+x + "K"

    if 1e5 <= amount < 1e7:
        x = str(truncate_float((amount / 1e7) * 100, 2))
        if not negative_flag:
            return x + "L"
        else:
            return '-'+x + "L"

    if amount >= 1e7:
        x = str(truncate_float(amount / 1e7, 2))
        if not negative_flag:
            return x + "Cr"
        else:
            return '-'+x + "Cr"
    
# Unit conversionss
def px2MM(val):
  # Sauce: https://www.figma.com/community/plugin/841435609952260079/Unit-Converter
#   return val * (25.4 / 72)
  return val * 0.264583333338

def mm2PX(val):
  # Sauce: https://www.figma.com/community/plugin/841435609952260079/Unit-Converter
  return val * 3.7795275591

def hex2RGB(val):
  return list(ImageColor.getcolor(val, "RGB"))

def px2pts(val):
    return val*0.75

#//*---remove empty strings
def remove_empty_strings(string):
    return string != ""
  
# remove pkl files
for f in glob.glob("*.pkl"):
  os.remove(f)
  
# reportpath=os.getcwd()+'/public/money-sign-reports/'
cwd = script_dir = os.path.abspath( os.path.dirname(__file__) )

logo = join(cwd,'assets','images','logo','1FBlack.png')
logo2 = join(cwd,'assets','images','logo','1FBlackPB.png')


def money_sign_pdf(json_data,final_pdf_name,save_path):
    pdf  = FPDF('L','mm',(px2MM(1080), px2MM(1920)))
    
    #//*-----File Cleaning----*//
    if os.path.exists("asset_chart.png"):
          os.remove("asset_chart.png")
      
    if os.path.exists("acutal_networth_chart.png"):
          os.remove("acutal_networth_chart.png")
          
    if os.path.exists("liabilities_chart.png"):
          os.remove("liabilities_chart.png")
    
    LGpkl_file = os.path.join(cwd,'assets','fonts','League_Spartan','static')
    test = os.listdir(LGpkl_file)
    for item in test:
        if item.endswith(".pkl"):
            os.remove(os.path.join(LGpkl_file, item))
    
    test = os.listdir(cwd)
    for item in test:
        if item.endswith(".pkl"):
            os.remove(item)
                        
    Pratapkl_file = os.path.join(cwd,'assets','fonts','Prata')
    test = os.listdir(Pratapkl_file)
    for item in test:
        if item.endswith(".pkl"):
            os.remove(os.path.join(Pratapkl_file, item))
            
    Inter_font = os.path.join(cwd,'assets','fonts','Inter','static')
    test = os.listdir(Inter_font)
    for item in test:
        if item.endswith(".pkl"):
            os.remove(os.path.join(Inter_font, item))
    
    print('\n\n\n cwd', cwd)
    pdf.set_auto_page_break(False)            
    try:
                
        pdf.add_font('LeagueSpartan-SemiBold', '', os.path.join(cwd, 'assets', 'fonts', 'League_Spartan','static', 'LeagueSpartan-SemiBold.ttf'))
        pdf.add_font('LeagueSpartan-Bold', '', os.path.join(cwd, 'assets', 'fonts', 'League_Spartan','static', 'LeagueSpartan-Bold.ttf'))
        pdf.add_font('LeagueSpartan-Regular', '', os.path.join(cwd, 'assets', 'fonts', 'League_Spartan','static', 'LeagueSpartan-Regular.ttf'))
        pdf.add_font('LeagueSpartan-Medium', '', os.path.join(cwd, 'assets', 'fonts', 'League_Spartan', 'static', 'LeagueSpartan-Medium.ttf'))
        pdf.add_font('LeagueSpartan-Light', '', os.path.join(cwd, 'assets', 'fonts', 'League_Spartan', 'static', 'LeagueSpartan-Light.ttf'))
        pdf.add_font('Prata', '', os.path.join(cwd, 'assets', 'fonts', 'Prata','Prata-Regular.ttf'))
        pdf.add_font('Inter-ExtraLight', '', os.path.join(cwd, 'assets', 'fonts', 'Inter','static','Inter-ExtraLight.ttf'))
    except:
        raise traceback.format_exc()
    
    # c_MoneyS = user_data['moneySign'].split(' ')
    try:
        money_sign = json_data['money_sign']['money_sign']
        c_MoneyS = money_sign.split(' ')
        c_MoneyS = c_MoneyS[-1].strip()
    except:
        print('User Does not have his Money Sign')
        raise traceback.format_exc() 


    
    money_signData={
        'Eagle':{
            'Front_P':{
                'Ms_image':'Eagle.png',
                'Vt_line':'#7C5FF2',
                'Date_c':'#C6B9FF'
            },
            'content':['#F3F6F9','#E6E0FF','#C6B9FF','#A792FF','#7C5FF2','#5641AA'],
            'Money_Sign':['#E6E0FF','#7C5FF2','Far-Sighted Eagle'],
            #//*-behav_bias = image,color,x-axis,y-axis,width,height
            'behav_bias':['Eagle_bb.svg','#7C5FF2',837,567,1083,519,'#A792FF'],
            'gen_profile':['#5641AA','#A792FF','#7C5FF2'],
            'fin_profile':['#E6E0FF']
        },
        'Horse':{
            'Front_P':{
                'Ms_image':'Horse.png',
                'Vt_line':'#4DC3A7',
                'Date_c':'#ACE4D7'
            },
            'content':['#F3F6F9','#DEF7F1','#ACE4D7','#82DBC6','#4DC3A7','#229479'],
            'Money_Sign':['#DEF7F1','#4DC3A7','Persistent Horse'],
            'behav_bias':['Horse_bb.svg','#82DBC6',1162,322,688,688,'#82DBC6'],
            'gen_profile':['#229479','#82DBC6','#4DC3A7'],
            'fin_profile':['#DEF7F1']
        },
        'Tiger':{
            'Front_P':{
                'Ms_image':'Tiger.png',
                'Vt_line':'#FFCA41',
                'Date_c':'#FFE6A8'
            },
            'content':['#F3F6F9','#FFF3DB','#FFE6A8','#FFD976','#FFCA41','#D2A530'],
            'Money_Sign':['#FFF3DB','#FFCA41','Tactical Tiger'],
            'behav_bias':['Tiger_bb.svg','#FFCA41',1170,330,680,680,'#FFD976'],
            'gen_profile':['#D2A530','#FFD976','#FFCA41'],
            'fin_profile':['#FFF3DB']
        },
        'Lion':{
            'Front_P':{
                'Ms_image':'Lion.png',
                'Vt_line':'#FFCA41',
                'Date_c':'#FFE6A8'
            },
            'content':['#F3F6F9','#FFF3DB','#FFE6A8','#FFD976','#FFCA41','#D2A530'],
            'Money_Sign':['#FFF3DB','#FFCA41','Opportunistic Lion'],
            'behav_bias':['Lion_bb.svg','#FFCA41',1177,337,673,673,'#FFD976'],
            'gen_profile':['#D2A530','#FFD976','#FFCA41'],
            'fin_profile':['#FFF3DB']
        },
        'Elephant':{
            'Front_P':{
                'Ms_image':'Elephant.png',
                'Vt_line':'#4DC3A7',
                'Date_c':'#ACE4D7'
            },
            'content':['#F3F6F9','#DEF7F1','#ACE4D7','#82DBC6','#4DC3A7','#229479'],
            'Money_Sign':['#DEF7F1','#4DC3A7','Virtuous Elephant'],
            'behav_bias':['Elephant_bb.svg','#4DC3A7',1177,377,673,673,'#82DBC6'],
            'gen_profile':['#229479','#82DBC6','#4DC3A7'],
            'fin_profile':['#DEF7F1']
        },
        'Turtle':{
            'Front_P':{
                'Ms_image':'Turtle.png',
                'Vt_line':'#649DE5',
                'Date_c':'#ADD0FB'
            },
            'content':['#F3F6F9','#DEEDFF','#ADD0FB','#90BEF8','#649DE5','#3D7DD0'],
            'Money_Sign':['#DEEDFF','#649DE5','Vigilant Turtle'],
            'behav_bias':['Turtle_bb.svg','#649DE5',1150,310,700,700,'#90BEF8'],
            'gen_profile':['#3D7DD0','#90BEF8','#649DE5'],
            'fin_profile':['#DEEDFF']
        },
        'Whale':{
            'Front_P':{
                'Ms_image':'Whale.png',
                'Vt_line':'#649DE5',
                'Date_c':'#ADD0FB'
            },
            'content':['#F3F6F9','#DEEDFF','#ADD0FB','#90BEF8','#649DE5','#3D7DD0'],
            'Money_Sign':['#DEEDFF','#649DE5','Enlightened Whale'],
            'behav_bias':['Whale_bb.svg','#649DE5',1177,337,673,673,'#90BEF8'],
            'gen_profile':['#3D7DD0','#90BEF8','#649DE5'],
            'fin_profile':['#DEEDFF']
        },
        'Shark':{
            'Front_P':{
                'Ms_image':'Shark.png',
                'Vt_line':'#7C5FF2',
                'Date_c':'#C6B9FF'
            },
            'content':['#F3F6F9','#E6E0FF','#C6B9FF','#A792FF','#7C5FF2','#5641AA'],
            'Money_Sign':['#E6E0FF','#7C5FF2','Stealthy Shark'],
            'behav_bias':['Shark_bb.svg','#7C5FF2',1170,330,680,680,'#A792FF'],
            'gen_profile':['#5641AA','#A792FF','#7C5FF2'],
            'fin_profile':['#E6E0FF']
        }
       
    }
    
    
    #//*----Pasing pdf_setting,Json Data, MoneySign Name,Money sign wise all Images,Backgrounds to function
    #//*-----Pages Sequence is based on the sequence of function Calling ---*//
    
    x = Banner(pdf,json_data,c_MoneyS,money_signData)
    if x==False:
        return None
    content(pdf,json_data,c_MoneyS,money_signData)
    fin_profile(pdf, json_data,c_MoneyS,money_signData)
    fbs(pdf,json_data,c_MoneyS,money_signData)
    money_signtm(pdf,json_data,c_MoneyS,money_signData)
    behave_bias(pdf,json_data,c_MoneyS,money_signData)
    gen_profile(pdf,json_data,c_MoneyS,money_signData)
    your_1_view_detail(pdf,json_data,c_MoneyS,money_signData)
    assets_chart(pdf,json_data,c_MoneyS,money_signData)
    liabilities_chart(pdf,json_data,c_MoneyS,money_signData)
    emergency_planning(pdf,json_data,c_MoneyS,money_signData)
    exp_lib_mang(pdf,json_data,c_MoneyS,money_signData)
    asset_allocation(pdf,json_data,c_MoneyS,money_signData)
    net_worth(pdf,json_data,c_MoneyS,money_signData)
    net_worth_projection(pdf,json_data,c_MoneyS,money_signData)
    assumptions(pdf,json_data,c_MoneyS,money_signData)
    bureao_report(pdf,json_data,c_MoneyS,money_signData)
    libility_management_1(pdf,json_data,c_MoneyS,money_signData)
    insurance_policy_eveluation(pdf,json_data,c_MoneyS,money_signData)
    insurance_policy_recommendation_summary(pdf,json_data,c_MoneyS,money_signData)
    mf_holding_eveluation(pdf,json_data,c_MoneyS,money_signData)
    fin_wellness_plan(pdf,json_data,c_MoneyS,money_signData)
    cashflow_plan(pdf,json_data,c_MoneyS,money_signData)
    term_insurance(pdf,json_data,c_MoneyS,money_signData)
    health_insurance(pdf,json_data,c_MoneyS,money_signData)
    equity_mutual_fund(pdf,json_data,c_MoneyS,money_signData)
    debt_mutual_fund(pdf,json_data,c_MoneyS,money_signData)
    hybrid_mutual_fund(pdf,json_data,c_MoneyS,money_signData)
    credit_card(pdf,json_data,c_MoneyS,money_signData)
    building_strong_credit_profile(pdf,json_data,c_MoneyS,money_signData)
    planning_your_taxes(pdf,json_data,c_MoneyS,money_signData)
    aval_tax_deduct_1(pdf,json_data,c_MoneyS,money_signData)
    aval_tax_deduct_2(pdf,json_data,c_MoneyS,money_signData)
    aval_tax_deduct_3(pdf,json_data,c_MoneyS,money_signData)
    aval_tax_deduct_4(pdf,json_data,c_MoneyS,money_signData)
    aval_tax_deduct_5(pdf,json_data,c_MoneyS,money_signData)
    capital_gains_1(pdf,json_data,c_MoneyS,money_signData)
    capital_gains_2(pdf,json_data,c_MoneyS,money_signData)
    capital_gains_3(pdf,json_data,c_MoneyS,money_signData)
    capital_gains_4(pdf,json_data,c_MoneyS,money_signData)
    planning_for_inheritance(pdf,json_data,c_MoneyS,money_signData)
    understanding_inheritance(pdf,json_data,c_MoneyS,money_signData)
    planning_your_esate_will(pdf,json_data,c_MoneyS,money_signData)
    disclaimer(pdf,json_data,c_MoneyS,money_signData)
    lastpage(pdf,json_data,c_MoneyS,money_signData)
    
    #//*---Calling again Content Page function at last because first need to get all Index number of pages then replace it to second Page(Content Page with no Indexing)
    content(pdf,json_data,c_MoneyS,money_signData)
    

    try:
        #//*---IF Saving path directory does not exist
        pdf.output('temp.pdf')
        
        # input_file = 'temp.pdf'
        pdf_output = PdfWriter()
        
        file = open('temp.pdf', 'rb')
        readpdf = PyPDF2.PdfReader(file)
        totalpages = len(readpdf.pages)
        
        pages_list = [0,-1]
        other_page = list(x for x in range(2,totalpages-1))
        pages_list = pages_list+other_page

        
        for i in pages_list:
            current_page = readpdf.pages[i]
            pdf_output.add_page(current_page)
        dir_name = save_path
        if not os.path.exists(dir_name):
            os.mkdir(dir_name)
        
        ts = dt.datetime.now()
        ts = str(ts.strftime("%d-%m-%Y"))
        output_file = str(final_pdf_name)+f'_{ts}'+'.pdf'
        opfile = join(save_path,output_file)
        pdf_output.write(opfile) 
        
        # #//*---setting up boto for aws s3
        # s3 = boto3.client(
        #     "s3",
        #     aws_access_key_id=os.environ.get('AWS_S3_ACCESS_KEY_ID'),
        #     aws_secret_access_key=os.environ.get('AWS_S3_SECRET_ACCESS_KEY'),
        #     region_name=os.environ.get('AWS_REGION_NAME')
        # )
        # s3.upload_fileobj(pdf_output.write(output_file), "your_bucket_name", output_file)
        
        
        
        file.close()
        
            #//*-----File Cleaning----*//
        if os.path.exists("asset_chart.png"):
            os.remove("asset_chart.png")
        
        if os.path.exists("acutal_networth_chart.png"):
            os.remove("acutal_networth_chart.png")
            
        if os.path.exists("liabilities_chart.png"):
            os.remove("liabilities_chart.png")
            
        if os.path.exists('temp.pdf'):
            os.remove('temp.pdf')
        return True
    except:
        raise traceback.format_exc()
    
#//*--------Function to Create a Page with heading and width----*//
def page_build(pdf,heading,width):
    pdf.add_page()
    pdf.set_fill_color(*hex2RGB('#FCF8ED'))
    pdf.rect(0, 0, px2MM(1920), px2MM(1080),'F')

    #//*----Featured List of Financial Products----*//
    pdf.set_xy(px2MM(120),px2MM(80)) 
    pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(60))
    pdf.set_text_color(*hex2RGB('#1A1A1D'))
    pdf.cell(px2MM(width), px2MM(84),heading,align='L')
    
    #//*---Top Black box
    pdf.set_fill_color(*hex2RGB('#000000'))
    pdf.rect(px2MM(0), px2MM(80), px2MM(15), px2MM(84),'F')     

    
    
#//*------Banner-----*//
def Banner(pdf,json_data,c_MoneyS,money_signData):
    try:
        # user_name =['name']
        user_name = json_data['meta']["name"]
        if user_name.strip()=="":
            print('No Name in PDF')
            return None
        # user_name = json_data['Name']
    except:
        return False
        
    # pdf = FPDF('L','mm','A4')
    pdf.add_page()
    pdf.set_fill_color(*hex2RGB('#000000'))
    pdf.rect(0, 0, px2MM(1120), px2MM(1080), 'F')
    
    MoneyS_color = money_signData[c_MoneyS]['Front_P']['Vt_line']
    Date_c = money_signData[c_MoneyS]['Front_P']['Date_c']
    Ms_Image = money_signData[c_MoneyS]['Front_P']['Ms_image']
    
    #/**--For Money sigh right banner

    pdf.set_fill_color(*hex2RGB(MoneyS_color))
    pdf.rect(px2MM(1120), px2MM(0), px2MM(800), px2MM(1080), 'F')
 
    pdf.image(join(cwd,'assets', 'images','money_sign_png',Ms_Image),px2MM(1120), px2MM(0), px2MM(800), px2MM(1080))
 
    pdf.set_fill_color(*hex2RGB('#000000'))
    pdf.rect(0, 0, px2MM(1120), px2MM(1080), 'F')
    
    
    #//*---1F logo--*/
    pdf.image(logo,px2MM(120), px2MM(80), px2MM(98), px2MM(113))
    
    pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(120))
    pdf.set_text_color(*hex2RGB('#FFFFFF'))
    pdf.set_xy(px2MM(120),px2MM(333))
    pdf.multi_cell(px2MM(796), px2MM(168),'Financial\nWellness Plan')

    
    # Test of User name and Date
    if len(user_name) > 24:
        name_y = 692
    else:
        name_y = 804

    # Test of User name and Date
    pdf.set_font('LeagueSpartan-Regular', size=px2pts(80))
    pdf.set_text_color(*hex2RGB('#FFFFFF'))
    pdf.set_xy(px2MM(120),px2MM(name_y))
    pdf.multi_cell(px2MM(924), px2MM(112),user_name.title(),align="L")
    y_after_name = mm2PX(pdf.get_y())


    pdf.set_font('LeagueSpartan-Light', size=px2pts(60))
    pdf.set_text_color(*hex2RGB(Date_c))
    Day=dt.datetime.now().strftime("%d")

    month=dt.datetime.now().strftime("%b")
    year=dt.datetime.now().strftime("%Y")

    if 4 <= int(Day) <= 20 or 24 <= int(Day) <= 30:
        suffix = "th"
    else:
        suffix = ["st", "nd", "rd"][int(Day) % 10 - 1]


    pdf.set_fill_color(*hex2RGB('#ffffff'))
    pdf.set_xy(px2MM(120),px2MM(y_after_name))
    pdf.cell(px2MM(60), px2MM(84),str(Day),'LR',align='R')
    
    x_after_day = pdf.get_x()

    pdf.set_font('LeagueSpartan-Light', size=px2pts(36))
    pdf.set_xy(px2MM(mm2PX(x_after_day)-5),px2MM(y_after_name- 10))
    pdf.cell(px2MM(32), px2MM(84),suffix,align='L')


    y_after_suffix = mm2PX(pdf.get_y()) #804
    d_x = pdf.get_x()
    d_x2 = pdf.get_x()
    pdf.set_font('LeagueSpartan-Light', size=px2pts(60))
    pdf.set_xy(px2MM(mm2PX(d_x2)),px2MM(y_after_name))
    pdf.cell(px2MM(100), px2MM(84),' '+str(month)+', '+str(year))
    #//*---Th suffix---*//


    height_of_rect = mm2PX(pdf.get_y())- name_y + 84

    #//*---Left Bottom Vertical Line
    pdf.set_xy(px2MM(0),px2MM(692))
    pdf.set_fill_color(*hex2RGB(MoneyS_color))
    pdf.rect(px2MM(0), px2MM(name_y), px2MM(20), px2MM(height_of_rect), 'F')
    
    
    
# //*----Contents----*//  

def content(pdf,json_data,c_MoneyS,money_signData):
    pdf.add_page()
    # pdf.rect()
    pdf.set_fill_color(*hex2RGB('#000000'))
    pdf.rect(0, 0, px2MM(1920), px2MM(1080), 'DF')
    
    #//*--Contents banner
    pdf.set_xy(px2MM(120),px2MM(80))
    pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(60))
    pdf.set_text_color(*hex2RGB('#FFFFFF'))
    pdf.cell(px2MM(600), px2MM(84), 'CONTENTS')
    
    #//*----- for vertical dash
    basy_y = 204
    y_gap = 140
    # h_y = 128.83
    fill_color = money_signData[c_MoneyS]['content']
    
    #//*----For Content headings and para
    
    cont_head_basey = 210    
    cont_para_basey = 266
    
    cont_headings = ['Your Financial Profile','Your 1 view','Your Financial Analysis','Your Financial Wellness Plan','Financial Products Featured List','Best Practices']

    cont_para = ['Financial Behaviour Score, MoneySign  , Generation Profile, Life stage','Snapshot, Detailed Snapshot',"Financial Metrics, Net Worth Projection, Liability Analysis, MF Holdings Evaluation","Key takeaways, Next 3 Months Action Plan", "Term Insurance Plans, Health Insurance Plans, Equity Mutual Funds, Credit Cards","Building a Strong Credit Profile, Planning Your Income Taxes, Capital Gains Taxation by Asset Type, Planning For Inheritance, Understanding Inheritance’s Tax Implications, Planning Your Estate and Will"] 
    index_no = [your_fin_prof_idx,your_1_view_idx,your_fin_analysis_idx,your_fw_plan_idx,fin_feat_product_list,best_practices_idx]
    
    for i in range(len(cont_headings)):

        if i == len(cont_headings)-1:
            basy_y = 914
            cont_head_basey = 904
            cont_para_basey = 960

        pdf.set_fill_color(*hex2RGB(fill_color[i]))
        pdf.rect(px2MM(120), px2MM(basy_y), px2MM(8), px2MM(100), 'F')
        basy_y += y_gap
        
        

        #//*---Contents for each vertical dash

        pdf.set_draw_color(*hex2RGB('#000000'))
        pdf.set_xy(px2MM(168),px2MM(cont_head_basey))
        pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(40))
        pdf.set_text_color(*hex2RGB('#FFFFFF'))
        pdf.set_line_width(px2MM(2))
        pdf.cell(px2MM(1500), px2MM(56),cont_headings[i])


        pdf.set_xy(px2MM(168),px2MM(cont_para_basey))
        pdf.set_font('LeagueSpartan-Medium', size=px2pts(24))
        pdf.set_text_color(*hex2RGB('#898B90'))
        pdf.multi_cell(px2MM(1500), px2MM(32),cont_para[i])

        #//*---Index Number----*//
        pdf.set_xy(px2MM(1675),px2MM(263+(i*140)))
        pdf.set_font('LeagueSpartan-Medium', size=px2pts(30))
        pdf.set_text_color(*hex2RGB('#898B90'))
        pdf.cell(px2MM(125), px2MM(42),str(index_no[i]),align='R')

        cont_head_basey+=y_gap
        cont_para_basey+=y_gap
    
    #//*--To print superscritp R 
    pdf.set_font('Inter-ExtraLight', size=16)
    pdf.set_text_color(*hex2RGB('#898B90'))
    pdf.set_xy(px2MM(550), px2MM(265))
    pdf.cell(px2MM(16), px2MM(34), '®')
     
    # pdf.set_font('LeagueSpartan-Medium', size=6)
    # pdf.set_text_color(*hex2RGB('#898B90'))
    # pdf.set_xy(px2MM(706), px2MM(273))  
    # pdf.cell(px2MM(16), px2MM(8),'TM') 
    
    # #//*--To print superscritp R 
    # pdf.set_font('LeagueSpartan-Light', size=26)
    # pdf.set_text_color(*hex2RGB('#898B90'))
    # pdf.set_xy(px2MM(707), px2MM(267))
    # pdf.cell(px2MM(16), px2MM(34), '®') 
      

#//*----Financial Behaviour Score----*//  

def fbs(pdf,json_data,c_MoneyS,money_signData):
    try:
        score = json_data['oneview']['fbs']
        if score==None:
            score = 0
    except:
        return None

    #//*---Page setup
    pdf.add_page()
    pdf.set_fill_color(*hex2RGB('#FCF8ED'))
    pdf.rect(0, 0, px2MM(1920), px2MM(1080), 'DF')
    
    #//*--Heading vertical line
    vl_color = money_signData[c_MoneyS]['content'][3]
    pdf.set_xy(px2MM(125),px2MM(78))
    pdf.set_fill_color(*hex2RGB(vl_color))
    pdf.rect(px2MM(0), px2MM(80), px2MM(15), px2MM(84), 'F') 
    
    #//*---heading statement
    pdf.set_xy(px2MM(120),px2MM(80))  
    pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(60))
    pdf.set_text_color(*hex2RGB('#1A1A1D'))
    pdf.cell(px2MM(659), px2MM(84),'Financial Behaviour Score') 
    
    #/*--Description--*/
    txt1 = '''Financial Behaviour Score is a numerical representation of your financial well-being, '''
    pdf.set_font('LeagueSpartan-Regular', size=px2pts(40))
    pdf.set_text_color(*hex2RGB('#000000'))
    pdf.set_xy(px2MM(940),px2MM(325))  
    pdf.multi_cell(px2MM(860), px2MM(56),txt1,align='L') 
    
    pdf.set_font('LeagueSpartan-Light', size=px2pts(36))
    pdf.set_text_color(*hex2RGB('#000000'))
    pdf.set_xy(px2MM(1655),px2MM(381))  
    pdf.multi_cell(px2MM(860), px2MM(56),'offering',align='L') 
    
    txt2 = '''an in-depth assessment of how closely your financial choices align with your personality, demography, generation, life constraints, and the macro-economic environment.'''
    pdf.set_xy(px2MM(940),px2MM(437))  
    pdf.multi_cell(px2MM(860), px2MM(56),txt2,align='L') 
    
    #//*----Desclamer---*//
     
    txt = '''Disclaimer: Financial Behaviour Score is part of 1 Finance's patent-pending holistic financial planning framework that is aimed at generating a wellness plan for the members to help them achieve financial well-being.'''

    pdf.set_xy(px2MM(941),px2MM(702))  
    pdf.set_font('LeagueSpartan-Regular', size=px2pts(24))
    pdf.set_text_color(*hex2RGB('#65676D'))
    pdf.multi_cell(px2MM(860), px2MM(32),txt,align='L') 
    
   

    #//*---Scale width is 640(excluding the curve corner) so 1%= 6.4 
    
    if score>=0 and score<=20:
        vl_x = 138+(score*5.75)
    elif score==21:
        vl_x = 272
    elif score>21 and score<=40:
        vl_x = 272+((score-20)*5.75)
    elif score==41:
        vl_x = 406
    elif score>41 and score<=60:
        vl_x = 406+((score-40)*5.75)
    elif score==61:
        vl_x = 540
    elif score>61 and score<=80:
        vl_x = 540+((score-60)*5.75)
    elif score==81:
        vl_x = 674
    elif score>81 and score<=100:
        vl_x = 674+((score-80)*5.75)
    else:
        vl_x=138+(score*6.64)

        
    if score>=0 and score<=21:
        rect_x = 120
        text_x = 165 
    elif score>=82 and score<=100:
        rect_x = 520
        text_x = 565
    else:
        rect_x = (score*6.64)-17
        text_x = 28+ (score*6.64)    
    # vl_x = 138+(score*6.6)
    
     #//*---Score---*//
    if score>=0 and score<=20:
        pdf.image(join(cwd,'assets','images','BehaviourMeter','meter_1_20.png'),px2MM(120), px2MM(627),px2MM(700), px2MM(134))
    elif score>20 and score<=40:
        pdf.image(join(cwd,'assets','images','BehaviourMeter','meter_20_40.png'),px2MM(120), px2MM(627),px2MM(700), px2MM(134))
    elif score>40 and score<=60:
        pdf.image(join(cwd,'assets','images','BehaviourMeter','meter_40_60.png'),px2MM(120), px2MM(627),px2MM(700), px2MM(134))
    elif score>60 and score<=80:
        pdf.image(join(cwd,'assets','images','BehaviourMeter','meter_60_80.png'),px2MM(120), px2MM(627),px2MM(700), px2MM(134))
    else :
        pdf.image(join(cwd,'assets','images','BehaviourMeter','meter_80_100.png'),px2MM(120), px2MM(627),px2MM(700), px2MM(134))
    
      
    #//*---Vertical Line of Score box
    pdf.set_fill_color(*hex2RGB('#FFFFFF'))
    # pdf.set_draw_color(*hex2RGB('#E9EAEE'))
    pdf.rect(px2MM(vl_x), px2MM(532), px2MM(13), px2MM(95), 'F') 
    
    #//*---Score Box
    pdf.set_fill_color(*hex2RGB('#FFFFFF'))
    pdf.rect(px2MM(rect_x),px2MM(284), px2MM(300), px2MM(248), 'F')
    pdf.set_xy(px2MM(text_x),px2MM(324)) 
    pdf.set_font('Prata', size=px2pts(120))
    pdf.set_text_color(*hex2RGB('#000000'))
    pdf.cell(px2MM(209), px2MM(168),str(int(score)),align='C')
    
    #//*---Scale---*/
    pdf.set_xy(px2MM(120),px2MM(782)) 
    pdf.set_font('LeagueSpartan-Medium', size=px2pts(39))
    pdf.set_text_color(*hex2RGB('#898B90'))
    pdf.cell(px2MM(600), px2MM(52),'0')
    
    pdf.set_xy(px2MM(761),px2MM(782)) 
    pdf.set_font('LeagueSpartan-Medium', size=px2pts(39))
    pdf.set_text_color(*hex2RGB('#898B90'))
    pdf.cell(px2MM(59), px2MM(52),'100')

    #//*-----Index Text of Page--**////
    index_text(pdf,'#1A1A1D')
#//*----MoneySign----*//  
  
def money_signtm(pdf,json_data,c_MoneyS,money_signData):
    try:
        # moneySing_desc =['description']
        moneySing_desc = json_data["money_sign"]['money_sign_desc']
    except:
        pass
    bg_color = money_signData[c_MoneyS]['Money_Sign'][0]
    vt_line_color = money_signData[c_MoneyS]['Money_Sign'][1]
    ms_name = json_data["money_sign"]['money_sign']
    texture = c_MoneyS+'_text.png'
    #//*---Page setup
    pdf.add_page()
    
    pdf.set_draw_color(*hex2RGB(bg_color))
    pdf.set_fill_color(*hex2RGB(bg_color))
    pdf.rect(0, 0, px2MM(1920), px2MM(1080), 'F') 
    
    #//*----Money Sign Background-****
    pdf.rect(px2MM(0), px2MM(0), px2MM(789.4), px2MM(1080))
    pdf.image(join(cwd,'assets', 'images','MoneySign',texture),px2MM(0), px2MM(0), px2MM(789.4), px2MM(1080))
  
    pdf.rect(px2MM(0), px2MM(0), px2MM(789.4), px2MM(1080))
    pdf.image(join(cwd,'assets', 'images','MoneySign',c_MoneyS+'_overlay.png'),px2MM(0), px2MM(0), px2MM(789.4), px2MM(1080))

    #//*--Purple vertical line
    # pdf.set_xy(px2MM(125),px2MM(78))
    pdf.set_fill_color(*hex2RGB(vt_line_color))
    pdf.rect(px2MM(0), px2MM(80), px2MM(15), px2MM(84), 'F') 
    
    #//*---heading statement
    pdf.set_xy(px2MM(120),px2MM(80))  
    pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(60))
    pdf.set_text_color(*hex2RGB('#1A1A1D'))
    pdf.cell(px2MM(600), px2MM(84),'MoneySign') 
    
    # #//*--To print superscritp TM  of heading
    # pdf.set_xy(px2MM(395), px2MM(83))
    # pdf.set_font('LeagueSpartan-SemiBold', size=26)
    # pdf.set_text_color(*hex2RGB('#1A1A1D'))
    # pdf.cell(px2MM(30), px2MM(42), 'TM') 
    
    #//*--To print superscritp R 
    pdf.set_font('Inter-ExtraLight', size=36)
    pdf.set_text_color(*hex2RGB('#1A1A1D'))
    pdf.set_xy(px2MM(400), px2MM(77))
    pdf.cell(px2MM(90), px2MM(84), '®')  
    
    #//*---Money Sign Logog---*//
    # pdf.rect(0, 0, px2MM(1120), px2MM(1080), 'F')
    pdf.image(join(cwd,'assets', 'images','MoneySign',c_MoneyS+'.png'),px2MM(120), px2MM(224), px2MM(700), px2MM(700))
    
    #//*---Money Sign Name
    pdf.set_xy(px2MM(290),px2MM(924))  
    pdf.set_font('Prata', size=px2pts(42))
    pdf.set_text_color(*hex2RGB('#000000'))
    pdf.cell(px2MM(400), px2MM(84),ms_name,align='C') 
    
    #//*----Description---*//

    desc = moneySing_desc.replace('<br><br>','')
    desc = desc.replace('<br>',' ')
    desc = desc.replace('\n','')
    pdf.set_draw_color(*hex2RGB('#E6E0FF'))
    pdf.set_font('LeagueSpartan-Regular',size=px2pts(30))
    pdf.set_text_color(*hex2RGB('#000000'))

    
    if c_MoneyS=='Eagle':
        pdf.set_xy(px2MM(940), px2MM(338))
        
    elif c_MoneyS=='Horse':
        pdf.set_xy(px2MM(940), px2MM(317))
        
    elif c_MoneyS=='Tiger':
        pdf.set_xy(px2MM(940), px2MM(296))
        
    elif c_MoneyS=='Lion':
        pdf.set_xy(px2MM(940), px2MM(275))
        
    elif c_MoneyS=='Elephant':
        pdf.set_xy(px2MM(940), px2MM(317))
        
    elif c_MoneyS=='Turtle':
        pdf.set_xy(px2MM(940), px2MM(233))
        
    elif c_MoneyS=='Whale':
        pdf.set_xy(px2MM(940), px2MM(275))
        
    elif c_MoneyS=='Shark':
        pdf.set_xy(px2MM(940), px2MM(275))
    else:
        pdf.set_xy(px2MM(940), px2MM(275))
   
    pdf.multi_cell(px2MM(860), px2MM(42),desc,align='L')     
    
    #//*----Desclaimer---*//
    desc_y = mm2PX(pdf.get_y())+33
    pdf.set_xy(px2MM(940), px2MM(mm2PX(pdf.get_y())+32))
    pdf.set_font('LeagueSpartan-Regular',size=px2pts(24))
    pdf.set_text_color(*hex2RGB('#4B4C51'))
    pdf.multi_cell(px2MM(860), px2MM(32),"""Disclaimer: MoneySign    is a personality assessment framework based on 1 Finance's""",align='L')
    # pdf.multi_cell(px2MM(860), px2MM(32),"""Disclaimer: MoneySign® is a personality assessment framework based on 1 Finance's""",align='L')
    
    #//*--To print superscritp R 
    pdf.set_xy(px2MM(1160), px2MM(desc_y))
    pdf.set_font('Inter-ExtraLight', size=18)
    pdf.set_text_color(*hex2RGB('#4B4C51'))
    pdf.cell(px2MM(16), px2MM(32), '®')  
    
    pdf.set_draw_color(*hex2RGB('#000000'))
    sec_l = pdf.get_y()
    pdf.set_xy(px2MM(940), px2MM(mm2PX(sec_l)+32))
    pdf.set_font('LeagueSpartan-SemiBold',size=px2pts(24))
    pdf.set_text_color(*hex2RGB('#4B4C51'))
    pdf.cell(px2MM(250), px2MM(32),"""patented technology""",align='L')
    
    pdf.set_xy(px2MM(1160), px2MM(mm2PX(sec_l)+32))
    pdf.set_font('LeagueSpartan-Regular',size=px2pts(24))
    pdf.set_text_color(*hex2RGB('#4B4C51'))
    pdf.cell(px2MM(670), px2MM(32),"""that implements one of the most scientifically validated models in""",align='L')
    
    pdf.set_xy(px2MM(940), px2MM(mm2PX(sec_l)+64))
    pdf.set_font('LeagueSpartan-Regular',size=px2pts(24))
    pdf.set_text_color(*hex2RGB('#4B4C51'))
    pdf.cell(px2MM(860), px2MM(32),"""psychology and helps in hyper-personalising the financial suggestions.""",align='L')

    #//*-----Index Text of Page--**////
    index_text(pdf,'#1A1A1D')
#//*----Behavioural Bias----*//
    
def behave_bias(pdf,json_data,c_MoneyS,money_signData):
    
    df=''
    try:
        behav_bias = pd.DataFrame.from_dict(json_data['money_sign']['money_sign_behavioural_bias'])
        if behav_bias.empty:
            return None
            
        behav_bias_keys = list(behav_bias['title'])
    except:
        return None
 
    
    page_data = money_signData[c_MoneyS]['behav_bias']
    m_image = page_data[0]
    m_color = page_data[1]
    rect_color = page_data[6]
    img_x = page_data[2]
    img_y = page_data[3]
    img_w = page_data[4]
    img_4 = page_data[5]
    ini = 0
    k = 2
    if len(behav_bias)>1:
        txt2 = """We have also identified some behavioural biases that you’re likely to display while making financial decisions, and should be conscious of:"""
    elif len(behav_bias)<2:
        txt2 = """We have also identified a behavioural bias that you’re likely to display while making financial decisions, and should be conscious of:"""

    for i in range(0,len(behav_bias),2):
        #//*---Page setup
        pdf.add_page()
        pdf.set_fill_color(*hex2RGB('#FCF8ED'))
        pdf.rect(0, 0, px2MM(1920), px2MM(1080),'F') 
        
        #//*---Cloud Images
        pdf.set_draw_color(*hex2RGB('#ffffff'))
        pdf.set_xy(px2MM(750), px2MM(520)) 
        # pdf.image(join(cwd,'assets', 'images','BehaviourBias','bias.png'),px2MM(750), px2MM(520), px2MM(1187.63), px2MM(570.77))
        pdf.image(join(cwd,'assets', 'images','BehaviourBias',m_image),px2MM(img_x), px2MM(img_y), px2MM(img_w), px2MM(img_4))
        
        
        #//*--Purple vertical line
        # pdf.set_xy(px2MM(125),px2MM(78))
        pdf.set_fill_color(*hex2RGB(m_color))
        pdf.rect(px2MM(0), px2MM(80), px2MM(15), px2MM(84), 'F') 
        
        
        #//*---heading 
        pdf.set_xy(px2MM(120),px2MM(80))  
        pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(60))
        pdf.set_text_color(*hex2RGB('#1A1A1D'))
        pdf.cell(px2MM(600), px2MM(84),'Behavioural Biases') 
        
        #//*---heading statement
        pdf.set_xy(px2MM(120),px2MM(204))  
        pdf.set_font('LeagueSpartan-Regular', size=px2pts(40))
        pdf.set_text_color(*hex2RGB('#000000'))
        pdf.multi_cell(px2MM(1300), px2MM(56),txt2) 
        
        
        #//*----Content-----*//
        h_bullet = 414
        h_heading = 396
        h_para = 472
        
        gap_bullet = 284
        gap_heading = 284
        gap_para = 284
        
        for j in range(ini,k):
            
            try:
                if behav_bias_keys[j]:
                    #//* bullet
                    pdf.set_xy(px2MM(120),px2MM(h_bullet))
                    pdf.set_fill_color(*hex2RGB(rect_color))  
                    pdf.rect(px2MM(120),px2MM(h_bullet),px2MM(20),px2MM(20),'F')
                
                #//*--heading
                pdf.set_xy(px2MM(165),px2MM(h_heading))
                pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(40))
                pdf.set_text_color(*hex2RGB('#000000'))
                pdf.multi_cell(px2MM(1255), px2MM(56),behav_bias_keys[j])
                
                #//*---para
                pdf.set_xy(px2MM(165),px2MM(h_para))
                pdf.set_font('LeagueSpartan-Regular', size=px2pts(30))
                pdf.set_text_color(*hex2RGB('#1A1A1D'))
                pdf.multi_cell(px2MM(1255), px2MM(42),behav_bias['desc'].iloc[j],align='L')  
                
                h_bullet += gap_bullet
                h_heading += gap_heading
                h_para += gap_para
            except:
                pass
        
        ini +=2
        k +=2 
        
        #//*-----Index Text of Page--**////
        index_text(pdf,'#1A1A1D')
   
#//*----Genration Profile----*//    
def gen_profile(pdf,json_data,c_MoneyS,money_signData):
    try:
        # df = pd.DataFrame.from_dict(json_data["Genration"])
        df = json_data["gen_profile"]['gen_profile']
        if df.strip()=="":
            return None
        
    except:
        return None

    #//*---Page setup----*//
    pdf.add_page()
    pdf.set_fill_color(*hex2RGB('#000000'))
    pdf.rect(0, 0, px2MM(1920), px2MM(1080),'F') 
    
    #//*--Purple vertical line
    # pdf.set_xy(px2MM(125),px2MM(78))
    pdf.set_fill_color(*hex2RGB('#FCF8ED'))
    pdf.rect(px2MM(0), px2MM(80), px2MM(15), px2MM(84), 'F') 
    
    #//*---heading statement
    pdf.set_xy(px2MM(120),px2MM(80))  
    pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(60))
    pdf.set_text_color(*hex2RGB('#FFFFFF'))
    pdf.cell(px2MM(600), px2MM(84),'Generation Profile')
    
    gen_color = money_signData[c_MoneyS]['gen_profile'][0]
    your_profile_color = money_signData[c_MoneyS]['gen_profile'][1]
    bullet_profile_color = money_signData[c_MoneyS]['gen_profile'][2]
    
    if df=='Generation 1':
        pdf.set_fill_color(*hex2RGB(gen_color))
        pdf.rect(px2MM(120), px2MM(204), px2MM(527), px2MM(915), 'F')
        pdf.image(join(cwd,'assets','images','genration profile','shade.png'),px2MM(120), px2MM(204), px2MM(527), px2MM(915))
        
        pdf.set_fill_color(*hex2RGB('#1A1A1D'))
        pdf.rect(px2MM(697), px2MM(204), px2MM(527), px2MM(915), 'F')
        
        pdf.set_fill_color(*hex2RGB('#1A1A1D'))
        pdf.rect(px2MM(1273), px2MM(204), px2MM(527), px2MM(915), 'F')
        sq_bullet_1 = bullet_profile_color
        sq_bullet_2 = '#313236'
        sq_bullet_3 = '#313236'
        
    elif df=='Generation 2':
        pdf.set_fill_color(*hex2RGB('#1A1A1D'))
        pdf.rect(px2MM(120), px2MM(204), px2MM(527), px2MM(915), 'F')
        
        pdf.set_fill_color(*hex2RGB(gen_color))
        pdf.rect(px2MM(697), px2MM(204), px2MM(527), px2MM(915), 'F')
        pdf.image(join(cwd,'assets','images','genration profile','shade.png'),px2MM(697), px2MM(204), px2MM(527), px2MM(915))
        
        pdf.set_fill_color(*hex2RGB('#1A1A1D'))
        pdf.rect(px2MM(1273), px2MM(204), px2MM(527), px2MM(915), 'F')
        sq_bullet = ['#313236',bullet_profile_color,'#313236']
        sq_bullet_1 = '#313236'
        sq_bullet_2 = bullet_profile_color
        sq_bullet_3 = '#313236'
    else:
        pdf.set_fill_color(*hex2RGB('#1A1A1D'))
        pdf.rect(px2MM(120), px2MM(204), px2MM(527), px2MM(915), 'F')
        
        pdf.set_fill_color(*hex2RGB('#1A1A1D'))
        pdf.rect(px2MM(697), px2MM(204), px2MM(527), px2MM(915), 'F')
        
        pdf.set_fill_color(*hex2RGB(gen_color))
        pdf.rect(px2MM(1273), px2MM(204), px2MM(527), px2MM(915), 'F')
        pdf.image(join(cwd,'assets','images','genration profile','shade.png'),px2MM(1273), px2MM(195), px2MM(527), px2MM(915))
        sq_bullet = ['#313236','#313236',bullet_profile_color]
        sq_bullet_1 = '#313236'
        sq_bullet_2 = '#313236'
        sq_bullet_3 = bullet_profile_color

        
    #//*---For base Rectangle---*//
    #//*-------------Card 1-----------*//
    #//*----For Heading (Genrations)---*//
    pdf.set_xy(px2MM(277),px2MM(244))  
    pdf.set_font('LeagueSpartan-Medium', size=px2pts(40))
    pdf.set_text_color(*hex2RGB('#FFFFFF'))
    pdf.cell(px2MM(215), px2MM(56),'Generation 1')

    
    #//*----Personality Traits---*/
    pdf.set_xy(px2MM(160),px2MM(330))  
    pdf.set_font('LeagueSpartan-Light', size=px2pts(24))
    pdf.set_text_color(*hex2RGB('#898B90'))
    
    # pdf.set_draw_color(*hex2RGB('#313236'))
    pdf.cell(px2MM(447), px2MM(35),'PERSONALITY TRAITS')
    pdf.image(join(cwd,'assets','images','genration profile','shade_line.png'),px2MM(160), px2MM(372), px2MM(447), px2MM(1))
    
        
    #//*--Point 1---*//
    
    pdf.set_fill_color(*hex2RGB(sq_bullet_1))
    pdf.rect(px2MM(160), px2MM(393), px2MM(10), px2MM(10),'F')
    
    pdf.set_xy(px2MM(190),px2MM(380))  
    pdf.set_font('LeagueSpartan-Light', size=px2pts(24))
    pdf.set_text_color(*hex2RGB('#E9EAEE'))
    pdf.set_draw_color(*hex2RGB('#313236'))
    pdf.multi_cell(px2MM(417), px2MM(30),'Primary bread-earner in family',align='L')
    
    pdf.set_fill_color(*hex2RGB(sq_bullet_1))
    pdf.rect(px2MM(160), px2MM(446), px2MM(10), px2MM(10),'F')
    
    pdf.set_xy(px2MM(190),px2MM(419))  
    pdf.set_font('LeagueSpartan-Light', size=px2pts(24))
    pdf.set_text_color(*hex2RGB('#E9EAEE'))
    pdf.set_draw_color(*hex2RGB('#313236'))
    pdf.multi_cell(px2MM(417), px2MM(30),'Work hard to provide for their loved ones despite limited education',align='L') 

    
    #//*----Financial Behaviour---*/
    pdf.set_xy(px2MM(160),px2MM(523))  
    pdf.set_font('LeagueSpartan-Light', size=px2pts(24))
    pdf.set_text_color(*hex2RGB('#898B90'))
    pdf.cell(px2MM(447), px2MM(35),'FINANCIAL BEHAVIOUR')
    pdf.image(join(cwd,'assets','images','genration profile','shade_line.png'),px2MM(160), px2MM(565), px2MM(447), px2MM(1))

    
    #//*--Point 3 to 4---*//
    pdf.set_fill_color(*hex2RGB(sq_bullet_1))
    pdf.rect(px2MM(160), px2MM(586), px2MM(10), px2MM(10),'F')
    
    pdf.set_xy(px2MM(190),px2MM(575))  
    pdf.set_font('LeagueSpartan-Light', size=px2pts(24))
    pdf.set_text_color(*hex2RGB('#E9EAEE'))
    pdf.set_draw_color(*hex2RGB('#313236'))
    pdf.multi_cell(px2MM(417), px2MM(30),'Earning for basic sustenance',align='L')
    
    pdf.set_fill_color(*hex2RGB(sq_bullet_1))
    pdf.rect(px2MM(160), px2MM(639), px2MM(10), px2MM(10),'F')
    
    pdf.set_xy(px2MM(190),px2MM(612))  
    pdf.set_font('LeagueSpartan-Light', size=px2pts(24))
    pdf.set_text_color(*hex2RGB('#E9EAEE'))
    pdf.set_draw_color(*hex2RGB('#313236'))
    pdf.multi_cell(px2MM(417), px2MM(30),'Prioritize stability and security over taking risks with their finances',align='L') 
    
    #//*----ASPIRATIONA---*/
    pdf.set_xy(px2MM(160),px2MM(716))  
    pdf.set_font('LeagueSpartan-Light', size=px2pts(24))
    pdf.set_text_color(*hex2RGB('#898B90'))
    pdf.cell(px2MM(447), px2MM(30),'ASPIRATIONS')
    pdf.image(join(cwd,'assets','images','genration profile','shade_line.png'),px2MM(160), px2MM(758), px2MM(447), px2MM(1))
    
    #//*--Point 5 to 3---*//
    pdf.set_fill_color(*hex2RGB(sq_bullet_1))
    pdf.rect(px2MM(160), px2MM(779), px2MM(10), px2MM(10),'F')
    
    pdf.set_xy(px2MM(190),px2MM(768))  
    pdf.set_font('LeagueSpartan-Light', size=px2pts(24))
    pdf.set_text_color(*hex2RGB('#E9EAEE'))
    pdf.set_draw_color(*hex2RGB('#313236'))
    pdf.multi_cell(px2MM(417), px2MM(30),'Providing social security to family',align='L')
    
    pdf.set_fill_color(*hex2RGB(sq_bullet_1))
    pdf.rect(px2MM(160), px2MM(816), px2MM(10), px2MM(10),'F')
    
    pdf.set_xy(px2MM(190),px2MM(805))  
    pdf.set_font('LeagueSpartan-Light', size=px2pts(24))
    pdf.set_text_color(*hex2RGB('#E9EAEE'))
    pdf.set_draw_color(*hex2RGB('#313236'))
    pdf.multi_cell(px2MM(417), px2MM(30),'Giving basic lifestyle to next generation',align='L') 

    #//*----Examples of Priorities---*/
    pdf.set_xy(px2MM(160),px2MM(877))  
    pdf.set_font('LeagueSpartan-Light', size=px2pts(24))
    pdf.set_text_color(*hex2RGB('#898B90'))
    pdf.cell(px2MM(457), px2MM(35),'EXAMPLE OF PRIORITIES')
    pdf.image(join(cwd,'assets','images','genration profile','shade_line.png'),px2MM(160), px2MM(919), px2MM(447), px2MM(1))
    
    #//*--Point 5 to 3---*//
    pdf.set_fill_color(*hex2RGB(sq_bullet_1))
    pdf.rect(px2MM(160), px2MM(956), px2MM(10), px2MM(10),'F')
    
    pdf.set_xy(px2MM(190),px2MM(929))  
    pdf.set_font('LeagueSpartan-Light', size=px2pts(24))
    pdf.set_text_color(*hex2RGB('#E9EAEE'))
    pdf.set_draw_color(*hex2RGB('#313236'))
    pdf.multi_cell(px2MM(417), px2MM(30),'Upgrading existing living facility to one with basic comfort and necessities')
    
    pdf.set_fill_color(*hex2RGB(sq_bullet_1))
    pdf.rect(px2MM(160), px2MM(1009), px2MM(10), px2MM(10),'F')
    
    pdf.set_xy(px2MM(190),px2MM(998))  
    pdf.set_font('LeagueSpartan-Light', size=px2pts(24))
    pdf.set_text_color(*hex2RGB('#E9EAEE'))
    pdf.set_draw_color(*hex2RGB('#313236'))
    pdf.multi_cell(px2MM(417), px2MM(32),'Living a dignified life in society') 
    
    
     #//*-------------Card 2-----------*//
    #//*----For Heading (Genrations)---*//
    pdf.set_xy(px2MM(850),px2MM(244))  
    pdf.set_font('LeagueSpartan-Medium', size=px2pts(40))
    pdf.set_text_color(*hex2RGB('#FFFFFF'))
    pdf.cell(px2MM(215), px2MM(56),'Generation 2')

    
    #//*----Personality Traits---*/
    pdf.set_xy(px2MM(737),px2MM(330))  
    pdf.set_font('LeagueSpartan-Light', size=px2pts(24))
    pdf.set_text_color(*hex2RGB('#898B90'))
    
    # pdf.set_draw_color(*hex2RGB('#313236'))
    pdf.cell(px2MM(447), px2MM(35),'PERSONALITY TRAITS')
    pdf.image(join(cwd,'assets','images','genration profile','shade_line.png'),px2MM(737), px2MM(372), px2MM(447), px2MM(1))
    
        
    #//*--Point 1---*//
    
    pdf.set_fill_color(*hex2RGB(sq_bullet_2))
    pdf.rect(px2MM(737), px2MM(393), px2MM(10), px2MM(10),'F')
    
    pdf.set_xy(px2MM(767),px2MM(380))  
    pdf.set_font('LeagueSpartan-Light', size=px2pts(24))
    pdf.set_text_color(*hex2RGB('#E9EAEE'))
    pdf.set_draw_color(*hex2RGB('#313236'))
    pdf.multi_cell(px2MM(417), px2MM(30),'Well-educated and skilled professional',align='L')
    
    pdf.set_fill_color(*hex2RGB(sq_bullet_2))
    pdf.rect(px2MM(737), px2MM(430), px2MM(10), px2MM(10),'F')
    
    pdf.set_xy(px2MM(767),px2MM(419))  
    pdf.set_font('LeagueSpartan-Light', size=px2pts(24))
    pdf.set_text_color(*hex2RGB('#E9EAEE'))
    pdf.set_draw_color(*hex2RGB('#313236'))
    pdf.multi_cell(px2MM(417), px2MM(30),'Focused on improving current lifestyle',align='L') 

    
    #//*----Financial Behaviour---*/
    pdf.set_xy(px2MM(737),px2MM(481))  
    pdf.set_font('LeagueSpartan-Light', size=px2pts(24))
    pdf.set_text_color(*hex2RGB('#898B90'))
    pdf.cell(px2MM(447), px2MM(35),'FINANCIAL BEHAVIOUR')
    pdf.image(join(cwd,'assets','images','genration profile','shade_line.png'),px2MM(737), px2MM(523), px2MM(447), px2MM(1))

    
    #//*--Point 3 to 4---*//
    pdf.set_fill_color(*hex2RGB(sq_bullet_2))
    pdf.rect(px2MM(737), px2MM(544), px2MM(10), px2MM(10),'F')
    
    pdf.set_xy(px2MM(767),px2MM(533))  
    pdf.set_font('LeagueSpartan-Light', size=px2pts(24))
    pdf.set_text_color(*hex2RGB('#E9EAEE'))
    pdf.set_draw_color(*hex2RGB('#313236'))
    pdf.multi_cell(px2MM(425), px2MM(30),'Save mindfully to build a reasonable corpus',align='L')
    
    pdf.set_fill_color(*hex2RGB(sq_bullet_2))
    pdf.rect(px2MM(737), px2MM(613), px2MM(10), px2MM(10),'F')
    
    pdf.set_xy(px2MM(767),px2MM(570))  
    pdf.set_font('LeagueSpartan-Light', size=px2pts(24))
    pdf.set_text_color(*hex2RGB('#E9EAEE'))
    pdf.set_draw_color(*hex2RGB('#313236'))
    pdf.multi_cell(px2MM(417), px2MM(30),'Prefer traditional investment options such as bank deposits, mutual funds, insurance plus investment plans etc.',align='L') 
    
    #//*----ASPIRATIONA---*/
    pdf.set_xy(px2MM(737),px2MM(696))  
    pdf.set_font('LeagueSpartan-Light', size=px2pts(24))
    pdf.set_text_color(*hex2RGB('#898B90'))
    pdf.cell(px2MM(447), px2MM(30),'ASPIRATIONS')
    pdf.image(join(cwd,'assets','images','genration profile','shade_line.png'),px2MM(737), px2MM(738), px2MM(447), px2MM(1))
    
    #//*--Point 5 to 3---*//
    pdf.set_fill_color(*hex2RGB(sq_bullet_2))
    pdf.rect(px2MM(737), px2MM(775), px2MM(10), px2MM(10),'F')
    
    pdf.set_xy(px2MM(767),px2MM(748))  
    pdf.set_font('LeagueSpartan-Light', size=px2pts(24))
    pdf.set_text_color(*hex2RGB('#E9EAEE'))
    pdf.set_draw_color(*hex2RGB('#313236'))
    pdf.multi_cell(px2MM(417), px2MM(30),'Providing a good lifestyle and education for future generations',align='L')
    
    pdf.set_fill_color(*hex2RGB(sq_bullet_2))
    pdf.rect(px2MM(737), px2MM(844), px2MM(10), px2MM(10),'F')
    
    pdf.set_xy(px2MM(767),px2MM(817))  
    pdf.set_font('LeagueSpartan-Light', size=px2pts(24))
    pdf.set_text_color(*hex2RGB('#E9EAEE'))
    pdf.set_draw_color(*hex2RGB('#313236'))
    pdf.multi_cell(px2MM(417), px2MM(30),'Achieving financial freedom to have more control over time',align='L') 

    #//*----Examples of Priorities---*/
    pdf.set_xy(px2MM(737),px2MM(911)) 
    pdf.set_font('LeagueSpartan-Light', size=px2pts(24))
    pdf.set_text_color(*hex2RGB('#898B90'))
    pdf.cell(px2MM(457), px2MM(35),'EXAMPLE OF PRIORITIES')
    pdf.image(join(cwd,'assets','images','genration profile','shade_line.png'),px2MM(737), px2MM(953), px2MM(447), px2MM(1))
    
    #//*--Point 5 to 3---*//
    pdf.set_fill_color(*hex2RGB(sq_bullet_2))
    pdf.rect(px2MM(737), px2MM(974), px2MM(10), px2MM(10),'F')
    
    pdf.set_xy(px2MM(767),px2MM(963))  
    pdf.set_font('LeagueSpartan-Light', size=px2pts(24))
    pdf.set_text_color(*hex2RGB('#E9EAEE'))
    pdf.set_draw_color(*hex2RGB('#313236'))
    pdf.multi_cell(px2MM(417), px2MM(30),'Creating secondary source of income')
    
    pdf.set_fill_color(*hex2RGB(sq_bullet_2))
    pdf.rect(px2MM(737), px2MM(1027), px2MM(10), px2MM(10),'F')
    
    pdf.set_xy(px2MM(767),px2MM(1000))  
    pdf.set_font('LeagueSpartan-Light', size=px2pts(24))
    pdf.set_text_color(*hex2RGB('#E9EAEE'))
    pdf.set_draw_color(*hex2RGB('#313236'))
    pdf.multi_cell(px2MM(417), px2MM(32),'Buying a quality car and a home with good amenities')
    
    #//*----------Card 3-------*//
    #//*----For Heading (Genrations)---*//
    pdf.set_xy(px2MM(1426),px2MM(244))  
    pdf.set_font('LeagueSpartan-Medium', size=px2pts(40))
    pdf.set_text_color(*hex2RGB('#FFFFFF'))
    pdf.cell(px2MM(215), px2MM(56),'Generation 3')

    
    #//*----Personality Traits---*/
    pdf.set_xy(px2MM(1313),px2MM(330))  
    pdf.set_font('LeagueSpartan-Light', size=px2pts(24))
    pdf.set_text_color(*hex2RGB('#898B90'))
    
    # pdf.set_draw_color(*hex2RGB('#313236'))
    pdf.cell(px2MM(447), px2MM(35),'PERSONALITY TRAITS')
    pdf.image(join(cwd,'assets','images','genration profile','shade_line.png'),px2MM(1313), px2MM(372), px2MM(447), px2MM(1))
    
        
    #//*--Point 1---*//
    
    pdf.set_fill_color(*hex2RGB(sq_bullet_3))
    pdf.rect(px2MM(1313), px2MM(409), px2MM(10), px2MM(10),'F')
    
    pdf.set_xy(px2MM(1343),px2MM(380))  
    pdf.set_font('LeagueSpartan-Light', size=px2pts(24))
    pdf.set_text_color(*hex2RGB('#E9EAEE'))
    pdf.set_draw_color(*hex2RGB('#313236'))
    pdf.multi_cell(px2MM(417), px2MM(30),'Early adopter of new trends and global products',align='L')
    
    pdf.set_fill_color(*hex2RGB(sq_bullet_3))
    pdf.rect(px2MM(1313), px2MM(478), px2MM(10), px2MM(10),'F')
    
    pdf.set_xy(px2MM(1343),px2MM(451))  
    pdf.set_font('LeagueSpartan-Light', size=px2pts(24))
    pdf.set_text_color(*hex2RGB('#E9EAEE'))
    pdf.set_draw_color(*hex2RGB('#313236'))
    pdf.multi_cell(px2MM(417), px2MM(30),'Willing to take high risks in pursuit of potential rewards',align='L') 

    
    #//*----Financial Behaviour---*/
    pdf.set_xy(px2MM(1313),px2MM(545))  
    pdf.set_font('LeagueSpartan-Light', size=px2pts(24))
    pdf.set_text_color(*hex2RGB('#898B90'))
    pdf.cell(px2MM(447), px2MM(35),'FINANCIAL BEHAVIOUR')
    pdf.image(join(cwd,'assets','images','genration profile','shade_line.png'),px2MM(1313), px2MM(587), px2MM(447), px2MM(1))

    
    #//*--Point 3 to 4---*//
    pdf.set_fill_color(*hex2RGB(sq_bullet_3))
    pdf.rect(px2MM(1313), px2MM(608), px2MM(10), px2MM(10),'F')
    
    pdf.set_xy(px2MM(1343),px2MM(597))  
    pdf.set_font('LeagueSpartan-Light', size=px2pts(24))
    pdf.set_text_color(*hex2RGB('#E9EAEE'))
    pdf.set_draw_color(*hex2RGB('#313236'))
    pdf.multi_cell(px2MM(425), px2MM(30),'Focused on building wealth',align='L')
    
    pdf.set_fill_color(*hex2RGB(sq_bullet_3))
    pdf.rect(px2MM(1313), px2MM(645), px2MM(10), px2MM(10),'F')
    
    pdf.set_xy(px2MM(1343),px2MM(634))  
    pdf.set_font('LeagueSpartan-Light', size=px2pts(24))
    pdf.set_text_color(*hex2RGB('#E9EAEE'))
    pdf.set_draw_color(*hex2RGB('#313236'))
    pdf.multi_cell(px2MM(417), px2MM(30),'Like experimenting with risky asset classes',align='L') 
    
    #//*----ASPIRATIONA---*/
    pdf.set_xy(px2MM(1313),px2MM(696))  
    pdf.set_font('LeagueSpartan-Light', size=px2pts(24))
    pdf.set_text_color(*hex2RGB('#898B90'))
    pdf.cell(px2MM(447), px2MM(30),'ASPIRATIONS')
    pdf.image(join(cwd,'assets','images','genration profile','shade_line.png'),px2MM(1313), px2MM(738), px2MM(447), px2MM(1))
    
    #//*--Point 5 to 3---*//
    pdf.set_fill_color(*hex2RGB(sq_bullet_3))
    pdf.rect(px2MM(1313), px2MM(759), px2MM(10), px2MM(10),'F')
    
    pdf.set_xy(px2MM(1343),px2MM(748))  
    pdf.set_font('LeagueSpartan-Light', size=px2pts(24))
    pdf.set_text_color(*hex2RGB('#E9EAEE'))
    pdf.set_draw_color(*hex2RGB('#313236'))
    pdf.multi_cell(px2MM(417), px2MM(30),'Interested in luxury purchases',align='L')
    
    pdf.set_fill_color(*hex2RGB(sq_bullet_3))
    pdf.rect(px2MM(1313), px2MM(812), px2MM(10), px2MM(10),'F')
    
    pdf.set_xy(px2MM(1343),px2MM(785))  
    pdf.set_font('LeagueSpartan-Light', size=px2pts(24))
    pdf.set_text_color(*hex2RGB('#E9EAEE'))
    pdf.set_draw_color(*hex2RGB('#313236'))
    pdf.multi_cell(px2MM(417), px2MM(30),'Driven to start new businesses and pursue hobbies as a profession',align='L') 

    #//*----Examples of Priorities---*/
    pdf.set_xy(px2MM(1313),px2MM(879)) 
    pdf.set_font('LeagueSpartan-Light', size=px2pts(24))
    pdf.set_text_color(*hex2RGB('#898B90'))
    pdf.cell(px2MM(457), px2MM(35),'EXAMPLE OF PRIORITIES')
    pdf.image(join(cwd,'assets','images','genration profile','shade_line.png'),px2MM(1313), px2MM(921), px2MM(447), px2MM(1))
    
    #//*--Point 5 to 3---*//
    pdf.set_fill_color(*hex2RGB(sq_bullet_3))
    pdf.rect(px2MM(1313), px2MM(958), px2MM(10), px2MM(10),'F')
    
    pdf.set_xy(px2MM(1343),px2MM(931))  
    pdf.set_font('LeagueSpartan-Light', size=px2pts(24))
    pdf.set_text_color(*hex2RGB('#E9EAEE'))
    pdf.set_draw_color(*hex2RGB('#313236'))
    pdf.multi_cell(px2MM(417), px2MM(30),'Growing investment portfolio by investing in alternative assets')
    
    pdf.set_fill_color(*hex2RGB(sq_bullet_3))
    pdf.rect(px2MM(1313), px2MM(1027), px2MM(10), px2MM(10),'F')
    
    pdf.set_xy(px2MM(1343),px2MM(1000))  
    pdf.set_font('LeagueSpartan-Light', size=px2pts(24))
    pdf.set_text_color(*hex2RGB('#E9EAEE'))
    pdf.set_draw_color(*hex2RGB('#313236'))
    pdf.multi_cell(px2MM(417), px2MM(32),'Staying informed and educated about financial trends and new products')
    
        
    #//**----For Your Profile box---*//
    if df=='Generation 1':
        pdf.set_fill_color(*hex2RGB(your_profile_color))
        pdf.rect(px2MM(120), px2MM(204), px2MM(117), px2MM(35),'F')
        pdf.set_xy(px2MM(132),px2MM(209))  
        pdf.set_font('LeagueSpartan-Medium', size=px2pts(18))
        pdf.set_text_color(*hex2RGB('#000000'))
        pdf.cell(px2MM(87), px2MM(25.2),'Your Profile') 
    
    elif df=='Generation 2':
        pdf.set_fill_color(*hex2RGB(your_profile_color))
        pdf.rect(px2MM(697), px2MM(204), px2MM(117), px2MM(35),'F')
        pdf.set_xy(px2MM(710),px2MM(209))  
        pdf.set_font('LeagueSpartan-Medium', size=px2pts(18))
        pdf.set_text_color(*hex2RGB('#000000'))
        pdf.cell(px2MM(87), px2MM(25),'Your Profile')
        
    elif df=='Generation 3':
        pdf.set_fill_color(*hex2RGB(your_profile_color))
        pdf.rect(px2MM(1273), px2MM(204), px2MM(117), px2MM(35),'F')
        pdf.set_xy(px2MM(1288),px2MM(209))  
        pdf.set_font('LeagueSpartan-Medium', size=px2pts(18))
        pdf.set_text_color(*hex2RGB('#000000'))
        pdf.cell(px2MM(87), px2MM(25.2),'Your Profile')  
    else:
        pass

    #//*-----Index Text of Page--**////
    index_text(pdf,'#FFFFFF')
#//*----Net Worth------*//

def net_worth(pdf,json_data,c_MoneyS,money_signData):
    try:
        df = json_data["networth"]
        # df2 = pd.DataFrame.from_dict(json_data['val_un_adv'])
    except:
        return None
    
   
    #//*---Page setup----*//
    pdf.add_page()
    pdf.set_fill_color(*hex2RGB('#FCF8ED'))
    pdf.rect(0, 0, px2MM(1920), px2MM(1080),'F') 
    
    #//*--Purple vertical line
    # pdf.set_xy(px2MM(125),px2MM(78))
    pdf.set_fill_color(*hex2RGB('#000000'))
    pdf.rect(px2MM(0), px2MM(80), px2MM(15), px2MM(84), 'F') 
    
    #//*---heading statement
    pdf.set_xy(px2MM(120),px2MM(80))  
    pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(60))
    pdf.set_text_color(*hex2RGB('#1A1A1D'))
    pdf.cell(px2MM(600), px2MM(84),'Net worth')
    
    #//*---What is Net worth
    pdf.set_xy(px2MM(400),px2MM(244))  
    pdf.set_font('LeagueSpartan-Regular', size=px2pts(40))
    # pdf.set_text_color(*hex2RGB('#1A1A1D'))
    text2 = '''Your net worth is simply the difference between'''
    pdf.cell(px2MM(790), px2MM(56),text2,align='C',markdown=True)
    
    pdf.set_xy(px2MM(1190),px2MM(244))  
    pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(40))
    # pdf.set_text_color(*hex2RGB('#1A1A1D'))
    text2 = '''what you own'''
    pdf.cell(px2MM(250), px2MM(56),text2,align='C',markdown=True)
    
    pdf.set_xy(px2MM(1440),px2MM(244))  
    pdf.set_font('LeagueSpartan-Regular',size=px2pts(40))
    # pdf.set_text_color(*hex2RGB('#1A1A1D'))
    text2 = '''(like'''
    pdf.cell(px2MM(70), px2MM(56),text2,align='C',markdown=True)
    
    pdf.set_xy(px2MM(390),px2MM(300))  
    pdf.set_font('LeagueSpartan-Regular', size=px2pts(40))
    # pdf.set_text_color(*hex2RGB('#1A1A1D'))
    text2 = '''your house, retirement funds, etc) and'''
    pdf.cell(px2MM(640), px2MM(56),text2,align='L',markdown=True)
    
    pdf.set_xy(px2MM(1030),px2MM(300))  
    pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(40))
    # pdf.set_text_color(*hex2RGB('#1A1A1D'))
    text2 = '''what you owe'''
    pdf.cell(px2MM(240), px2MM(56),text2,align='L',markdown=True)
    
    pdf.set_xy(px2MM(1275),px2MM(300))  
    pdf.set_font('LeagueSpartan-Regular', size=px2pts(40))
    # pdf.set_text_color(*hex2RGB('#1A1A1D'))
    text2 = '''(your liabilities'''
    pdf.cell(px2MM(250), px2MM(56),text2,align='L',markdown=True)
    
    pdf.set_xy(px2MM(390),px2MM(356))  
    pdf.set_font('LeagueSpartan-Regular', size=px2pts(40))
    # pdf.set_text_color(*hex2RGB('#1A1A1D'))
    text2 = '''such as mortgage, credit card debt and so forth).'''
    pdf.cell(px2MM(1130), px2MM(56),text2,align='C',markdown=True)
    
    #//*---Rect---*//
    
    #//*--White rect dynamic x
    white_rect = ('Total Assets','Total Liabilities','Networth')
    white_rect_x = 140
    white_rect_x_gap = 560
    white_rect_text_x = 180
    white_rect_text_x_gap = 560
    
    #//*--Color rect dynamic x
    color_rect = ('#7C5FF2','#FFCA41','#4DC3A7')
    tot_assets = '₹ ' +str(format_cash2(float(df['assets'])))
    tot_liab = '₹ ' +str(format_cash2(float(df['liabilities'])))
    tot_networth = '₹ ' +str(format_cash2(float(df['networth'])))
    
    #//*---Total networth = Total_Assets - Total_Liabilities
    color_rect_val = (tot_assets,tot_liab,tot_networth)
    color_rect_x = 235
    color_rect_x_gap = 560
    color_rect_text_x = 275
    color_rect_text_x_gap = 560
    
    pdf.set_draw_color(*hex2RGB('#E9EAEE'))
    pdf.set_line_width(px2MM(0.2))
    for i in range(3):
        #//*---White rectangle with text
        pdf.set_fill_color(*hex2RGB('#FFFFFF'))
        pdf.rect(px2MM(white_rect_x),px2MM(492),px2MM(520),px2MM(173),'FD')
        
        pdf.set_xy(px2MM(white_rect_text_x),px2MM(532))  
        pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(40))
        pdf.set_text_color(*hex2RGB('#000000'))
        pdf.cell(px2MM(440), px2MM(56),white_rect[i],border=0,align='C')
        white_rect_x+=white_rect_x_gap
        white_rect_text_x+=white_rect_text_x_gap
        
        #//*---Color Rect with text---*//
        pdf.set_fill_color(*hex2RGB(color_rect[i]))
        pdf.rect(px2MM(color_rect_x),px2MM(618),px2MM(330),px2MM(158),'F')
        
        pdf.set_xy(px2MM(color_rect_text_x),px2MM(658))  
        pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(56))
        pdf.set_text_color(*hex2RGB('#FFFFFF'))
        pdf.cell(px2MM(250), px2MM(78),color_rect_val[i],border=0,align='C')
        color_rect_x+=color_rect_x_gap
        color_rect_text_x+=color_rect_text_x_gap
      
    #//*---For circle operator symbol 
    
    white_circle1_x = 639                              
    common_gap = 564 
                              
    color_circle_x = 653                                
    
    opt_x = 667.33  
    opt_val = ('-','=')
    opt_height=(3.33,13.33) 
    
                           
    for i in range(2):
        #//*---white outer circle---*//
        pdf.set_fill_color(*hex2RGB('#FFFFFF'))
        pdf.circle(x=px2MM(white_circle1_x),y=px2MM(539),r=px2MM(80),style='F')
        
        #//*---Color Inner circle---*//
        pdf.set_fill_color(*hex2RGB('#FCF8ED'))
        pdf.circle(x=px2MM(color_circle_x),y=px2MM(553),r=px2MM(52),style='F')
        
        #//*---For operator
        pdf.set_xy(px2MM(opt_x),px2MM(572.33))  
        pdf.set_font('LeagueSpartan-Light', size=px2pts(70))
        pdf.set_text_color(*hex2RGB('#000000'))
        pdf.cell(px2MM(23.33), px2MM(opt_height[i]),opt_val[i],border=0,align='C')
        
        white_circle1_x+=common_gap
        color_circle_x+=common_gap
        opt_x+=common_gap
        
    #//*----For Value under Adivisoary---*//
    
        pdf.set_xy(px2MM(705),px2MM(892))  
        pdf.set_font('LeagueSpartan-Regular', size=px2pts(40))
        pdf.set_text_color(*hex2RGB('#000000'))
        pdf.cell(px2MM(374), px2MM(56),'Value Under Advisory:',border=0,align='L')
        
        pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(40))
        pdf.set_xy(px2MM(1090),px2MM(892)) 
        val_ud_adv = '₹ '+str(format_cash2(float(df['value_under_advisory'])))
        pdf.cell(px2MM(374), px2MM(56),val_ud_adv,border=0,align='L')
        
        pdf.set_xy(px2MM(250),px2MM(968)) 
        pdf.set_text_color(*hex2RGB('#1A1A1D'))
        pdf.set_font('LeagueSpartan-Light', size=px2pts(24))
        pdf.cell(px2MM(1419), px2MM(32),'This includes total of your assets and liabilities.',border=0,align='C')
        
        #//*-----Index Text of Page--**////
        index_text(pdf,'#1A1A1D')
        
#//*----Expense and Liability Management------*//

def exp_lib_mang(pdf,json_data,c_MoneyS,money_signData):
    try:
        exp_lib_mang = json_data["ratios"]
        if exp_lib_mang=={}:
           return None 
    except:
        return None
    
    exp_lib_mang_keys = list(exp_lib_mang.keys())
   
    #//*---Page setup----*//
    pdf.add_page()
    pdf.set_fill_color(*hex2RGB('#FCF8ED'))
    pdf.rect(0, 0, px2MM(1920), px2MM(1080),'F') 
    
    pdf.image(join(cwd,'assets','images','backgrounds','doubleLine.png'),px2MM(1449),px2MM(0),px2MM(471),px2MM(1080))
    
    #//*--Purple vertical line
    # pdf.set_xy(px2MM(125),px2MM(78))
    pdf.set_fill_color(*hex2RGB('#000000'))
    pdf.rect(px2MM(0), px2MM(80), px2MM(15), px2MM(84), 'F')
    
    #//*----Purple Rectange of Heading Expense and Liability Management
    pdf.set_fill_color(*hex2RGB('#000000'))
    pdf.rect(px2MM(1310), px2MM(81), px2MM(490), px2MM(82),'F')
    
    pdf.set_xy(px2MM(1330),px2MM(101)) 
    pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(30))
    pdf.set_text_color(*hex2RGB('#ffffff'))
    pdf.cell(px2MM(450), px2MM(42),'Expense and Liability Management')
     
    #//*---heading statement
    pdf.set_xy(px2MM(120),px2MM(80))  
    pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(60))
    pdf.set_text_color(*hex2RGB('#1A1A1D'))
    pdf.cell(px2MM(589), px2MM(84),'Your Financial Analysis')
    

    # all_statements = df['Comments']
    #//*----6 Boxes--*//
    main_box_x =main_box_x1 = 120
    heading_label_x = heading_label_x1 = 160
    x_common_gap = x1_common_gap = 577
    score_box_x = score_box_x1 = 160
    score_x = score_x1 = 170
    ideal_range_x = ideal_range_x1 = 403
    all_stat_x = all_stat_x1 = 160
    
    
    # ideal_min = df["Ideal Range"]
    pdf.set_draw_color(*hex2RGB('#E9EAEE'))
    pdf.set_line_width(px2MM(0.2))
    for i in range(3):
        
        #//*---vor horizontol Boxes Row 1
        
        pdf.set_fill_color(*hex2RGB('#ffffff'))
        pdf.rect(px2MM(main_box_x), px2MM(204), px2MM(527), px2MM(362),'FD')
        
        #//*----Box Headings----*//
        pdf.set_xy(px2MM(heading_label_x),px2MM(244)) 
        pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(28))
        pdf.set_text_color(*hex2RGB('#2E3034'))
        pdf.cell(px2MM(447), px2MM(39),exp_lib_mang[exp_lib_mang_keys[i]]['title'],align='L')
        
        #//*----Color Score Box---*//
        
        if exp_lib_mang[exp_lib_mang_keys[i]]['color']=='green':
            pdf.set_fill_color(*hex2RGB('#71EBB8'))
        else:
            pdf.set_fill_color(*hex2RGB('#FF937B'))
            
        pdf.rect(px2MM(score_box_x), px2MM(313), px2MM(102), px2MM(52),'F')
        
        pdf.set_xy(px2MM(score_x),px2MM(318)) 
        pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(30))
        pdf.set_text_color(*hex2RGB('#1A1A1D'))
        pdf.cell(px2MM(82), px2MM(42),str(round(float(exp_lib_mang[exp_lib_mang_keys[i]]['total'])*100))+'%',align='C')
        
        #//*-----Ideal Ranges
        pdf.set_xy(px2MM(ideal_range_x),px2MM(323)) 
        pdf.set_font('LeagueSpartan-Regular', size=px2pts(24))
        pdf.set_text_color(*hex2RGB('#313236'))
        if exp_lib_mang[exp_lib_mang_keys[i]]['ideal_range'] =="":
            pdf.cell(px2MM(204), px2MM(32),'',align='C')
        else:
            # val = exp_lib_mang[exp_lib_mang_keys[i]]['ideal_range'].split('-')
            val = exp_lib_mang[exp_lib_mang_keys[i]]['ideal_range']
            # val = " - ".join(list(str(format_cash2(float(x))) for x in val))
            pdf.cell(px2MM(204), px2MM(32),'Ideal: '+val,align='C')
        
        #//*----Statements----*//
        pdf.set_xy(px2MM(all_stat_x),px2MM(395)) 
        pdf.set_font('LeagueSpartan-Light', size=px2pts(24))
        pdf.set_text_color(*hex2RGB('#000000'))
        pdf.multi_cell(px2MM(447), px2MM(32),exp_lib_mang[exp_lib_mang_keys[i]]['comment'],align='L')
        
        main_box_x+= x_common_gap
        score_box_x+=x_common_gap
        score_x+=x_common_gap
        ideal_range_x+=x_common_gap
        heading_label_x+=x_common_gap
        all_stat_x+=x_common_gap
        
        
    #//*----Lower 3 boxes----*//    
    main_box_x1 = 120
    heading_label_x1 = 160
    x1_common_gap = 577
    score_box_x1 = 160
    score_x1 = 170
    ideal_range_x1 = 403
    all_stat_x1 = 160
    
    pdf.set_draw_color(*hex2RGB('#E9EAEE'))
    pdf.set_line_width(px2MM(0.2))    
    for i in range(3,6):
        
        #//*---vor horizontol Boxes Row 2
        
        pdf.set_fill_color(*hex2RGB('#ffffff'))
        pdf.rect(px2MM(main_box_x1), px2MM(616), px2MM(527), px2MM(362),'FD')
        
        #//*----Box Headings----*//
        pdf.set_xy(px2MM(heading_label_x1),px2MM(656)) 
        pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(28))
        pdf.set_text_color(*hex2RGB('#2E3034'))
        pdf.cell(px2MM(447), px2MM(39),exp_lib_mang[exp_lib_mang_keys[i]]['title'],align='L')
        
        #//*----Color Score Box---*//
        
        if exp_lib_mang[exp_lib_mang_keys[i]]['color']=='green':
            pdf.set_fill_color(*hex2RGB('#71EBB8'))
        else:
            pdf.set_fill_color(*hex2RGB('#FF937B'))
            
        pdf.rect(px2MM(score_box_x1), px2MM(715), px2MM(102), px2MM(52),'F')
        
        pdf.set_xy(px2MM(score_x1),px2MM(720)) 
        pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(30))
        pdf.set_text_color(*hex2RGB('#1A1A1D'))
        pdf.cell(px2MM(82), px2MM(42),str(round(float(exp_lib_mang[exp_lib_mang_keys[i]]['total'])*100))+'%',align='C')
        
        #//*-----Ideal Ranges
        pdf.set_xy(px2MM(ideal_range_x1),px2MM(735)) 
        pdf.set_font('LeagueSpartan-Regular', size=px2pts(24))
        pdf.set_text_color(*hex2RGB('#313236'))
        if exp_lib_mang[exp_lib_mang_keys[i]]['ideal_range'] =="":
            pdf.cell(px2MM(204), px2MM(32),'',align='C')
        else:
            val = exp_lib_mang[exp_lib_mang_keys[i]]['ideal_range']
            # val = exp_lib_mang[exp_lib_mang_keys[i]]['ideal_range'].split('-')
            # val = " - ".join(list(str(format_cash2(float(x))) for x in val))
            pdf.cell(px2MM(204), px2MM(32),'Ideal: '+val,align='C')
        
        #//*----Statements----*//
        pdf.set_xy(px2MM(all_stat_x1),px2MM(807)) 
        pdf.set_font('LeagueSpartan-Light', size=px2pts(24))
        pdf.set_text_color(*hex2RGB('#000000'))
        pdf.multi_cell(px2MM(447), px2MM(32),exp_lib_mang[exp_lib_mang_keys[i]]['comment'],align='L')
        
        main_box_x1+= x1_common_gap
        score_box_x1+=x1_common_gap
        score_x1+=x1_common_gap
        ideal_range_x1+=x1_common_gap
        heading_label_x1+=x1_common_gap
        all_stat_x1+=x1_common_gap
        
    pdf.set_xy(px2MM(250),px2MM(1019)) 
    pdf.set_font('LeagueSpartan-Regular', size=px2pts(18))
    pdf.set_text_color(*hex2RGB('#1A1A1D'))
    txt = '''Disclaimer: The red shade denotes a value that falls outside of the suggested range for a given metric, while a green shade indicates a value that falls within that suggested range.'''
    pdf.cell(px2MM(1420), px2MM(21.09),txt,align='C')
    
    #//*-----Index Text of Page--**////
    index_text(pdf,'#1A1A1D')
    global your_fin_analysis_idx
    if your_fin_analysis_idx == 0:
        your_fin_analysis_idx = pdf.page_no()
       
       
#//*----Asset Alocation------*//

def asset_allocation(pdf,json_data,c_MoneyS,money_signData):
    try:
        asset_alloc = json_data['asset_allocation']
        if asset_alloc=={}:
            return None
    except:
        return None
    
    asset_alloc_keys = list(asset_alloc.keys())
    
    #//*---Page setup----*//
    pdf.add_page()
    pdf.set_fill_color(*hex2RGB('#FCF8ED'))
    pdf.rect(0, 0, px2MM(1920), px2MM(1080),'F') 
    
    pdf.image(join(cwd,'assets','images','backgrounds','doubleLine.png'),px2MM(1449),px2MM(0),px2MM(471),px2MM(1080))
    
    #//*--Purple vertical line
    pdf.set_fill_color(*hex2RGB('#000000'))
    pdf.rect(px2MM(0), px2MM(80), px2MM(15), px2MM(84), 'F')
    
    #//*----Black Rectange of Heading Expense and Liability Management
    pdf.set_fill_color(*hex2RGB('#000000'))
    pdf.rect(px2MM(1554), px2MM(81), px2MM(246), px2MM(82),'F')
    
    pdf.set_xy(px2MM(1574),px2MM(101)) 
    pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(30))
    pdf.set_text_color(*hex2RGB('#ffffff'))
    pdf.cell(px2MM(206), px2MM(42),'Asset Allocation')
     
    #//*---heading statement
    pdf.set_xy(px2MM(120),px2MM(80))  
    pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(60))
    pdf.set_text_color(*hex2RGB('#1A1A1D'))
    pdf.cell(px2MM(589), px2MM(84),'Your Financial Analysis')

    # all_statements = df['Comments']
    #//*----6 Boxes--*//
    main_box_x =main_box_x1 = 120
    heading_label_x = heading_label_x1 = 160
    x_common_gap = x1_common_gap = 577
    score_box_x = score_box_x1 = 160
    score_x = score_x1 = 170
    ideal_range_x = ideal_range_x1 = 403
    all_stat_x = all_stat_x1 = 160

    pdf.set_draw_color(*hex2RGB('#E9EAEE'))
    pdf.set_line_width(px2MM(0.2))
    for i in range(3):
        
        #//*---vor horizontol Boxes Row 1
        
        pdf.set_fill_color(*hex2RGB('#ffffff'))
        pdf.rect(px2MM(main_box_x), px2MM(204), px2MM(527), px2MM(362),'FD')
        
        #//*----Box Headings----*//
        pdf.set_xy(px2MM(heading_label_x),px2MM(244)) 
        pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(28))
        pdf.set_text_color(*hex2RGB('#2E3034'))
        pdf.cell(px2MM(447), px2MM(39),asset_alloc[asset_alloc_keys[i]]['title'],align='L')
        
        #//*----Color Score Box---*//

        if asset_alloc[asset_alloc_keys[i]]['color']=='green':
            pdf.set_fill_color(*hex2RGB('#71EBB8'))
        else:
            pdf.set_fill_color(*hex2RGB('#FF937B'))
            
        pdf.rect(px2MM(score_box_x), px2MM(313), px2MM(102), px2MM(52),'F')
        
        pdf.set_xy(px2MM(score_x),px2MM(318)) 
        pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(30))
        pdf.set_text_color(*hex2RGB('#1A1A1D'))
        pdf.cell(px2MM(82), px2MM(42),str(round(float(asset_alloc[asset_alloc_keys[i]]['total'])*100))+'%',align='C')
        
        #//*-----Ideal Ranges
        pdf.set_xy(px2MM(ideal_range_x),px2MM(323)) 
        pdf.set_font('LeagueSpartan-Regular', size=px2pts(24))
        pdf.set_text_color(*hex2RGB('#313236'))
        if asset_alloc[asset_alloc_keys[i]]['ideal_range']=="":
            pdf.cell(px2MM(204), px2MM(32),'',align='C')
        else:
            val = asset_alloc[asset_alloc_keys[i]]['ideal_range']
            pdf.cell(px2MM(204), px2MM(32),'Ideal: '+val,align='C')
        
        #//*----Statements----*//
        pdf.set_xy(px2MM(all_stat_x),px2MM(395)) 
        pdf.set_font('LeagueSpartan-Light', size=px2pts(24))
        pdf.set_text_color(*hex2RGB('#000000'))
        pdf.multi_cell(px2MM(447), px2MM(32),asset_alloc[asset_alloc_keys[i]]['comment'],align='L')
        
        main_box_x+= x_common_gap
        score_box_x+=x_common_gap
        score_x+=x_common_gap
        ideal_range_x+=x_common_gap
        heading_label_x+=x_common_gap
        all_stat_x+=x_common_gap
        
        
    #//*----Lower 3 boxes----*//    
    main_box_x1 = 120
    heading_label_x1 = 160
    x1_common_gap = 577
    score_box_x1 = 160
    score_x1 = 170
    ideal_range_x1 = 403
    all_stat_x1 = 160
    
    pdf.set_draw_color(*hex2RGB('#E9EAEE'))
    pdf.set_line_width(px2MM(0.2))    
    for i in range(3,5):
        
        #//*---vor horizontol Boxes Row 1
        
        pdf.set_fill_color(*hex2RGB('#ffffff'))
        pdf.rect(px2MM(main_box_x1), px2MM(616), px2MM(527), px2MM(362),'FD')
        
        #//*----Box Headings----*//
        pdf.set_xy(px2MM(heading_label_x1),px2MM(656)) 
        pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(28))
        pdf.set_text_color(*hex2RGB('#2E3034'))
        pdf.cell(px2MM(447), px2MM(39),asset_alloc[asset_alloc_keys[i]]['title'],align='L')
        
        #//*----Color Score Box---*//
        # if float(df['Actual'][i])>=ideal_mean[i] and float(df['Actual'][i])<=ideal_max[i]:
        #     pdf.set_fill_color(*hex2RGB('#71EBB8'))
        # else:
        #     pdf.set_fill_color(*hex2RGB('#FF937B'))
        
        if asset_alloc[asset_alloc_keys[i]]['color']=='green':
            pdf.set_fill_color(*hex2RGB('#71EBB8'))
        else:
            pdf.set_fill_color(*hex2RGB('#FF937B'))
            
        pdf.rect(px2MM(score_box_x1), px2MM(715), px2MM(102), px2MM(52),'F')
        
        pdf.set_xy(px2MM(score_x1),px2MM(720)) 
        pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(30))
        pdf.set_text_color(*hex2RGB('#1A1A1D'))
        pdf.cell(px2MM(82), px2MM(42),str(round(float(asset_alloc[asset_alloc_keys[i]]['total'])*100))+'%',align='C')
        
        #//*-----Ideal Ranges
        pdf.set_xy(px2MM(ideal_range_x1),px2MM(735)) 
        pdf.set_font('LeagueSpartan-Regular', size=px2pts(24))
        pdf.set_text_color(*hex2RGB('#313236'))
        if asset_alloc[asset_alloc_keys[i]]['ideal_range']=="":
            pdf.cell(px2MM(204), px2MM(32),'',align='C')
        else:
            val = asset_alloc[asset_alloc_keys[i]]['ideal_range']
            # val = asset_alloc[asset_alloc_keys[i]]['ideal_range'].split('-')
            # val = " - ".join(list(str(format_cash2(float(x))) for x in val))
            pdf.cell(px2MM(204), px2MM(32),'Ideal: '+val,align='C')
        
        #//*----Statements----*//
        pdf.set_xy(px2MM(all_stat_x1),px2MM(807)) 
        pdf.set_font('LeagueSpartan-Light', size=px2pts(24))
        pdf.set_text_color(*hex2RGB('#000000'))
        pdf.multi_cell(px2MM(447), px2MM(32),asset_alloc[asset_alloc_keys[i]]['comment'],align='L')
        
        main_box_x1+= x1_common_gap
        score_box_x1+=x1_common_gap
        score_x1+=x1_common_gap
        ideal_range_x1+=x1_common_gap
        heading_label_x1+=x1_common_gap
        all_stat_x1+=x1_common_gap
        
    pdf.set_xy(px2MM(250),px2MM(1019)) 
    pdf.set_font('LeagueSpartan-Regular', size=px2pts(18))
    pdf.set_text_color(*hex2RGB('#1A1A1D'))
    txt = '''Disclaimer: The red shade denotes a value that falls outside of the suggested range for a given metric, while a green shade indicates a value that falls within that suggested range.'''
    pdf.cell(px2MM(1420), px2MM(21.09),txt,align='C')
    
    #//*-----Index Text of Page--**////
    index_text(pdf,'#1A1A1D')
    global your_fin_analysis_idx
    if your_fin_analysis_idx == 0:
        your_fin_analysis_idx = pdf.page_no()
             
#//*----Emergency Planning------*//

def emergency_planning(pdf,json_data,c_MoneyS,money_signData):
    try:
        emergency = json_data['emergency']
        if emergency=={}:
            return None      
    except:
        return None
    
   
    #//*---Page setup----*//
    pdf.add_page()
    pdf.set_fill_color(*hex2RGB('#FCF8ED'))
    pdf.rect(0, 0, px2MM(1920), px2MM(1080),'F') 
    
    pdf.image(join(cwd,'assets','images','backgrounds','doubleLine.png'),px2MM(1449),px2MM(0),px2MM(471),px2MM(1080))
    
    #//*--Purple vertical line
    # pdf.set_xy(px2MM(125),px2MM(78))
    pdf.set_fill_color(*hex2RGB('#000000'))
    
    pdf.rect(px2MM(0), px2MM(80), px2MM(15), px2MM(84), 'F')
    
    #//*----Black Rectange of Heading Expense and Liability Management
    pdf.set_fill_color(*hex2RGB('#000000'))
    pdf.rect(px2MM(1499), px2MM(81), px2MM(301), px2MM(82),'F')
    
    pdf.set_xy(px2MM(1519),px2MM(101)) 
    pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(30))
    pdf.set_text_color(*hex2RGB('#ffffff'))
    pdf.cell(px2MM(206), px2MM(42),'Emergency Planning')
     
    #//*---heading statement
    pdf.set_xy(px2MM(120),px2MM(80))  
    pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(60))
    pdf.set_text_color(*hex2RGB('#1A1A1D'))
    pdf.cell(px2MM(589), px2MM(84),'Your Financial Analysis')
    
    
    # all_statements = df['Comments']
    #//*----6 Boxes--*//
    main_box_x =main_box_x1 = 120
    heading_label_x = heading_label_x1 = 160
    x_common_gap = x1_common_gap = 577
    score_box_x = score_box_x1 = 160
    score_x = score_x1 = 170
    ideal_range_x = ideal_range_x1 = 403
    all_stat_x = all_stat_x1 = 160
    
    
    # ideal_min = df['Ideal']
    emergency_keys = list(emergency.keys())
    pdf.set_draw_color(*hex2RGB('#E9EAEE'))
    pdf.set_line_width(px2MM(0.2))
    for i in range(len(emergency)):
        #//*---vor horizontol Boxes Row 1
        
        pdf.set_fill_color(*hex2RGB('#ffffff'))
        pdf.rect(px2MM(main_box_x), px2MM(264), px2MM(527), px2MM(362),'FD')
        
        #//*----Box Headings----*//
        pdf.set_xy(px2MM(heading_label_x),px2MM(304)) 
        pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(28))
        pdf.set_text_color(*hex2RGB('#2E3034'))
        pdf.cell(px2MM(447), px2MM(39),emergency[emergency_keys[i]]['title'],align='L')
        
        #//*----Color Score Box---*//


        if emergency[emergency_keys[i]]['color'].lower() =='green':
            pdf.set_fill_color(*hex2RGB('#71EBB8'))
        else:
            pdf.set_fill_color(*hex2RGB('#FF937B'))
            
        pdf.rect(px2MM(score_box_x), px2MM(373), px2MM(119), px2MM(52),'F')
        
        pdf.set_xy(px2MM(score_x),px2MM(378)) 
        pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(30))
        pdf.set_text_color(*hex2RGB('#1A1A1D'))
        # pdf.cell(px2MM(70), px2MM(42),str(int(df["Actual"][i]))+str(df["unit2"][i]),align='C')
        if emergency[emergency_keys[i]]['total'] == "0" or emergency[emergency_keys[i]]['total'] == "0.0":
            pdf.cell(px2MM(90), px2MM(42),'₹ 0.0L',align='C')
        else:
            pdf.cell(px2MM(90), px2MM(42),'₹ '+str(format_cash2(float(emergency[emergency_keys[i]]['total']))),align='C')
        
        #//*-----Ideal Ranges
        pdf.set_xy(px2MM(ideal_range_x),px2MM(383)) 
        pdf.set_font('LeagueSpartan-Regular', size=px2pts(24))
        pdf.set_text_color(*hex2RGB('#313236'))
        # ideal_range = str(ideal_min[i]).split('-')
        # pdf.cell(px2MM(204), px2MM(32),'Ideal: '+str(ideal_min[i])+str(df['unit1'][i]),align='C')
        if emergency[emergency_keys[i]]['ideal_range']=="":
            pdf.cell(px2MM(204), px2MM(32),'',align='C')
        else:
            val = emergency[emergency_keys[i]]['ideal_range'].split('-')
            val = " - ".join(list('₹ '+str(format_cash2(float(x))) for x in val))
            pdf.cell(px2MM(204), px2MM(32),'Ideal: '+val,align='C')
        
        #//*----Statements----*//
        pdf.set_xy(px2MM(all_stat_x),px2MM(455)) 
        pdf.set_font('LeagueSpartan-Light', size=px2pts(24))
        pdf.set_text_color(*hex2RGB('#000000'))
        pdf.multi_cell(px2MM(447), px2MM(32),emergency[emergency_keys[i]]['comment'],align='L')
        
        main_box_x+= x_common_gap
        score_box_x+=x_common_gap
        score_x+=x_common_gap
        ideal_range_x+=x_common_gap
        heading_label_x+=x_common_gap
        all_stat_x+=x_common_gap
        
        
    #//*----Lower 3 boxes----*//    
    main_box_x1 = 120
    heading_label_x1 = 160
    x1_common_gap = 577
    score_box_x1 = 160
    score_x1 = 170
    ideal_range_x1 = 403
    all_stat_x1 = 160
    
    pdf.set_xy(px2MM(250),px2MM(1019)) 
    pdf.set_font('LeagueSpartan-Regular', size=px2pts(18))
    pdf.set_text_color(*hex2RGB('#1A1A1D'))
    txt = '''Disclaimer: The red shade denotes a value that falls outside of the suggested range for a given metric, while a green shade indicates a value that falls within that suggested range.'''
    pdf.cell(px2MM(1420), px2MM(21.09),txt,align='C')
    
    #//*-----Index Text of Page--**////
    index_text(pdf,'#1A1A1D')
    global your_fin_analysis_idx
    your_fin_analysis_idx = pdf.page_no()
    
#//*-------Assets(pIEcHART)-----*//    
def assets_chart(pdf,json_data,c_MoneyS,money_signData):
    try:
        # df = pd.DataFrame.from_dict(json_data["Snapshot of Holding - Asset"])
        df_table = pd.DataFrame.from_dict(json_data["assets"]['table'])
        df_pie = pd.DataFrame.from_dict(json_data["assets"]['pie'])
    except:
        return None
    
    if df_pie.empty:
        return None
        
    flag = False
    

    for i in range(len(df_pie['percentage'])):
        if float(df_pie['percentage'].iloc[i]) > 0:
            flag = True
            
    if flag == False:
        return None
    
    start = 0
    stop = 8
    
    for pg in range(0,len(df_table),9):
        
        pdf.add_page()
        pdf.set_fill_color(*hex2RGB('#FCF8ED'))
        pdf.rect(0, 0, px2MM(1920), px2MM(1080),'F')
        
        pdf.set_fill_color(*hex2RGB('#000000'))
        pdf.rect(0, 0, px2MM(964), px2MM(1080),'F')
        
        
        if pg ==0:
            #//*----Donut Pie Chart---*//
            font_path = join(cwd,'assets','fonts','Prata')
            font_files = font_manager.findSystemFonts(fontpaths=font_path)
            for font_file in font_files:
                font_manager.fontManager.addfont(font_file)

        # set font

            labels = df_pie["particular"]
            sizes = df_pie['percentage']
            # sizes = list(float(x) for x in sizes)
            
            aut_size = list(str(x) for x in sizes)
            h_size = df_pie['percentage'].tolist()

            free_colors = ['#A792FF','#82DBC6','#90BEF8','#FFC27E','#FFD976']
            w = 1.08
            wed_height = []
            for i in h_size:
                if h_size == 0:
                    wed_height.append(1)
                elif w == 1:
                    w = 1.08
                    wed_height.append(w)
                elif w ==1.08:
                    w = 1
                    wed_height.append(w)
                else :
                    w = 1
                    wed_height.append(w)

            colors = free_colors[0:len(df_pie)]
            df_pie['colors'] = colors
            fig, ax0 = plt.subplots(figsize=(6.8, 6.8))
            font = {'family': 'prata','color':  'black','weight': 'normal','size': 24,}
            wedges, plt_labels, junk = ax0.pie(sizes, colors = colors,startangle=90,wedgeprops = {"edgecolor" : "black",'linewidth': 2.5,'antialiased': True,'width':1},autopct=autopct_generator(9),textprops=font)
            # wed_height = [1.08,1,1.08,1,1.08,1,1.08,1]
            plt.rcParams['font.family'] = 'prata'
            
            for i in range(len(wedges)):
                wedges[i].set_radius(wed_height[i])
            
            # plt.pie(sizes, colors = colors, autopct='%1.0f%%', startangle=90,pctdistance=0.6,textprops={'size': '14'})

            centre_circle = plt.Circle((0,0),0.2,color='black')
            fig = plt.gcf()
            fig.patch.set_facecolor('black')
            fig.gca().add_artist(centre_circle)
            
        #//*---------------------------------**----------------------*//
        # plt.show()
            plt.tight_layout()
            plt.savefig('asset_chart.png',dpi=450)
            
        #//*---Adding Pie Chart to Page---*//
        pdf.image('asset_chart.png',px2MM(80), px2MM(195), px2MM(600), px2MM(600))
            
        #//*----Legends---*//
        
        circle_y = 794
        common_gap = 42
        text_y = 788
 
        for i in range(0,len(df_pie)):
                
            pdf.set_fill_color(*hex2RGB(df_pie['colors'].iloc[i]))
            pdf.circle(x=px2MM(227),y=px2MM(circle_y),r=px2MM(20),style='F')
            
            pdf.set_xy(px2MM(267),px2MM(text_y)) 
            pdf.set_font('LeagueSpartan-Regular', size=px2pts(24))
            pdf.set_text_color(*hex2RGB('#ffffff'))
            pdf.multi_cell(px2MM(250), px2MM(32),str(df_pie["particular"].iloc[i])+':',align='L')
            
            pdf.set_xy(px2MM(517),px2MM(text_y))
            pdf.cell(px2MM(80), px2MM(32),"{:.0f}".format(int(round(float(df_pie['percentage'].iloc[i]))))+'%',align='R')
            
            #//*---Adding double gap to the next value if the current test exceeds the width
            if len(df_pie["particular"].iloc[i])>24:
                circle_y+=common_gap
                text_y+=common_gap
                
            circle_y+=common_gap
            text_y+=common_gap
        
        #//*----Assets----*//
        pdf.set_xy(px2MM(120),px2MM(80)) 
        pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(60))
        pdf.set_text_color(*hex2RGB('#ffffff'))
        pdf.multi_cell(px2MM(600), px2MM(84),'Assets',align='L')
        
        #//*---Assets Date----*//

        Day=dt.datetime.now().strftime("%d")
        month=dt.datetime.now().strftime("%b")
        year=dt.datetime.now().strftime("%Y")

        if 4 <= int(Day) <= 20 or 24 <= int(Day) <= 30:
            suffix = "th"
        else:
            suffix = ["st", "nd", "rd"][int(Day) % 10 - 1]
            
        pdf.set_xy(px2MM(314),px2MM(106)) 
        pdf.set_font('LeagueSpartan-Regular', size=px2pts(24))
        pdf.set_text_color(*hex2RGB('#ffffff'))
        pdf.cell(px2MM(100), px2MM(32),f'As on {str(Day)}',align='R')


        x_after_day = mm2PX(pdf.get_x())-5

        pdf.set_font('LeagueSpartan-Light', size=px2pts(18))
        pdf.set_xy(px2MM(x_after_day),px2MM(106))
        pdf.cell(px2MM(22), px2MM(20),suffix,align='L')
        
        x_after_day = mm2PX(pdf.get_x())
        pdf.set_xy(px2MM(x_after_day),px2MM(106)) 
        pdf.set_font('LeagueSpartan-Regular', size=px2pts(24))
        pdf.set_text_color(*hex2RGB('#ffffff'))
        pdf.cell(px2MM(300), px2MM(32),f'{str(month)} {str(year)}',align='L')

        #//*---Existing Assets ----*//
        pdf.set_fill_color(*hex2RGB('#000000'))
        pdf.rect(px2MM(1424), px2MM(81), px2MM(376), px2MM(82),'F')
        
        pdf.set_xy(px2MM(1444),px2MM(101)) 
        pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(30))
        pdf.set_text_color(*hex2RGB('#ffffff'))
        ext_value = '₹ '+str(format_cash2(float(json_data['assets']['total']['market_value'])))
        pdf.cell(px2MM(336), px2MM(42),f'Existing Assets: {ext_value}',align='C')
        
        #//*-----Assets Table---*//
        #//*----Col1 Assets
        pdf.set_fill_color(*hex2RGB('#ffffff'))
        pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(24))
        pdf.set_draw_color(*hex2RGB('#E9EAEE'))
        pdf.set_line_width(px2MM(0.2))
        pdf.rect(px2MM(690), px2MM(317), px2MM(297), px2MM(72),'FD')
        
        pdf.set_xy(px2MM(710),px2MM(337)) 
        pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(24))
        pdf.set_text_color(*hex2RGB('#1A1A1D'))
        pdf.cell(px2MM(257), px2MM(32),'Asset',align='L')
        
        #//*----Col2 %
    
        pdf.rect(px2MM(987), px2MM(317), px2MM(100), px2MM(72),'FD')
        pdf.set_xy(px2MM(1007),px2MM(337)) 
        pdf.cell(px2MM(60), px2MM(32),'%',align='C')
        
        #//*----Col3 Assets Class
        pdf.rect(px2MM(1087), px2MM(317), px2MM(293), px2MM(72),'FD')
        
        pdf.set_xy(px2MM(1107),px2MM(337)) 
        pdf.cell(px2MM(253), px2MM(32),'Asset Class',align='L')
        
        #//*----Col4 Market Value
        pdf.rect(px2MM(1380), px2MM(317), px2MM(177), px2MM(72),'FD')
        
        pdf.set_xy(px2MM(1400),px2MM(337)) 
        pdf.cell(px2MM(137), px2MM(32),'Market Value',align='R')
        
        #//*----Col5 Monthly Investments
        pdf.rect(px2MM(1557), px2MM(317), px2MM(243), px2MM(72),'FD')
        
        pdf.set_xy(px2MM(1577),px2MM(337)) 
        pdf.cell(px2MM(203), px2MM(32),'Monthly Investment',align='R')
        
        #//*---Dynamic y axis---
        rect_y = 389
        rect_gap = 62
        state_y = 404
        state_gap = 62


        y_high = pdf.get_y()+20
        col = '#F3F6F9'
        for i in range(start,stop):
            try:
                if not df_table["asset"].iloc[i]:
                    break
            except IndexError:
                break
            
            #//*-----Assets Table---*//
            #//*----Col1 Assets
            if col == '#F3F6F9':
                col = '#FFFFFF'
            else:
                col = '#F3F6F9'
            
            if i==start:
                col = '#F3F6F9'
            
            pdf.set_fill_color(*hex2RGB(col))
                
                
            pdf.set_font('LeagueSpartan-Regular', size=px2pts(24))
            pdf.set_draw_color(*hex2RGB('#E9EAEE'))
            pdf.rect(px2MM(690), px2MM(rect_y), px2MM(297), px2MM(62),'FD')
            
            
            pdf.set_xy(px2MM(710),px2MM(state_y)) 
            pdf.set_font('LeagueSpartan-Regular', size=px2pts(24))
            pdf.set_text_color(*hex2RGB('#1A1A1D'))
            pdf.cell(px2MM(257), px2MM(32),str(df_table['asset'][i]),align='L')
            
            #//*----Col2 %
        
            pdf.rect(px2MM(987), px2MM(rect_y), px2MM(100), px2MM(62),'FD')
            pdf.set_xy(px2MM(1007),px2MM(state_y)) 
            pdf.cell(px2MM(60), px2MM(32),"{:.0f}".format(float(df_table['percentage'][i]))+'%',align='C')
            
            #//*----Col3 Assets Class
            pdf.rect(px2MM(1087), px2MM(rect_y), px2MM(293), px2MM(62),'FD')
            
            pdf.set_xy(px2MM(1107),px2MM(state_y)) 
            pdf.cell(px2MM(253), px2MM(32),str((df_table['asset_class'][i])),align='L')
            
            #//*----Col4 Market Value
            pdf.rect(px2MM(1380), px2MM(rect_y), px2MM(177), px2MM(62),'FD')
            
            pdf.set_xy(px2MM(1400),px2MM(state_y)) 
            if df_table['market_value'][i] == '':
                pdf.cell(px2MM(137), px2MM(32),'-',align='R')
            else:
                pdf.cell(px2MM(137), px2MM(32),'₹ '+str(format_cash2(float(df_table['market_value'][i]))),align='R')
                
            #//*----Col5 Monthly Investments
            pdf.rect(px2MM(1557), px2MM(rect_y), px2MM(243), px2MM(62),'FD')
            
            pdf.set_xy(px2MM(1577),px2MM(state_y))
            
            if df_table['monthly_investments'][i] == '' or int(float(df_table['monthly_investments'][i])) == 0:
                pdf.cell(px2MM(203), px2MM(32),'₹ 0.0K',align='R')
            else:
                pdf.cell(px2MM(203), px2MM(32),'₹ '+str(format_cash2(float(df_table['monthly_investments'][i]))),align='R')
            
            rect_y+=rect_gap
            state_y+=state_gap  
            y_high = pdf.get_y()
            
        start+=8
        stop+=8
        
        # tot_rect_y = mm2PX(pdf.get_y())+47
        # tot_text_y = mm2PX(pdf.get_y())+56
        tot_rect_y = rect_y
        tot_text_y = state_y
        #//*-----Index Text of Page--**////
        index_text(pdf,'#1A1A1D')
    tot_text_y -=4
    #//*----Total Line
    pdf.set_fill_color(*hex2RGB('#ffffff'))
    pdf.set_line_width(px2MM(1))
    pdf.set_draw_color(*hex2RGB('#B9BABE'))   
    pdf.set_font('LeagueSpartan-Regular', size=px2pts(24))

    pdf.rect(px2MM(690), px2MM(tot_rect_y), px2MM(1110), px2MM(1),'FD')
    pdf.set_draw_color(*hex2RGB('#FFFFFF'))
    
    pdf.set_draw_color(*hex2RGB('#E9EAEE'))
    pdf.set_line_width(px2MM(0.2))
    pdf.rect(px2MM(690), px2MM(tot_rect_y+1), px2MM(1110), px2MM(52),'FD')
    
    
    pdf.set_xy(px2MM(710),px2MM(tot_text_y)) 
    pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(24))
    pdf.set_text_color(*hex2RGB('#1A1A1D'))
    pdf.cell(px2MM(257), px2MM(32),'Total',align='L')
    
    #//*----Col2 %
    pdf.set_xy(px2MM(1007),px2MM(tot_text_y)) 
    pdf.cell(px2MM(60), px2MM(32),' ',align='C')
    
    #//*----Col3 Assets Class
    pdf.set_xy(px2MM(1107),px2MM(tot_text_y)) 
    pdf.cell(px2MM(253), px2MM(32),' ',align='L')
    
    #//*----Col4 Market Value
    pdf.set_xy(px2MM(1400),px2MM(tot_text_y)) 
    pdf.cell(px2MM(137), px2MM(32),'₹ '+ str(format_cash2(float(json_data['assets']['total']['market_value']))),align='R')
        
    #//*----Col5 Monthly Investments
    pdf.set_xy(px2MM(1577),px2MM(tot_text_y))
    if int(float(json_data['assets']['total']['monthly_investments'])) == 0 or json_data['assets']['total']['monthly_investments']=='':
        pdf.cell(px2MM(203), px2MM(32),'₹ 0.0K',align='R')
    else:
        pdf.cell(px2MM(203), px2MM(32),'₹ '+str(format_cash2(float(json_data['assets']['total']['monthly_investments']))),align='R')
        
        
        
#//**--------------Liability Pie Chart----------------------*/
def liabilities_chart(pdf,json_data,c_MoneyS,money_signData):
    try:
        df_table = pd.DataFrame.from_dict(json_data["liabilities"]['table'])
        df_pie = pd.DataFrame.from_dict(json_data["liabilities"]['pie'])
    except:
        return None 
    
    if df_pie.empty:
        return None
    flag = False 
    for i in range(len(df_pie['percentage'])):
        if int(float(df_pie['percentage'].iloc[i])*100) > 0:
            flag = True

                    
    if flag == False:
        return None

    emi_total = sum([eval(i) for i in df_table['emi'].tolist()])
    
    start = 0
    stop = 8
    
    for pg in range(0,len(df_table),8):
        
        pdf.add_page()
        pdf.set_fill_color(*hex2RGB('#FCF8ED'))
        pdf.rect(0, 0, px2MM(1920), px2MM(1080),'F')
        pdf.set_draw_color(*hex2RGB('#E9EAEE'))
        pdf.set_fill_color(*hex2RGB('#000000'))
        pdf.rect(0, 0, px2MM(964), px2MM(1080),'F')
        
        
        if pg ==0:
            
            font_path = join(cwd,'assets','fonts','Prata')
            font_files = font_manager.findSystemFonts(fontpaths=font_path)
            for font_file in font_files:
                font_manager.fontManager.addfont(font_file)
            # #//*----Donut Pie Chart---*//
            free_colors = ['#FFD976','#ffffff','#A792FF','#82DBC6','#90BEF8','#FFC27E','#FFD976','#3D7DD0']
            colors = free_colors[0:len(df_pie)]
            
            df_pie['colors'] = colors
            df_pie_chart = df_pie

            df_pie_chart = df_pie_chart.replace(0,np.nan, regex=True)
            df_pie_chart = df_pie_chart.dropna()
            
            labels = df_pie_chart['particular']
            sizes = df_pie_chart['percentage']

            fig, ax0 = plt.subplots(figsize=(6.8, 6.8))
            wed_height = [1,0.9,1,1.08,1,1.08,1,1.08,1]
            font = {'family': 'prata','color':  'black','weight': 'normal','size': 24,}
            wedges, plt_labels, junk = ax0.pie(sizes, colors = df_pie_chart['colors'],startangle=90,wedgeprops = {"edgecolor" : "black",'linewidth': 2,'antialiased': True},autopct=autopct_generator(9),textprops=font)
            
            for i in range(len(wedges)):
                wedges[i].set_radius(wed_height[i])
                
            centre_circle = plt.Circle((0,0),0.2,color='black')
            fig = plt.gcf()
            fig.patch.set_facecolor('black')
            fig.gca().add_artist(centre_circle)
            plt.tight_layout()
            plt.savefig('liabilities_chart.png',dpi=650)

        #//*----Adding PIE image
        pdf.image('liabilities_chart.png',px2MM(80), px2MM(300), px2MM(600), px2MM(600))
            
        #//*---Description----*//
        pdf.set_xy(px2MM(120),px2MM(184)) 
        pdf.set_font('LeagueSpartan-Regular', size=px2pts(24))
        pdf.set_text_color(*hex2RGB('#FFFFFF'))
        desc = """Good liabilities generally are productive, with favourable rates and terms, while bad ones are for non-essential expenses, have high rates, or unfavourable terms. Prioritising the repayment of bad liabilities is wise, as they cost more in the long run."""
        pdf.multi_cell(px2MM(812), px2MM(32),desc,align='L')
        
        
        #//*----Legends---*//
        circle_y = 899
        common_gap = 42
        text_y = 893

        labels = df_pie['particular']
        colors = free_colors[0:len(labels)]
        sizes = df_pie['percentage']
        for i in range(0,len(labels)):
            pdf.set_fill_color(*hex2RGB(df_pie['colors'].iloc[i]))
            pdf.circle(x=px2MM(213),y=px2MM(circle_y),r=px2MM(20),style='F')
            
            pdf.set_xy(px2MM(253),px2MM(text_y)) 
            pdf.set_font('LeagueSpartan-Regular', size=px2pts(24))
            pdf.set_text_color(*hex2RGB('#ffffff'))
            pdf.cell(px2MM(200), px2MM(32),labels[i]+':',align='L')
            
            pdf.set_xy(px2MM(484),px2MM(text_y))
            pdf.cell(px2MM(56), px2MM(32),str(int(round(sizes[i])))+'%',align='R')
                
            circle_y+=common_gap
            text_y+=common_gap
            
        #//*----Snapshot of Holding - Liability----*//
        pdf.set_xy(px2MM(120),px2MM(80)) 
        pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(60))
        pdf.set_text_color(*hex2RGB('#ffffff'))
        pdf.cell(px2MM(224), px2MM(84),'Liabilities',align='L')

        Day=dt.datetime.now().strftime("%d")
        month=dt.datetime.now().strftime("%b")
        year=dt.datetime.now().strftime("%Y")

        if 4 <= int(Day) <= 20 or 24 <= int(Day) <= 30:
            suffix = "th"
        else:
            suffix = ["st", "nd", "rd"][int(Day) % 10 - 1]
            
        pdf.set_xy(px2MM(394),px2MM(106)) 
        pdf.set_font('LeagueSpartan-Regular', size=px2pts(24))
        pdf.set_text_color(*hex2RGB('#ffffff'))
        pdf.cell(px2MM(100), px2MM(32),f'As on {str(Day)}',align='R')


        x_after_day = mm2PX(pdf.get_x())-5

        pdf.set_font('LeagueSpartan-Light', size=px2pts(18))
        pdf.set_xy(px2MM(x_after_day),px2MM(106))
        pdf.cell(px2MM(22), px2MM(20),suffix,align='L')
        
        x_after_day = mm2PX(pdf.get_x())
        pdf.set_xy(px2MM(x_after_day),px2MM(106)) 
        pdf.set_font('LeagueSpartan-Regular', size=px2pts(24))
        pdf.set_text_color(*hex2RGB('#ffffff'))
        pdf.cell(px2MM(300), px2MM(32),f'{str(month)} {str(year)}',align='L')
        
        #//*---Existing Liabilities ----*//
        pdf.set_fill_color(*hex2RGB('#000000'))
        pdf.rect(px2MM(1385), px2MM(81), px2MM(415), px2MM(82),'F')
        
        pdf.set_xy(px2MM(1405),px2MM(101)) 
        pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(30))
        pdf.set_text_color(*hex2RGB('#ffffff'))
        pdf.cell(px2MM(375), px2MM(42),'Existing Liabilities:'+' ₹ '+str(format_cash2(float(json_data['liabilities']['total']))),align='C')

        
        #//*-----Liability Table---*//
        #//*----Col1 Liabilities
        pdf.set_fill_color(*hex2RGB('#ffffff'))
        pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(24))
        pdf.set_draw_color(*hex2RGB('#E9EAEE'))
        pdf.set_line_width(px2MM(0.2))
        pdf.rect(px2MM(690), px2MM(390), px2MM(230), px2MM(104),'F')
        
        pdf.set_xy(px2MM(710),px2MM(411)) 
        pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(24))
        pdf.set_text_color(*hex2RGB('#1A1A1D'))
        pdf.cell(px2MM(190), px2MM(64),'Liabilities',align='L')
        
        #//*----Col2Category
    
        pdf.rect(px2MM(920), px2MM(390), px2MM(140), px2MM(104),'FD')
        pdf.set_xy(px2MM(940),px2MM(411)) 
        pdf.cell(px2MM(100), px2MM(64),'Category',align='C')
        
        #//*----Col3 Account Age in Months
        pdf.rect(px2MM(1060), px2MM(390), px2MM(170), px2MM(104),'FD')
        pdf.set_xy(px2MM(1070),px2MM(411)) 
        pdf.multi_cell(px2MM(140), px2MM(32),'Account Age in Months',align='R')
        
        #//*----Pending Months
        pdf.rect(px2MM(1230), px2MM(390), px2MM(130), px2MM(104),'FD')
        pdf.set_xy(px2MM(1240),px2MM(411)) 
        pdf.multi_cell(px2MM(100), px2MM(32),'Pending Months',align='R')
        
        #//*----Outstanding Amount
        pdf.rect(px2MM(1360), px2MM(390), px2MM(170), px2MM(104),'FD')
        pdf.set_xy(px2MM(1370),px2MM(411)) 
        pdf.multi_cell(px2MM(140), px2MM(32),'Outstanding Amount',align='R')
        
        #//*----EMI
        pdf.rect(px2MM(1530), px2MM(390), px2MM(140), px2MM(104),'FD')
        pdf.set_xy(px2MM(1550),px2MM(411)) 
        pdf.multi_cell(px2MM(100), px2MM(64),'EMI',align='R')
        
        #//*----Interest Rate
        pdf.rect(px2MM(1670), px2MM(390), px2MM(130), px2MM(104),'FD')
        pdf.set_xy(px2MM(1690),px2MM(411)) 
        pdf.multi_cell(px2MM(90), px2MM(32),'Interest Rate',align='R')
    
        
        #//*---Dynamic y axis---
        rect_y = 494
        rect_gap = 62
        state_y = 509
        state_gap = 62 

        y_high = pdf.get_y()+20
        for i in range(start,stop):
            try:
                if not df_table["liability"].iloc[i]:
                    break
            except IndexError:
                break
            
            #//*-----Liability Table---*//
            #//*----Col1 Liability
            if i%2==1:
                pdf.set_fill_color(*hex2RGB('#ffffff'))
            else:
                pdf.set_fill_color(*hex2RGB('#F3F6F9'))
                
            pdf.set_font('LeagueSpartan-Regular', size=px2pts(24))
            pdf.set_draw_color(*hex2RGB('#E9EAEE'))
            pdf.set_text_color(*hex2RGB('#1A1A1D'))
            
            pdf.rect(px2MM(690), px2MM(rect_y), px2MM(230), px2MM(62),'FD')
            pdf.set_xy(px2MM(710),px2MM(state_y)) 
            pdf.cell(px2MM(190), px2MM(32),str(df_table['liability'].iloc[i]),align='L')
            
            #//*----Col2---*//
            pdf.rect(px2MM(920), px2MM(rect_y), px2MM(140), px2MM(62),'FD')
            pdf.set_xy(px2MM(940),px2MM(state_y)) 
            pdf.cell(px2MM(100), px2MM(32),str(df_table['liability_category'].iloc[i]),align='C')
            
            #//*----Col3---*//
            
            pdf.rect(px2MM(1060), px2MM(rect_y), px2MM(170), px2MM(62),'FD')
            pdf.set_xy(px2MM(1080),px2MM(state_y)) 
            if df_table['account_age_in_months'].iloc[i] =="" or df_table['account_age_in_months'].iloc[i] ==0:
                pdf.cell(px2MM(130), px2MM(32),' ',align='R')
            else:
                pdf.cell(px2MM(130), px2MM(32),str(df_table['account_age_in_months'].iloc[i]),align='R')
            
            #//*----Col4---*//
            
            pdf.rect(px2MM(1230), px2MM(rect_y), px2MM(130), px2MM(62),'FD')
            pdf.set_xy(px2MM(1250),px2MM(state_y)) 
            if df_table['pending_months'].iloc[i] == '' or df_table['pending_months'].iloc[i]==0:
                pdf.cell(px2MM(90), px2MM(32),' ',align='R')
            else:
                pdf.cell(px2MM(90), px2MM(32),str(int(df_table['pending_months'].iloc[i])),align='R')
                
            #//*----Col5---*//
            
            pdf.rect(px2MM(1360), px2MM(rect_y), px2MM(170), px2MM(62),'FD')
            pdf.set_xy(px2MM(1380),px2MM(state_y))

            if int(float(df_table['outstanding_amount'].iloc[i])) == 0 or df_table['outstanding_amount'].iloc[i]=='': 
                pdf.cell(px2MM(130), px2MM(32),'₹ 0.0L ',align='R')
            else:
                pdf.cell(px2MM(130),px2MM(32), '₹ '+str(format_cash2(round(float(df_table['outstanding_amount'].iloc[i]),1))),align='R')

            
            #//*----Col6---*//
            
            pdf.rect(px2MM(1530), px2MM(rect_y), px2MM(140), px2MM(62),'FD')
            pdf.set_xy(px2MM(1550),px2MM(state_y)) 
            if df_table['emi'].iloc[i] == 0 or df_table['emi'].iloc[i] =='':
                pdf.cell(px2MM(100), px2MM(32),' ',align='R')
            else:
                pdf.cell(px2MM(100), px2MM(32),'₹ '+str(format_cash2(round(float(df_table['emi'].iloc[i]),1))),align='R')
                
            #//*----Col7---*//
            
            pdf.rect(px2MM(1670), px2MM(rect_y), px2MM(130), px2MM(62),'FD')
            pdf.set_xy(px2MM(1690),px2MM(state_y)) 
            if df_table['interest_rate'].iloc[i] == 0 or df_table['interest_rate'].iloc[i] =='':
                pdf.cell(px2MM(100), px2MM(32),' ',align='R')
            else:
                pdf.cell(px2MM(100), px2MM(32),str(round(df_table['interest_rate'].iloc[i],1))+'%',align='R')
            
            rect_y+=rect_gap
            state_y+=state_gap
            y_high = pdf.get_y()
            
        start+=8
        stop+=8

        tot_rect_y = rect_y
        tot_text_y = state_y
        
        #//*-----Index Text of Page--**////
        index_text(pdf,'#1A1A1D')
    tot_text_y -= 0
    #//*-----------Total Line---------------*//
    pdf.set_fill_color(*hex2RGB('#ffffff'))
    pdf.set_line_width(px2MM(1))
    pdf.set_draw_color(*hex2RGB('#B9BABE'))
    pdf.set_font('LeagueSpartan-SemiBold',size=px2pts(24))
    pdf.rect(px2MM(690), px2MM(tot_rect_y), px2MM(1110), px2MM(1),'FD')
    
    pdf.set_draw_color(*hex2RGB('#E9EAEE'))
    pdf.set_line_width(px2MM(0.2))
    
    pdf.rect(px2MM(690), px2MM(tot_rect_y+1), px2MM(1110), px2MM(62),'FD')
    pdf.set_xy(px2MM(710),px2MM(tot_text_y)) 
    pdf.cell(px2MM(190), px2MM(32),'Total',align='L')
    
    #//*----Col2---*//
    pdf.set_xy(px2MM(940),px2MM(tot_text_y)) 
    pdf.cell(px2MM(100), px2MM(32),' ',align='C')
    
    #//*----Col3---*//
    pdf.set_xy(px2MM(1080),px2MM(tot_text_y)) 
    pdf.cell(px2MM(130), px2MM(32),' ',align='R')
    
    #//*----Col4---*//
    pdf.set_xy(px2MM(1250),px2MM(tot_text_y)) 
    pdf.cell(px2MM(90), px2MM(32),' ',align='R')
        
    #//*----Col5---*//
    pdf.set_xy(px2MM(1380),px2MM(tot_text_y))
    pdf.cell(px2MM(130),px2MM(32),' ₹ '+str(format_cash2(round(float(json_data['liabilities']['total']),1))),align='R')

    #//*----Col6---*//
    pdf.set_xy(px2MM(1550),px2MM(tot_text_y)) 
    pdf.cell(px2MM(100), px2MM(32),' ₹ '+str(format_cash2(round(emi_total,1))),align='R')
        
    #//*----Col7---*//
    pdf.set_xy(px2MM(1690),px2MM(tot_text_y)) 
    pdf.cell(px2MM(100), px2MM(32),' ',align='R')
          
#//*---------------------------------------------------------------     
        
def autopct_generator(limit):
    def inner_autopct(pct):
        return ('%.0f' % pct)+'%' if pct > limit else ''
    return inner_autopct  
        
#//*-------Net Worth Projection-----*//    
def net_worth_projection(pdf,json_data,c_MoneyS,money_signData):
    try:
        df = pd.DataFrame.from_dict(json_data["networth"]['networth_projection']['table'])
        if df.empty:
            return None 
    except:
        return None   
    ini = 0
    stps = 28
    ini_2 = ini
    last_val = 28
    for tab in range(ini,len(df),stps):
        #//*---Page setup----*//
        pdf.add_page()
        pdf.set_fill_color(*hex2RGB('#FCF8ED'))
        pdf.rect(0, 0, px2MM(1920), px2MM(1080),'F')
        
        #//*-----Index Text of Page--**////
        index_text(pdf,'#1A1A1D')

        pdf.image(join(cwd,'assets','images','backgrounds','doubleLine.png'),px2MM(1449),px2MM(0),px2MM(471),px2MM(1080))
        
        #//*----Net Worth Projection----*//
        pdf.set_xy(px2MM(120),px2MM(80)) 
        pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(60))
        pdf.set_text_color(*hex2RGB('#1A1A1D'))
        pdf.cell(px2MM(535), px2MM(84),'Net Worth Projection',align='L')
        
        #//*-----Table White rect
        pdf.set_fill_color(*hex2RGB('#FFFFFF'))

        year_rect_x = 886
        year_state_x = 906
        val_rect_y=val_rect_y2 = 289
        val_state_y=val_state_y2 = 299
        
        current_rect_x = 1025
        current_state_x = 1045
        
        project_rect_x = 1164
        project_state_x = 1184

        pdf.set_draw_color(*hex2RGB('#E9EAEE'))
        pdf.set_line_width(px2MM(0.2))
        pdf.rect(px2MM(846), px2MM(204), px2MM(954), px2MM(755),'FD')
        #//*-----Table Headings---*//
        
        #//*---Table 1
        #//*--Col 1
        pdf.set_draw_color(*hex2RGB('#E9EAEE'))
        pdf.set_fill_color(*hex2RGB('#F3F6F9'))
        pdf.set_line_width(px2MM(0.2))
        pdf.rect(px2MM(886), px2MM(244), px2MM(139), px2MM(45),'FD')
        pdf.set_font('LeagueSpartan-Medium', size=px2pts(18))
        pdf.set_text_color(*hex2RGB('#000000'))
        
        pdf.set_xy(px2MM(906),px2MM(254))
        pdf.cell(px2MM(99), px2MM(25),'Year',align='C')
        
        #//*--Col 2
        pdf.rect(px2MM(1025), px2MM(244), px2MM(139), px2MM(45),'FD')
        pdf.set_xy(px2MM(1045),px2MM(254))
        pdf.cell(px2MM(99), px2MM(25),'CNWT (Cr)',align='C')
        
        #//*--Col 3

        pdf.rect(px2MM(1164), px2MM(244), px2MM(139), px2MM(45),'FD')
        pdf.set_xy(px2MM(1184),px2MM(254))
        pdf.cell(px2MM(99), px2MM(25),'NWTEP (Cr)',align='C')
        
        #//*---Table 2
        if len(df)-ini_2 > 14:
            #//*--Col 1
            pdf.set_fill_color(*hex2RGB('#F3F6F9'))
            pdf.set_line_width(px2MM(0.2))
            pdf.rect(px2MM(1343), px2MM(244), px2MM(139), px2MM(45),'FD')
            pdf.set_font('LeagueSpartan-Medium', size=px2pts(18))
            
            pdf.set_xy(px2MM(1363),px2MM(254))
            pdf.set_text_color(*hex2RGB('#000000'))
            pdf.cell(px2MM(99), px2MM(25),'Year',align='C')
            
            #//*--Col 2
            pdf.rect(px2MM(1482), px2MM(244), px2MM(139), px2MM(45),'FD')
            pdf.set_xy(px2MM(1502),px2MM(254))
            pdf.cell(px2MM(99), px2MM(25),'CNWT (Cr)',align='C')
            
            #//*--Col 3
            pdf.rect(px2MM(1621), px2MM(244), px2MM(139), px2MM(45),'FD')
            pdf.set_xy(px2MM(1641),px2MM(254))
            pdf.cell(px2MM(99), px2MM(25),'NWTEP (Cr)',align='C')
            
        
        #//**--Table x and y settings---**//
        common_gap = 45
        
        #//*---Table value---*//
        
        for i in range(ini_2,last_val):
            try:
        
            #//*----Col 1
                if i%2==0:
                    pdf.set_fill_color(*hex2RGB('#ffffff'))
                else:
                    pdf.set_fill_color(*hex2RGB('#F3F6F9'))
                pdf.set_draw_color(*hex2RGB('#E9EAEE'))
                pdf.set_line_width(px2MM(0.2))
                pdf.set_font('LeagueSpartan-Regular', size=px2pts(18))
                pdf.set_xy(px2MM(year_state_x),px2MM(val_state_y))
                pdf.set_text_color(*hex2RGB('#000000'))
                if (df['year'][i]):
                    pdf.rect(px2MM(year_rect_x), px2MM(val_rect_y), px2MM(139), px2MM(45),'FD') 
                pdf.cell(px2MM(99), px2MM(25),"{0:.0f}".format(df['year'][i]),align='C')
                
                #//*---Col 2
                pdf.rect(px2MM(current_rect_x), px2MM(val_rect_y), px2MM(139), px2MM(45),'FD') 
                pdf.set_xy(px2MM(current_state_x),px2MM(val_state_y))
                pdf.cell(px2MM(99), px2MM(25),"{0:.1f}".format(df['cnwt'][i]),align='C')
                
                #//*---Col 3
                pdf.rect(px2MM(project_rect_x), px2MM(val_rect_y), px2MM(139), px2MM(45),'FD') 
                pdf.set_xy(px2MM(project_state_x),px2MM(val_state_y))
                pdf.cell(px2MM(99), px2MM(25),"{0:.1f}".format(df['nwtet'][i]),align='C')
                
                val_rect_y+=common_gap
                val_state_y+=common_gap
                
                if i==ini_2+27:

                    ini_2 += 28
                    last_val+=28
  
                #//*---Reiniatilizing x and y axis for 2nd side table
                if i == ini_2+13:
                    year_rect_x = 1343
                    year_state_x = 1363
                    current_rect_x = 1482
                    current_state_x = 1502
                    project_rect_x = 1621
                    project_state_x = 1641
                    
                    val_rect_y = val_rect_y2
                    val_state_y= val_state_y2

            except:
                pass
                
        if tab ==0:
            #//*------Line Graph---*//
            font_dir = [join(cwd, 'assets','fonts','League_Spartan','static')]
            font_files2 = font_manager.findSystemFonts(fontpaths=font_dir)
            for files in font_files2:
                font_manager.fontManager.addfont(files)

            fig,ax = plt.subplots()
            min_year = df['year'].min()
            max_year = df['year'].max()
            
            a = df['year'].astype(int)
            b = df['nwtet'].astype(float)
            c = df['cnwt'].astype(float)  
 
            #//*----------case 1-------------------*//
            pp = math.ceil(len(df['year'])/8)
            
            color_b = '#FF7051'
            color_a =  '#43D195'
            
            ax = sns.lineplot(x = a,y=b)
            plt.plot(a,b,color=color_a,ms = 3 ,lw = 1)
            plt.plot(a,c,color=color_b,ms = 3 ,lw = 1)
            # plt.plot(min(a),min(b),color='black',ms = 5)
            ax.yaxis.set_major_formatter(tick.FuncFormatter(y_fmt))
    
            plt.xlabel('')
            plt.ylabel('')

            min_year = df['year'].min()
            max_year = df['year'].max()
 
            min_ideal = min(b)
            max_ideal = max(b)
            
            if max(c)>max(b):
                max_ideal = max(c)
            else:
                max_ideal = max(b)
                
            if min(c) <= min(b):
                min_ideal = min(c)
            else:
                min_ideal = min(b)
                
            z = max_ideal/3
            max_ideal = max_ideal +z
                
            plt.xlim(min(a)-1,max(a))
            plt.ylim(int(min_ideal)-1,max_ideal)
            # plt.ylim(min_ideal,max_ideal)
            
  
            #//*----setting Shade of NWTEP as green and CNWT as red statically---*//  
                  
            max_b = max(b)
            red_lp = np.linspace(max_b,min_ideal-1,100)
            for i in red_lp:
                plt.fill_between(a,i,b,color= '#FFD4CB',alpha=0.03) 
            
            NbData = len(a)  
            max_a = max(a)    
            red_MaxBL = [[MaxBL] * NbData for MaxBL in range(max_a)]
            Max = [np.asarray(red_MaxBL[x]) for x in range(max_a)]
            
            for x in range (math.ceil(max(b)),max_a):
                plt.fill_between(a,Max[x],b, facecolor='white', alpha=1) 
                
            # plt.fill_between(a,c,b,color= '#D4FFED',alpha=.9,interpolate=True)
            plt.fill_between(a,c,b,where=b>c,color= '#D4FFED',alpha=.9,interpolate=True)
            plt.fill_between(a,c,b,where=b<c,color= '#FFD4CB',alpha=.6,interpolate=True)
   
                    
                

            # #//*----Color shading case 3
            # if max(c)>max(b):
            #     ln_color_b = '#FFD4CB' #red
            #     ln_color_a =  '#D4FFED' #green

            #     plt.fill_between(a,c,color=ln_color_a,alpha=1) 
            #     plt.fill_between(a,c,b,color=ln_color_b,alpha=0.7) 
            # else:
            #     ln_color_a =  '#D4FFED' #green
            #     ln_color_b = '#FFD4CB' #red
                
            #     plt.fill_between(a,b,color=ln_color_b,alpha=1) 
            #     plt.fill_between(a,b,c,color=ln_color_a,alpha=0.7)
            
            # color_b = '#D4FFED' #green
            # color_a =  '#FFD4CB'  #red
    

            # plt.fill_between(a,b,color=color_b,alpha=1) 
            # plt.fill_between(a,c,color=color_a,alpha=0.7)  
            
            #//*----Circle marker at starting
 
            plt.plot(min(a),min(b), 'o',markerfacecolor='none', ms=10, markeredgecolor='black')  
            plt.plot(min(a),min(b),linewidth=4, marker ='.',color='#000000')   
                  
            # # //*-------Case2 of Ticks------*//

            pp = math.ceil(len(a)/8)
            rem = pp%8
            arg = np.arange(start=df['year'].min(), stop=df['year'].max(), step=pp)
            plt.xticks(np.arange(min(a)+1, max(a),pp))
            
            if len(a)>0 and len(a)<=8:
                plt.xticks(np.arange(min(a), max(a), 1))
            elif len(a)>8 and len(a)<=16:
                plt.xticks(np.arange(min(a), max(a), 2))
            elif len(a)>16 and len(a)<=24:
                plt.xticks(np.arange(min(a), max(a), 3))
            elif len(a)>24 and len(a)<=32:
                plt.xticks(np.arange(min(a), max(a)+1, 4))
            elif len(a)>32 and len(a)<=40:
                plt.xticks(np.arange(min(a), max(a)+1,5))
            else:
                plt.xticks(np.arange(min(a), max(a)+1,math.ceil(len(a)/8)))
                

            #//*---X tick Rotation
            plt.yticks(fontname = "Arial")  
            plt.xticks(fontname = "Arial")  
            ax.tick_params(axis='x', labelrotation = 00)
            ax.tick_params(axis='both',labelsize=10,colors='#65676D')
            ax.tick_params(axis='y',labelsize=10)
            ax.grid(color='#DCDCDC', linestyle='-', linewidth=0.15)
            ax.yaxis.grid(True) 
            ax.xaxis.grid(True)
            ax.spines[['right', 'top','left','bottom']].set_visible(False)
            plt.tick_params(left = False,bottom = False)
        
            plt.savefig('acutal_networth_chart.png',dpi=250)
        
        #//*----Legend and Graph plotting---*//
        pdf.set_fill_color(*hex2RGB('#ffffff'))
        pdf.set_draw_color(*hex2RGB('#E9EAEE'))
        pdf.set_line_width(px2MM(0.2))
        pdf.rect(px2MM(120),px2MM(204),px2MM(686),px2MM(762),'FD')    
        pdf.image('acutal_networth_chart.png',px2MM(160),px2MM(206),px2MM(606),px2MM(400))

        pdf.set_fill_color(*hex2RGB(color_b))
        pdf.rect(px2MM(169),px2MM(629),px2MM(12),px2MM(12),'F')   
        
        pdf.set_font('LeagueSpartan-Medium', size=px2pts(24))
        pdf.set_xy(px2MM(196),px2MM(619))
        pdf.set_text_color(*hex2RGB('#000000'))
        pdf.cell(px2MM(559), px2MM(32),'Current Net Worth Trajectory (CNWT)',align='L')
        
        pdf.set_font('LeagueSpartan-Regular', size=px2pts(24))
        pdf.set_xy(px2MM(196),px2MM(656))
        pdf.set_text_color(*hex2RGB('#898B90'))
        # max_curr = "{0:.2f}".format(float(df['cnwt'].max()))
        max_curr = str(format_cash2(float(json_data['networth']['networth_projection']['retirement_cnwt'])))
        # today = datetime.now()
        # mnth = today.strftime("%B")
        mnth = json_data['networth']['networth_projection']['retirement_month_year']
        # data1 = mnth+' '+str(int(max_year))+' | ₹'+max_curr+' Cr'
        data1 = mnth+' | ₹ '+max_curr
        pdf.cell(px2MM(559), px2MM(32),data1,align='L')
        
        pdf.set_font('LeagueSpartan-Light', size=px2pts(18))
        pdf.set_xy(px2MM(196),px2MM(693))
        pdf.set_text_color(*hex2RGB('#000000'))
        pdf.cell(px2MM(559), px2MM(25),'Assumes that you maintain your current financial habits until retirement.',align='L')
        
        #//*------
        pdf.set_fill_color(*hex2RGB(color_a))
        pdf.rect(px2MM(169),px2MM(758),px2MM(12),px2MM(12),'F')   
        
        pdf.set_font('LeagueSpartan-Medium', size=px2pts(24))
        pdf.set_xy(px2MM(196),px2MM(748))
        pdf.set_text_color(*hex2RGB('#000000'))
        # pdf.cell(px2MM(295), px2MM(32),str(df['Current net worth Trajectory'].max()),align='C') 
        pdf.cell(px2MM(559), px2MM(32),'Net worth Trajectory With Effective Planning (NWTEP)',align='L')
        
        pdf.set_font('LeagueSpartan-Regular', size=px2pts(24))
        pdf.set_xy(px2MM(196),px2MM(785))
        pdf.set_text_color(*hex2RGB('#898B90'))
        # pdf.cell(px2MM(295), px2MM(32),str(df['Current net worth Trajectory'].max()),align='C')
        max_net_worth = str(format_cash2(float(json_data['networth']['networth_projection']['retirement_nwtet']))) 

        data2 = mnth+' | ₹ '+max_net_worth
        pdf.cell(px2MM(559), px2MM(32),data2,align='L')
        
        pdf.set_font('LeagueSpartan-Light', size=px2pts(18))
        pdf.set_xy(px2MM(196),px2MM(822))
        pdf.set_text_color(*hex2RGB('#000000'))
        pdf.multi_cell(px2MM(559), px2MM(25),'''Assumes that your finances are aligned with your personality by following the ideal guidance provided on the 'Your Financial Analysis' pages on the following aspects: expense and liability management, asset allocation, and emergency planning.''',align='L')
        
        ini = 29
        stps = 29 
        
def y_fmt(x, y):
    return f'₹ {int(x)}Cr'.format(x)
#//*---Structure for Term and Health Insurance---*// 
def term_health_features(pdf,df,pg_name):
    
    def add_base_page(pdf): 
        pdf.add_page()
        pdf.set_fill_color(*hex2RGB('#FCF8ED'))
        pdf.rect(0, 0, px2MM(1920), px2MM(1080),'F')
        
        # black rectangle
        pdf.set_fill_color(*hex2RGB('#000000'))
        pdf.rect(px2MM(0), px2MM(80), px2MM(15), px2MM(84), 'F')

        #//*----Featured List of Financial Products----*//
        pdf.set_xy(px2MM(120),px2MM(80)) 
        pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(60))
        pdf.set_text_color(*hex2RGB('#1A1A1D'))
        pdf.cell(px2MM(877), px2MM(84),'Financial Products Featured List',align='L')
        
        #//*---Top Black box
        pdf.set_fill_color(*hex2RGB('#313236'))
        pdf.set_font('LeagueSpartan-Medium', size=px2pts(24))
        pdf.set_text_color(*hex2RGB('#ffffff'))
        
        if pg_name == 'Term Insurance Plans':
        
            pdf.rect(px2MM(126), px2MM(204), px2MM(242), px2MM(42),'F')
            pdf.set_xy(px2MM(141),px2MM(209)) 
            pdf.cell(px2MM(212), px2MM(32),"Term Insurance Plans",align='L')
        elif  pg_name == 'Health Insurance Plans':
            pdf.rect(px2MM(126), px2MM(204), px2MM(259), px2MM(42),'F')
            pdf.set_xy(px2MM(141),px2MM(209)) 
            pdf.cell(px2MM(229), px2MM(32),"Health Insurance Plans",align='L')
        
            
        # //*---Col 1
        
        pdf.set_fill_color(*hex2RGB('#ffffff'))
        pdf.set_draw_color(*hex2RGB('#E9EAEE'))
        pdf.rect(px2MM(126), px2MM(246), px2MM(558), px2MM(72),'FD')
        
        pdf.set_xy(px2MM(146),px2MM(266)) 
        pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(24))
        pdf.set_text_color(*hex2RGB('#1A1A1D'))
        pdf.multi_cell(px2MM(200), px2MM(32),'Plan Details',align='L')
        
        #//*---Col 2
        pdf.rect(px2MM(684), px2MM(246), px2MM(558), px2MM(72),'FD')
        pdf.set_xy(px2MM(704),px2MM(266)) 
        pdf.multi_cell(px2MM(230), px2MM(32),'Strength',align='L')
    
        #//*---Col 3
        pdf.rect(px2MM(1242), px2MM(246), px2MM(558), px2MM(72),'FD')
        pdf.set_xy(px2MM(1262),px2MM(266)) 
        pdf.multi_cell(px2MM(524), px2MM(32),'Weakness',align='L')
    
    #//*---Desclaimer Function---*//
    def add_disclaimer(pdf):
        pdf.set_fill_color(*hex2RGB('#FCF8ED'))
        pdf.rect(px2MM(0), px2MM(1006), px2MM(1920), px2MM(40),'F')
        desclaimer = "Disclaimer: The above featured list is based on 1 Finance's proprietary research."
        pdf.set_text_color(*hex2RGB('#000000'))
        pdf.set_font('LeagueSpartan-Regular', size=px2pts(24))
        pdf.set_xy(px2MM(120),px2MM(1008))      
        pdf.multi_cell(px2MM(1680), px2MM(32),desclaimer,align='C')
        
        #//*-----Index Text of Page--**////
        index_text(pdf,'#1A1A1D')
        global fin_feat_product_list
        
        if fin_feat_product_list == 0:
            fin_feat_product_list = pdf.page_no()
        
    #//*---Function to check height of rectangle
    def get_rect_h(rem,strength,weakness):
        pdf.set_text_color(*hex2RGB('#FCF8ED'))
        pdf.set_text_color(*hex2RGB('#000000'))
        rem_pro = 0
        rem_con = 0
        
        s_text_h = rem
        w_text_h = rem
        
        for j in range(len(strength)):
            pdf.set_font('LeagueSpartan-Regular', size=px2pts(24))
            pdf.set_xy(px2MM(744),px2MM(s_text_h)) 
            pdf.multi_cell(px2MM(478), px2MM(32),strength[j],align='L')
            
            s_text_h = mm2PX(pdf.get_y()) 
        
        rem_pro = mm2PX(pdf.get_y())
        
        for j in range(len(weakness)):
            pdf.set_font('LeagueSpartan-Regular', size=px2pts(24))
            pdf.set_xy(px2MM(1302),px2MM(w_text_h)) 
            pdf.multi_cell(px2MM(478), px2MM(32),weakness[j],align='L')
            w_text_h = mm2PX(pdf.get_y()) 
        
        rem_con = mm2PX(pdf.get_y())
        
        if rem_pro >=rem_con:
            rect_h = rem_pro-rem+20
        else:
            rect_h = rem_con-rem+20
            
        if rem+rect_h > 975:
            pdf.set_fill_color(*hex2RGB('#FCF8ED'))
            pdf.rect(px2MM(126), px2MM(rem+1), px2MM(1674), px2MM(rect_h),'F')
            add_disclaimer(pdf)
               
        return rect_h+20
     
    add_base_page(pdf)
    rem = mm2PX(pdf.get_y())+20
    
    #//*---Disclaimer----*//
    add_disclaimer(pdf)
    
    row_color = '#F3F6F9'    
    pdf.set_fill_color(*hex2RGB(row_color))
    for i in range(len(df)):
        
        rect_h = get_rect_h(rem,df['pros'].iloc[i],df['cons'].iloc[i])
        text_y = rem+20
        
        if rem+rect_h > 975:
            add_base_page(pdf)
            rem = mm2PX(pdf.get_y())+20
            
            #//*---Disclaimer----*//
            add_disclaimer(pdf)
            
            row_color = '#F3F6F9'
            pdf.set_fill_color(*hex2RGB(row_color))
            rect_h = get_rect_h(rem,df['pros'].iloc[i],df['cons'].iloc[i])
            text_y = rem+20
            
        # print('rect height : ',rect_h)
        
        pdf.set_text_color(*hex2RGB('#65676D'))
        pdf.set_font('LeagueSpartan-Regular', size=px2pts(24))
        
        #//*---Column 1 Value
        pdf.set_fill_color(*hex2RGB(row_color))
        pdf.rect(px2MM(126), px2MM(rem), px2MM(558), px2MM(rect_h),'FD')
        pdf.set_xy(px2MM(146),px2MM(text_y)) 
        pdf.multi_cell(px2MM(120), px2MM(32),'Insurer - ',align='L')
        
        pdf.set_text_color(*hex2RGB('#1A1A1D'))
        pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(24))
        pdf.set_xy(px2MM(240),px2MM(text_y)) 
        pdf.multi_cell(px2MM(404), px2MM(32),df['insurer'].iloc[i],align='L')
        
        plan = mm2PX(pdf.get_y())
        
        pdf.set_text_color(*hex2RGB('#65676D'))
        pdf.set_font('LeagueSpartan-Regular', size=px2pts(24))
        pdf.set_xy(px2MM(146),px2MM(plan+12)) 
        pdf.multi_cell(px2MM(100), px2MM(32),'Plan -',align='L')
        
        pdf.set_text_color(*hex2RGB('#1A1A1D'))
        pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(24)) 
        pdf.set_xy(px2MM(210),px2MM(plan+12)) 
        pdf.multi_cell(px2MM(404), px2MM(32),df['plan'].iloc[i],align='L')
        
        #//*----Column 2 Value
        pdf.set_font('LeagueSpartan-Regular', size=px2pts(24))
        pdf.rect(px2MM(684), px2MM(rem), px2MM(558), px2MM(rect_h),'FD')
        
        strength = list(filter(remove_empty_strings, df['pros'].iloc[i]))
        for j in range(len(strength)):
            if strength[j]=='':
                continue
            if j==0:
                pdf.set_fill_color(*hex2RGB('#000000'))
                pdf.circle(x=px2MM(724),y=px2MM(text_y+15),r=px2MM(6),style='F')
                
                
                pdf.set_xy(px2MM(744),px2MM(text_y)) 
                pdf.multi_cell(px2MM(478), px2MM(32),strength[j],align='L')
                
            else:
                p_y = mm2PX(pdf.get_y())
                pdf.set_fill_color(*hex2RGB('#000000'))
                pdf.circle(x=px2MM(724),y=px2MM(p_y+15),r=px2MM(6),style='F')
        
                pdf.set_xy(px2MM(744),px2MM(p_y)) 
                pdf.multi_cell(px2MM(478), px2MM(32),strength[j],align='L')
                
        #//*----Column 3 Value
        pdf.set_fill_color(*hex2RGB(row_color))
        pdf.rect(px2MM(1242), px2MM(rem), px2MM(558), px2MM(rect_h),'FD')
        weakness = list(filter(remove_empty_strings, df['cons'].iloc[i]))
        for j in range(0,len(weakness)):

            if weakness[j]=='':
                continue
            if j==0:
                pdf.set_fill_color(*hex2RGB('#000000'))
                pdf.circle(x=px2MM(1282),y=px2MM(text_y+15),r=px2MM(6),style='F')
                
                pdf.set_font('LeagueSpartan-Regular', size=px2pts(24))
                pdf.set_xy(px2MM(1302),px2MM(text_y)) 
                pdf.multi_cell(px2MM(478), px2MM(32),weakness[j],align='L')
                
            else:
                p_y = mm2PX(pdf.get_y())
                pdf.set_fill_color(*hex2RGB('#000000'))
                pdf.circle(x=px2MM(1282),y=px2MM(p_y+15),r=px2MM(6),style='F')
                pdf.set_xy(px2MM(1302),px2MM(p_y))      
                pdf.multi_cell(px2MM(478), px2MM(32),weakness[j],align='L')
                
        pdf.set_fill_color(*hex2RGB(row_color))  
        
        if row_color == '#F3F6F9':
            row_color = '#FFFFFF'
        else:
            row_color = '#F3F6F9'
            
        rem = rem+rect_h
            
        # //*----Black VerticaL Line
        pdf.set_fill_color(*hex2RGB('#313236'))
        pdf.rect(px2MM(120), px2MM(204), px2MM(6), px2MM(rem-204),'F')
        
        # pdf.set_fill_color(*hex2RGB('#FCF8ED'))
        # pdf.rect(px2MM(120), px2MM(1006), px2MM(1690), px2MM(40),'FD')
        # desclaimer = "Disclaimer: The above featured list is based on 1 Finance's proprietary research."
        # pdf.set_text_color(*hex2RGB('#000000'))
        # pdf.set_font('LeagueSpartan-Light', size=px2pts(24))
        # pdf.set_xy(px2MM(120),px2MM(1008))      
        # pdf.multi_cell(px2MM(1680), px2MM(32),desclaimer,align='C')
            
  
#//*------Health Term Insurance -----*//        
def term_insurance(pdf,json_data,c_MoneyS,money_signData):
    
    #//*------For Term Insurance Plans-----*//
    df = pd.DataFrame.from_dict(json_data["featured_list"]['term_insurance']['table'])
    pg_name = 'Term Insurance Plans'
    term_health_features(pdf,df,pg_name)
    
   
    
#//*------Health Insurance Plans-----*//

def health_insurance(pdf,json_data,c_MoneyS,money_signData):

    df = pd.DataFrame.from_dict(json_data["featured_list"]['health_insurance']['table'])
    pg_name = 'Health Insurance Plans'
    term_health_features(pdf,df,pg_name)
    
#//*----New Mutual Fund-------------*//

def equity_mutual_fund(pdf,json_data,c_MoneyS,money_signData):
    equity = json_data["featured_list"]['equity_mutual_funds']
    for key, items in equity.items():

        df = pd.DataFrame.from_dict(json_data["featured_list"]['equity_mutual_funds'][key])
        fund_scheme = list(df['fund_scheme'])
        strength = list(df['strengths'])
        weakness = list(df['weakness'])   
        
        def add_equity_page(pdf,key): 
            pdf.add_page()
            pdf.set_fill_color(*hex2RGB('#FCF8ED'))
            pdf.rect(0, 0, px2MM(1920), px2MM(1080),'F')
            
            # black rectangle
            pdf.set_fill_color(*hex2RGB('#000000'))
            pdf.rect(px2MM(0), px2MM(80), px2MM(15), px2MM(84), 'F')

            #//*----Featured List of Financial Products----*//
            pdf.set_xy(px2MM(120),px2MM(80)) 
            pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(60))
            pdf.set_text_color(*hex2RGB('#1A1A1D'))
            pdf.cell(px2MM(877), px2MM(84),'Financial Products Featured List',align='L')
            
            #//*---Top Black box
            pdf.set_fill_color(*hex2RGB('#313236'))
            pdf.set_font('LeagueSpartan-Medium', size=px2pts(24))
            pdf.set_text_color(*hex2RGB('#ffffff'))
            
            if key == 'large_cap':
                val = 'Large Cap Index'
                pdf.rect(px2MM(126), px2MM(204), px2MM(419), px2MM(42),'F')
                pdf.set_xy(px2MM(141),px2MM(209)) 
                pdf.cell(px2MM(389), px2MM(32),"Equity Mutual Funds - "+val,align='L')
            else:
                val = 'Flexicap fund'
                pdf.rect(px2MM(126), px2MM(204), px2MM(388), px2MM(42),'F')
                pdf.set_xy(px2MM(141),px2MM(209)) 
                pdf.cell(px2MM(358), px2MM(32),"Equity Mutual Funds - "+val,align='L')
                
            # //*---Col 1
            
            pdf.set_fill_color(*hex2RGB('#ffffff'))
            pdf.set_draw_color(*hex2RGB('#E9EAEE'))
            pdf.rect(px2MM(126), px2MM(246), px2MM(558), px2MM(72),'FD')
            
            pdf.set_xy(px2MM(146),px2MM(266)) 
            pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(24))
            pdf.set_text_color(*hex2RGB('#1A1A1D'))
            pdf.multi_cell(px2MM(200), px2MM(32),'Fund Scheme',align='L')
            
            #//*---Col 2
            pdf.rect(px2MM(684), px2MM(246), px2MM(558), px2MM(72),'FD')
            pdf.set_xy(px2MM(704),px2MM(266)) 
            pdf.multi_cell(px2MM(230), px2MM(32),'Strength',align='L')
        
            #//*---Col 3
            pdf.rect(px2MM(1242), px2MM(246), px2MM(558), px2MM(72),'FD')
            pdf.set_xy(px2MM(1262),px2MM(266)) 
            pdf.multi_cell(px2MM(524), px2MM(32),'Weakness',align='L')
            
        #//*---Desclaimer Function---*//
        def add_disclaimer(pdf):
            pdf.set_fill_color(*hex2RGB('#FCF8ED'))
            pdf.rect(px2MM(0), px2MM(1006), px2MM(1920), px2MM(40),'F')
            desclaimer = "Disclaimer: All the above schemes are Growth-Direct plans. The above featured list is based on 1 Finance's proprietary research."
            pdf.set_text_color(*hex2RGB('#000000'))
            pdf.set_font('LeagueSpartan-Regular', size=px2pts(24))
            pdf.set_xy(px2MM(120),px2MM(1008))      
            pdf.multi_cell(px2MM(1680), px2MM(32),desclaimer,align='C')
            
            #//*-----Index Text of Page--**////
            index_text(pdf,'#1A1A1D')
            global fin_feat_product_list
            
            if fin_feat_product_list == 0:
                fin_feat_product_list = pdf.page_no()
        
        
        #//*---Function to check height of rectangle
        def get_rect_h(rem,strength,weakness):
            pdf.set_text_color(*hex2RGB('#FCF8ED'))
            # pdf.set_text_color(*hex2RGB('#000000'))
            rem_pro = 0
            rem_con = 0
            
            s_text_h = rem
            w_text_h = rem
            
            for j in range(len(strength)):
                pdf.set_font('LeagueSpartan-Regular', size=px2pts(24))
                pdf.set_xy(px2MM(744),px2MM(s_text_h)) 
                pdf.multi_cell(px2MM(518), px2MM(32),strength[j],align='L')
                
                s_text_h = mm2PX(pdf.get_y()) 
            
            rem_pro = mm2PX(pdf.get_y())
            
            for j in range(len(weakness)):
                pdf.set_font('LeagueSpartan-Regular', size=px2pts(24))
                pdf.set_xy(px2MM(1302),px2MM(w_text_h)) 
                pdf.multi_cell(px2MM(518), px2MM(32),weakness[j],align='L')
                w_text_h = mm2PX(pdf.get_y()) 
            
            rem_con = mm2PX(pdf.get_y())
            
            if rem_pro >=rem_con:
                rect_h = rem_pro-rem+20
            else:
                rect_h = rem_con-rem+20
                
            if rem+rect_h +20 > 950:
                pdf.set_fill_color(*hex2RGB('#FCF8ED'))
                pdf.rect(px2MM(126), px2MM(rem+1), px2MM(1674), px2MM(rect_h),'F')
                add_disclaimer(pdf)
                
            return rect_h+20
        
        
        row_color = '#F3F6F9'    
        pdf.set_fill_color(*hex2RGB(row_color))
        add_equity_page(pdf,key)
        rem = mm2PX(pdf.get_y())+20
        
        #//*--Disclaimer--*//
        add_disclaimer(pdf)
        
        
        
        for i in range(len(df)):
            
            rect_h = get_rect_h(rem,df['strengths'].iloc[i],df['weakness'].iloc[i])
            text_y = rem+20
            
            if rem+rect_h > 880:
                add_equity_page(pdf,key)
                rem = mm2PX(pdf.get_y())+20
                #//*---Disclaimer----*//
                add_disclaimer(pdf)
                
                row_color = '#F3F6F9'
                pdf.set_fill_color(*hex2RGB(row_color))
                rect_h = get_rect_h(rem,df['strengths'].iloc[i],df['weakness'].iloc[i])
                text_y = rem+20
                
            # print('rect height : ',rect_h)
            
            pdf.set_text_color(*hex2RGB('#1A1A1D'))
            pdf.set_font('LeagueSpartan-Regular', size=px2pts(24))
            
            #//*---Column 1 Value
            pdf.set_fill_color(*hex2RGB(row_color))
            pdf.rect(px2MM(126), px2MM(rem), px2MM(558), px2MM(rect_h),'FD')
            pdf.set_xy(px2MM(146),px2MM(text_y)) 
            pdf.multi_cell(px2MM(518), px2MM(32),df['fund_scheme'].iloc[i],align='L')
            
            #//*----Column 2 Value
            pdf.rect(px2MM(684), px2MM(rem), px2MM(558), px2MM(rect_h),'FD')
            strength = list(filter(remove_empty_strings, df['strengths'].iloc[i]))
            for j in range(len(strength)):
                if strength[j]=='':
                    continue
                if j==0:
                    pdf.set_fill_color(*hex2RGB('#000000'))
                    pdf.circle(x=px2MM(724),y=px2MM(text_y+15),r=px2MM(6),style='F')
                    
                    
                    pdf.set_xy(px2MM(744),px2MM(text_y)) 
                    pdf.multi_cell(px2MM(478), px2MM(32),strength[j],align='L')
                    
                else:
                    p_y = mm2PX(pdf.get_y())
                    pdf.set_fill_color(*hex2RGB('#000000'))
                    pdf.circle(x=px2MM(724),y=px2MM(p_y+15),r=px2MM(6),style='F')
            
                    pdf.set_xy(px2MM(744),px2MM(p_y)) 
                    pdf.multi_cell(px2MM(478), px2MM(32),strength[j],align='L')
                    
            #//*----Column 3 Value
            pdf.set_fill_color(*hex2RGB(row_color))
            pdf.rect(px2MM(1242), px2MM(rem), px2MM(558), px2MM(rect_h),'FD')
            weakness = list(filter(remove_empty_strings, df['weakness'].iloc[i]))
            for j in range(0,len(weakness)):

                if weakness[j]=='':
                    continue
                if j==0:
                    pdf.set_fill_color(*hex2RGB('#000000'))
                    pdf.circle(x=px2MM(1282),y=px2MM(text_y+15),r=px2MM(6),style='F')
                    
                    pdf.set_font('LeagueSpartan-Regular', size=px2pts(24))
                    pdf.set_xy(px2MM(1302),px2MM(text_y)) 
                    pdf.multi_cell(px2MM(478), px2MM(32),weakness[j],align='L')
                    
                else:
                    p_y = mm2PX(pdf.get_y())
                    pdf.set_fill_color(*hex2RGB('#000000'))
                    pdf.circle(x=px2MM(1282),y=px2MM(p_y+15),r=px2MM(6),style='F')
                    pdf.set_xy(px2MM(1302),px2MM(p_y))      
                    pdf.multi_cell(px2MM(478), px2MM(32),weakness[j],align='L')
                    
            pdf.set_fill_color(*hex2RGB(row_color))  
            
            if row_color == '#F3F6F9':
                row_color = '#FFFFFF'
            else:
                row_color = '#F3F6F9'
                
            rem = rem+rect_h
                
            # //*----Black VerticaL Line
            pdf.set_fill_color(*hex2RGB('#313236'))
            pdf.rect(px2MM(120), px2MM(204), px2MM(6), px2MM(rem-204),'F')
            
            # desclaimer = "Disclaimer: All the above schemes are Growth-Direct plans. The above featured list is based on 1 Finance's proprietary research. "
            # pdf.set_text_color(*hex2RGB('#000000'))
            # pdf.set_font('LeagueSpartan-Light', size=px2pts(24))
            # pdf.set_xy(px2MM(120),px2MM(1008))      
            # pdf.multi_cell(px2MM(1680), px2MM(32),desclaimer,align='C')
                

            
            
#//*-----Bureau Report Summary---*//
def bureao_report(pdf,json_data,c_MoneyS,money_signData):
    try:
        csa = json_data["bureau_report_summary"]['credit_score_analysis']
        cft = pd.DataFrame.from_dict(json_data["bureau_report_summary"]['credit_facilities_taken'])
    except:
        print(sys.exc_info())
        return None
    
    if csa['score'].strip()=="":
        return None
    
    if cft.empty:
        return None
    try:
        # cft = cft.update(cft.select_dtypes(include=np.number).applymap('{:,g}'.format))
        type_facility = cft["type_of_facility"].tolist()
        tot_record = cft["total_records"].tolist()
        active_acc = cft["active_accounts"].tolist()
        clsd_acc = cft["closed_accounts"].tolist()
        acc_neg_hist = cft["accounts_with_negative_history"].tolist()
        
        
        total_record = sum(list(filter(lambda i: isinstance(i, (int,float)), tot_record)))
        total_active_account = sum(list(filter(lambda i: isinstance(i, (int,float)), active_acc)))
        total_closed_accounts = sum(list(filter(lambda i: isinstance(i, (int,float)), clsd_acc)))
        total_acc_neg_hist = sum(list(filter(lambda i: isinstance(i, (int,float)), acc_neg_hist)))
    except:
        return None

    
    pdf.add_page()
    pdf.set_fill_color(*hex2RGB('#FCF8ED'))
    pdf.rect(0, 0, px2MM(1920), px2MM(1080),'F')

    #//*----Featured List of Financial Products----*//
    pdf.set_xy(px2MM(120),px2MM(80)) 
    pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(60))
    pdf.set_text_color(*hex2RGB('#1A1A1D'))
    pdf.cell(px2MM(877), px2MM(84),'Bureau Report Summary',align='L')
    
    #//*---Top Black box
    pdf.set_fill_color(*hex2RGB('#000000'))
    pdf.rect(px2MM(0), px2MM(80), px2MM(15), px2MM(84),'F')
    
    #//*---Credit Score Analysis
    pdf.set_fill_color(*hex2RGB('#313236'))
    pdf.rect(px2MM(126), px2MM(204), px2MM(243), px2MM(42),'F')
    
    
    
    pdf.set_xy(px2MM(141),px2MM(209)) 
    pdf.set_font('LeagueSpartan-Medium', size=px2pts(24))
    pdf.set_text_color(*hex2RGB('#ffffff'))
    pdf.cell(px2MM(213), px2MM(32),'Credit Score Analysis',align='C')
    
    #//*---Table Header----*//
    pdf.set_fill_color(*hex2RGB('#ffffff'))
    pdf.set_draw_color(*hex2RGB('#E9EAEE'))
    pdf.set_line_width(px2MM(0.2))
    #//*---Col 1
    pdf.rect(px2MM(126), px2MM(246), px2MM(240), px2MM(72),'FD')
    pdf.set_xy(px2MM(146),px2MM(266)) 
    pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(24))
    pdf.set_text_color(*hex2RGB('#1A1A1D'))
    pdf.cell(px2MM(200), px2MM(32),'Your Credit Score',align='C')
     #//*---Col 2
    pdf.rect(px2MM(366), px2MM(246), px2MM(320), px2MM(72),'FD')
    pdf.set_xy(px2MM(386),px2MM(266)) 
    pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(24))
    pdf.cell(px2MM(600), px2MM(32),'Our Evaluation',align='L')
     #//*---Col 3
    pdf.rect(px2MM(686), px2MM(246), px2MM(1114), px2MM(72),'FD')
    pdf.set_xy(px2MM(706),px2MM(266)) 
    pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(24))
    pdf.cell(px2MM(1074), px2MM(32),'Comments',align='L')
    
    
    #//*---Table Value---*//
    if len(csa["commentary"]) > 92:
        rect_h = 104
        text_h = 64
        comm_h = 32
        bl_hight1 = 218
    else:
        comm_h = 32
        rect_h = 72
        text_h = 32
        bl_hight1 = 186
        
    pdf.set_fill_color(*hex2RGB('#313236'))
    pdf.rect(px2MM(120), px2MM(204), px2MM(6), px2MM(bl_hight1),'F')
    
    
    pdf.set_fill_color(*hex2RGB('#F3F6F9'))
    pdf.set_draw_color(*hex2RGB('#E9EAEE'))
    pdf.set_line_width(px2MM(0.2))
    #//*---Col 1
    pdf.rect(px2MM(126), px2MM(318), px2MM(240), px2MM(rect_h),'FD')
    pdf.set_xy(px2MM(146),px2MM(338)) 
    pdf.set_font('LeagueSpartan-Regular', size=px2pts(24))
    pdf.set_text_color(*hex2RGB('#1A1A1D'))
    pdf.cell(px2MM(200), px2MM(text_h),str(csa['score']),align='C')
     #//*---Col 2
    pdf.rect(px2MM(366), px2MM(318), px2MM(320), px2MM(rect_h),'FD')
    pdf.set_xy(px2MM(386),px2MM(338)) 
    pdf.set_font('LeagueSpartan-Regular', size=px2pts(24))
    pdf.multi_cell(px2MM(600), px2MM(text_h),csa['our_evaluation'],align='L')
     #//*---Col 3
    pdf.rect(px2MM(686), px2MM(318), px2MM(1114), px2MM(rect_h),'FD')
    # pdf.set_fill_color(*hex2RGB('#000000'))
    # pdf.circle(x=px2MM(728),y=px2MM(351),r=px2MM(5),style='F')
    pdf.set_xy(px2MM(706),px2MM(338)) 
    pdf.set_font('LeagueSpartan-Regular', size=px2pts(24))
    pdf.multi_cell(px2MM(1074), px2MM(comm_h),csa["commentary"],align='L')
    
    
    
    #//*---Credit Facilities Taken---*//
    
    pdf.set_fill_color(*hex2RGB('#313236'))
    pdf.rect(px2MM(126), px2MM(502), px2MM(248), px2MM(42),'F')
    
    pdf.set_xy(px2MM(141),px2MM(507)) 
    pdf.set_font('LeagueSpartan-Medium', size=px2pts(24))
    pdf.set_text_color(*hex2RGB('#ffffff'))
    pdf.cell(px2MM(218), px2MM(32),'Credit Facilities Taken',align='C')
    
    bl_hight = 114
    
    
    #//*---Table Header----*//
    pdf.set_fill_color(*hex2RGB('#ffffff'))
    pdf.set_draw_color(*hex2RGB('#E9EAEE'))
    pdf.set_line_width(px2MM(0.2))
    pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(24))
    
    #//*---Col 1
    pdf.rect(px2MM(126), px2MM(544), px2MM(290), px2MM(72),'FD')
    pdf.set_xy(px2MM(146),px2MM(564)) 
    pdf.set_text_color(*hex2RGB('#1A1A1D'))
    pdf.cell(px2MM(250), px2MM(32),'Type of Facility',align='L')
    #//*---Col 2
    pdf.rect(px2MM(416), px2MM(544), px2MM(290), px2MM(72),'FD')
    pdf.set_xy(px2MM(436),px2MM(564)) 
    pdf.cell(px2MM(250), px2MM(32),'Total Records',align='C')
    #//*---Col 3
    pdf.rect(px2MM(706), px2MM(544), px2MM(290), px2MM(72),'FD')
    pdf.set_xy(px2MM(726),px2MM(564)) 
    pdf.cell(px2MM(250), px2MM(32),'Active Accounts',align='C')
    #//*---Col 4
    pdf.rect(px2MM(996), px2MM(544), px2MM(290), px2MM(72),'FD')
    pdf.set_xy(px2MM(1016),px2MM(564)) 
    pdf.cell(px2MM(250), px2MM(32),'Closed Accounts',align='C')
    #//*---Col 5
    pdf.rect(px2MM(1286), px2MM(544), px2MM(514), px2MM(72),'FD')
    pdf.set_xy(px2MM(1306),px2MM(564)) 
    pdf.cell(px2MM(474), px2MM(32),'Accounts with Negative History',align='C')
    
    for i in range(len(type_facility)):
        #//*---Table Header----*//
        if i%2==0:
            pdf.set_fill_color(*hex2RGB('#F3F6F9'))
            pdf.set_font('LeagueSpartan-Regular', size=px2pts(24))
        else:
            pdf.set_font('LeagueSpartan-Regular', size=px2pts(24))
            pdf.set_fill_color(*hex2RGB('#ffffff'))
            
        if i==len(type_facility):
            pdf.set_fill_color(*hex2RGB('#ffffff'))
            pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(24))
            
        pdf.set_draw_color(*hex2RGB('#E9EAEE'))
        pdf.set_line_width(px2MM(0.5))
        #//*---Col 1
        pdf.rect(px2MM(126), px2MM(616+(i*52)), px2MM(290), px2MM(52),'FD')
        pdf.set_xy(px2MM(146),px2MM(626+(i*52))) 
        
        pdf.set_text_color(*hex2RGB('#1A1A1D'))
        pdf.cell(px2MM(250), px2MM(32),str(type_facility[i]),align='L')
        #//*---Col 2
        pdf.rect(px2MM(416), px2MM(616+(i*52)), px2MM(290), px2MM(52),'FD')
        pdf.set_xy(px2MM(436),px2MM(626+(i*52)))
        # if tot_record[i] == 0 or tot_record[i] == '':
        #     pdf.cell(px2MM(250), px2MM(32),'0',align='C')
        # else:
        pdf.cell(px2MM(250), px2MM(32),str(tot_record[i]),align='C')
        #//*---Col 3
        pdf.rect(px2MM(706), px2MM(616+(i*52)), px2MM(290), px2MM(52),'FD')
        pdf.set_xy(px2MM(726),px2MM(626+(i*52))) 
        # if active_acc[i] ==0 or active_acc[i] == '':
        #     pdf.cell(px2MM(250), px2MM(32),'0',align='C')
        # else: 
        pdf.cell(px2MM(250), px2MM(32),str(active_acc[i]),align='C')
        #//*---Col 4
        pdf.rect(px2MM(996), px2MM(616+(i*52)), px2MM(290), px2MM(52),'FD')
        pdf.set_xy(px2MM(1016),px2MM(626+(i*52))) 
        # if clsd_acc[i] ==0:
        #     pdf.cell(px2MM(250), px2MM(32),'0',align='C')
        # else:
        pdf.cell(px2MM(250), px2MM(32),str(clsd_acc[i]),align='C')
        #//*---Col 5
        pdf.rect(px2MM(1286), px2MM(616+(i*52)), px2MM(514), px2MM(52),'FD')
        pdf.set_xy(px2MM(1306),px2MM(626+(i*52))) 
        # if acc_neg_hist[i]==0:
        #     pdf.cell(px2MM(474), px2MM(32),'0',align='C')
        # else:
        pdf.cell(px2MM(474), px2MM(32),str(acc_neg_hist[i]),align='C')
        
    
        bl_hight+=52
    
    #//*---Total----*// 
    tot_height = pdf.get_y()   
    pdf.set_fill_color(*hex2RGB('#B9BABE'))
    pdf.set_draw_color(*hex2RGB('#B9BABE'))
    pdf.set_line_width(px2MM(1))
    pdf.rect(px2MM(126), px2MM(mm2PX(tot_height)+43), px2MM(1674), px2MM(1),'FD') 
    
    
    pdf.set_fill_color(*hex2RGB('#ffffff'))
    pdf.set_draw_color(*hex2RGB('#E9EAEE'))
    pdf.set_line_width(px2MM(0.2))
    pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(24))
    pdf.rect(px2MM(126), px2MM(mm2PX(tot_height)+44), px2MM(1674), px2MM(52),'FD') 
    
    pdf.set_text_color(*hex2RGB('#1A1A1D'))
    #//*---Col 1

    # pdf.rect(px2MM(126), px2MM(mm2PX(tot_height)+43), px2MM(290), px2MM(52),'F') 
    pdf.set_xy(px2MM(146),px2MM(mm2PX(tot_height)+53)) 
    pdf.cell(px2MM(290), px2MM(32),'Total',align='L')
    #//*---Col 2
    # pdf.rect(px2MM(416), px2MM(mm2PX(tot_height)+43), px2MM(290), px2MM(52),'F')
    pdf.set_xy(px2MM(416),px2MM(mm2PX(tot_height)+53)) 
    pdf.cell(px2MM(290), px2MM(32),str(total_record),align='C')
    #//*---Col 3
    # pdf.rect(px2MM(706), px2MM(mm2PX(tot_height)+43), px2MM(290), px2MM(52),'F')
    pdf.set_xy(px2MM(706),px2MM(mm2PX(tot_height)+53)) 
    pdf.cell(px2MM(290), px2MM(32),str(total_active_account),align='C')
    #//*---Col 4
    # pdf.rect(px2MM(996), px2MM(mm2PX(tot_height)+43), px2MM(290), px2MM(52),'F')
    pdf.set_xy(px2MM(996),px2MM(mm2PX(tot_height)+53)) 
    pdf.cell(px2MM(290), px2MM(32),str(total_closed_accounts),align='C')
    #//*---Col 5
    # pdf.rect(px2MM(1286), px2MM(mm2PX(tot_height)+43), px2MM(514), px2MM(52),'F')
    pdf.set_xy(px2MM(1286),px2MM(mm2PX(tot_height)+53)) 
    pdf.cell(px2MM(514), px2MM(32),str(total_acc_neg_hist),align='C')
    
    bl_hight+=52
    rem = mm2PX(pdf.get_y())+43
    bl_hight = rem-502
    pdf.set_fill_color(*hex2RGB('#313236'))
    pdf.rect(px2MM(120), px2MM(502), px2MM(6), px2MM(bl_hight),'F')
    
    #//*-----Index Text of Page--**////
    index_text(pdf,'#1A1A1D')
            
#//*-----Liability Management 1---*//
def libility_management_1(pdf,json_data,c_MoneyS,money_signData):
    try:
        aff_check = pd.DataFrame.from_dict(json_data["liability_management"]['table'])
        aff_check_total = json_data["liability_management"]['total']
        aff_comment = json_data["liability_management"]['comments']
    except:
        print(sys.exc_info())
        return None
    
    if aff_check.empty:
        return None
    
    try:
        lib_type = aff_check["liability_type"].tolist()
        outstanding = aff_check['current_liability_distribution_outstanding_percentage'].tolist()
        out_emi = aff_check['current_liability_distribution_emi_percentage'].tolist()
        balance = aff_check['suggested_loan_size_range'].tolist()
        bal_emi = aff_check['suggested_emi_range'].tolist()
        
    except:
        return None
    pdf.add_page()
    pdf.set_fill_color(*hex2RGB('#FCF8ED'))
    pdf.rect(0, 0, px2MM(1920), px2MM(1080),'F')
    
    #//*-----Index Text of Page--**////
    index_text(pdf,'#1A1A1D')

    #//*----Featured List of Financial Products----*//
    pdf.set_xy(px2MM(120),px2MM(80)) 
    pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(60))
    pdf.set_text_color(*hex2RGB('#1A1A1D'))
    pdf.cell(px2MM(877), px2MM(84),'Liability Management',align='L')
    
    #//*---Top Black box
    pdf.set_fill_color(*hex2RGB('#000000'))
    pdf.rect(px2MM(0), px2MM(80), px2MM(15), px2MM(84),'F')    
    
    
    pdf.set_fill_color(*hex2RGB('#313236'))
    pdf.rect(px2MM(126), px2MM(204), px2MM(224), px2MM(42),'F')
    
    pdf.set_xy(px2MM(141),px2MM(209)) 
    pdf.set_font('LeagueSpartan-Medium', size=px2pts(24))
    pdf.set_text_color(*hex2RGB('#ffffff'))
    pdf.cell(px2MM(194), px2MM(32),'Affordability Check',align='C')
    
    bl_height = 146
    #//*------Affordability Check----*//
    #//*---Col 1
    pdf.set_draw_color(*hex2RGB('#E9EAEE'))
    pdf.set_line_width(px2MM(0.2))
    pdf.set_fill_color(*hex2RGB('#ffffff'))
    pdf.rect(px2MM(126), px2MM(246), px2MM(290), px2MM(104),'FD')

    pdf.set_xy(px2MM(146),px2MM(280)) 
    pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(24))
    pdf.set_text_color(*hex2RGB('#1A1A1D'))
    pdf.cell(px2MM(194), px2MM(32),'Liability Type',align='L')

    #//*----Col 1/1-----*//
    pdf.rect(px2MM(416), px2MM(246), px2MM(692), px2MM(52),'FD')
    pdf.set_xy(px2MM(436),px2MM(256)) 
    pdf.cell(px2MM(652), px2MM(32),'Current Liability Distribution',align='C')
    
    #//*----Col 1/1-1-----*//
    pdf.rect(px2MM(416), px2MM(298), px2MM(346), px2MM(52),'FD')
    pdf.set_xy(px2MM(436),px2MM(308)) 
    pdf.cell(px2MM(306), px2MM(32),'Outstanding',align='C')
    
    #//*----Col 1/1-2-----*//
    pdf.rect(px2MM(762), px2MM(298), px2MM(346), px2MM(52),'FD')
    pdf.set_xy(px2MM(782),px2MM(308)) 
    pdf.cell(px2MM(306), px2MM(32),'EMI',align='C')
    
     #//*----Col 2/1-----*//
    pdf.rect(px2MM(1108), px2MM(246), px2MM(692), px2MM(52),'FD')
    pdf.set_xy(px2MM(1128),px2MM(256)) 
    pdf.cell(px2MM(652), px2MM(32),'Suggested Range',align='C')
    
    #//*----Col 2/1-1-----*//
    pdf.rect(px2MM(1108), px2MM(298), px2MM(346), px2MM(52),'FD')
    pdf.set_xy(px2MM(1128),px2MM(308)) 
    pdf.cell(px2MM(306), px2MM(32),'Loan Size',align='C')
    
    #//*----Col 2/1-2-----*//
    pdf.rect(px2MM(1454), px2MM(298), px2MM(346), px2MM(52),'FD')
    pdf.set_xy(px2MM(1474),px2MM(308)) 
    pdf.cell(px2MM(306), px2MM(32),'EMI',align='C')
    
    
    #//*---Table Data---*//
    
    
    rect_y = 350
    text_y = 365
    common_gap = 62
    
    for i in range(len(lib_type)):
        if i%2==0:
            pdf.set_fill_color(*hex2RGB('#F3F6F9'))
        else:
            pdf.set_fill_color(*hex2RGB('#ffffff'))
            
        pdf.set_draw_color(*hex2RGB('#E9EAEE'))
        pdf.set_line_width(px2MM(0.2))
        pdf.rect(px2MM(126), px2MM(rect_y), px2MM(290), px2MM(62),'FD')

        #//*--Col 1---*/
        pdf.set_xy(px2MM(146),px2MM(text_y)) 
        pdf.set_font('LeagueSpartan-Regular', size=px2pts(24))
        pdf.set_text_color(*hex2RGB('#1A1A1D'))
        pdf.cell(px2MM(194), px2MM(32),str(lib_type[i]),align='L')
        
        #//*---Col 2---*/
        pdf.rect(px2MM(416), px2MM(rect_y), px2MM(346), px2MM(62),'FD')
        pdf.set_xy(px2MM(436),px2MM(text_y)) 
        # pdf.cell(px2MM(306), px2MM(32),outstanding[i]+'%',align='C')
        pdf.cell(px2MM(306), px2MM(32),'₹ '+str(format_cash2(float(outstanding[i]))),align='C')
        
        #//*---Col 3---*/
        pdf.rect(px2MM(762), px2MM(rect_y), px2MM(346), px2MM(62),'FD')
        pdf.set_xy(px2MM(782),px2MM(text_y)) 
        
        if int(float(out_emi[i]))==0:
            pdf.cell(px2MM(306), px2MM(32),'₹ 0.0K',align='C')
        else:
            pdf.cell(px2MM(306), px2MM(32),'₹ '+str(format_cash(float(out_emi[i]))),align='C')
        
        #//*---Col 4---*/
        pdf.rect(px2MM(1108), px2MM(rect_y), px2MM(346), px2MM(62),'FD')
        pdf.set_xy(px2MM(1128),px2MM(text_y)) 
        val =  balance[i].split('to')
        val = " to ".join(list(str(format_cash2(float(x))) for x in val))
        # pdf.cell(px2MM(306), px2MM(32),balance[i],align='C')
        pdf.cell(px2MM(306), px2MM(32),val,align='C')
        
        #//*---Col 5---*/
        pdf.rect(px2MM(1454), px2MM(rect_y), px2MM(346), px2MM(62),'FD')
        pdf.set_xy(px2MM(1474),px2MM(text_y)) 
        val =  bal_emi[i].split('to')
        val = " to ".join(list(str(format_cash(float(x))) for x in val))
        # pdf.cell(px2MM(306), px2MM(32),bal_emi[i],align='C')
        pdf.cell(px2MM(306), px2MM(32),val,align='C')
        
        rect_y+=common_gap
        text_y+=common_gap
        bl_height+=common_gap
        
    # locale.setlocale(locale.LC_MONETARY, 'en_IN')
    text_y -= 5    
    #//*---Total-----*//
        
    pdf.set_draw_color(*hex2RGB('#B9BABE'))
    pdf.set_fill_color(*hex2RGB('#B9BABE'))
    pdf.set_line_width(px2MM(1))
    pdf.rect(px2MM(126), px2MM(rect_y), px2MM(1674), px2MM(1),'FD')
    
    pdf.set_draw_color(*hex2RGB('#E9EAEE'))
    pdf.set_line_width(px2MM(0.2))
    pdf.set_fill_color(*hex2RGB('#FFFFFF'))
    pdf.rect(px2MM(126), px2MM(rect_y+1), px2MM(1674), px2MM(52),'FD')

    #//*--Col 1---*/
    pdf.set_xy(px2MM(146),px2MM(text_y)) 
    pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(24))
    pdf.set_text_color(*hex2RGB('#1A1A1D'))
    pdf.cell(px2MM(194), px2MM(32),'Total',align='L')
    
    #//*---Col 2---*/

    pdf.set_xy(px2MM(436),px2MM(text_y)) 
    val1 = str(locale.currency(float(aff_check_total['current_liability_distribution_outstanding_percentage']), grouping=True))
    val1 = val1.split('.')[0]
    val1 = '₹ '+str(format_cash2(float(aff_check_total['current_liability_distribution_outstanding_percentage'])))
    pdf.cell(px2MM(306), px2MM(32),val1,align='C')
    
    #//*---Col 3---*/
    pdf.set_xy(px2MM(782),px2MM(text_y)) 
    if int(float(aff_check_total['current_liability_distribution_emi_percentage']))==0:
        val2 = '₹ 0.0K'
    else:
        val2 = '₹ '+str(format_cash(float(aff_check_total['current_liability_distribution_emi_percentage'])))
    pdf.cell(px2MM(306), px2MM(32),val2,align='C')
    
    #//*---Col 4---*/
    pdf.set_xy(px2MM(1128),px2MM(text_y))
    val =  aff_check_total['suggested_loan_size_range'].split('to')
    val = " to ".join(list(str(format_cash2(float(x))) for x in val))
    pdf.cell(px2MM(306), px2MM(32),val,align='C')
    
    #//*---Col 5---*/
    pdf.set_xy(px2MM(1474),px2MM(text_y)) 
    val =  aff_check_total['suggested_emi_range'].split('to')
    val = " to ".join(list(str(format_cash(float(x))) for x in val))
    pdf.cell(px2MM(306), px2MM(32),val,align='C')
    
    
    try:
        statement = aff_comment 
    except:
        return None 
    
    if statement ==[]:
        return None 
    
    flag = 'True'
    
    for i in statement :
        if i == "" or i == None:
            flag = 'False'
        else:
            flag = 'True'
            break
    
    if flag == 'False':
        return None
    
    
    #//*---Long Black vertical line
    pdf.set_fill_color(*hex2RGB('#313236'))
    pdf.rect(px2MM(120), px2MM(204), px2MM(6), px2MM(bl_height+53),'F')
    
    comment_y = pdf.get_y()
    
    pdf.set_font('LeagueSpartan-Regular', size=px2pts(40))
    pdf.set_text_color(*hex2RGB('#000000'))
    pdf.set_xy(px2MM(120),px2MM(mm2PX(comment_y)+122)) 
    pdf.cell(px2MM(170), px2MM(56),'Comments',align='L')
    
    for_stat = 682
    
    
    for i in range(len(statement)):    
        if statement[i] == "":
            continue   
        pdf.set_fill_color(*hex2RGB('#000000'))
        pdf.rect(px2MM(120), px2MM(for_stat+20), px2MM(10), px2MM(10),'F')
        pdf.set_font('LeagueSpartan-Regular', size=px2pts(30))
        pdf.set_text_color(*hex2RGB('#1A1A1D'))
        pdf.set_xy(px2MM(150),px2MM(for_stat)) 
        pdf.cell(px2MM(1304), px2MM(42),statement[i],align='L')
        
        for_stat+=52
    
    

#//*----Financial Product List (Debt Mutual Funds)-----*//
def debt_mutual_fund(pdf,json_data,c_MoneyS,money_signData):
    
    
    debt = json_data["featured_list"]['debt_mutual_fund']
    
    for key, items in debt.items():

        df = pd.DataFrame.from_dict(json_data["featured_list"]['debt_mutual_fund'][key])
        fund_scheme = list(df['name'])
        investment_horizon = list(df['investment_horizon'])
        strength = list(df['strength'])
        weakness = list(df['weakness'])   
        
        def add_debt_page(pdf,key): 
            pdf.add_page()
            pdf.set_fill_color(*hex2RGB('#FCF8ED'))
            pdf.rect(0, 0, px2MM(1920), px2MM(1080),'F')
            
            # black rectangle
            pdf.set_fill_color(*hex2RGB('#000000'))
            pdf.rect(px2MM(0), px2MM(80), px2MM(15), px2MM(84), 'F')

            #//*----Featured List of Financial Products----*//
            pdf.set_xy(px2MM(120),px2MM(80)) 
            pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(60))
            pdf.set_text_color(*hex2RGB('#1A1A1D'))
            pdf.cell(px2MM(877), px2MM(84),'Financial Products Featured List',align='L')
            
            #//*---Top Black box
            pdf.set_fill_color(*hex2RGB('#313236'))
            pdf.set_font('LeagueSpartan-Medium', size=px2pts(24))
            pdf.set_text_color(*hex2RGB('#ffffff'))
            
            if key == 'liquid_funds':
                val = 'Liquid Funds'
                pdf.rect(px2MM(126), px2MM(204), px2MM(368), px2MM(42),'F')
                pdf.set_xy(px2MM(141),px2MM(209)) 
                pdf.cell(px2MM(343), px2MM(32),"Debt Mutual Funds - "+val,align='L')
            elif key == 'short_term':
                val = 'Short Term'
                pdf.rect(px2MM(126), px2MM(204), px2MM(350), px2MM(42),'F')
                pdf.set_xy(px2MM(141),px2MM(209)) 
                pdf.cell(px2MM(335), px2MM(32),"Debt Mutual Funds - "+val,align='L')
            elif key == 'dynamic_bond':
                val = 'Dynamic Bond'
                pdf.rect(px2MM(126), px2MM(204), px2MM(383), px2MM(42),'F')
                pdf.set_xy(px2MM(141),px2MM(209)) 
                pdf.cell(px2MM(358), px2MM(32),"Debt Mutual Funds - "+val,align='L')
            else:
                val = 'Liquid Funds'
                pdf.rect(px2MM(126), px2MM(204), px2MM(368), px2MM(42),'F')
                pdf.set_xy(px2MM(141),px2MM(209)) 
                pdf.cell(px2MM(343), px2MM(32),"Debt Mutual Funds - "+val,align='L')
                
                
            # //*---Col 1
            pdf.set_line_width(px2MM(0.2))
            pdf.set_fill_color(*hex2RGB('#ffffff'))
            pdf.set_draw_color(*hex2RGB('#E9EAEE'))
            pdf.rect(px2MM(126), px2MM(246), px2MM(558), px2MM(72),'FD')
            
            pdf.set_xy(px2MM(146),px2MM(266)) 
            pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(24))
            pdf.set_text_color(*hex2RGB('#1A1A1D'))
            pdf.multi_cell(px2MM(200), px2MM(32),'Scheme Details',align='L')
            
            #//*---Col 2
            pdf.rect(px2MM(684), px2MM(246), px2MM(558), px2MM(72),'FD')
            pdf.set_xy(px2MM(704),px2MM(266)) 
            pdf.multi_cell(px2MM(230), px2MM(32),'Strength',align='L')
        
            #//*---Col 3
            pdf.rect(px2MM(1242), px2MM(246), px2MM(558), px2MM(72),'FD')
            pdf.set_xy(px2MM(1262),px2MM(266)) 
            pdf.multi_cell(px2MM(524), px2MM(32),'Weakness',align='L')
            
        #//*---Desclaimer Function---*//
        def add_disclaimer(pdf):
            pdf.set_fill_color(*hex2RGB('#FCF8ED'))
            pdf.rect(px2MM(0), px2MM(1006), px2MM(1920), px2MM(40),'F')
            desclaimer = "Disclaimer: All the above schemes are Growth-Direct plans. The above featured list is based on 1 Finance's proprietary research. "
            pdf.set_text_color(*hex2RGB('#000000'))
            pdf.set_font('LeagueSpartan-Regular', size=px2pts(24))
            pdf.set_xy(px2MM(120),px2MM(1008))      
            pdf.multi_cell(px2MM(1680), px2MM(32),desclaimer,align='C')
            
            #//*-----Index Text of Page--**////
            index_text(pdf,'#1A1A1D')
            global fin_feat_product_list
            if fin_feat_product_list == 0:
                fin_feat_product_list = pdf.page_no()
        
        
        #//*---Function to check height of rectangle
        def get_rect_h(rem,strength,weakness):
            pdf.set_text_color(*hex2RGB('#FCF8ED'))
            # pdf.set_text_color(*hex2RGB('#000000'))
            rem_pro = 0
            rem_con = 0
            
            s_text_h = rem
            w_text_h = rem
            
            for j in range(len(strength)):
                pdf.set_font('LeagueSpartan-Regular', size=px2pts(24))
                pdf.set_xy(px2MM(744),px2MM(s_text_h)) 
                pdf.multi_cell(px2MM(478), px2MM(32),strength[j],align='L')
                
                s_text_h = mm2PX(pdf.get_y()) 
            
            rem_pro = mm2PX(pdf.get_y())
            
            for j in range(len(weakness)):
                pdf.set_font('LeagueSpartan-Regular', size=px2pts(24))
                pdf.set_xy(px2MM(1302),px2MM(w_text_h)) 
                pdf.multi_cell(px2MM(478), px2MM(32),weakness[j],align='L')
                w_text_h = mm2PX(pdf.get_y()) 
            
            rem_con = mm2PX(pdf.get_y())
            
            if rem_pro >=rem_con:
                rect_h = rem_pro-rem+20
            else:
                rect_h = rem_con-rem+20
                
            if rem+rect_h +20 > 950:
                pdf.set_fill_color(*hex2RGB('#FCF8ED'))
                pdf.rect(px2MM(126), px2MM(rem+1), px2MM(1674), px2MM(rect_h),'F')
                add_disclaimer(pdf)
                
            return rect_h+20
        
        
        row_color = '#F3F6F9'    
        pdf.set_fill_color(*hex2RGB(row_color))
        add_debt_page(pdf,key)
        rem = mm2PX(pdf.get_y())+20
        
        #//*--Disclaimer--*//
        add_disclaimer(pdf)
        
        
        for i in range(len(df)):
            
            rect_h = get_rect_h(rem,df['strength'].iloc[i],df['weakness'].iloc[i])
            text_y = rem+20
            
            if rem+rect_h > 950:
                add_debt_page(pdf,key)
                rem = mm2PX(pdf.get_y())+20
                #//*---Disclaimer--*//
                add_disclaimer(pdf)
                
                row_color = '#F3F6F9'
                pdf.set_fill_color(*hex2RGB(row_color))
                rect_h = get_rect_h(rem,df['strength'].iloc[i],df['weakness'].iloc[i])
                text_y = rem+20
                
            # print('rect height : ',rect_h)
            
            pdf.set_text_color(*hex2RGB('#65676D'))
            pdf.set_font('LeagueSpartan-Regular', size=px2pts(24))
            pdf.set_line_width(px2MM(0.2))
            #//*---Column 1 Value
            pdf.set_fill_color(*hex2RGB(row_color))
            pdf.rect(px2MM(126), px2MM(rem), px2MM(558), px2MM(rect_h),'FD')
            pdf.set_xy(px2MM(146),px2MM(text_y)) 
            pdf.multi_cell(px2MM(90), px2MM(32),'Name - ',align='L')
            
            pdf.set_text_color(*hex2RGB('#1A1A1D'))
            pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(24))
            pdf.set_xy(px2MM(226),px2MM(text_y)) 
            pdf.multi_cell(px2MM(478), px2MM(32),df['name'].iloc[i],align='L')
            
            inv_hozn = mm2PX(pdf.get_y())
            
            pdf.set_text_color(*hex2RGB('#65676D'))
            pdf.set_font('LeagueSpartan-Regular', size=px2pts(24))
            pdf.set_xy(px2MM(146),px2MM(inv_hozn+12)) 
            pdf.multi_cell(px2MM(220), px2MM(32),'Investment horizon - ',align='L')
            
            pdf.set_text_color(*hex2RGB('#1A1A1D'))
            pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(24)) 
            pdf.set_xy(px2MM(356),px2MM(inv_hozn+12)) 
            pdf.multi_cell(px2MM(348), px2MM(32),df['investment_horizon'].iloc[i],align='L')
            
            #//*----Column 2 Value
            pdf.set_font('LeagueSpartan-Regular', size=px2pts(24))
            pdf.rect(px2MM(684), px2MM(rem), px2MM(558), px2MM(rect_h),'FD')
            
            strength = list(filter(remove_empty_strings, df['strength'].iloc[i]))
            
            for j in range(len(strength)):
                if strength[j]=='':
                    continue
                if j==0:
                    pdf.set_fill_color(*hex2RGB('#000000'))
                    pdf.circle(x=px2MM(724),y=px2MM(text_y+15),r=px2MM(6),style='F')
                    
                    
                    pdf.set_xy(px2MM(744),px2MM(text_y)) 
                    pdf.multi_cell(px2MM(478), px2MM(32),strength[j],align='L')
                    
                else:
                    p_y = mm2PX(pdf.get_y())
                    pdf.set_fill_color(*hex2RGB('#000000'))
                    pdf.circle(x=px2MM(724),y=px2MM(p_y+15),r=px2MM(6),style='F')
            
                    pdf.set_xy(px2MM(744),px2MM(p_y)) 
                    pdf.multi_cell(px2MM(478), px2MM(32),strength[j],align='L')
                    
            #//*----Column 3 Value
            pdf.set_fill_color(*hex2RGB(row_color))
            pdf.rect(px2MM(1242), px2MM(rem), px2MM(558), px2MM(rect_h),'FD')
            
            weakness = list(filter(remove_empty_strings, df['weakness'].iloc[i]))
            for j in range(0,len(weakness)):

                if weakness[j]=='':
                    continue
                if j==0:
                    pdf.set_fill_color(*hex2RGB('#000000'))
                    pdf.circle(x=px2MM(1282),y=px2MM(text_y+15),r=px2MM(6),style='F')
                    
                    pdf.set_font('LeagueSpartan-Regular', size=px2pts(24))
                    pdf.set_xy(px2MM(1302),px2MM(text_y)) 
                    pdf.multi_cell(px2MM(478), px2MM(32),weakness[j],align='L')
                    
                else:
                    p_y = mm2PX(pdf.get_y())
                    pdf.set_fill_color(*hex2RGB('#000000'))
                    pdf.circle(x=px2MM(1282),y=px2MM(p_y+15),r=px2MM(6),style='F')
                    pdf.set_xy(px2MM(1302),px2MM(p_y))      
                    pdf.multi_cell(px2MM(478), px2MM(32),weakness[j],align='L')
                    
            pdf.set_fill_color(*hex2RGB(row_color))  
            
            if row_color == '#F3F6F9':
                row_color = '#FFFFFF'
            else:
                row_color = '#F3F6F9'
                
            rem = rem+rect_h
                
            # //*----Black VerticaL Line
            pdf.set_fill_color(*hex2RGB('#313236'))
            pdf.rect(px2MM(120), px2MM(204), px2MM(6), px2MM(rem-204),'F')
            
            # desclaimer = "Disclaimer: All the above schemes are Growth-Direct plans. The above featured list is based on 1 Finance's proprietary research. "
            # pdf.set_text_color(*hex2RGB('#000000'))
            # pdf.set_font('LeagueSpartan-Light', size=px2pts(24))
            # pdf.set_xy(px2MM(120),px2MM(1008))      
            # pdf.multi_cell(px2MM(1680), px2MM(32),desclaimer,align='C')
                
            
            
            
#//*----Financial Product List (Hybrid Mutual Funds)-----*//
def hybrid_mutual_fund(pdf,json_data,c_MoneyS,money_signData):
    hybrid = json_data["featured_list"]['hybrid_mutual_fund']
    
    for key, items in hybrid.items():

        df = pd.DataFrame.from_dict(json_data["featured_list"]['hybrid_mutual_fund'][key])  
        
        def add_hybrid_page(pdf,key): 
            pdf.add_page()
            pdf.set_fill_color(*hex2RGB('#FCF8ED'))
            pdf.rect(0, 0, px2MM(1920), px2MM(1080),'F')
            
            # black rectangle
            pdf.set_fill_color(*hex2RGB('#000000'))
            pdf.rect(px2MM(0), px2MM(80), px2MM(15), px2MM(84), 'F')

            #//*----Featured List of Financial Products----*//
            pdf.set_xy(px2MM(120),px2MM(80)) 
            pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(60))
            pdf.set_text_color(*hex2RGB('#1A1A1D'))
            pdf.cell(px2MM(877), px2MM(84),'Financial Products Featured List',align='L')
            
            #//*---Top Black box
            pdf.set_fill_color(*hex2RGB('#313236'))
            pdf.set_font('LeagueSpartan-Medium', size=px2pts(24))
            pdf.set_text_color(*hex2RGB('#ffffff'))
            
            if key == 'balanced_advantage':
                val = 'Balanced Advantage'
                pdf.rect(px2MM(126), px2MM(204), px2MM(470), px2MM(42),'F')
                pdf.set_xy(px2MM(141),px2MM(209)) 
                pdf.cell(px2MM(454), px2MM(32),"Hybrid Mutual Funds - "+val,align='L')
            elif key == 'aggressive_hybrid':
                val = 'Aggressive Hybrid'
                pdf.rect(px2MM(126), px2MM(204), px2MM(443), px2MM(42),'F')
                pdf.set_xy(px2MM(141),px2MM(209)) 
                pdf.cell(px2MM(428), px2MM(32),"Hybrid Mutual Funds - "+val,align='L')
            else:
                val = 'Balanced Advantage'
                pdf.rect(px2MM(126), px2MM(204), px2MM(470), px2MM(42),'F')
                pdf.set_xy(px2MM(141),px2MM(209)) 
                pdf.cell(px2MM(454), px2MM(32),"Hybrid Mutual Funds - "+val,align='L')
                
            # //*---Col 1
            
            pdf.set_fill_color(*hex2RGB('#ffffff'))
            pdf.set_draw_color(*hex2RGB('#E9EAEE'))
            pdf.set_line_width(px2MM(0.2))
            pdf.rect(px2MM(126), px2MM(246), px2MM(558), px2MM(72),'FD')
            
            pdf.set_xy(px2MM(146),px2MM(266)) 
            pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(24))
            pdf.set_text_color(*hex2RGB('#1A1A1D'))
            pdf.multi_cell(px2MM(200), px2MM(32),'Scheme Details',align='L')
            
            #//*---Col 2
            pdf.rect(px2MM(684), px2MM(246), px2MM(558), px2MM(72),'FD')
            pdf.set_xy(px2MM(704),px2MM(266)) 
            pdf.multi_cell(px2MM(230), px2MM(32),'Strength',align='L')
        
            #//*---Col 3
            pdf.rect(px2MM(1242), px2MM(246), px2MM(558), px2MM(72),'FD')
            pdf.set_xy(px2MM(1262),px2MM(266)) 
            pdf.multi_cell(px2MM(524), px2MM(32),'Weakness',align='L')
            
        #//*---Desclaimer Function---*//
        def add_disclaimer(pdf):
            pdf.set_fill_color(*hex2RGB('#FCF8ED'))
            pdf.rect(px2MM(0), px2MM(1006), px2MM(1920), px2MM(40),'F')
            desclaimer = "Disclaimer: All the above schemes are Growth-Direct plans. The above featured list is based on 1 Finance's proprietary research. "
            pdf.set_text_color(*hex2RGB('#000000'))
            pdf.set_font('LeagueSpartan-Regular', size=px2pts(24))
            pdf.set_xy(px2MM(120),px2MM(1008))      
            pdf.multi_cell(px2MM(1680), px2MM(32),desclaimer,align='C')
            
            #//*-----Index Text of Page--**////
            index_text(pdf,'#1A1A1D')
        
        
        #//*---Function to check height of rectangle
        def get_rect_h(rem,strength,weakness):
            pdf.set_text_color(*hex2RGB('#FCF8ED'))
            # pdf.set_text_color(*hex2RGB('#000000'))
            rem_pro = 0
            rem_con = 0
            
            s_text_h = rem
            w_text_h = rem
            
            for j in range(len(strength)):
                pdf.set_font('LeagueSpartan-Regular', size=px2pts(24))
                pdf.set_xy(px2MM(744),px2MM(s_text_h)) 
                pdf.multi_cell(px2MM(478), px2MM(32),strength[j],align='L')
                
                s_text_h = mm2PX(pdf.get_y()) 
            
            rem_pro = mm2PX(pdf.get_y())
            
            for j in range(len(weakness)):
                pdf.set_font('LeagueSpartan-Regular', size=px2pts(24))
                pdf.set_xy(px2MM(1302),px2MM(w_text_h)) 
                pdf.multi_cell(px2MM(478), px2MM(32),weakness[j],align='L')
                w_text_h = mm2PX(pdf.get_y()) 
            
            rem_con = mm2PX(pdf.get_y())
            
            if rem_pro >=rem_con:
                rect_h = rem_pro-rem+20
            else:
                rect_h = rem_con-rem+20
                
            if rem+rect_h +20 > 950:
                pdf.set_fill_color(*hex2RGB('#FCF8ED'))
                pdf.rect(px2MM(126), px2MM(rem+2), px2MM(1674), px2MM(rect_h),'F')
                add_disclaimer(pdf)
                
            return rect_h+20
        
        
        row_color = '#F3F6F9'    
        pdf.set_fill_color(*hex2RGB(row_color))
        add_hybrid_page(pdf,key)
        rem = mm2PX(pdf.get_y())+20
        
        #//*---Disclaimer---*/
        add_disclaimer(pdf)
        
        
        for i in range(len(df)):
            
            rect_h = get_rect_h(rem,df['strength'].iloc[i],df['weakness'].iloc[i])
            text_y = rem+20
            
            if rem+rect_h > 950:
                add_hybrid_page(pdf,key)
                rem = mm2PX(pdf.get_y())+20
                add_disclaimer(pdf)
                
                row_color = '#F3F6F9'
                pdf.set_fill_color(*hex2RGB(row_color))
                rect_h = get_rect_h(rem,df['strength'].iloc[i],df['weakness'].iloc[i])
                text_y = rem+20
                
            # print('rect height : ',rect_h)
            
            pdf.set_text_color(*hex2RGB('#65676D'))
            pdf.set_font('LeagueSpartan-Regular', size=px2pts(24))
            pdf.set_line_width(px2MM(0.2))
            #//*---Column 1 Value
            pdf.set_fill_color(*hex2RGB(row_color))
            pdf.rect(px2MM(126), px2MM(rem), px2MM(558), px2MM(rect_h),'FD')
            pdf.set_xy(px2MM(146),px2MM(text_y)) 
            pdf.multi_cell(px2MM(90), px2MM(32),'Name - ',align='L')
            
            pdf.set_text_color(*hex2RGB('#1A1A1D'))
            pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(24))
            pdf.set_xy(px2MM(226),px2MM(text_y)) 
            pdf.multi_cell(px2MM(478), px2MM(32),df['name'].iloc[i],align='L')
            
            inv_hozn = mm2PX(pdf.get_y())
            
            pdf.set_text_color(*hex2RGB('#65676D'))
            pdf.set_font('LeagueSpartan-Regular', size=px2pts(24))
            pdf.set_xy(px2MM(146),px2MM(inv_hozn+12)) 
            pdf.multi_cell(px2MM(220), px2MM(32),'Investment horizon - ',align='L')
            
            pdf.set_text_color(*hex2RGB('#1A1A1D'))
            pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(24)) 
            pdf.set_xy(px2MM(356),px2MM(inv_hozn+12)) 
            pdf.multi_cell(px2MM(348), px2MM(32),df['investment_horizon'].iloc[i],align='L')
            
            #//*----Column 2 Value
            pdf.set_font('LeagueSpartan-Regular', size=px2pts(24))
            pdf.rect(px2MM(684), px2MM(rem), px2MM(558), px2MM(rect_h),'FD')
            
            strength = list(filter(remove_empty_strings, df['strength'].iloc[i]))
            
            for j in range(len(strength)):
                if strength[j]=='':
                    continue
                if j==0:
                    pdf.set_fill_color(*hex2RGB('#000000'))
                    pdf.circle(x=px2MM(724),y=px2MM(text_y+15),r=px2MM(6),style='F')
                    
                    
                    pdf.set_xy(px2MM(744),px2MM(text_y)) 
                    pdf.multi_cell(px2MM(478), px2MM(32),strength[j],align='L')
                    
                else:
                    p_y = mm2PX(pdf.get_y())
                    pdf.set_fill_color(*hex2RGB('#000000'))
                    pdf.circle(x=px2MM(724),y=px2MM(p_y+15),r=px2MM(6),style='F')
            
                    pdf.set_xy(px2MM(744),px2MM(p_y)) 
                    pdf.multi_cell(px2MM(478), px2MM(32),strength[j],align='L')
                    
            #//*----Column 3 Value
            pdf.set_fill_color(*hex2RGB(row_color))
            pdf.rect(px2MM(1242), px2MM(rem), px2MM(558), px2MM(rect_h),'FD')
            
            weakness = list(filter(remove_empty_strings, df['weakness'].iloc[i]))
            for j in range(0,len(weakness)):

                if weakness[j]=='':
                    continue
                if j==0:
                    pdf.set_fill_color(*hex2RGB('#000000'))
                    pdf.circle(x=px2MM(1282),y=px2MM(text_y+15),r=px2MM(6),style='F')
                    
                    pdf.set_font('LeagueSpartan-Regular', size=px2pts(24))
                    pdf.set_xy(px2MM(1302),px2MM(text_y)) 
                    pdf.multi_cell(px2MM(478), px2MM(32),weakness[j],align='L')
                    
                else:
                    p_y = mm2PX(pdf.get_y())
                    pdf.set_fill_color(*hex2RGB('#000000'))
                    pdf.circle(x=px2MM(1282),y=px2MM(p_y+15),r=px2MM(6),style='F')
                    pdf.set_xy(px2MM(1302),px2MM(p_y))      
                    pdf.multi_cell(px2MM(478), px2MM(32),weakness[j],align='L')
                    
            pdf.set_fill_color(*hex2RGB(row_color))  
            
            if row_color == '#F3F6F9':
                row_color = '#FFFFFF'
            else:
                row_color = '#F3F6F9'
                
            rem = rem+rect_h
                
            # //*----Black VerticaL Line
            pdf.set_fill_color(*hex2RGB('#313236'))
            pdf.rect(px2MM(120), px2MM(204), px2MM(6), px2MM(rem-204),'F')
            
            # desclaimer = "Disclaimer: All the above schemes are Growth-Direct plans. The above featured list is based on 1 Finance's proprietary research. "
            # pdf.set_text_color(*hex2RGB('#000000'))
            # pdf.set_font('LeagueSpartan-Light', size=px2pts(24))
            # pdf.set_xy(px2MM(120),px2MM(1008))      
            # pdf.multi_cell(px2MM(1680), px2MM(32),desclaimer,align='C')
                
            
            
#//*----Financial Product List (Credit Cards)-----*//
def credit_card(pdf,json_data,c_MoneyS,money_signData):
    
    df = pd.DataFrame.from_dict(json_data["featured_list"]['credit_card'])
    
    def add_credit_card_page(pdf): 
        pdf.add_page()
        pdf.set_fill_color(*hex2RGB('#FCF8ED'))
        pdf.rect(0, 0, px2MM(1920), px2MM(1080),'F')
        
        # black rectangle
        pdf.set_fill_color(*hex2RGB('#000000'))
        pdf.rect(px2MM(0), px2MM(80), px2MM(15), px2MM(84), 'F')

        #//*----Featured List of Financial Products----*//
        pdf.set_xy(px2MM(120),px2MM(80)) 
        pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(60))
        pdf.set_text_color(*hex2RGB('#1A1A1D'))
        pdf.cell(px2MM(877), px2MM(84),'Financial Products Featured List',align='L')
        
        #//*---Top Black box
        pdf.set_fill_color(*hex2RGB('#313236'))
        pdf.set_font('LeagueSpartan-Medium', size=px2pts(24))
        pdf.set_text_color(*hex2RGB('#ffffff'))

        pdf.rect(px2MM(126), px2MM(204), px2MM(158), px2MM(42),'F')
        pdf.set_xy(px2MM(141),px2MM(209)) 
        pdf.cell(px2MM(128), px2MM(32),"Card Details",align='L')
            
        # //*---Col 1
        
        pdf.set_fill_color(*hex2RGB('#ffffff'))
        pdf.set_draw_color(*hex2RGB('#E9EAEE'))
        pdf.set_line_width(px2MM(0.2))
        pdf.rect(px2MM(126), px2MM(246), px2MM(558), px2MM(72),'FD')
        
        pdf.set_xy(px2MM(146),px2MM(266)) 
        pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(24))
        pdf.set_text_color(*hex2RGB('#1A1A1D'))
        pdf.multi_cell(px2MM(200), px2MM(32),'Card Details',align='L')
        
        #//*---Col 2
        pdf.rect(px2MM(684), px2MM(246), px2MM(558), px2MM(72),'FD')
        pdf.set_xy(px2MM(704),px2MM(266)) 
        pdf.multi_cell(px2MM(230), px2MM(32),'Strength',align='L')
    
        #//*---Col 3
        pdf.rect(px2MM(1242), px2MM(246), px2MM(558), px2MM(72),'FD')
        pdf.set_xy(px2MM(1262),px2MM(266)) 
        pdf.multi_cell(px2MM(524), px2MM(32),'Weakness',align='L')
        
    #//*---Desclaimer Function---*//
    def add_disclaimer(pdf):
        pdf.set_fill_color(*hex2RGB('#FCF8ED'))
        pdf.rect(px2MM(120), px2MM(1006), px2MM(1690), px2MM(40),'F')
        desclaimer = "Disclaimer: The above featured list is based on 1 Finance's proprietary research. "
        pdf.set_text_color(*hex2RGB('#000000'))
        pdf.set_font('LeagueSpartan-Regular', size=px2pts(24))
        pdf.set_xy(px2MM(120),px2MM(1008))      
        pdf.multi_cell(px2MM(1680), px2MM(32),desclaimer,align='C')
    
    
    #//*---Function to check height of rectangle
    def get_rect_h(rem,df,i):
        pdf.set_text_color(*hex2RGB('#FCF8ED'))
        # pdf.set_text_color(*hex2RGB('#000000'))
        rem_det = 0
        rem_pro = 0
        rem_con = 0
        
        det_text_h = rem+20
        s_text_h = rem+20
        w_text_h = rem+20
        
        strength,weakness = df['strength'].iloc[i],df['weakness'].iloc[i]
        
        pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(24))
        pdf.set_text_color(*hex2RGB('#1A1A1D'))
        pdf.set_xy(px2MM(146),px2MM(det_text_h)) 
        pdf.multi_cell(px2MM(518), px2MM(32),df['card_name'].iloc[i],align='L')
        
        det_y = mm2PX(pdf.get_y())
        pdf.set_font('LeagueSpartan-Regular', size=px2pts(24))
        pdf.set_text_color(*hex2RGB('#65676D'))
        pdf.set_xy(px2MM(146),px2MM(det_y)) 
        pdf.multi_cell(px2MM(518), px2MM(32),df['card_type'].iloc[i],align='L')
        
        det_y = mm2PX(pdf.get_y())+36
        pdf.set_font('LeagueSpartan-Regular', size=px2pts(24))
        pdf.set_text_color(*hex2RGB('#65676D'))
        pdf.set_xy(px2MM(146),px2MM(det_y)) 
        pdf.multi_cell(px2MM(518), px2MM(32),'Annual Fee - ',align='L')
        
        pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(24))
        pdf.set_text_color(*hex2RGB('#1A1A1D'))
        pdf.set_xy(px2MM(280),px2MM(det_y)) 
        pdf.multi_cell(px2MM(390), px2MM(32),df['annual_fee'].iloc[i],border='0',align='L')
        
        det_y = mm2PX(pdf.get_y())+16
        pdf.set_font('LeagueSpartan-Regular', size=px2pts(24))
        pdf.set_text_color(*hex2RGB('#65676D'))
        pdf.set_xy(px2MM(146),px2MM(det_y)) 
        pdf.multi_cell(px2MM(518), px2MM(32),'Eligibility -',align='L')
        
        pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(24))
        pdf.set_text_color(*hex2RGB('#1A1A1D'))
        pdf.set_xy(px2MM(260),px2MM(det_y)) 
        eligibility = list(df['eligibility'].iloc[i].split(' '))
        elg_text = ''
        elg_rem = []
        for k in range(len(eligibility)):
            elg_text = elg_text+' '+eligibility[k]
            if mm2PX(pdf.get_string_width(elg_text)) > 410:
                elg_text = elg_text.replace(eligibility[k],'')
                elg_rem = eligibility[k:]
                break
        pdf.multi_cell(px2MM(420), px2MM(32),elg_text,align='L')
        if not elg_rem ==[]:
            pdf.set_xy(px2MM(146),px2MM(det_y+32)) 
            pdf.multi_cell(px2MM(518), px2MM(32),' '.join(elg_rem),align='L')
        
        det_y = mm2PX(pdf.get_y())+16
        pdf.set_font('LeagueSpartan-Regular', size=px2pts(24))
        pdf.set_text_color(*hex2RGB('#65676D'))
        pdf.set_xy(px2MM(146),px2MM(det_y)) 
        pdf.multi_cell(px2MM(518), px2MM(32),'Best suited for -',align='L')
        
        pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(24))
        pdf.set_text_color(*hex2RGB('#1A1A1D'))
        best_suited_for = list(df['best_suited_for'].iloc[i].split(' '))
        suited_text = ''
        suited_rem = []
        for k in range(len(best_suited_for)):
            suited_text = suited_text+' '+best_suited_for[k]
            if mm2PX(pdf.get_string_width(suited_text)) > 300:
                suited_text = suited_text.replace(best_suited_for[k],'')
                suited_rem = best_suited_for[k:]
                break
        pdf.set_xy(px2MM(310),px2MM(det_y)) 
        # pdf.multi_cell(px2MM(350), px2MM(32),df['best_suited_for'].iloc[i],align='L')
        pdf.multi_cell(px2MM(350), px2MM(32),suited_text,align='L')
        
        if not suited_rem == []:
            pdf.set_xy(px2MM(146),px2MM(det_y+32)) 
            pdf.multi_cell(px2MM(518), px2MM(32),' '.join(suited_rem),align='L')
            
        
        det_y = mm2PX(pdf.get_y())+16
        pdf.set_font('LeagueSpartan-Regular', size=px2pts(24))
        pdf.set_text_color(*hex2RGB('#65676D'))
        pdf.set_xy(px2MM(146),px2MM(det_y)) 
        pdf.multi_cell(px2MM(450), px2MM(32),'Best reward points (RP) conversion rate -',align='L')
        
        pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(24))
        pdf.set_text_color(*hex2RGB('#1A1A1D'))
        pdf.set_xy(px2MM(555),px2MM(det_y)) 
        pdf.multi_cell(px2MM(125), px2MM(32),df['best_reward_points_conversion_rate'].iloc[i],border='0',align='L')
    
        
        rem_det = mm2PX(pdf.get_y())-rem+20
        for j in range(len(strength)):
            pdf.set_font('LeagueSpartan-Regular', size=px2pts(24))
            pdf.set_xy(px2MM(744),px2MM(s_text_h)) 
            pdf.multi_cell(px2MM(478), px2MM(32),strength[j],align='L')
            
            s_text_h = mm2PX(pdf.get_y()) 
        
        rem_pro = mm2PX(pdf.get_y())-rem
        
        for j in range(len(weakness)):
            pdf.set_font('LeagueSpartan-Regular', size=px2pts(24))
            pdf.set_xy(px2MM(1302),px2MM(w_text_h)) 
            pdf.multi_cell(px2MM(478), px2MM(32),weakness[j],align='L')
            w_text_h = mm2PX(pdf.get_y()) 
        
        rem_con = mm2PX(pdf.get_y())-rem
        h_list = [rem_det,rem_pro,rem_con]
        rect_h = max(h_list)
        
        if rem+rect_h +20 > 950:
            pdf.set_fill_color(*hex2RGB('#FCF8ED'))
            pdf.rect(px2MM(126), px2MM(rem+2), px2MM(1674), px2MM(rect_h),'F')
            add_disclaimer(pdf)
        
        
            
        return rect_h+20
    
    for i in range(len(df)):
        
        row_color = '#F3F6F9'    
        pdf.set_fill_color(*hex2RGB(row_color))
        add_credit_card_page(pdf)
        rem = mm2PX(pdf.get_y())+20
        
        add_disclaimer(pdf)
        
        rect_h = get_rect_h(rem,df,i)
        text_y = rem+20
        
        pdf.set_text_color(*hex2RGB('#65676D'))
        pdf.set_font('LeagueSpartan-Regular', size=px2pts(24))
        pdf.set_line_width(px2MM(0.2))
        #//*---Column 1 Value

        pdf.set_fill_color(*hex2RGB(row_color))
        pdf.rect(px2MM(126), px2MM(rem), px2MM(558), px2MM(rect_h),'FD')
        
        pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(24))
        pdf.set_text_color(*hex2RGB('#1A1A1D'))
        pdf.set_xy(px2MM(146),px2MM(rem+20)) 
        pdf.multi_cell(px2MM(518), px2MM(32),df['card_name'].iloc[i],align='L')
        
        det_y = mm2PX(pdf.get_y())
        pdf.set_font('LeagueSpartan-Regular', size=px2pts(24))
        pdf.set_text_color(*hex2RGB('#65676D'))
        pdf.set_xy(px2MM(146),px2MM(det_y)) 
        pdf.multi_cell(px2MM(518), px2MM(32),df['card_type'].iloc[i],align='L')
        
        det_y = mm2PX(pdf.get_y())+36
        pdf.set_font('LeagueSpartan-Regular', size=px2pts(24))
        pdf.set_text_color(*hex2RGB('#65676D'))
        pdf.set_xy(px2MM(146),px2MM(det_y)) 
        pdf.multi_cell(px2MM(518), px2MM(32),'Annual Fee - ',align='L')
        
        pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(24))
        pdf.set_text_color(*hex2RGB('#1A1A1D'))
        pdf.set_xy(px2MM(280),px2MM(det_y)) 
        pdf.multi_cell(px2MM(390), px2MM(32),df['annual_fee'].iloc[i],border='0',align='L')
        
        det_y = mm2PX(pdf.get_y())+16
        pdf.set_font('LeagueSpartan-Regular', size=px2pts(24))
        pdf.set_text_color(*hex2RGB('#65676D'))
        pdf.set_xy(px2MM(146),px2MM(det_y)) 
        pdf.multi_cell(px2MM(518), px2MM(32),'Eligibility -',align='L')
        
        pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(24))
        pdf.set_text_color(*hex2RGB('#1A1A1D'))
        pdf.set_xy(px2MM(260),px2MM(det_y)) 
        eligibility = list(df['eligibility'].iloc[i].split(' '))
        elg_text = ''
        elg_rem = []
        for k in range(len(eligibility)):
            elg_text = elg_text+' '+eligibility[k]
            if mm2PX(pdf.get_string_width(elg_text)) > 410:
                elg_text = elg_text.replace(eligibility[k],'')
                elg_rem = eligibility[k:]
                break
        pdf.multi_cell(px2MM(420), px2MM(32),elg_text,align='L')
        if not elg_rem ==[]:
            pdf.set_xy(px2MM(146),px2MM(det_y+32)) 
            pdf.multi_cell(px2MM(518), px2MM(32),' '.join(elg_rem),align='L')
        
        det_y = mm2PX(pdf.get_y())+16
        pdf.set_font('LeagueSpartan-Regular', size=px2pts(24))
        pdf.set_text_color(*hex2RGB('#65676D'))
        pdf.set_xy(px2MM(146),px2MM(det_y)) 
        pdf.multi_cell(px2MM(518), px2MM(32),'Best suited for -',align='L')
        
        pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(24))
        pdf.set_text_color(*hex2RGB('#1A1A1D'))
        best_suited_for = list(df['best_suited_for'].iloc[i].split(' '))
        suited_text = ''
        suited_rem = []
        for k in range(len(best_suited_for)):
            suited_text = suited_text+' '+best_suited_for[k]
            if mm2PX(pdf.get_string_width(suited_text)) > 340:
                suited_text = suited_text.replace(best_suited_for[k],'')
                suited_rem = best_suited_for[k:]
                break
        pdf.set_xy(px2MM(310),px2MM(det_y)) 
        # pdf.multi_cell(px2MM(350), px2MM(32),df['best_suited_for'].iloc[i],align='L')
        pdf.multi_cell(px2MM(350), px2MM(32),suited_text,align='L')
        
        if not suited_rem == []:
            pdf.set_xy(px2MM(146),px2MM(det_y+32)) 
            pdf.multi_cell(px2MM(518), px2MM(32),' '.join(suited_rem),align='L')
            
        
        det_y = mm2PX(pdf.get_y())+16
        pdf.set_font('LeagueSpartan-Regular', size=px2pts(24))
        pdf.set_text_color(*hex2RGB('#65676D'))
        pdf.set_xy(px2MM(146),px2MM(det_y)) 
        pdf.multi_cell(px2MM(450), px2MM(32),'Best reward points (RP) conversion rate -',align='L')
        
        pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(24))
        pdf.set_text_color(*hex2RGB('#1A1A1D'))
        pdf.set_xy(px2MM(555),px2MM(det_y)) 
        pdf.multi_cell(px2MM(125), px2MM(32),df['best_reward_points_conversion_rate'].iloc[i],border='0',align='L')
    
        
        #//*----Column 2 Value
        pdf.set_font('LeagueSpartan-Regular', size=px2pts(24))
        pdf.rect(px2MM(684), px2MM(rem), px2MM(558), px2MM(rect_h),'FD')
        
        strength = list(filter(remove_empty_strings, df['strength'].iloc[i]))
        for j in range(len(strength)):
            if strength[j]=='':
                continue
            if j==0:
                pdf.set_fill_color(*hex2RGB('#000000'))
                pdf.circle(x=px2MM(724),y=px2MM(text_y+15),r=px2MM(6),style='F')
                
                
                pdf.set_xy(px2MM(744),px2MM(text_y)) 
                pdf.multi_cell(px2MM(478), px2MM(32),strength[j],align='L')
                
            else:
                p_y = mm2PX(pdf.get_y())
                pdf.set_fill_color(*hex2RGB('#000000'))
                pdf.circle(x=px2MM(724),y=px2MM(p_y+15),r=px2MM(6),style='F')
        
                pdf.set_xy(px2MM(744),px2MM(p_y)) 
                pdf.multi_cell(px2MM(478), px2MM(32),strength[j],align='L')
                
        #//*----Column 3 Value
        pdf.set_fill_color(*hex2RGB(row_color))
        pdf.rect(px2MM(1242), px2MM(rem), px2MM(558), px2MM(rect_h),'FD')
        weakness = list(filter(remove_empty_strings, df['weakness'].iloc[i]))
        for j in range(0,len(weakness)):

            if weakness[j]=='':
                continue
            if j==0:
                pdf.set_fill_color(*hex2RGB('#000000'))
                pdf.circle(x=px2MM(1282),y=px2MM(text_y+15),r=px2MM(6),style='F')
                
                pdf.set_font('LeagueSpartan-Regular', size=px2pts(24))
                pdf.set_xy(px2MM(1302),px2MM(text_y)) 
                pdf.multi_cell(px2MM(478), px2MM(32),weakness[j],align='L')
                
            else:
                p_y = mm2PX(pdf.get_y())
                pdf.set_fill_color(*hex2RGB('#000000'))
                pdf.circle(x=px2MM(1282),y=px2MM(p_y+15),r=px2MM(6),style='F')
                pdf.set_xy(px2MM(1302),px2MM(p_y))      
                pdf.multi_cell(px2MM(478), px2MM(32),weakness[j],align='L')
                
        pdf.set_fill_color(*hex2RGB(row_color))  
            
        # //*----Black VerticaL Line
        pdf.set_fill_color(*hex2RGB('#313236'))
        pdf.rect(px2MM(120), px2MM(204), px2MM(6), px2MM(rect_h+114),'F')
        
        # desclaimer = "Disclaimer: The above featured list is based on 1 Finance's proprietary research.  "
        # pdf.set_text_color(*hex2RGB('#000000'))
        # pdf.set_font('LeagueSpartan-Regular', size=px2pts(24))
        # pdf.set_xy(px2MM(110),px2MM(1008))      
        # pdf.multi_cell(px2MM(1680), px2MM(32),desclaimer,align='C')
            
        #//*-----Index Text of Page--**////
        index_text(pdf,'#1A1A1D')

def aval_tax_deduct_1(pdf,json_data,c_MoneyS,money_signData):
    pdf.add_page()
    pdf.set_fill_color(*hex2RGB('#FCF8ED'))
    pdf.rect(0, 0, px2MM(1920), px2MM(1080),'F')

    #//*----Available Tax Deductions----*//
    pdf.set_xy(px2MM(120),px2MM(80)) 
    pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(60))
    pdf.set_text_color(*hex2RGB('#000000'))
    pdf.cell(px2MM(641), px2MM(84),'Available Tax Deductions',align='L')
    
    pdf.set_xy(px2MM(791),px2MM(106)) 
    pdf.set_font('LeagueSpartan-Regular', size=px2pts(24))
    pdf.set_text_color(*hex2RGB('#000000'))
    pdf.cell(px2MM(237), px2MM(32),'(as per Old Tax Regime)',align='L')
    
    
    #//*----Content----*//
    bx_height = [72,296,136,200]
    bx_x =[204,276,572,708]
    
    #//*---Table rectangle
    for i in range(4):
        pdf.set_draw_color(*hex2RGB('#E9EAEE'))
        pdf.set_line_width(px2MM(0.5))
        if i%2==0:
            pdf.set_fill_color(*hex2RGB('#FFFFFF'))
        else:
            pdf.set_fill_color(*hex2RGB('#F3F6F9'))
        pdf.rect(px2MM(120), px2MM(bx_x[i]), px2MM(150), px2MM(bx_height[i]),'FD')
        pdf.rect(px2MM(270), px2MM(bx_x[i]), px2MM(885), px2MM(bx_height[i]),'FD')
        pdf.rect(px2MM(1155), px2MM(bx_x[i]), px2MM(345), px2MM(bx_height[i]),'FD')
        pdf.rect(px2MM(1500), px2MM(bx_x[i]), px2MM(300), px2MM(bx_height[i]),'FD')
        
    #//*----Table heading---*//
    pdf.set_text_color(*hex2RGB('#1A1A1D'))
    pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(24))
    
    pdf.set_xy(px2MM(140),px2MM(224)) 
    pdf.cell(px2MM(110), px2MM(32),'Section',align='L')
    
    pdf.set_xy(px2MM(290),px2MM(224)) 
    pdf.cell(px2MM(845), px2MM(32),'Income Tax Deduction on',align='L')
    
    pdf.set_xy(px2MM(1175),px2MM(224)) 
    pdf.cell(px2MM(305), px2MM(32),'Allowed Limit',align='L')
    
    pdf.set_xy(px2MM(1520),px2MM(224)) 
    pdf.cell(px2MM(260), px2MM(32),'Applicable For',align='L')
    
    #//*---COL 1 VALUE---*//
    
    pdf.set_text_color(*hex2RGB('#1A1A1D'))
    pdf.set_font('LeagueSpartan-Regular', size=px2pts(24))
    
    pdf.set_xy(px2MM(140),px2MM(296)) 
    pdf.cell(px2MM(110), px2MM(32),'80C',align='L')
    
    pdf.set_xy(px2MM(140),px2MM(592)) 
    pdf.cell(px2MM(110), px2MM(32),'80CCC',align='L')
    
    pdf.set_xy(px2MM(140),px2MM(728)) 
    pdf.cell(px2MM(110), px2MM(32),'80CCD(1)',align='L')
    
    #//*---COL 2 VALUE---*//

    cir_x = [306,407,439,471,503,535,607,838,870]
    pdf.set_fill_color(*hex2RGB('#000000'))
    for i in range(9):
        pdf.circle(x=px2MM(310), y=px2MM(cir_x[i]), r=px2MM(5), style='F')
        
    pdf.set_xy(px2MM(330),px2MM(296)) 
    pdf.multi_cell(px2MM(815), px2MM(32),'''Investment in PPF, National Saving Certificate (NSC), Sukanya Samriddhi Yojana (SSY), ULIP, ELSS, 5-year tax-saving FD, Senior Citizen Savings Scheme (SCSS), infrastructure bonds''',align='L')
    pdf.set_xy(px2MM(330),px2MM(392)) 
    pdf.multi_cell(px2MM(815), px2MM(32),'''Employee’s share of PF contribution''',align='L')    
    pdf.set_xy(px2MM(330),px2MM(424)) 
    pdf.multi_cell(px2MM(815), px2MM(32),'''Life Insurance premium payment''',align='L')
    pdf.set_xy(px2MM(330),px2MM(456)) 
    pdf.multi_cell(px2MM(815), px2MM(32),'''Children’s Tuition Fee''',align='L')
    pdf.set_xy(px2MM(330),px2MM(488)) 
    pdf.multi_cell(px2MM(815), px2MM(32),'''Principal repayment of home loan''',align='L') 
    pdf.set_xy(px2MM(330),px2MM(520)) 
    pdf.multi_cell(px2MM(815), px2MM(32),'''Stamp duty and registration charges for purchase of property.''',align='L')
    
    pdf.set_xy(px2MM(330),px2MM(592)) 
    pdf.multi_cell(px2MM(815), px2MM(32),'''For LIC or other insurer pension annuity plan deposits from Section 10 funds (23AAB)''',align='L')
    
    pdf.set_xy(px2MM(290),px2MM(728)) 
    pdf.multi_cell(px2MM(845), px2MM(32),'''Employee contribution under section 80CCD(1) towards National Pension Scheme (NPS) account or the Atal Pension Yojana (APY) account. Maximum deduction is the lesser of:''',align='L')
    pdf.set_xy(px2MM(330),px2MM(824)) 
    pdf.multi_cell(px2MM(815), px2MM(32),'''10% of salary (for employees)''',align='L')
    pdf.set_xy(px2MM(330),px2MM(856)) 
    pdf.multi_cell(px2MM(815), px2MM(32),'''20% of gross total income (for self-employed)''',align='L')
   
    
    #//*----Column 3 Value---*//
    
    pdf.set_xy(px2MM(1175),px2MM(296)) 
    pdf.cell(px2MM(305), px2MM(32),'Rs 1.5L ',align='L')
    pdf.set_xy(px2MM(1175),px2MM(296+32)) 
    pdf.multi_cell(px2MM(315), px2MM(32),'(aggregate of sections 80CCD, 80CCC, 80C)',align='L')
    
    pdf.set_xy(px2MM(1175),px2MM(592)) 
    pdf.cell(px2MM(305), px2MM(32),'Rs 1.5L',align='L')
    pdf.set_xy(px2MM(1175),px2MM(592+32)) 
    pdf.multi_cell(px2MM(315), px2MM(32),'(aggregate of sections 80CCD, 80CCC, 80C)',align='L')
    
    
    pdf.set_xy(px2MM(1175),px2MM(728)) 
    pdf.cell(px2MM(305), px2MM(32),'Rs 1.5L',align='L')
    pdf.set_xy(px2MM(1175),px2MM(728+32)) 
    pdf.multi_cell(px2MM(315), px2MM(32),'(aggregate of sections 80CCD, 80CCC, 80C)',align='L')
    
    
    #//*---COL 4 VALUE---*//
    
    pdf.set_text_color(*hex2RGB('#1A1A1D'))
    pdf.set_font('LeagueSpartan-Regular', size=px2pts(24))
    
    pdf.set_xy(px2MM(1520),px2MM(296)) 
    pdf.multi_cell(px2MM(260), px2MM(32),'Individuals, Hindu Undivided Families',align='L')
    
    pdf.set_xy(px2MM(1520),px2MM(592)) 
    pdf.cell(px2MM(260), px2MM(32),'Individuals',align='L')
    
    pdf.set_xy(px2MM(1520),px2MM(728)) 
    pdf.cell(px2MM(260), px2MM(32),'Individuals',align='L')
    
    #//*-----Index Text of Page--**////
    index_text(pdf,'#1A1A1D')
    

#//*----Available Tax Deductions(Page 2)----*//

def aval_tax_deduct_2(pdf,json_data,c_MoneyS,money_signData):
    pdf.add_page()
    pdf.set_fill_color(*hex2RGB('#FCF8ED'))
    pdf.rect(0, 0, px2MM(1920), px2MM(1080),'F')

    #//*----Available Tax Deductions----*//
    pdf.set_xy(px2MM(120),px2MM(80)) 
    pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(60))
    pdf.set_text_color(*hex2RGB('#000000'))
    pdf.cell(px2MM(641), px2MM(84),'Available Tax Deductions',align='L')
    
    pdf.set_xy(px2MM(791),px2MM(106)) 
    pdf.set_font('LeagueSpartan-Regular', size=px2pts(24))
    pdf.set_text_color(*hex2RGB('#000000'))
    pdf.cell(px2MM(237), px2MM(32),'(as per Old Tax Regime)',align='L')
    
    
    #//*----Content----*//
    # bx_height = [72,72,72,72,104,200,104,72]
    # bx_x =[204,276,348,420,492,596,796,900]
    
    bx_height = [72,200,72,104,168,168]
    bx_x =[204,276,476,548,652,820]
    
    #//*---Table rectangle
    for i in range(6):
        pdf.set_line_width(px2MM(0.5))
        pdf.set_draw_color(*hex2RGB('#E9EAEE'))
        if i%2==0:
            pdf.set_fill_color(*hex2RGB('#FFFFFF'))
        else:
            pdf.set_fill_color(*hex2RGB('#F3F6F9'))
            
        pdf.rect(px2MM(120), px2MM(bx_x[i]), px2MM(150), px2MM(bx_height[i]),'FD')
        pdf.rect(px2MM(270), px2MM(bx_x[i]), px2MM(885), px2MM(bx_height[i]),'FD')
        pdf.rect(px2MM(1155), px2MM(bx_x[i]), px2MM(345), px2MM(bx_height[i]),'FD')
        pdf.rect(px2MM(1500), px2MM(bx_x[i]), px2MM(300), px2MM(bx_height[i]),'FD')
        
    #//*----Table heading---*//
    pdf.set_text_color(*hex2RGB('#1A1A1D'))
    pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(24))
    
    pdf.set_xy(px2MM(140),px2MM(224)) 
    pdf.cell(px2MM(110), px2MM(32),'Section',align='L')
    
    pdf.set_xy(px2MM(290),px2MM(224)) 
    pdf.cell(px2MM(845), px2MM(32),'Income Tax Deduction on',align='L')
    
    pdf.set_xy(px2MM(1175),px2MM(224)) 
    pdf.cell(px2MM(305), px2MM(32),'Allowed Limit',align='L')
    
    pdf.set_xy(px2MM(1520),px2MM(224)) 
    pdf.cell(px2MM(260), px2MM(32),'Applicable For',align='L')  
    
    #//*---COL 1 VALUE---*//
    
    pdf.set_text_color(*hex2RGB('#1A1A1D'))
    pdf.set_font('LeagueSpartan-Regular', size=px2pts(24))
    
    pdf.set_xy(px2MM(140),px2MM(296)) 
    pdf.cell(px2MM(110), px2MM(32),'80CCD (2)',align='L')
    
    pdf.set_xy(px2MM(140),px2MM(496)) 
    pdf.cell(px2MM(110), px2MM(32),'80CCD(1B)',align='L')
    
    pdf.set_xy(px2MM(140),px2MM(568)) 
    pdf.cell(px2MM(110), px2MM(32),'80CCH',align='L')
    
    pdf.set_xy(px2MM(140),px2MM(672)) 
    pdf.cell(px2MM(110), px2MM(32),'80D',align='L')
    
    pdf.set_xy(px2MM(140),px2MM(840)) 
    pdf.cell(px2MM(110), px2MM(32),'80DD',align='L')
    

    
    
    # #//*---COL 2 VALUE---*//
    
    cir_x = [311,511,583,685,718,919,951]
    pdf.set_fill_color(*hex2RGB('#000000'))
    for i in range(len(cir_x)):
        pdf.circle(x=px2MM(310), y=px2MM(cir_x[i]), r=px2MM(6), style='F')

    pdf.set_xy(px2MM(330),px2MM(296)) 
    pdf.cell(px2MM(805), px2MM(32),'Employer contribution to NPS account',align='L')
    
    pdf.set_xy(px2MM(330),px2MM(496)) 
    pdf.cell(px2MM(805), px2MM(32),'Additional contribution to NPS',align='L')
    
    pdf.set_xy(px2MM(330),px2MM(568)) 
    pdf.cell(px2MM(805), px2MM(32),'Contribution to Agniveer corpus fund (applicable from Nov 2022)',align='L')
    
    pdf.set_xy(px2MM(330),px2MM(672)) 
    pdf.cell(px2MM(805), px2MM(32),'Medical Insurance – self, spouse, children',align='L')
    
    pdf.set_xy(px2MM(330),px2MM(704)) 
    pdf.cell(px2MM(805), px2MM(32),'Medical Insurance – parents',align='L')
    
    pdf.set_xy(px2MM(290),px2MM(840)) 
    pdf.multi_cell(px2MM(845), px2MM(32),'Medical treatment for handicapped dependents or payment to specified scheme for maintenance of handicapped dependent.',align='L')
    
    pdf.set_xy(px2MM(330),px2MM(904)) 
    pdf.cell(px2MM(805), px2MM(32),'Disability is 40% or more but less than 80%',align='L')
    
    pdf.set_xy(px2MM(330),px2MM(936)) 
    pdf.cell(px2MM(805), px2MM(32),'Disability is 80% or more',align='L')
    
    #//*---COL 3 VALUE---*//
    
    cir_x = [342,404,686,718,854,886]
    pdf.set_fill_color(*hex2RGB('#000000'))
    for i in range(len(cir_x)):
        pdf.circle(x=px2MM(1195), y=px2MM(cir_x[i]), r=px2MM(6), style='F')

    pdf.set_xy(px2MM(1175),px2MM(296)) 
    pdf.cell(px2MM(305), px2MM(32),'Maximum up:',align='L')
    
    pdf.set_xy(px2MM(1215),px2MM(328)) 
    pdf.multi_cell(px2MM(265), px2MM(32),'14% of salary (for Central Govt. employees)',align='L')
    
    pdf.set_xy(px2MM(1215),px2MM(392)) 
    pdf.multi_cell(px2MM(265), px2MM(32),'10% of salary (for other employees)',align='L')
    
    pdf.set_xy(px2MM(1175),px2MM(496)) 
    pdf.multi_cell(px2MM(305), px2MM(32),'₹ 50K',align='L')
    
    pdf.set_xy(px2MM(1175),px2MM(568)) 
    pdf.cell(px2MM(305), px2MM(32),'No limit',align='L')
    
    pdf.set_xy(px2MM(1215),px2MM(672)) 
    pdf.cell(px2MM(265), px2MM(32),'₹ 25K',align='L')
    
    pdf.set_xy(px2MM(1215),px2MM(704)) 
    pdf.multi_cell(px2MM(265), px2MM(32),' ₹ 25K (parents <60 years) & ₹ 50K (parents >=60 years)',align='L')
    
    pdf.set_xy(px2MM(1215),px2MM(840)) 
    pdf.multi_cell(px2MM(265), px2MM(32),'₹ 75K',align='L')
    pdf.set_xy(px2MM(1215),px2MM(872)) 
    pdf.cell(px2MM(265), px2MM(32),'₹ 1.25L',align='L')
    
    # #//*---COL 4 VALUE---*//

    pdf.set_xy(px2MM(1520),px2MM(296)) 
    pdf.cell(px2MM(260), px2MM(32),'Individuals',align='L')
    
    pdf.set_xy(px2MM(1520),px2MM(496)) 
    pdf.cell(px2MM(260), px2MM(32),'Individuals',align='L')
    
    pdf.set_xy(px2MM(1520),px2MM(568)) 
    pdf.multi_cell(px2MM(260), px2MM(32),'Individuals enrolled in Agneepath scheme',align='L')
    
    pdf.set_xy(px2MM(1520),px2MM(672)) 
    pdf.multi_cell(px2MM(260), px2MM(32),'Individuals and HUFs',align='L')
    
    pdf.set_xy(px2MM(1520),px2MM(840)) 
    pdf.multi_cell(px2MM(260), px2MM(32),'individuals and HUFs who have a handicapped dependent',align='L')
    
    
    #//*-----Index Text of Page--**////
    index_text(pdf,'#1A1A1D')
    
#//*----Available Tax Deductions(Page 3)----*// 
def aval_tax_deduct_3(pdf,json_data,c_MoneyS,money_signData):
    pdf.add_page()
    pdf.set_fill_color(*hex2RGB('#FCF8ED'))
    pdf.rect(0, 0, px2MM(1920), px2MM(1080),'F')

    #//*----Available Tax Deductions----*//
    pdf.set_xy(px2MM(120),px2MM(80)) 
    pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(60))
    pdf.set_text_color(*hex2RGB('#000000'))
    pdf.cell(px2MM(641), px2MM(84),'Available Tax Deductions',align='L')
    
    pdf.set_xy(px2MM(791),px2MM(106)) 
    pdf.set_font('LeagueSpartan-Regular', size=px2pts(24))
    pdf.set_text_color(*hex2RGB('#000000'))
    pdf.cell(px2MM(237), px2MM(32),'(as per Old Tax Regime)',align='L')
    
    
    #//*----Content----*//
    bx_height = [72,168,104,104,136,104,104]
    bx_x =[204,276,444,548,652,788,892]
    
    #//*---Table rectangle
    for i in range(len(bx_x)):
        pdf.set_draw_color(*hex2RGB('#E9EAEE'))
        pdf.set_line_width(px2MM(0.5))
        if i%2==0:
            pdf.set_fill_color(*hex2RGB('#FFFFFF'))
        else:
            pdf.set_fill_color(*hex2RGB('#F3F6F9'))
            
        pdf.rect(px2MM(120), px2MM(bx_x[i]), px2MM(150), px2MM(bx_height[i]),'FD')
        pdf.rect(px2MM(270), px2MM(bx_x[i]), px2MM(885), px2MM(bx_height[i]),'FD')
        pdf.rect(px2MM(1155), px2MM(bx_x[i]), px2MM(345), px2MM(bx_height[i]),'FD')
        pdf.rect(px2MM(1500), px2MM(bx_x[i]), px2MM(300), px2MM(bx_height[i]),'FD')
        
    #//*----Table heading---*//
    pdf.set_text_color(*hex2RGB('#1A1A1D'))
    pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(24))
    
    pdf.set_xy(px2MM(140),px2MM(224)) 
    pdf.cell(px2MM(110), px2MM(32),'Section',align='L')
    
    pdf.set_xy(px2MM(290),px2MM(224)) 
    pdf.cell(px2MM(845), px2MM(32),'Income Tax Deduction on',align='L')
    
    pdf.set_xy(px2MM(1175),px2MM(224)) 
    pdf.cell(px2MM(305), px2MM(32),'Allowed Limit',align='L')
    
    pdf.set_xy(px2MM(1520),px2MM(224))
    pdf.cell(px2MM(260), px2MM(32),'Applicable For',align='L')    
    
    #//*---COL 1 VALUE---*//
    
    pdf.set_text_color(*hex2RGB('#1A1A1D'))
    pdf.set_font('LeagueSpartan-Regular', size=px2pts(24))
    
    pdf.set_xy(px2MM(140),px2MM(296)) 
    pdf.cell(px2MM(110), px2MM(32),'80DDB',align='L')
    
    pdf.set_xy(px2MM(140),px2MM(464)) 
    pdf.cell(px2MM(110), px2MM(32),'80E',align='L')
    
    pdf.set_xy(px2MM(140),px2MM(568)) 
    pdf.cell(px2MM(110), px2MM(32),'80EE',align='L')
    
    pdf.set_xy(px2MM(140),px2MM(672)) 
    pdf.cell(px2MM(110), px2MM(32),'80EEA',align='L')
    
    pdf.set_xy(px2MM(140),px2MM(808)) 
    pdf.cell(px2MM(110), px2MM(32),'80EEB',align='L')
    
    pdf.set_xy(px2MM(140),px2MM(912)) 
    pdf.cell(px2MM(110), px2MM(32),'80G',align='L')
    
    #//*---COL 2 VALUE---*//
    cir_x = (342,374,478,582,686,822,926)
    pdf.set_fill_color(*hex2RGB('#000000'))
    for i in range(len(cir_x)):
        pdf.circle(x=px2MM(310), y=px2MM(cir_x[i]), r=px2MM(6), style='F')
    
    pdf.set_xy(px2MM(290),px2MM(296)) 
    pdf.multi_cell(px2MM(845), px2MM(32),'Medical expenditure on self or dependent relative for diseases specified in rule 11DD',align='L')
    pdf.set_xy(px2MM(330),px2MM(296+32)) 
    pdf.cell(px2MM(805), px2MM(32),'For less than 60 years old',align='L')
    pdf.set_xy(px2MM(330),px2MM(296+64)) 
    pdf.cell(px2MM(805), px2MM(32),'For more than 60 years old',align='L')
    
    pdf.set_xy(px2MM(330),px2MM(464)) 
    pdf.multi_cell(px2MM(805), px2MM(32),'''Interest on education loan''',align='L')
    
    pdf.set_xy(px2MM(330),px2MM(568)) 
    pdf.multi_cell(px2MM(805), px2MM(32),'''IInterest on home loan for first-time homeowners, available for loans sanctioned between 01-Apr-2016 and 31-Mar-2017''',align='L')
    
    pdf.set_xy(px2MM(330),px2MM(672)) 
    pdf.multi_cell(px2MM(805), px2MM(32),'''Interest on home loan (over and above Rs. 2,00,000 deduction under 24B, allowing taxpayers to deduct total of Rs. 3,50,000 for interest on home loan) for loans sanctioned between 01-Apr-2019 and 31-Mar-2022''',align='L')
    
    
    pdf.set_xy(px2MM(330),px2MM(808)) 
    pdf.multi_cell(px2MM(805), px2MM(32),'''Interest on  loan taken between 01-Apr-2019 and 31-Mar-2023 for purchase of electric vehicle''',align='L')
    
    pdf.set_xy(px2MM(330),px2MM(912)) 
    pdf.multi_cell(px2MM(805), px2MM(32),'''Contributions to certain relief funds and charitable institutions''',align='L')
    

    #//*---COL 3 VALUE---*//
    cir_x = (310,374)
    pdf.set_fill_color(*hex2RGB('#000000'))
    for i in range(len(cir_x)):
        pdf.circle(x=px2MM(1195), y=px2MM(cir_x[i]), r=px2MM(6), style='F')
        
    pdf.set_xy(px2MM(1215),px2MM(296)) 
    pdf.multi_cell(px2MM(265), px2MM(32),'Lower of ₹ 40K or amount actually paid',align='L')
    pdf.set_xy(px2MM(1215),px2MM(360)) 
    pdf.multi_cell(px2MM(265), px2MM(32),'Lower of ₹ 1L or amount actually paid',align='L')
    
    pdf.set_xy(px2MM(1175),px2MM(464)) 
    pdf.multi_cell(px2MM(305), px2MM(32),'Interest paid for a period of 8 years',align='L')
    
    pdf.set_xy(px2MM(1175),px2MM(568)) 
    pdf.multi_cell(px2MM(305), px2MM(32),'₹ 50K',align='L')
    
    pdf.set_xy(px2MM(1175),px2MM(672)) 
    pdf.multi_cell(px2MM(305), px2MM(32),'₹ 1.5L',align='L')

    pdf.set_xy(px2MM(1175),px2MM(808)) 
    pdf.multi_cell(px2MM(305), px2MM(32),'₹ 1.5L',align='L')

    pdf.set_xy(px2MM(1175),px2MM(912)) 
    pdf.multi_cell(px2MM(305), px2MM(32),'50% or 100% of the donation amount can be claimed.',align='L')

    
    #//*---COL 4 VALUE---*//
    
    pdf.set_xy(px2MM(1520),px2MM(296)) 
    pdf.cell(px2MM(260), px2MM(32),'Individuals and HUFs',align='L')
    
    pdf.set_xy(px2MM(1520),px2MM(464)) 
    pdf.multi_cell(px2MM(260), px2MM(32),'Individuals',align='L')
    
    pdf.set_xy(px2MM(1520),px2MM(568)) 
    pdf.cell(px2MM(260), px2MM(32),'Individuals',align='L')
    
    pdf.set_xy(px2MM(1520),px2MM(672)) 
    pdf.multi_cell(px2MM(260), px2MM(32),'Individuals',align='L')
    
    pdf.set_xy(px2MM(1520),px2MM(808)) 
    pdf.cell(px2MM(260), px2MM(32),'Individuals',align='L')
    
    pdf.set_xy(px2MM(1520),px2MM(912)) 
    pdf.multi_cell(px2MM(260), px2MM(32),'Individuals, HUFs, companies',align='L')

    #//*-----Index Text of Page--**////
    index_text(pdf,'#1A1A1D')
    
#//*----Available Tax Deductions(Page 4)----*// 
def aval_tax_deduct_4(pdf,json_data,c_MoneyS,money_signData):
    pdf.add_page()
    pdf.set_fill_color(*hex2RGB('#FCF8ED'))
    pdf.rect(0, 0, px2MM(1920), px2MM(1080),'F')

    #//*----Available Tax Deductions----*//
    pdf.set_xy(px2MM(120),px2MM(80)) 
    pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(60))
    pdf.set_text_color(*hex2RGB('#000000'))
    pdf.cell(px2MM(641), px2MM(84),'Available Tax Deductions',align='L')
    
    pdf.set_xy(px2MM(791),px2MM(106)) 
    pdf.set_font('LeagueSpartan-Regular', size=px2pts(24))
    pdf.set_text_color(*hex2RGB('#000000'))
    pdf.cell(px2MM(237), px2MM(32),'(as per Old Tax Regime)',align='L')
    
    
    #//*----Content----*//
    bx_height = [72,200,136,104,104,104,104]
    bx_x =[204,276,476,612,716,820,924]
    
    #//*---Table rectangle
    for i in range(len(bx_x)):
        pdf.set_draw_color(*hex2RGB('#E9EAEE'))
        pdf.set_line_width(px2MM(0.5))
        if i%2==0:
            pdf.set_fill_color(*hex2RGB('#FFFFFF'))
        else:
            pdf.set_fill_color(*hex2RGB('#F3F6F9'))
            
        pdf.rect(px2MM(120), px2MM(bx_x[i]), px2MM(150), px2MM(bx_height[i]),'FD')
        pdf.rect(px2MM(270), px2MM(bx_x[i]), px2MM(885), px2MM(bx_height[i]),'FD')
        pdf.rect(px2MM(1155), px2MM(bx_x[i]), px2MM(345), px2MM(bx_height[i]),'FD')
        pdf.rect(px2MM(1500), px2MM(bx_x[i]), px2MM(300), px2MM(bx_height[i]),'FD')
        
    #//*----Table heading---*//
    pdf.set_text_color(*hex2RGB('#1A1A1D'))
    pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(24))
    
    pdf.set_xy(px2MM(140),px2MM(224)) 
    pdf.cell(px2MM(110), px2MM(32),'Section',align='L')
    
    pdf.set_xy(px2MM(290),px2MM(224)) 
    pdf.cell(px2MM(845), px2MM(32),'Income Tax Deduction on',align='L')
    
    pdf.set_xy(px2MM(1175),px2MM(224)) 
    pdf.cell(px2MM(305), px2MM(32),'Allowed Limit',align='L')
    
    pdf.set_xy(px2MM(1520),px2MM(224))
    pdf.cell(px2MM(260), px2MM(32),'Applicable For',align='L')    
    
    #//*---COL 1 VALUE---*//
    
    pdf.set_text_color(*hex2RGB('#1A1A1D'))
    pdf.set_font('LeagueSpartan-Regular', size=px2pts(24))
    
    pdf.set_xy(px2MM(140),px2MM(296)) 
    pdf.cell(px2MM(110), px2MM(32),'80GG',align='L')
    
    pdf.set_xy(px2MM(140),px2MM(496)) 
    pdf.cell(px2MM(110), px2MM(32),'80GGA',align='L')
    
    pdf.set_xy(px2MM(140),px2MM(632)) 
    pdf.cell(px2MM(110), px2MM(32),'80GGC',align='L')
    
    pdf.set_xy(px2MM(140),px2MM(736)) 
    pdf.cell(px2MM(110), px2MM(32),'80RRB',align='L')
    
    pdf.set_xy(px2MM(140),px2MM(840)) 
    pdf.cell(px2MM(110), px2MM(32),'80TTA (1)',align='L')
    
    pdf.set_xy(px2MM(140),px2MM(944)) 
    pdf.cell(px2MM(110), px2MM(32),'80TTB',align='L')
    
    #//*---COL 2 VALUE---*//
    cir_x = (310,510,646,750,854,958)
    pdf.set_fill_color(*hex2RGB('#000000'))
    for i in range(len(cir_x)):
        pdf.circle(x=px2MM(310), y=px2MM(cir_x[i]), r=px2MM(6), style='F')
    
    
    pdf.set_xy(px2MM(330),px2MM(296)) 
    pdf.cell(px2MM(805), px2MM(32),'For rent paid when HRA is not received from an employer',align='L')
    
    pdf.set_xy(px2MM(330),px2MM(496)) 
    pdf.multi_cell(px2MM(805), px2MM(32),"""Donation for scientific, social science,  research, or rural development to specific universities, colleges or research association""",align='L')
    
    pdf.set_xy(px2MM(330),px2MM(632)) 
    pdf.multi_cell(px2MM(805), px2MM(32),'''Contribution by individuals to political parties''',align='L')
    
    pdf.set_xy(px2MM(330),px2MM(736)) 
    pdf.multi_cell(px2MM(805), px2MM(32),'''Royalty on patents''',align='L')

    pdf.set_xy(px2MM(330),px2MM(840)) 
    pdf.multi_cell(px2MM(805), px2MM(32),'''Interest income from savings account''',align='L')

    pdf.set_xy(px2MM(330),px2MM(940)) 
    pdf.multi_cell(px2MM(805), px2MM(32),'''Exemption of interest from banks, post offices, etc.''',align='L')
    
    #//*---COL 3 VALUE---*//
    cir_x = (342,406,438)
    pdf.set_fill_color(*hex2RGB('#000000'))
    for i in range(len(cir_x)):
        pdf.circle(x=px2MM(1195), y=px2MM(cir_x[i]), r=px2MM(6), style='F')
        
    pdf.set_xy(px2MM(1175),px2MM(296)) 
    pdf.multi_cell(px2MM(305), px2MM(32),'Least of:',align='L')    
    pdf.set_xy(px2MM(1215),px2MM(328)) 
    pdf.multi_cell(px2MM(265), px2MM(32),'Rent paid minus 10% of total income',align='L')
    pdf.set_xy(px2MM(1215),px2MM(392)) 
    pdf.multi_cell(px2MM(265), px2MM(32),'₹ 5K per month',align='L')
    pdf.set_xy(px2MM(1215),px2MM(424)) 
    pdf.multi_cell(px2MM(305), px2MM(32),'25% of total income',align='L')
    
    pdf.set_xy(px2MM(1175),px2MM(496)) 
    pdf.multi_cell(px2MM(305), px2MM(32),'No limit',align='L')
    
    pdf.set_xy(px2MM(1175),px2MM(632)) 
    pdf.multi_cell(px2MM(305), px2MM(32),'Amount contributed (not allowed if paid in cash)',align='L')

    pdf.set_xy(px2MM(1175),px2MM(736)) 
    pdf.multi_cell(px2MM(305), px2MM(32),'₹ 3L',align='L')

    pdf.set_xy(px2MM(1175),px2MM(840)) 
    pdf.multi_cell(px2MM(305), px2MM(32),'₹ 10K',align='L')
    
    pdf.set_xy(px2MM(1175),px2MM(944)) 
    pdf.multi_cell(px2MM(305), px2MM(32),'Maximum up to ₹ 50K',align='L')

    
    #//*---COL 4 VALUE---*//
    
    pdf.set_xy(px2MM(1520),px2MM(296)) 
    pdf.multi_cell(px2MM(260), px2MM(32),'Individuals not receiving HRA',align='L')
    
    pdf.set_xy(px2MM(1520),px2MM(496)) 
    pdf.multi_cell(px2MM(265), px2MM(32),'All individuals except those having Income from business and profession',align='L')
    
    pdf.set_xy(px2MM(1520),px2MM(632)) 
    pdf.cell(px2MM(260), px2MM(32),'Individuals',align='L')
    
    pdf.set_xy(px2MM(1520),px2MM(736)) 
    pdf.multi_cell(px2MM(260), px2MM(32),'Resident individual who is a patentee',align='L')
    
    pdf.set_xy(px2MM(1520),px2MM(840)) 
    pdf.multi_cell(px2MM(260), px2MM(32),'Individuals and HUFs (except senior citizens)',align='L')
    
    pdf.set_xy(px2MM(1520),px2MM(944)) 
    pdf.multi_cell(px2MM(260), px2MM(32),'Senior citizens (above 60 years)',align='L')

    #//*-----Index Text of Page--**////
    index_text(pdf,'#1A1A1D')
    
def aval_tax_deduct_5(pdf,json_data,c_MoneyS,money_signData):
    pdf.add_page()
    pdf.set_fill_color(*hex2RGB('#FCF8ED'))
    pdf.rect(0, 0, px2MM(1920), px2MM(1080),'F')

    pdf.set_xy(px2MM(791),px2MM(106)) 
    pdf.set_font('LeagueSpartan-Regular', size=px2pts(24))
    pdf.set_text_color(*hex2RGB('#000000'))
    pdf.cell(px2MM(237), px2MM(32),'(as per Old Tax Regime)',align='L')

    pdf.set_draw_color(*hex2RGB('#E9EAEE'))

    #//----Availablel Tax Deductions 1st Page----//
    #//----Headings----//
    pdf.set_xy(px2MM(120),px2MM(80)) 
    pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(60))
    pdf.set_text_color(*hex2RGB('#000000'))
    pdf.cell(px2MM(641), px2MM(84),'Available Tax Deductions',align='L')

    # Section
    
    pdf.set_fill_color(*hex2RGB('#FFFFFF'))
    pdf.set_text_color(*hex2RGB('#1A1A1D'))
    pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(24))
    pdf.rect(px2MM(120), px2MM(204), px2MM(150), px2MM(72),'FD')
    pdf.set_xy(px2MM(140),px2MM(224))
    pdf.cell(px2MM(110), px2MM(32),"Section",align='L')

    # Section Col Values
    vals = ['80U','24B','10(13A)','10(5)']
    rect_top = [276,412,484,716]
    rect_height = [136,72,232,264]
    set_xy_top = [296,432,504,736]


    for i in range(len(vals)):
        fill_color = "#FFFFFF" if i % 2 != 0 else '#F3F6F9'
        pdf.set_fill_color(*hex2RGB(fill_color))
        pdf.set_font('LeagueSpartan-Regular', size=px2pts(24))
        pdf.rect(px2MM(120), px2MM(rect_top[i]), px2MM(150), px2MM(rect_height[i]),'FD')
        pdf.set_xy(px2MM(140),px2MM(set_xy_top[i]))
        pdf.cell(px2MM(150), px2MM(32),vals[i],align='L')
    pdf.set_fill_color(*hex2RGB('#FFFFFF')) # Set Fill Color Back To White
    

    # Income Tax Deduction on
    pdf.set_text_color(*hex2RGB('#1A1A1D'))
    pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(24))
    pdf.rect(px2MM(270), px2MM(204), px2MM(885), px2MM(72),'FD')
    pdf.set_xy(px2MM(290),px2MM(224))
    pdf.cell(px2MM(845), px2MM(32),"Income Tax Deduction on",align='L')


    vals = ["Individual suffering from physical disability (including blindness) or mental retardation",
            "Individual suffering from severe disability"]
    for i in range(4):
        fill_color = "#FFFFFF" if i % 2 != 0 else '#F3F6F9'
        pdf.set_fill_color(*hex2RGB(fill_color))
        pdf.set_font('LeagueSpartan-Regular', size=px2pts(24))
        pdf.rect(px2MM(270), px2MM(rect_top[i]), px2MM(885), px2MM(rect_height[i]),'FD')
    pdf.set_fill_color(*hex2RGB('#FFFFFF')) # Set Fill Color Back To White

    pdf.set_fill_color(*hex2RGB('#000000'))
    # Kaala Timba
    pdf.circle(x=px2MM(290+5), y=px2MM(296 + 20), r=px2MM(5), style='F')
    pdf.set_xy(px2MM(290 + 20),px2MM(296+5))
    pdf.multi_cell(px2MM(810), px2MM(32),vals[0],align='L')

    pdf.circle(x=px2MM(290+5), y=px2MM(296 + 64 + 20), r=px2MM(5), style='F')
    pdf.set_xy(px2MM(290 + 20),px2MM(296 + 64 + 5))
    pdf.multi_cell(px2MM(810), px2MM(32),vals[1],align='L')

    
    # Tab 2
    val = ["Tax exemption on interest paid on home loan"]
    pdf.circle(x=px2MM(290+5), y=px2MM(432 + 20), r=px2MM(5), style='F')
    pdf.set_xy(px2MM(290 + 20),px2MM(432 + 5)) 
    pdf.multi_cell(px2MM(810), px2MM(32),val[0],align='L')
    
    val = ["House Rent Allowance (HRA)"]
    pdf.circle(x=px2MM(290+5), y=px2MM(504 + 20), r=px2MM(5), style='F')
    pdf.set_xy(px2MM(290 + 20),px2MM(504 + 5)) 
    pdf.multi_cell(px2MM(810), px2MM(32),val[0],align='L')


    # Tab 3
    # Val 1
    val = "Leave Travel Allowance (LTA)"
    pdf.circle(x=px2MM(290+5), y=px2MM(736 + 20), r=px2MM(5), style='F')
    pdf.set_xy(px2MM(290 + 20),px2MM(736+5))    
    pdf.multi_cell(px2MM(810), px2MM(32),val,align='L')

    pdf.set_fill_color(*hex2RGB('#FFFFFF')) # Set Fill Color Back To White


    # Allowed Limit
    pdf.set_text_color(*hex2RGB('#1A1A1D'))
    pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(24))
    pdf.rect(px2MM(1155), px2MM(204), px2MM(345), px2MM(72),'FD')
    pdf.set_xy(px2MM(1175),px2MM(224))
    pdf.cell(px2MM(260), px2MM(32),"Allowed Limit",align='L')

    # Allowed Limit Vals 
    for i in range(4):
        fill_color = "#FFFFFF" if i % 2 != 0 else '#F3F6F9'
        pdf.set_fill_color(*hex2RGB(fill_color))
        pdf.set_font('LeagueSpartan-Regular', size=px2pts(24))
        pdf.rect(px2MM(1155), px2MM(rect_top[i]), px2MM(345), px2MM(rect_height[i]),'FD')
        # pdf.set_xy(px2MM(1175),px2MM(set_xy_top[i]))
        # pdf.multi_cell(px2MM(305), px2MM(32),vals[i],align='L')
    pdf.set_fill_color(*hex2RGB('#FFFFFF')) # Set Fill Color Back To White
    pdf.set_fill_color(*hex2RGB('#000000'))


    vals = ["₹ 75K","₹ 1.25L"]
    # Kaala Timba
    # pdf.circle(x=px2MM(1175+5), y=px2MM(296 + 20), r=px2MM(5), style='F')
    pdf.circle(x=px2MM(1175+10), y=px2MM(296 + 20), r=px2MM(5), style='F')
    # pdf.set_xy(px2MM(1175 + 20),px2MM(296+5))
    pdf.set_xy(px2MM(1175 + 25),px2MM(296+5))
    pdf.multi_cell(px2MM(810), px2MM(32),vals[0],align='L')
    
    pdf.circle(x=px2MM(1175 + 10), y=px2MM(296 + 32 +20), r=px2MM(5), style='F')
    pdf.set_xy(px2MM(1175 + 25),px2MM(296+ 32 +5))
    pdf.multi_cell(px2MM(810), px2MM(32),vals[1],align='L')
    
    # Tab 3
    # pdf.set_xy(px2MM(1175 + 20),px2MM(432))
    pdf.set_xy(px2MM(1175 + 5),px2MM(432))
    pdf.multi_cell(px2MM(810), px2MM(32),"₹ 2L",align='L')


    # Tab 3
    pdf.set_xy(px2MM(1175 + 5),px2MM(504))
    pdf.multi_cell(px2MM(810), px2MM(32),"Least of:",align='L')

    # Point 2
    pdf.circle(x=px2MM(1175+10), y=px2MM(504 + 32 +20), r=px2MM(5), style='F')
    pdf.set_xy(px2MM(1175 + 25),px2MM(504 + 32 + 5))
    pdf.multi_cell(px2MM(810), px2MM(32),"Actual HRA received",align='L')
    
    # Point 3
    pdf.circle(x=px2MM(1175+10), y=px2MM(504 + 32 + 32 + 20), r=px2MM(5), style='F')
    pdf.set_xy(px2MM(1175 + 25),px2MM(504 + 32 +32 + 5))
    pdf.multi_cell(px2MM(305), px2MM(32),"40%/50% of Basic+DA for non-metro/ metro city",align='L')
    
    
    # Point 4
    pdf.circle(x=px2MM(1175+10), y=px2MM(504 + 32 + 32 + 64 + 20), r=px2MM(5), style='F')
    pdf.set_xy(px2MM(1175 + 25),px2MM(504 + 32 + 32 + 64 + 5))
    pdf.multi_cell(px2MM(305), px2MM(32),"Actual rent paid minus 10% \nof Basic+DA",align='L')


    # Tab 4
    # Point 1
    val = "Only for self/family's domestic travel by train \n(max upto 1st class AC \nFare by shortest route) or flight (max upto economy class fare)"
    val2 = "Allowed 2 times in 4 years"

    pdf.circle(x=px2MM(1175+10), y=px2MM(736 + 20 - 5), r=px2MM(5), style='F')
    pdf.set_xy(px2MM(1175 + 25),px2MM(736 + 5 - 5))
    pdf.multi_cell(px2MM(305), px2MM(32),val,align='L')
    
    # Point 2
    pdf.circle(x=px2MM(1175+10), y=px2MM(736 + 216 - 5), r=px2MM(5), style='F')
    pdf.set_xy(px2MM(1175 + 25),px2MM(736 + 198 + 5 - 5))
    pdf.multi_cell(px2MM(305), px2MM(32),val2,align='L')


    
    pdf.set_fill_color(*hex2RGB('#FFFFFF')) # Set Fill Color Back To White
    # Applicable For
    pdf.set_text_color(*hex2RGB('#1A1A1D'))
    pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(24))
    pdf.rect(px2MM(1500), px2MM(204), px2MM(300), px2MM(72),'FD')
    pdf.set_xy(px2MM(1520),px2MM(224))
    pdf.cell(px2MM(269), px2MM(32),"Allowed Limit",align='L')

    # Applicable For Values
    # vals = ['Individuals and HUFs','Individuals','Individuals']
    for i in range(4):
        fill_color = "#FFFFFF" if i % 2 != 0 else '#F3F6F9'
        pdf.set_fill_color(*hex2RGB(fill_color))
        pdf.set_font('LeagueSpartan-Regular', size=px2pts(24))
        pdf.rect(px2MM(1500), px2MM(rect_top[i]), px2MM(300), px2MM(rect_height[i]),'FD')
        # pdf.set_xy(px2MM(1520),px2MM(set_xy_top[i]))
        # pdf.cell(px2MM(260), px2MM(32),"Hello World",align='L')
    pdf.set_fill_color(*hex2RGB('#FFFFFF')) # Set Fill Color Back To White

    # Tab 1

    set_xy_top = [296,432,504,736]

    pdf.set_xy(px2MM(1520 + 5),px2MM(296))
    pdf.multi_cell(px2MM(260), px2MM(32),"Individuals with disabilities",align='L')
    
    pdf.set_xy(px2MM(1520 + 5),px2MM(432))
    pdf.multi_cell(px2MM(260), px2MM(32),"Individuals",align='L')

    pdf.set_xy(px2MM(1520 + 5),px2MM(504))
    pdf.multi_cell(px2MM(260), px2MM(32),"Individual receiving HRA from employer",align='L')
    
    pdf.set_xy(px2MM(1520 + 5),px2MM(736))
    pdf.multi_cell(px2MM(260), px2MM(32),"Individual receiving LTA from employer",align='L')
    
    #//*-----Index Text of Page--**////
    index_text(pdf,'#1A1A1D')
 
 
 
 
#//*----Capital Gains Taxation by Asset Type (Page 1)-----*//   
 
def capital_gains_1(pdf,json_data,c_MoneyS,money_signData):
    pdf.add_page()
    pdf.set_fill_color(*hex2RGB('#FCF8ED'))
    pdf.rect(0, 0, px2MM(1920), px2MM(1080),'F')
    
    # black rectangle
    pdf.set_fill_color(*hex2RGB('#000000'))
    pdf.rect(px2MM(0), px2MM(80), px2MM(15), px2MM(84), 'F')

    #//*----Headings----//
    pdf.set_xy(px2MM(120),px2MM(80)) 
    pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(60))
    pdf.set_text_color(*hex2RGB('#000000'))
    pdf.cell(px2MM(960), px2MM(84),'Capital Gains Taxation by Asset Type',align='L')
    
    #//*---Black vertical line--*-//
    pdf.set_fill_color(*hex2RGB('#313236'))
    pdf.rect(px2MM(120), px2MM(204), px2MM(6), px2MM(414),'F')
    
    pdf.rect(px2MM(126), px2MM(204), px2MM(100), px2MM(42),'F')
    pdf.set_xy(px2MM(141),px2MM(209)) 
    pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(24))
    pdf.set_text_color(*hex2RGB('#FFFFFF'))
    pdf.cell(px2MM(80), px2MM(32),'Equity',align='L')
    
    
    #//*----Equity Table-----------------*//
    #//*---Columns---*//
    pdf.set_draw_color(*hex2RGB('#E9EAEE'))
    pdf.set_line_width(px2MM(0.2))
    pdf.set_fill_color(*hex2RGB('#FFFFFF'))
    pdf.rect(px2MM(126), px2MM(246), px2MM(457), px2MM(104),'FD')
    
    pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(24))
    pdf.set_text_color(*hex2RGB('#1A1A1D'))
    pdf.set_xy(px2MM(146),px2MM(266)) 
    pdf.cell(px2MM(80), px2MM(64),'Asset Type',align='L')
    
    pdf.rect(px2MM(583), px2MM(246), px2MM(608.5), px2MM(52),'FD')
    pdf.set_xy(px2MM(603),px2MM(256)) 
    pdf.cell(px2MM(568.5), px2MM(32),'Long-term Capital Gains (LTCG)',align='C')
    
    pdf.rect(px2MM(1191.5), px2MM(246), px2MM(608.5), px2MM(52),'FD')
    pdf.set_xy(px2MM(1211.5),px2MM(256)) 
    pdf.cell(px2MM(568.5), px2MM(32),'Short-term Capital Gains (STCG)',align='C')
    
    pdf.rect(px2MM(583), px2MM(298), px2MM(189), px2MM(52),'FD')
    pdf.set_xy(px2MM(603),px2MM(308)) 
    pdf.cell(px2MM(155), px2MM(32),'Holding Period',align='C')
    
    pdf.rect(px2MM(772), px2MM(298), px2MM(419.5), px2MM(52),'FD')
    pdf.set_xy(px2MM(792),px2MM(308)) 
    pdf.cell(px2MM(379.5), px2MM(32),'Tax Rate',align='C')
    
    pdf.rect(px2MM(1191.5), px2MM(298), px2MM(189), px2MM(52),'FD')
    pdf.set_xy(px2MM(1211.5),px2MM(308)) 
    pdf.cell(px2MM(155), px2MM(32),'Holding Period',align='C')
    
    pdf.rect(px2MM(1380.5), px2MM(298), px2MM(419.5), px2MM(52),'FD')
    pdf.set_xy(px2MM(1400.5),px2MM(308)) 
    pdf.cell(px2MM(379.5), px2MM(32),'Tax Rate',align='C')
    
    #//*----Row 1(1/1)----*//
    pdf.set_fill_color(*hex2RGB('#F3F6F9'))
    pdf.rect(px2MM(126), px2MM(350), px2MM(240), px2MM(144),'FD')
    pdf.set_font('LeagueSpartan-Regular', size=px2pts(24))
    pdf.set_xy(px2MM(146),px2MM(406)) 
    pdf.cell(px2MM(210), px2MM(32),'Domestic shares',align='L')
    
    pdf.rect(px2MM(366), px2MM(350), px2MM(217), px2MM(72),'FD')
    pdf.set_xy(px2MM(386),px2MM(370)) 
    pdf.cell(px2MM(177), px2MM(32),'Listed',align='L')
    
    pdf.rect(px2MM(583), px2MM(350), px2MM(189), px2MM(72),'FD')
    pdf.set_xy(px2MM(603),px2MM(370)) 
    pdf.cell(px2MM(155), px2MM(32),'> 1 year',align='C')
    
    pdf.rect(px2MM(772), px2MM(350), px2MM(419.5), px2MM(72),'FD')
    pdf.set_xy(px2MM(792),px2MM(370)) 
    pdf.cell(px2MM(379.5), px2MM(32),'10% on LTCG > ₹ 1 lakh/year',align='C')
    
    pdf.rect(px2MM(1191.5), px2MM(350), px2MM(189), px2MM(72),'FD')
    pdf.set_xy(px2MM(1211.5),px2MM(370)) 
    pdf.cell(px2MM(155), px2MM(32),'< 1 year',align='C')
    
    pdf.rect(px2MM(1380.5), px2MM(350), px2MM(419.5), px2MM(72),'FD')
    pdf.set_xy(px2MM(1400.5),px2MM(370)) 
    pdf.cell(px2MM(379.5), px2MM(32),'15%',align='C')
    
    #//*--row (1/2)---*//
    pdf.rect(px2MM(366), px2MM(422), px2MM(217), px2MM(72),'FD')
    pdf.set_xy(px2MM(386),px2MM(442)) 
    pdf.cell(px2MM(177), px2MM(32),'Unlisted',align='L')
    
    pdf.rect(px2MM(583), px2MM(422), px2MM(189), px2MM(72),'FD')
    pdf.set_xy(px2MM(603),px2MM(442)) 
    pdf.cell(px2MM(155), px2MM(32),'> 2 years',align='C')
    
    pdf.rect(px2MM(772), px2MM(422), px2MM(419.5), px2MM(72),'FD')
    pdf.set_xy(px2MM(792),px2MM(442)) 
    pdf.cell(px2MM(379.5), px2MM(32),'20% with indexation',align='C')
    
    pdf.rect(px2MM(1191.5), px2MM(422), px2MM(189), px2MM(72),'FD')
    pdf.set_xy(px2MM(1211.5),px2MM(442)) 
    pdf.cell(px2MM(155), px2MM(32),'< 2 years',align='C')
    
    pdf.rect(px2MM(1380.5), px2MM(422), px2MM(419.5), px2MM(72),'FD')
    pdf.set_xy(px2MM(1400.5),px2MM(442)) 
    pdf.cell(px2MM(379.5), px2MM(32),'As per income tax slab',align='C')
    
    #//*--row (2)---*//
    pdf.set_fill_color(*hex2RGB('#FFFFFF'))
    pdf.rect(px2MM(126), px2MM(494), px2MM(457), px2MM(62),'FD')
    pdf.set_xy(px2MM(146),px2MM(509)) 
    pdf.cell(px2MM(366), px2MM(32),'Equity mutual funds',align='L')
    
    pdf.rect(px2MM(583), px2MM(494), px2MM(189), px2MM(62),'FD')
    pdf.set_xy(px2MM(603),px2MM(509)) 
    pdf.cell(px2MM(155), px2MM(32),'> 1 year',align='C')
    
    pdf.rect(px2MM(772), px2MM(494), px2MM(419.5), px2MM(62),'FD')
    pdf.set_xy(px2MM(792),px2MM(509)) 
    pdf.cell(px2MM(379.5), px2MM(32),'10% on LTCG > ₹ 1 lakh/year',align='C')
    
    pdf.rect(px2MM(1191.5), px2MM(494), px2MM(189), px2MM(62),'FD')
    pdf.set_xy(px2MM(1211.5),px2MM(509)) 
    pdf.cell(px2MM(155), px2MM(32),'< 1 year',align='C')
    
    pdf.rect(px2MM(1380.5), px2MM(494), px2MM(419.5), px2MM(62),'FD')
    pdf.set_xy(px2MM(1400.5),px2MM(509)) 
    pdf.cell(px2MM(379.5), px2MM(32),'15%',align='C')
    
    #//*--row (3)---*//
    pdf.set_fill_color(*hex2RGB('#F3F6F9'))
    pdf.rect(px2MM(126), px2MM(556), px2MM(457), px2MM(62),'FD')
    pdf.set_xy(px2MM(146),px2MM(571)) 
    pdf.cell(px2MM(366), px2MM(32),'Foreign shares',align='L')
    
    pdf.rect(px2MM(583), px2MM(556), px2MM(189), px2MM(62),'FD')
    pdf.set_xy(px2MM(603),px2MM(571)) 
    pdf.cell(px2MM(155), px2MM(32),'> 2 years',align='C')
    
    pdf.rect(px2MM(772), px2MM(556), px2MM(419.5), px2MM(62),'FD')
    pdf.set_xy(px2MM(792),px2MM(571)) 
    pdf.cell(px2MM(379.5), px2MM(32),'20% with indexation',align='C')
    
    pdf.rect(px2MM(1191.5), px2MM(556), px2MM(189), px2MM(62),'FD')
    pdf.set_xy(px2MM(1211.5),px2MM(571)) 
    pdf.cell(px2MM(155), px2MM(32),'< 2 years',align='C')
    
    pdf.rect(px2MM(1380.5), px2MM(556), px2MM(419.5), px2MM(62),'FD')
    pdf.set_xy(px2MM(1400.5),px2MM(571)) 
    pdf.cell(px2MM(379.5), px2MM(32),'As per income tax slab',align='C')
    
    
    #//*----Real Estate Table-----------------*//
    #//*---Black vertical line--*-//
    pdf.set_fill_color(*hex2RGB('#313236'))
    pdf.rect(px2MM(120), px2MM(658), px2MM(6), px2MM(208),'F')
    
    pdf.rect(px2MM(126), px2MM(658), px2MM(147), px2MM(42),'F')
    pdf.set_xy(px2MM(141),px2MM(663)) 
    pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(24))
    pdf.set_text_color(*hex2RGB('#FFFFFF'))
    pdf.cell(px2MM(120), px2MM(32),'Real Estate',align='L')
    
    
    #//*---Columns---*//
    pdf.set_draw_color(*hex2RGB('#E9EAEE'))
    pdf.set_line_width(px2MM(0.2))
    pdf.set_fill_color(*hex2RGB('#FFFFFF'))
    pdf.rect(px2MM(126), px2MM(700), px2MM(457), px2MM(104),'FD')
    
    pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(24))
    pdf.set_text_color(*hex2RGB('#1A1A1D'))
    pdf.set_xy(px2MM(146),px2MM(720)) 
    pdf.cell(px2MM(80), px2MM(64),'Asset Type',align='L')
    
    pdf.rect(px2MM(583), px2MM(700), px2MM(608.5), px2MM(52),'FD')
    pdf.set_xy(px2MM(603),px2MM(710)) 
    pdf.cell(px2MM(568.5), px2MM(32),'Long-term Capital Gains (LTCG)',align='C')
    
    pdf.rect(px2MM(1191.5), px2MM(700), px2MM(608.5), px2MM(52),'FD')
    pdf.set_xy(px2MM(1211.5),px2MM(710)) 
    pdf.cell(px2MM(568.5), px2MM(32),'Short-term Capital Gains (STCG)',align='C')
    
    pdf.rect(px2MM(583), px2MM(752), px2MM(189), px2MM(52),'FD')
    pdf.set_xy(px2MM(603),px2MM(762)) 
    pdf.cell(px2MM(155), px2MM(32),'Holding Period',align='C')
    
    pdf.rect(px2MM(772), px2MM(752), px2MM(419.5), px2MM(52),'FD')
    pdf.set_xy(px2MM(792),px2MM(762)) 
    pdf.cell(px2MM(379.5), px2MM(32),'Tax Rate',align='C')
    
    pdf.rect(px2MM(1191.5), px2MM(752), px2MM(189), px2MM(52),'FD')
    pdf.set_xy(px2MM(1211.5),px2MM(762)) 
    pdf.cell(px2MM(155), px2MM(32),'Holding Period',align='C')
    
    pdf.rect(px2MM(1380.5), px2MM(752), px2MM(419.5), px2MM(52),'FD')
    pdf.set_xy(px2MM(1400.5),px2MM(762)) 
    pdf.cell(px2MM(379.5), px2MM(32),'Tax Rate',align='C')
    
    #//*--row (1)---*//
    pdf.set_fill_color(*hex2RGB('#F3F6F9'))
    pdf.set_font('LeagueSpartan-Regular', size=px2pts(24))
    
    pdf.rect(px2MM(126), px2MM(804), px2MM(457), px2MM(62),'FD')
    pdf.set_xy(px2MM(146),px2MM(819)) 
    pdf.cell(px2MM(366), px2MM(32),'Residential/commercial',align='L')
    
    pdf.rect(px2MM(583), px2MM(804), px2MM(189), px2MM(62),'FD')
    pdf.set_xy(px2MM(603),px2MM(819)) 
    pdf.cell(px2MM(155), px2MM(32),'> 2 years',align='C')
    
    pdf.rect(px2MM(772), px2MM(804), px2MM(419.5), px2MM(62),'FD')
    pdf.set_xy(px2MM(792),px2MM(819)) 
    pdf.cell(px2MM(379.5), px2MM(32),'20% with indexation',align='C')
    
    pdf.rect(px2MM(1191.5), px2MM(804), px2MM(189), px2MM(62),'FD')
    pdf.set_xy(px2MM(1211.5),px2MM(819)) 
    pdf.cell(px2MM(155), px2MM(32),'< 2 years',align='C')
    
    pdf.rect(px2MM(1380.5), px2MM(804), px2MM(419.5), px2MM(62),'FD')
    pdf.set_xy(px2MM(1400.5),px2MM(819)) 
    pdf.cell(px2MM(379.5), px2MM(32),'As per income tax slab',align='C')
    
    desc_list = ["Capital gains tax exemptions under the Income Tax Act:","1. Section 112A (Grandfather provision): Exempts LTCG tax on equity shares/units bought before 31st January 2018, adjusting the cost to the values as of 1st February 2018.","2. Section 54: Allows tax exemption on LTCG from the sale of a house, provided capital gains are reinvested in a new residential property.","3. Section 54EC: Offers tax exemption on gains when proceeds from housing property sales are reinvested in specific bonds issued by NHAI or REC.","4. Section 54F: Grants tax exemption on gains from the sale of long-term capital assets other than house property, when sales proceeds are reinvested in a new residential property."]
    
    for i in range(len(desc_list)):
        pdf.set_font('LeagueSpartan-Light', size=px2pts(18))
        pdf.set_xy(px2MM(120),px2MM(906+(i*25))) 
        pdf.cell(px2MM(1680), px2MM(25),desc_list[i],align='L')
        
    #//*-----Index Text of Page--**////
    index_text(pdf,'#1A1A1D')
          
    
#//*----Capital Gains Taxation by Asset Type (Page 2)-----*//  
 
def capital_gains_2(pdf,json_data,c_MoneyS,money_signData):
    pdf.add_page()
    pdf.set_fill_color(*hex2RGB('#FCF8ED'))
    pdf.rect(0, 0, px2MM(1920), px2MM(1080),'F')
    
    # black rectangle
    pdf.set_fill_color(*hex2RGB('#000000'))
    pdf.rect(px2MM(0), px2MM(80), px2MM(15), px2MM(84), 'F')

    #//*----Headings----//
    pdf.set_xy(px2MM(120),px2MM(80)) 
    pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(60))
    pdf.set_text_color(*hex2RGB('#000000'))
    pdf.cell(px2MM(960), px2MM(84),'Capital Gains Taxation by Asset Type',align='L')
    
    #//*---Black vertical line--*-//
    pdf.set_fill_color(*hex2RGB('#313236'))
    pdf.rect(px2MM(120), px2MM(204), px2MM(6), px2MM(394),'F')
    
    pdf.rect(px2MM(126), px2MM(204), px2MM(80), px2MM(42),'F')
    pdf.set_xy(px2MM(141),px2MM(209)) 
    pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(24))
    pdf.set_text_color(*hex2RGB('#FFFFFF'))
    pdf.cell(px2MM(70), px2MM(32),'Debt',align='L')
    
    
    #//*----Debt Table-----------------*//
    #//*---Columns---*//
    pdf.set_draw_color(*hex2RGB('#E9EAEE'))
    pdf.set_line_width(px2MM(0.2))
    pdf.set_fill_color(*hex2RGB('#FFFFFF'))
    pdf.rect(px2MM(126), px2MM(246), px2MM(457), px2MM(104),'FD')
    
    pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(24))
    pdf.set_text_color(*hex2RGB('#1A1A1D'))
    pdf.set_xy(px2MM(146),px2MM(266)) 
    pdf.cell(px2MM(80), px2MM(64),'Asset Type',align='L')
    
    pdf.rect(px2MM(583), px2MM(246), px2MM(608.5), px2MM(52),'FD')
    pdf.set_xy(px2MM(603),px2MM(256)) 
    pdf.cell(px2MM(568.5), px2MM(32),'Long-term Capital Gains (LTCG)',align='C')
    
    pdf.rect(px2MM(1191.5), px2MM(246), px2MM(608.5), px2MM(52),'FD')
    pdf.set_xy(px2MM(1211.5),px2MM(256)) 
    pdf.cell(px2MM(568.5), px2MM(32),'Short-term Capital Gains (STCG)',align='C')
    
    pdf.rect(px2MM(583), px2MM(298), px2MM(189), px2MM(52),'FD')
    pdf.set_xy(px2MM(603),px2MM(308)) 
    pdf.cell(px2MM(155), px2MM(32),'Holding Period',align='C')
    
    pdf.rect(px2MM(772), px2MM(298), px2MM(419.5), px2MM(52),'FD')
    pdf.set_xy(px2MM(792),px2MM(308)) 
    pdf.cell(px2MM(379.5), px2MM(32),'Tax Rate',align='C')
    
    pdf.rect(px2MM(1191.5), px2MM(298), px2MM(189), px2MM(52),'FD')
    pdf.set_xy(px2MM(1211.5),px2MM(308)) 
    pdf.cell(px2MM(155), px2MM(32),'Holding Period',align='C')
    
    pdf.rect(px2MM(1380.5), px2MM(298), px2MM(419.5), px2MM(52),'FD')
    pdf.set_xy(px2MM(1400.5),px2MM(308)) 
    pdf.cell(px2MM(379.5), px2MM(32),'Tax Rate',align='C')
    
    
    #//*--row (1)---*//
    pdf.set_fill_color(*hex2RGB('#F3F6F9'))
    pdf.set_font('LeagueSpartan-Regular', size=px2pts(24))
    
    pdf.rect(px2MM(126), px2MM(350), px2MM(457), px2MM(72),'FD')
    pdf.set_xy(px2MM(146),px2MM(370)) 
    pdf.cell(px2MM(417), px2MM(32),'Debt mutual funds',align='L')
    
    pdf.rect(px2MM(583), px2MM(350), px2MM(189), px2MM(72),'FD')
    pdf.set_xy(px2MM(603),px2MM(370)) 
    pdf.cell(px2MM(155), px2MM(32),'Any',align='C')
    
    pdf.rect(px2MM(772), px2MM(350), px2MM(419.5), px2MM(72),'FD')
    pdf.set_xy(px2MM(792),px2MM(370)) 
    pdf.cell(px2MM(379.5), px2MM(32),'As per income tax slab',align='C')
    
    pdf.rect(px2MM(1191.5), px2MM(350), px2MM(189), px2MM(72),'FD')
    pdf.set_xy(px2MM(1211.5),px2MM(370)) 
    pdf.cell(px2MM(155), px2MM(32),'Any',align='C')
    
    pdf.rect(px2MM(1380.5), px2MM(350), px2MM(419.5), px2MM(72),'FD')
    pdf.set_xy(px2MM(1400.5),px2MM(370)) 
    pdf.cell(px2MM(379.5), px2MM(32),'As per income tax slab',align='C')
    
    #//*--row (2)---*//
    pdf.set_fill_color(*hex2RGB('#FFFFFF'))
    pdf.set_font('LeagueSpartan-Regular', size=px2pts(24))
    
    pdf.rect(px2MM(126), px2MM(422), px2MM(457), px2MM(104),'FD')
    pdf.set_xy(px2MM(146),px2MM(458)) 
    pdf.cell(px2MM(417), px2MM(32),'Listed/Zero coupon bonds',align='L')
    
    pdf.rect(px2MM(583), px2MM(422), px2MM(189), px2MM(104),'FD')
    pdf.set_xy(px2MM(603),px2MM(458)) 
    pdf.cell(px2MM(155), px2MM(32),'> 1 year',align='C')
    
    pdf.rect(px2MM(772), px2MM(422), px2MM(419.5), px2MM(104),'FD')
    pdf.set_xy(px2MM(792),px2MM(442)) 
    pdf.multi_cell(px2MM(379.5), px2MM(32),'20% with indexation, 10% w/o indexation, whichever is lower',align='C')
    
    pdf.rect(px2MM(1191.5), px2MM(422), px2MM(189), px2MM(104),'FD')
    pdf.set_xy(px2MM(1211.5),px2MM(458)) 
    pdf.cell(px2MM(155), px2MM(32),'< 1 year',align='C')
    
    pdf.rect(px2MM(1380.5), px2MM(422), px2MM(419.5), px2MM(104),'FD')
    pdf.set_xy(px2MM(1400.5),px2MM(458)) 
    pdf.cell(px2MM(379.5), px2MM(32),'As per income tax slab',align='C')
    
    #//*--row (3)---*//
    pdf.set_fill_color(*hex2RGB('#F3F6F9'))
    
    pdf.rect(px2MM(126), px2MM(526), px2MM(457), px2MM(72),'FD')
    pdf.set_xy(px2MM(146),px2MM(546)) 
    pdf.cell(px2MM(417), px2MM(32),'Unlisted bonds',align='L')
    
    pdf.rect(px2MM(583), px2MM(526), px2MM(189), px2MM(72),'FD')
    pdf.set_xy(px2MM(603),px2MM(546)) 
    pdf.cell(px2MM(155), px2MM(32),'> 3 years',align='C')
    
    pdf.rect(px2MM(772), px2MM(526), px2MM(419.5), px2MM(72),'FD')
    pdf.set_xy(px2MM(792),px2MM(546)) 
    pdf.multi_cell(px2MM(379.5), px2MM(32),'20% with indexation',align='C')
    
    pdf.rect(px2MM(1191.5), px2MM(526), px2MM(189), px2MM(72),'FD')
    pdf.set_xy(px2MM(1211.5),px2MM(546)) 
    pdf.cell(px2MM(155), px2MM(32),'< 3 years',align='C')
    
    pdf.rect(px2MM(1380.5), px2MM(526), px2MM(419.5), px2MM(72),'FD')
    pdf.set_xy(px2MM(1400.5),px2MM(546)) 
    pdf.cell(px2MM(379.5), px2MM(32),'As per income tax slab',align='C')
    
    
    #//*---Black vertical line--*-//
    pdf.set_fill_color(*hex2RGB('#313236'))
    pdf.rect(px2MM(120), px2MM(204), px2MM(6), px2MM(394),'F')
    
    pdf.rect(px2MM(126), px2MM(204), px2MM(80), px2MM(42),'F')
    pdf.set_xy(px2MM(141),px2MM(209)) 
    pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(24))
    pdf.set_text_color(*hex2RGB('#FFFFFF'))
    pdf.cell(px2MM(70), px2MM(32),'Debt',align='L')
    
    
    #//*----Passive Income Assets Table-----------------*//
    #//*---Black vertical line--*-//
    pdf.set_fill_color(*hex2RGB('#313236'))
    pdf.rect(px2MM(120), px2MM(638), px2MM(6), px2MM(218),'F')
    
    pdf.rect(px2MM(126), px2MM(638), px2MM(257), px2MM(42),'F')
    pdf.set_xy(px2MM(141),px2MM(643)) 
    pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(24))
    pdf.set_text_color(*hex2RGB('#FFFFFF'))
    pdf.cell(px2MM(250), px2MM(32),'Passive Income Assets',align='L')
    
    #//*---Columns---*//
    pdf.set_draw_color(*hex2RGB('#E9EAEE'))
    pdf.set_line_width(px2MM(0.2))
    pdf.set_fill_color(*hex2RGB('#FFFFFF'))
    pdf.rect(px2MM(126), px2MM(680), px2MM(457), px2MM(104),'FD')
    
    pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(24))
    pdf.set_text_color(*hex2RGB('#1A1A1D'))
    pdf.set_xy(px2MM(146),px2MM(700)) 
    pdf.cell(px2MM(80), px2MM(64),'Asset Type',align='L')
    
    pdf.rect(px2MM(583), px2MM(680), px2MM(608.5), px2MM(52),'FD')
    pdf.set_xy(px2MM(603),px2MM(690)) 
    pdf.cell(px2MM(568.5), px2MM(32),'Long-term Capital Gains (LTCG)',align='C')
    
    pdf.rect(px2MM(1191.5), px2MM(680), px2MM(608.5), px2MM(52),'FD')
    pdf.set_xy(px2MM(1211.5),px2MM(690)) 
    pdf.cell(px2MM(568.5), px2MM(32),'Short-term Capital Gains (STCG)',align='C')
    
    pdf.rect(px2MM(583), px2MM(732), px2MM(189), px2MM(52),'FD')
    pdf.set_xy(px2MM(603),px2MM(742)) 
    pdf.cell(px2MM(155), px2MM(32),'Holding Period',align='C')
    
    pdf.rect(px2MM(772), px2MM(732), px2MM(419.5), px2MM(52),'FD')
    pdf.set_xy(px2MM(792),px2MM(742)) 
    pdf.cell(px2MM(379.5), px2MM(32),'Tax Rate',align='C')
    
    pdf.rect(px2MM(1191.5), px2MM(732), px2MM(189), px2MM(52),'FD')
    pdf.set_xy(px2MM(1211.5),px2MM(742)) 
    pdf.cell(px2MM(155), px2MM(32),'Holding Period',align='C')
    
    pdf.rect(px2MM(1380.5), px2MM(732), px2MM(419.5), px2MM(52),'FD')
    pdf.set_xy(px2MM(1400.5),px2MM(742)) 
    pdf.cell(px2MM(379.5), px2MM(32),'Tax Rate',align='C')
    
    
    #//*--row (1)---*//
    pdf.set_fill_color(*hex2RGB('#F3F6F9'))
    pdf.set_font('LeagueSpartan-Regular', size=px2pts(24))
    
    pdf.rect(px2MM(126), px2MM(784), px2MM(457), px2MM(72),'FD')
    pdf.set_xy(px2MM(146),px2MM(804)) 
    pdf.cell(px2MM(417), px2MM(32),'REITs/InvITs',align='L')
    
    pdf.rect(px2MM(583), px2MM(784), px2MM(189), px2MM(72),'FD')
    pdf.set_xy(px2MM(603),px2MM(804)) 
    pdf.cell(px2MM(155), px2MM(32),'> 3 years',align='C')
    
    pdf.rect(px2MM(772), px2MM(784), px2MM(419.5), px2MM(72),'FD')
    pdf.set_xy(px2MM(792),px2MM(804)) 
    pdf.cell(px2MM(379.5), px2MM(32),'10% on LTCG > ₹ 1 lakh/year',align='C')
    
    pdf.rect(px2MM(1191.5), px2MM(784), px2MM(189), px2MM(72),'FD')
    pdf.set_xy(px2MM(1211.5),px2MM(804)) 
    pdf.cell(px2MM(155), px2MM(32),'< 3 years',align='C')
    
    pdf.rect(px2MM(1380.5), px2MM(784), px2MM(419.5), px2MM(72),'FD')
    pdf.set_xy(px2MM(1400.5),px2MM(804)) 
    pdf.cell(px2MM(379.5), px2MM(32),'15%',align='C')
    
    #//*-----Index Text of Page--**////
    index_text(pdf,'#1A1A1D')
    
    
#//*----Capital Gains Taxation by Asset Type (Page 3)-----*//      
def capital_gains_3(pdf,json_data,c_MoneyS,money_signData):
    pdf.add_page()
    pdf.set_fill_color(*hex2RGB('#FCF8ED'))
    pdf.rect(0, 0, px2MM(1920), px2MM(1080),'F')
    
    # black rectangle
    pdf.set_fill_color(*hex2RGB('#000000'))
    pdf.rect(px2MM(0), px2MM(80), px2MM(15), px2MM(84), 'F')

    #//*----Headings----//
    pdf.set_xy(px2MM(120),px2MM(80)) 
    pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(60))
    pdf.set_text_color(*hex2RGB('#000000'))
    pdf.cell(px2MM(960), px2MM(84),'Capital Gains Taxation by Asset Type',align='L')
    
    #//*---Black vertical line--*-//
    pdf.set_fill_color(*hex2RGB('#313236'))
    pdf.rect(px2MM(120), px2MM(204), px2MM(6), px2MM(572),'F')
    
    pdf.rect(px2MM(126), px2MM(204), px2MM(100), px2MM(42),'F')
    pdf.set_xy(px2MM(141),px2MM(209)) 
    pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(24))
    pdf.set_text_color(*hex2RGB('#FFFFFF'))
    pdf.cell(px2MM(80), px2MM(32),'Others',align='L')
    
    
    #//*----Equity Table-----------------*//
    #//*---Columns---*//
    pdf.set_draw_color(*hex2RGB('#E9EAEE'))
    pdf.set_line_width(px2MM(0.2))
    pdf.set_fill_color(*hex2RGB('#FFFFFF'))
    pdf.rect(px2MM(126), px2MM(246), px2MM(457), px2MM(104),'FD')
    
    pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(24))
    pdf.set_text_color(*hex2RGB('#1A1A1D'))
    pdf.set_xy(px2MM(146),px2MM(266)) 
    pdf.cell(px2MM(80), px2MM(64),'Asset Type',align='L')
    
    pdf.rect(px2MM(583), px2MM(246), px2MM(608.5), px2MM(52),'FD')
    pdf.set_xy(px2MM(603),px2MM(256)) 
    pdf.cell(px2MM(568.5), px2MM(32),'Long-term Capital Gains (LTCG)',align='C')
    
    pdf.rect(px2MM(1191.5), px2MM(246), px2MM(608.5), px2MM(52),'FD')
    pdf.set_xy(px2MM(1211.5),px2MM(256)) 
    pdf.cell(px2MM(568.5), px2MM(32),'Short-term Capital Gains (STCG)',align='C')
    
    pdf.rect(px2MM(583), px2MM(298), px2MM(189), px2MM(52),'FD')
    pdf.set_xy(px2MM(603),px2MM(308)) 
    pdf.cell(px2MM(155), px2MM(32),'Holding Period',align='C')
    
    pdf.rect(px2MM(772), px2MM(298), px2MM(419.5), px2MM(52),'FD')
    pdf.set_xy(px2MM(792),px2MM(308)) 
    pdf.cell(px2MM(379.5), px2MM(32),'Tax Rate',align='C')
    
    pdf.rect(px2MM(1191.5), px2MM(298), px2MM(189), px2MM(52),'FD')
    pdf.set_xy(px2MM(1211.5),px2MM(308)) 
    pdf.cell(px2MM(155), px2MM(32),'Holding Period',align='C')
    
    pdf.rect(px2MM(1380.5), px2MM(298), px2MM(419.5), px2MM(52),'FD')
    pdf.set_xy(px2MM(1400.5),px2MM(308)) 
    pdf.cell(px2MM(379.5), px2MM(32),'Tax Rate',align='C')
    
    #//*----Row 1(1/1)----*//
    pdf.set_fill_color(*hex2RGB('#F3F6F9'))
    pdf.rect(px2MM(126), px2MM(350), px2MM(240), px2MM(240),'FD')
    pdf.set_font('LeagueSpartan-Regular', size=px2pts(24))
    pdf.set_xy(px2MM(146),px2MM(406)) 
    pdf.multi_cell(px2MM(210), px2MM(32),'Hybrid mutual funds (<35% equity)/Market-Linked Debentures',align='L')
    
    pdf.rect(px2MM(366), px2MM(350), px2MM(217), px2MM(104),'FD')
    pdf.set_xy(px2MM(386),px2MM(370)) 
    pdf.multi_cell(px2MM(200), px2MM(32),'Purchased before 1st April,  2023',align='L')
    
    pdf.rect(px2MM(583), px2MM(350), px2MM(189), px2MM(104),'FD')
    pdf.set_xy(px2MM(603),px2MM(386)) 
    pdf.cell(px2MM(155), px2MM(32),'> 3 years',align='C')
    
    pdf.rect(px2MM(772), px2MM(350), px2MM(419.5), px2MM(104),'FD')
    pdf.set_xy(px2MM(792),px2MM(386)) 
    pdf.cell(px2MM(379.5), px2MM(32),'20% with indexation',align='C')
    
    pdf.rect(px2MM(1191.5), px2MM(350), px2MM(189), px2MM(104),'FD')
    pdf.set_xy(px2MM(1211.5),px2MM(386)) 
    pdf.cell(px2MM(155), px2MM(32),'< 3 years',align='C')
    
    pdf.rect(px2MM(1380.5), px2MM(350), px2MM(419.5), px2MM(104),'FD')
    pdf.set_xy(px2MM(1400.5),px2MM(386)) 
    pdf.cell(px2MM(379.5), px2MM(32),'As per income tax slab',align='C')
    
    #//*--row (1/2)---*//
    pdf.rect(px2MM(366), px2MM(454), px2MM(217), px2MM(136),'FD')
    pdf.set_xy(px2MM(386),px2MM(474)) 
    pdf.multi_cell(px2MM(177), px2MM(32),'Purchased on or after 1st April, 2023',align='L')
    
    pdf.rect(px2MM(583), px2MM(454), px2MM(189), px2MM(136),'FD')
    pdf.set_xy(px2MM(603),px2MM(506)) 
    pdf.cell(px2MM(155), px2MM(32),'Any',align='C')
    
    pdf.rect(px2MM(772), px2MM(454), px2MM(419.5), px2MM(136),'FD')
    pdf.set_xy(px2MM(792),px2MM(506)) 
    pdf.cell(px2MM(379.5), px2MM(32),'As per income tax slab',align='C')
    
    pdf.rect(px2MM(1191.5), px2MM(454), px2MM(189), px2MM(136),'FD')
    pdf.set_xy(px2MM(1211.5),px2MM(506)) 
    pdf.cell(px2MM(155), px2MM(32),'Any',align='C')
    
    pdf.rect(px2MM(1380.5), px2MM(454), px2MM(419.5), px2MM(136),'FD')
    pdf.set_xy(px2MM(1400.5),px2MM(506)) 
    pdf.cell(px2MM(379.5), px2MM(32),'As per income tax slab',align='C')
    
    #//*--row (2)-row(4)---*//
    asset_type= ['Hybrid mutual funds (35% - 65% equity)','Collectibles/Antiques','Cryptocurrencies/NFTs']
    long_hold = ['> 3 years','> 3 years','Any']
    long_tax = ['20% with indexation','20% with indexation','30% for all holding periods']
    short_hold = ['< 3 years','< 3 years','Any']
    short_tax = ['As per income tax slab','As per income tax slab','30% for all holding periods']
    pdf.set_font('LeagueSpartan-Regular', size=px2pts(24))
    
    col = '#FFFFFF'
    
    for i in range(len(asset_type)):
        pdf.set_fill_color(*hex2RGB(col))
        pdf.rect(px2MM(126), px2MM(590+(i*62)), px2MM(457), px2MM(62),'FD')
        pdf.set_xy(px2MM(146),px2MM(605+(i*62))) 
        pdf.cell(px2MM(366), px2MM(32),asset_type[i],align='L')
        
        pdf.rect(px2MM(583), px2MM(590+(i*62)), px2MM(189), px2MM(62),'FD')
        pdf.set_xy(px2MM(603),px2MM(605+(i*62))) 
        pdf.cell(px2MM(155), px2MM(32),long_hold[i],align='C')
        
        pdf.rect(px2MM(772), px2MM(590+(i*62)), px2MM(419.5), px2MM(62),'FD')
        pdf.set_xy(px2MM(792),px2MM(605+(i*62))) 
        pdf.cell(px2MM(379.5), px2MM(32),long_tax[i],align='C')
        
        pdf.rect(px2MM(1191.5), px2MM(590+(i*62)), px2MM(189), px2MM(62),'FD')
        pdf.set_xy(px2MM(1211.5),px2MM(605+(i*62))) 
        pdf.cell(px2MM(155), px2MM(32),short_hold[i],align='C')
        
        pdf.rect(px2MM(1380.5), px2MM(590+(i*62)), px2MM(419.5), px2MM(62),'FD')
        pdf.set_xy(px2MM(1400.5),px2MM(605+(i*62))) 
        pdf.cell(px2MM(379.5), px2MM(32),short_tax[i],align='C')
        
        if col == '#F3F6F9':
            col = '#FFFFFF'
        else:
            col = '#F3F6F9'
        
    #//*-----Index Text of Page--**////
    index_text(pdf,'#1A1A1D')
    
#//*----Capital Gains Taxation by Asset Type (Page 4)-----*//      
def capital_gains_4(pdf,json_data,c_MoneyS,money_signData):
    pdf.add_page()
    pdf.set_fill_color(*hex2RGB('#FCF8ED'))
    pdf.rect(0, 0, px2MM(1920), px2MM(1080),'F')
    
    # black rectangle
    pdf.set_fill_color(*hex2RGB('#000000'))
    pdf.rect(px2MM(0), px2MM(80), px2MM(15), px2MM(84), 'F')

    #//*----Headings----//
    pdf.set_xy(px2MM(120),px2MM(80)) 
    pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(60))
    pdf.set_text_color(*hex2RGB('#000000'))
    pdf.cell(px2MM(960), px2MM(84),'Capital Gains Taxation by Asset Type',align='L')
    
    #//*---Black vertical line--*-//
    pdf.set_fill_color(*hex2RGB('#313236'))
    pdf.rect(px2MM(120), px2MM(204), px2MM(6), px2MM(466),'F')
    
    pdf.rect(px2MM(126), px2MM(204), px2MM(100), px2MM(42),'F')
    pdf.set_xy(px2MM(141),px2MM(209)) 
    pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(24))
    pdf.set_text_color(*hex2RGB('#FFFFFF'))
    pdf.cell(px2MM(80), px2MM(32),'Others',align='L')
    
    
    #//*----Equity Table-----------------*//
    #//*---Columns---*//
    pdf.set_draw_color(*hex2RGB('#E9EAEE'))
    pdf.set_line_width(px2MM(0.2))
    pdf.set_fill_color(*hex2RGB('#FFFFFF'))
    pdf.rect(px2MM(126), px2MM(246), px2MM(457), px2MM(104),'FD')
    
    pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(24))
    pdf.set_text_color(*hex2RGB('#1A1A1D'))
    pdf.set_xy(px2MM(146),px2MM(266)) 
    pdf.cell(px2MM(80), px2MM(64),'Asset Type',align='L')
    
    pdf.rect(px2MM(583), px2MM(246), px2MM(608.5), px2MM(52),'FD')
    pdf.set_xy(px2MM(603),px2MM(256)) 
    pdf.cell(px2MM(568.5), px2MM(32),'Long-term Capital Gains (LTCG)',align='C')
    
    pdf.rect(px2MM(1191.5), px2MM(246), px2MM(608.5), px2MM(52),'FD')
    pdf.set_xy(px2MM(1211.5),px2MM(256)) 
    pdf.cell(px2MM(568.5), px2MM(32),'Short-term Capital Gains (STCG)',align='C')
    
    pdf.rect(px2MM(583), px2MM(298), px2MM(189), px2MM(52),'FD')
    pdf.set_xy(px2MM(603),px2MM(308)) 
    pdf.cell(px2MM(155), px2MM(32),'Holding Period',align='C')
    
    pdf.rect(px2MM(772), px2MM(298), px2MM(419.5), px2MM(52),'FD')
    pdf.set_xy(px2MM(792),px2MM(308)) 
    pdf.cell(px2MM(379.5), px2MM(32),'Tax Rate',align='C')
    
    pdf.rect(px2MM(1191.5), px2MM(298), px2MM(189), px2MM(52),'FD')
    pdf.set_xy(px2MM(1211.5),px2MM(308)) 
    pdf.cell(px2MM(155), px2MM(32),'Holding Period',align='C')
    
    pdf.rect(px2MM(1380.5), px2MM(298), px2MM(419.5), px2MM(52),'FD')
    pdf.set_xy(px2MM(1400.5),px2MM(308)) 
    pdf.cell(px2MM(379.5), px2MM(32),'Tax Rate',align='C')
    
    #//*----Row 1(1/1)----*//
    pdf.set_fill_color(*hex2RGB('#F3F6F9'))
    pdf.rect(px2MM(126), px2MM(350), px2MM(240), px2MM(176),'FD')
    pdf.set_font('LeagueSpartan-Regular', size=px2pts(24))
    pdf.set_xy(px2MM(146),px2MM(406)) 
    pdf.multi_cell(px2MM(210), px2MM(32),'Sovereign Gold Bonds',align='L')
    
    pdf.rect(px2MM(366), px2MM(350), px2MM(217), px2MM(72),'FD')
    pdf.set_xy(px2MM(386),px2MM(370)) 
    pdf.multi_cell(px2MM(200), px2MM(32),'Held till Maturity',align='L')
    
    pdf.rect(px2MM(583), px2MM(350), px2MM(189), px2MM(72),'FD')
    pdf.set_xy(px2MM(603),px2MM(370)) 
    pdf.cell(px2MM(155), px2MM(32),'Any',align='C')
    
    pdf.rect(px2MM(772), px2MM(350), px2MM(419.5), px2MM(72),'FD')
    pdf.set_xy(px2MM(792),px2MM(370)) 
    pdf.cell(px2MM(379.5), px2MM(32),'Exempt',align='C')
    
    pdf.rect(px2MM(1191.5), px2MM(350), px2MM(189), px2MM(72),'FD')
    pdf.set_xy(px2MM(1211.5),px2MM(370)) 
    pdf.cell(px2MM(155), px2MM(32),'Any',align='C')
    
    pdf.rect(px2MM(1380.5), px2MM(350), px2MM(419.5), px2MM(72),'FD')
    pdf.set_xy(px2MM(1400.5),px2MM(370)) 
    pdf.cell(px2MM(379.5), px2MM(32),'Exempt',align='C')
    
    #//*--row (1/2)---*//
    pdf.rect(px2MM(366), px2MM(422), px2MM(217), px2MM(104),'FD')
    pdf.set_xy(px2MM(386),px2MM(442)) 
    pdf.multi_cell(px2MM(200), px2MM(32),'Sold in secondary market',align='L')
    
    pdf.rect(px2MM(583), px2MM(422), px2MM(189), px2MM(104),'FD')
    pdf.set_xy(px2MM(603),px2MM(458)) 
    pdf.cell(px2MM(155), px2MM(32),'> 3 years',align='C')
    
    pdf.rect(px2MM(772), px2MM(422), px2MM(419.5), px2MM(104),'FD')
    pdf.set_xy(px2MM(792),px2MM(458)) 
    pdf.cell(px2MM(379.5), px2MM(32),'20% with indexation',align='C')
    
    pdf.rect(px2MM(1191.5), px2MM(422), px2MM(189), px2MM(104),'FD')
    pdf.set_xy(px2MM(1211.5),px2MM(458)) 
    pdf.cell(px2MM(155), px2MM(32),'< 3 years',align='C')
    
    pdf.rect(px2MM(1380.5), px2MM(422), px2MM(419.5), px2MM(104),'FD')
    pdf.set_xy(px2MM(1400.5),px2MM(458)) 
    pdf.cell(px2MM(379.5), px2MM(32),'As per income tax slab',align='C')
    
    
    #//*----Row 2(2/1)----*//
    pdf.set_fill_color(*hex2RGB('#FFFFFF'))
    pdf.rect(px2MM(126), px2MM(526), px2MM(240), px2MM(144),'FD')
    pdf.set_font('LeagueSpartan-Regular', size=px2pts(24))
    pdf.set_xy(px2MM(146),px2MM(566)) 
    pdf.multi_cell(px2MM(210), px2MM(32),'Non-Convertible Debentures',align='L')
    
    pdf.rect(px2MM(366), px2MM(526), px2MM(217), px2MM(72),'FD')
    pdf.set_xy(px2MM(386),px2MM(546)) 
    pdf.multi_cell(px2MM(200), px2MM(32),'Listed',align='L')
    
    pdf.rect(px2MM(583), px2MM(526), px2MM(189), px2MM(72),'FD')
    pdf.set_xy(px2MM(603),px2MM(546)) 
    pdf.cell(px2MM(155), px2MM(32),'> 1 year',align='C')
    
    pdf.rect(px2MM(772), px2MM(526), px2MM(419.5), px2MM(72),'FD')
    pdf.set_xy(px2MM(792),px2MM(546)) 
    pdf.cell(px2MM(379.5), px2MM(32),'10% without indexation',align='C')
    
    pdf.rect(px2MM(1191.5), px2MM(526), px2MM(189), px2MM(72),'FD')
    pdf.set_xy(px2MM(1211.5),px2MM(546)) 
    pdf.cell(px2MM(155), px2MM(32),'< 1 year',align='C')
    
    pdf.rect(px2MM(1380.5), px2MM(526), px2MM(419.5), px2MM(72),'FD')
    pdf.set_xy(px2MM(1400.5),px2MM(546)) 
    pdf.cell(px2MM(379.5), px2MM(32),'As per income tax slab',align='C')
    
    #//*--row (2/2)---*//
    pdf.rect(px2MM(366), px2MM(598), px2MM(217), px2MM(72),'FD')
    pdf.set_xy(px2MM(386),px2MM(618)) 
    pdf.multi_cell(px2MM(200), px2MM(32),'Unlisted',align='L')
    
    pdf.rect(px2MM(583), px2MM(598), px2MM(189), px2MM(72),'FD')
    pdf.set_xy(px2MM(603),px2MM(618)) 
    pdf.cell(px2MM(155), px2MM(32),'> 3 years',align='C')
    
    pdf.rect(px2MM(772), px2MM(598), px2MM(419.5), px2MM(72),'FD')
    pdf.set_xy(px2MM(792),px2MM(618)) 
    pdf.cell(px2MM(379.5), px2MM(32),'20% with indexation',align='C')
    
    pdf.rect(px2MM(1191.5), px2MM(598 ), px2MM(189), px2MM(72),'FD')
    pdf.set_xy(px2MM(1211.5),px2MM(618)) 
    pdf.cell(px2MM(155), px2MM(32),'< 3 years',align='C')
    
    pdf.rect(px2MM(1380.5), px2MM(598 ), px2MM(419.5), px2MM(72),'FD')
    pdf.set_xy(px2MM(1400.5),px2MM(618)) 
    pdf.cell(px2MM(379.5), px2MM(32),'As per income tax slab',align='C')
    
    #//*-----Index Text of Page--**////
    index_text(pdf,'#1A1A1D')
              
#//*---------------------Planning For Inheritance------------------*//

def planning_for_inheritance(pdf,json_data,c_MoneyS,money_signData):
    pdf.add_page()
    pdf.set_fill_color(*hex2RGB('#FCF8ED'))
    pdf.rect(0, 0, px2MM(1920), px2MM(1080),'F')

    #//*----Headings----//
    pdf.set_xy(px2MM(120),px2MM(80)) 
    pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(60))
    pdf.set_text_color(*hex2RGB('#1A1A1D'))
    pdf.cell(px2MM(614), px2MM(84),'Planning For Inheritance',align='L')
    
    #//*---Black vertical line--*-//
    pdf.set_fill_color(*hex2RGB('#000000'))
    pdf.rect(px2MM(0), px2MM(80), px2MM(15), px2MM(84),'F')
    
    header_text = "Inheritance planning is your proactive roadmap, preparing you today for the sensitive journey of inheriting wealth in the future. It's about equipping yourself from now itself with the knowledge to manage future responsibilities wisely, fostering financial stability and peace of mind, all while maintaining family harmony and strengthening ties. Here's what to keep in mind:"
    
    pdf.set_xy(px2MM(120),px2MM(224)) 
    pdf.set_font('LeagueSpartan-Regular', size=px2pts(30))
    pdf.set_text_color(*hex2RGB('#1A1A1D'))
    pdf.multi_cell(px2MM(1680), px2MM(42),header_text,align='L')
    
    y_rect = [408,470,532,594,656]
    y_text = [390,452,514,576,638]
    
    stat_text = ["In India, wealth can be inherited in four primary ways: via a will, succession laws, gifting, or through a trust.","While a will and succession laws come into play after the owner's death, gifts can be given during their lifetime.","Trusts involve a legal entity managing wealth for beneficiaries, often requiring specialized legal advice.","Always retain documents and records of the assets purchased by the previous generation.","For real estate purchased before 1st April 2001, obtain a valuation certificate from a registered property valuer for fair market value as of 1st April 2001."]
    
    pdf.set_font('LeagueSpartan-Regular', size=px2pts(30))
    pdf.set_text_color(*hex2RGB('#1A1A1D'))
    pdf.set_fill_color(*hex2RGB('#000000'))
    for i in range(len(stat_text)):
        pdf.rect(px2MM(120), px2MM(y_rect[i]), px2MM(10), px2MM(10),'F')
        pdf.set_xy(px2MM(150),px2MM(y_text[i])) 
        pdf.multi_cell(px2MM(1650), px2MM(42),stat_text[i],align='L')
        
    #//*-----Index Text of Page--**////
    index_text(pdf,'#1A1A1D')
    
    
#//*---------------------Understanding Inheritance’s Tax Implications------------------*//

def understanding_inheritance(pdf,json_data,c_MoneyS,money_signData)        :
    pdf.add_page()
    pdf.set_fill_color(*hex2RGB('#FCF8ED'))
    pdf.rect(0, 0, px2MM(1920), px2MM(1080),'F')

    #//*----Headings----//
    pdf.set_xy(px2MM(120),px2MM(80)) 
    pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(60))
    pdf.set_text_color(*hex2RGB('#1A1A1D'))
    pdf.cell(px2MM(1200), px2MM(84),'Understanding Inheritance’s Tax Implications',align='L')
    
    #//*---Black vertical line--*-//
    pdf.set_fill_color(*hex2RGB('#000000'))
    pdf.rect(px2MM(0), px2MM(80), px2MM(15), px2MM(84),'F')
    
    header_text = "Although inheritance may not have any immediate tax consequences in India, it is critical to understand the scenarios that can trigger tax liabilities. Here are some essential details:"
    
    pdf.set_xy(px2MM(120),px2MM(224)) 
    pdf.set_font('LeagueSpartan-Regular', size=px2pts(30))
    pdf.set_text_color(*hex2RGB('#1A1A1D'))
    pdf.multi_cell(px2MM(1680), px2MM(42),header_text,align='L')
    
    
    y_rect = [366,428,532,594]
    y_text = [348,514,576]
    
    stat_text = ["There’s no inheritance tax in India. However, capital gains tax is applicable when selling inherited assets.","If filing ITR 3, disclosing the inherited assets is advisable.","If the deceased earned during any part of the financial year in which they passed away, their ITR should be filed."]
    
    
    pdf.set_font('LeagueSpartan-Regular', size=px2pts(30))
    pdf.set_text_color(*hex2RGB('#1A1A1D'))
    pdf.set_fill_color(*hex2RGB('#000000'))
    for i in range(len(y_rect)):
        pdf.rect(px2MM(120), px2MM(y_rect[i]), px2MM(10), px2MM(10),'F')
        
    for i in range(len(stat_text)):
        pdf.set_xy(px2MM(150),px2MM(y_text[i])) 
        pdf.multi_cell(px2MM(1650), px2MM(42),stat_text[i],align='L')
        
        
    #//*---For 2nd Cooment----*//
    pdf.set_xy(px2MM(150),px2MM(410)) 
    pdf.cell(px2MM(1500), px2MM(42),"While calculating capital gains, cost will be the price paid by the previous owner at the time of purchase. Refer to the",align='L')    
        
    pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(30))
    pdf.set_xy(px2MM(1595),px2MM(410)) 
    pdf.cell(px2MM(150), px2MM(42),"‘Capital Gains",align='L')  
    
    pdf.set_xy(px2MM(150),px2MM(452)) 
    pdf.multi_cell(px2MM(400), px2MM(42),"Taxation by Asset Type’",align='L')   
    
    pdf.set_font('LeagueSpartan-Regular', size=px2pts(30))
    pdf.set_xy(px2MM(465),px2MM(452)) 
    pdf.multi_cell(px2MM(600), px2MM(42),"table for more details.",align='L')  
        
    #//*-----Index Text of Page--**////
    index_text(pdf,'#1A1A1D')
    
    
    
    
#//*--------------------------MF Holdings Evaluation (new)------------------------------------------------//

def mf_holding_eveluation(pdf,json_data,c_MoneyS,money_signData):
    try:
        tab_val1 = json_data["mf_holding_evaluation"]['table']
        if tab_val1==[]:
            return None

        tab_total = json_data["mf_holding_evaluation"]['total']
        tab_total['scheme_name']='Total'
        tab_total['plan']=''
        tab_total['category']=''
        tab_total['scheme_type']=''
        tab_total['fund_evaluation_quality']=''
        tab_val1.append(tab_total)
        mf_hold = pd.DataFrame.from_dict(tab_val1)

        # mf_hold_total = json_data["mf_holding_evaluation"]['total']
        mf_comment1 = json_data["mf_holding_evaluation"]['comments1']
        mf_comment2 = json_data["mf_holding_evaluation"]['comments2']
    except:
        return None
    
    #//*---LOOP FOR TABLE---*//
    if mf_hold.empty:
        return None   
    
    def mf_page_create(pdf):
        pdf.add_page()
        pdf.set_fill_color(*hex2RGB('#FCF8ED'))
        pdf.rect(0, 0, px2MM(1920), px2MM(1080),'F')

        #//*----Featured List of Financial Products----*//
        pdf.set_xy(px2MM(120),px2MM(80)) 
        pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(60))
        pdf.set_text_color(*hex2RGB('#1A1A1D'))
        pdf.cell(px2MM(602), px2MM(84),'MF Holdings Evaluation',align='L')
        
        #//*---Top Black box
        pdf.set_fill_color(*hex2RGB('#000000'))
        pdf.rect(px2MM(0), px2MM(80), px2MM(15), px2MM(84),'F') 
        
        
        #//*---Table Column Name--*//
        pdf.set_fill_color(*hex2RGB('#FFFFFF'))
        pdf.set_draw_color(*hex2RGB('#E9EAEE'))
        pdf.set_line_width(px2MM(0.2))
        
        #//*--Col 1
        pdf.rect(px2MM(120), px2MM(204), px2MM(396), px2MM(104),'FD')
        pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(24))
        pdf.set_text_color(*hex2RGB('#1A1A1D'))
        pdf.set_xy(px2MM(140),px2MM(224)) 
        pdf.cell(px2MM(356), px2MM(64),'Scheme Name',align='L')
        
        #//*--Col 2
        pdf.rect(px2MM(516), px2MM(204), px2MM(160), px2MM(104),'FD')
        pdf.set_xy(px2MM(536),px2MM(224)) 
        pdf.cell(px2MM(120), px2MM(64),'Plan',align='L')
        
        #//*--Col 3
        pdf.rect(px2MM(676), px2MM(204), px2MM(160), px2MM(104),'FD')
        pdf.set_xy(px2MM(696),px2MM(224)) 
        pdf.cell(px2MM(120), px2MM(64),'Category',align='L')
        
        #//*--Col 4
        pdf.rect(px2MM(836), px2MM(204), px2MM(230), px2MM(104),'FD')
        pdf.set_xy(px2MM(856),px2MM(224)) 
        pdf.cell(px2MM(190), px2MM(64),'Scheme Type',align='L')
        
        #//*--Col 5
        pdf.rect(px2MM(1066), px2MM(204), px2MM(212), px2MM(104),'FD')
        pdf.set_xy(px2MM(1086),px2MM(224)) 
        pdf.cell(px2MM(172), px2MM(64),'Current Value',align='R')
        
        #//*--Col 6
        pdf.rect(px2MM(1278), px2MM(204), px2MM(320), px2MM(52),'FD')
        pdf.set_xy(px2MM(1298),px2MM(214)) 
        pdf.cell(px2MM(280), px2MM(32),'Fund Evaluation',align='C')
        
        pdf.rect(px2MM(1278), px2MM(256), px2MM(160), px2MM(52),'FD')
        pdf.set_xy(px2MM(1298),px2MM(266)) 
        pdf.cell(px2MM(120), px2MM(32),'Score*',align='C')
        
        pdf.rect(px2MM(1438), px2MM(256), px2MM(160), px2MM(52),'FD')
        pdf.set_xy(px2MM(1458),px2MM(266)) 
        pdf.cell(px2MM(120), px2MM(32),'Quality',align='C')
        
        #//*--Col 7
        pdf.rect(px2MM(1598), px2MM(204), px2MM(202), px2MM(104),'FD')
        pdf.set_xy(px2MM(1618),px2MM(224)) 
        pdf.multi_cell(px2MM(162), px2MM(32),'Excess Annual Expense**',align='R')
        
    mf_page_create(pdf)    
    rect_y = mm2PX(pdf.get_y())+20
    text_y = rect_y+15
    col = '#F3F6F9'
    
    for i in range(len(mf_hold)):

        if 1080-rect_y < 124:
            mf_page_create(pdf) 
            rect_y = mm2PX(pdf.get_y())+20
            text_y = rect_y+15
            col = '#F3F6F9'
            
        h_rect = 62
        h_text = w1 = w2 = 32
        gp = 15
        
        if len(mf_hold["scheme_name"].iloc[i]) > 32 or len(mf_hold["scheme_type"].iloc[i]) > 15:
            h_rect = 94
            h_text = 64
            h_text = w1 = w2 = 64
            
        if len(mf_hold["scheme_name"].iloc[i]) > 32:
            w1 = 32
        if len(mf_hold["scheme_type"].iloc[i]) > 15:
            w2 = 32
           
        pdf.set_font('LeagueSpartan-Regular', size=px2pts(24))
        pdf.set_text_color(*hex2RGB('#1A1A1D'))
        
        if i == (len(mf_hold)-1):
            pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(24))
            pdf.set_fill_color(*hex2RGB('#B9BABE'))
            pdf.rect(px2MM(120), px2MM(rect_y), px2MM(1680), px2MM(1),'FD')
            pdf.set_fill_color(*hex2RGB('#FFFFFF'))
            pdf.set_draw_color(*hex2RGB('#E9EAEE'))
            pdf.set_line_width(px2MM(0.2))
            pdf.rect(px2MM(120), px2MM(rect_y+1), px2MM(1680), px2MM(h_rect),'FD')
        else: 
            pdf.set_fill_color(*hex2RGB(col))
            pdf.rect(px2MM(120), px2MM(rect_y), px2MM(396), px2MM(h_rect),'FD')
            pdf.rect(px2MM(516), px2MM(rect_y), px2MM(160), px2MM(h_rect),'FD')
            pdf.rect(px2MM(676), px2MM(rect_y), px2MM(160), px2MM(h_rect),'FD')
            pdf.rect(px2MM(836), px2MM(rect_y), px2MM(230), px2MM(h_rect),'FD')
            pdf.rect(px2MM(1066), px2MM(rect_y), px2MM(212), px2MM(h_rect),'FD')
            pdf.rect(px2MM(1278), px2MM(rect_y), px2MM(160), px2MM(h_rect),'FD')
            pdf.rect(px2MM(1438), px2MM(rect_y), px2MM(160), px2MM(h_rect),'FD')
            pdf.rect(px2MM(1598), px2MM(rect_y), px2MM(202), px2MM(h_rect),'FD')
        
        #//*--Col 1
        pdf.set_xy(px2MM(140),px2MM(text_y)) 
        pdf.multi_cell(px2MM(345), px2MM(w1),mf_hold["scheme_name"].iloc[i],align='L')
        
        #//*--Col 2
        pdf.set_xy(px2MM(536),px2MM(text_y)) 
        pdf.multi_cell(px2MM(120), px2MM(h_text),mf_hold["plan"].iloc[i],align='L')
        
        #//*--Col 3
        pdf.set_xy(px2MM(696),px2MM(text_y)) 
        pdf.multi_cell(px2MM(135), px2MM(h_text),mf_hold["category"].iloc[i],align='L')
        
        #//*--Col 4
        pdf.set_xy(px2MM(856),px2MM(text_y)) 
        pdf.multi_cell(px2MM(190), px2MM(w2),mf_hold["scheme_type"].iloc[i],align='L')
        
        #//*--Col 5
        locale.setlocale(locale.LC_MONETARY, 'en_IN') 
        pdf.set_xy(px2MM(1086),px2MM(text_y)) 
        if mf_hold["current_value"].iloc[i] == '-':
            pdf.multi_cell(px2MM(172), px2MM(h_text),'',align='R')
        else:   
            val1 = str(locale.currency(float(mf_hold["current_value"].iloc[i]), grouping=True))
            val1 = val1.split('.')[0] 
            val1 = '₹ '+str(format_cash2(float(mf_hold["current_value"].iloc[i])))
            # val1 = locale.currency(mf_hold["current_value"].iloc[i], grouping=True)
            pdf.multi_cell(px2MM(172), px2MM(h_text),val1,align='R')
        
        #//*--Col 6    
        pdf.set_xy(px2MM(1298),px2MM(text_y)) 
        if mf_hold["fund_evaluation_score"].iloc[i] == 0 and  i == (len(mf_hold)-1):
            pdf.multi_cell(px2MM(120), px2MM(h_text),' ',align='C')
        elif mf_hold["fund_evaluation_score"].iloc[i]=="" and not i == (len(mf_hold)-1):
            pdf.multi_cell(px2MM(120), px2MM(h_text),"-",align='C')
        else:
            pdf.multi_cell(px2MM(120), px2MM(h_text),str(int(mf_hold["fund_evaluation_score"].iloc[i])),align='C')
        
        pdf.set_xy(px2MM(1458),px2MM(text_y)) 
        if mf_hold["fund_evaluation_quality"].iloc[i]=="" and i == (len(mf_hold)-1):
            pdf.multi_cell(px2MM(120), px2MM(h_text)," ",align='C')
            
        elif mf_hold["fund_evaluation_quality"].iloc[i]=="" and not i == (len(mf_hold)-1):
            pdf.multi_cell(px2MM(120), px2MM(h_text),"-",align='C')
            
        else:
            pdf.multi_cell(px2MM(120), px2MM(h_text),mf_hold["fund_evaluation_quality"].iloc[i],align='C')
        
        #//*--Col 7
        pdf.set_xy(px2MM(1618),px2MM(text_y)) 
        if int(float(mf_hold["excess_annual_expense"].iloc[i])) == 0 or mf_hold["excess_annual_expense"].iloc[i]== "":
            pdf.multi_cell(px2MM(162), px2MM(h_text),'₹ 0.0K',align='R')
            
        else:
            val = '₹ '+str(format_cash3(float(mf_hold["excess_annual_expense"].iloc[i])))
            pdf.multi_cell(px2MM(162), px2MM(h_text),val,align='R')

        if col == '#F3F6F9':
            col = '#FFFFFF'
        else:
            col = '#F3F6F9'
            
        rect_y=mm2PX(pdf.get_y())+gp
        text_y=rect_y+15
        
        
        #//*-----Index Text of Page--**////
        rem = mm2PX(pdf.get_y())+25
    
        pdf.set_xy(px2MM(1870), px2MM(1018))  
        pdf.set_font('LeagueSpartan-Light', size=px2pts(30))
        pdf.set_text_color(*hex2RGB('#1A1A1D'))
        pdf.cell(px2MM(20), px2MM(42),str(pdf.page_no()),align='R')

    
    #//*---1st Comment----*//
    if (1080-rem) > 150:  
        for i in range (len(mf_comment1)):
            pdf.set_font('LeagueSpartan-Light', size=px2pts(18))
            pdf.set_text_color(*hex2RGB('#000000'))
            pdf.set_xy(px2MM(140),px2MM(rem+40+(i*35))) 
            pdf.cell(px2MM(1680), px2MM(25),mf_comment1[i],align='L')
    else:
        pdf.add_page()
        pdf.set_fill_color(*hex2RGB('#FCF8ED'))
        pdf.rect(0, 0, px2MM(1920), px2MM(1080),'F')

        #//*----Featured List of Financial Products----*//
        pdf.set_xy(px2MM(120),px2MM(80)) 
        pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(60))
        pdf.set_text_color(*hex2RGB('#1A1A1D'))
        pdf.cell(px2MM(602), px2MM(84),'MF Holdings Evaluation',align='L')
        
        #//*---Top Black box
        pdf.set_fill_color(*hex2RGB('#000000'))
        pdf.rect(px2MM(0), px2MM(80), px2MM(15), px2MM(84),'F') 
        
        #//*-----Index Text of Page--**////
        rem = mm2PX(pdf.get_y())+140
        index_text(pdf,'#1A1A1D') 
        for i in range (len(mf_comment1)):
            pdf.set_font('LeagueSpartan-Light', size=px2pts(18))
            pdf.set_text_color(*hex2RGB('#000000'))
            pdf.set_xy(px2MM(140),px2MM(rem+(i*35))) 
            pdf.cell(px2MM(1680), px2MM(25),mf_comment1[i],align='L')
     
            
    rem1 = mm2PX(pdf.get_y())        
    cnt = 0
    for i in range(len(mf_comment2)):
        x = len(mf_comment2[i])/130
        if x > int(x):
            x = int(x)+1
        cnt+=x
  
    #//*----Second Comment--------*//
    
    # //*--check if space is avaliable or not to print comment
    if 1080-rem1 >  (156+(cnt*42)):
        pdf.set_font('LeagueSpartan-Regular', size=px2pts(40))
        pdf.set_text_color(*hex2RGB('#000000'))
        pdf.set_xy(px2MM(120),px2MM(rem1+60)) 
        pdf.cell(px2MM(170), px2MM(56),'Comments',align='L')

        cm_cnt = 0
        pdf.set_font('LeagueSpartan-Regular', size=px2pts(30))
        pdf.set_text_color(*hex2RGB('#000000'))
        rem1+=146
        

        #//*---Point1------*//
        start_point = mf_comment2[0].split('MF commission analyser')
        pdf.set_font('LeagueSpartan-Regular', size=px2pts(30))
        pdf.set_text_color(*hex2RGB('#000000'))
        pdf.set_fill_color(*hex2RGB('#000000'))
        pdf.rect(px2MM(120),px2MM(rem1+18),px2MM(10),px2MM(10),'F')
        
        #//*---Check if link variable exist or not
        if "MF commission analyser" in mf_comment2[0]:
            pdf.set_xy(px2MM(150),px2MM(rem1))
            # pdf.multi_cell(px2MM(1650), px2MM(42),"""70% of your MF investments (by value) are in Regular plans. As a result, you might pay 0.62% of your investment value in excess commissions every year. By switching to direct plans you can enhance your returns by that much. Explore our""",align='L')
            pdf.multi_cell(px2MM(1650), px2MM(42),start_point[0],align='L')
            
            pdf.set_text_color(*hex2RGB('#3366CC'))
            pdf.set_font('LeagueSpartan-Regular','U',size=px2pts(30))
            pdf.set_xy(px2MM(1500),px2MM(rem1+42))
            pdf.cell(px2MM(300), px2MM(42),"""MF commission analyser""",align='L',link='https://1finance.co.in/calculator/mutual-funds')
            
            pdf.set_text_color(*hex2RGB('#000000'))
            pdf.set_font('LeagueSpartan-Regular',size=px2pts(30))
            pdf.set_xy(px2MM(150),px2MM(rem1+84))
            pdf.cell(px2MM(1650), px2MM(42),""" to estimate excess commissions paid by you till date.""",align='L')
            
            rem2 = mm2PX(pdf.get_y())+42
        else:
            pdf.set_xy(px2MM(150),px2MM(rem1))
            pdf.multi_cell(px2MM(1650), px2MM(42),str(mf_comment2[0]),align='L')
        
            rem2 = mm2PX(pdf.get_y())
        
        #//*---Point 2
        start_point2 = mf_comment2[1].split('MF scoring and ranking page.')
        pdf.set_font('LeagueSpartan-Regular', size=px2pts(30))
        pdf.set_text_color(*hex2RGB('#000000'))
        pdf.set_fill_color(*hex2RGB('#000000'))
        pdf.rect(px2MM(120),px2MM(rem2+28),px2MM(10),px2MM(10),'F')
        pdf.set_xy(px2MM(150),px2MM(rem2+10))
        # pdf.multi_cell(px2MM(1650), px2MM(42),"""54% of your equity MFs (by value) are high quality. Consider removing low/medium quality funds from your portfolio. Our equity MF featured list is available at the end of this report, and an evaluation of all equity MFs is available on our""",align='L')
        pdf.multi_cell(px2MM(1650), px2MM(42),start_point2[0],align='L')
        
        pdf.set_text_color(*hex2RGB('#3366CC'))
        pdf.set_font('LeagueSpartan-Regular','U',size=px2pts(30))
        pdf.set_xy(px2MM(1435),px2MM(rem2+52))
        pdf.cell(px2MM(350), px2MM(42),"""MF scoring and ranking.""",align='L',link='https://1finance.co.in/mutual-product-scoring?page=featured')
        
        #//*-----Index Text of Page--**////
        index_text(pdf,'#1A1A1D')
    else:
        pdf.add_page()
        pdf.set_fill_color(*hex2RGB('#FCF8ED'))
        pdf.rect(0, 0, px2MM(1920), px2MM(1080),'F')

        #//*----Featured List of Financial Products----*//
        pdf.set_xy(px2MM(120),px2MM(80)) 
        pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(60))
        pdf.set_text_color(*hex2RGB('#1A1A1D'))
        pdf.cell(px2MM(602), px2MM(84),'MF Holdings Evaluation',align='L')
        
        #//*---Top Black box
        pdf.set_fill_color(*hex2RGB('#000000'))
        pdf.rect(px2MM(0), px2MM(80), px2MM(15), px2MM(84),'F') 
        
        rem1 = mm2PX(pdf.get_y())+60
        
        pdf.set_font('LeagueSpartan-Regular', size=px2pts(40))
        pdf.set_text_color(*hex2RGB('#000000'))
        pdf.set_xy(px2MM(120),px2MM(rem1+60)) 
        pdf.cell(px2MM(170), px2MM(56),'Comments',align='L')

        rem1 = mm2PX(pdf.get_y())+60
        
        start_point = mf_comment2[0].split('MF commission analyser')
        pdf.set_font('LeagueSpartan-Regular', size=px2pts(30))
        pdf.set_text_color(*hex2RGB('#000000'))
        pdf.set_fill_color(*hex2RGB('#000000'))
        pdf.rect(px2MM(120),px2MM(rem1+18),px2MM(10),px2MM(10),'F')
        
        if "MF commission analyser" in mf_comment2[0]:
            pdf.set_xy(px2MM(150),px2MM(rem1))
            pdf.multi_cell(px2MM(1650), px2MM(42),start_point[0],align='L')
            
            pdf.set_text_color(*hex2RGB('#3366CC'))
            pdf.set_font('LeagueSpartan-Regular','U',size=px2pts(30))
            pdf.set_xy(px2MM(1500),px2MM(rem1+42))
            pdf.cell(px2MM(300), px2MM(42),"""MF commission analyser""",align='L',link='https://1finance.co.in/calculator/mutual-funds')
            
            pdf.set_text_color(*hex2RGB('#000000'))
            pdf.set_font('LeagueSpartan-Regular',size=px2pts(30))
            pdf.set_xy(px2MM(150),px2MM(rem1+84))
            pdf.cell(px2MM(1650), px2MM(42),""" to estimate excess commissions paid by you till date.""",align='L')
            
            rem2 = mm2PX(pdf.get_y())+42
        else:
            pdf.set_xy(px2MM(150),px2MM(rem1))
            pdf.multi_cell(px2MM(1650), px2MM(42),str(mf_comment2[0]),align='L')
        
            rem2 = mm2PX(pdf.get_y())
        
        #//*---Point 2
        start_point2 = mf_comment2[1].split('MF scoring and ranking page.')
        pdf.set_font('LeagueSpartan-Regular', size=px2pts(30))
        pdf.set_text_color(*hex2RGB('#000000'))
        pdf.set_fill_color(*hex2RGB('#000000'))
        pdf.rect(px2MM(120),px2MM(rem2+28),px2MM(10),px2MM(10),'F')
        pdf.set_xy(px2MM(150),px2MM(rem2+10))
        # pdf.multi_cell(px2MM(1650), px2MM(42),"""54% of your equity MFs (by value) are high quality. Consider removing low/medium quality funds from your portfolio. Our equity MF featured list is available at the end of this report, and an evaluation of all equity MFs is available on our""",align='L')
        pdf.multi_cell(px2MM(1650), px2MM(42),start_point2[0],align='L')
        
        pdf.set_font('LeagueSpartan-Regular','U',size=px2pts(30))
        pdf.set_text_color(*hex2RGB('#3366CC'))
        pdf.set_xy(px2MM(1435),px2MM(rem2+52))
        pdf.cell(px2MM(350), px2MM(42),"""MF scoring and ranking.""",align='L',link='https://1finance.co.in/mutual-product-scoring?page=featured')

        #//*-----Index Text of Page--**////
        index_text(pdf,'#1A1A1D') 
        


#//*----------------------------------**------------------------------------------------//

# #//*-----Insurance Policy Evaluation----*//
def insurance_policy_eveluation(pdf,json_data,c_MoneyS,money_signData):
    try:
        tab_val1 = json_data["insurance_policy_evaluation"]['table']
        if tab_val1==[]:
            return None

        tab_total = json_data["insurance_policy_evaluation"]['total']
        tab_total['policy_name']='Total'
        tab_total['plan_type']=''
        tab_total['start_date']=''
        tab_total['policy_tenure']=''
        tab_total['suggested_action']=''
        tab_val1.append(tab_total)
        insurance_policy = pd.DataFrame.from_dict(tab_val1)

    except:
        return None
    
    #//*---LOOP FOR TABLE---*//
    if insurance_policy.empty:
        return None   
    
    def insur_page_create(pdf):
        pdf.add_page()
        pdf.set_fill_color(*hex2RGB('#FCF8ED'))
        pdf.rect(0, 0, px2MM(1920), px2MM(1080),'F')

        #//*----Featured List of Financial Products----*//
        pdf.set_xy(px2MM(120),px2MM(80)) 
        pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(60))
        pdf.set_text_color(*hex2RGB('#1A1A1D'))
        pdf.cell(px2MM(602), px2MM(84),'Insurance Policy Evaluation',align='L')
        
        #//*---Top Black box
        pdf.set_fill_color(*hex2RGB('#000000'))
        pdf.rect(px2MM(0), px2MM(80), px2MM(15), px2MM(84),'F') 
        
        
        #//*---Table Column Name--*//
        pdf.set_fill_color(*hex2RGB('#FFFFFF'))
        pdf.set_draw_color(*hex2RGB('#E9EAEE'))
        pdf.set_line_width(px2MM(0.2))
        
        #//*--Heading Col 
        pdf.rect(px2MM(120), px2MM(204), px2MM(910), px2MM(52),'FD')
        pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(20))
        pdf.set_text_color(*hex2RGB('#1A1A1D'))
        pdf.set_xy(px2MM(140),px2MM(216)) 
        pdf.cell(px2MM(870), px2MM(28),'Policy Details',align='C')
        
        pdf.rect(px2MM(1030), px2MM(204), px2MM(770), px2MM(52),'FD')
        pdf.set_xy(px2MM(1050),px2MM(216)) 
        pdf.cell(px2MM(730), px2MM(28),'Policy Evaluation',align='C')
        
        #//*--Col 1
        pdf.rect(px2MM(120), px2MM(256), px2MM(210), px2MM(124),'FD')
        pdf.set_xy(px2MM(140),px2MM(304)) 
        pdf.cell(px2MM(170), px2MM(28),'Policy Name',align='L')
        
        #//*--Col 2
        pdf.rect(px2MM(330), px2MM(256), px2MM(170), px2MM(124),'FD')
        pdf.set_xy(px2MM(350),px2MM(304)) 
        pdf.cell(px2MM(130), px2MM(28),'Plan Type',align='L')
        
        #//*--Col 3
        pdf.rect(px2MM(500), px2MM(256), px2MM(130), px2MM(124),'FD')
        pdf.set_xy(px2MM(520),px2MM(304)) 
        pdf.cell(px2MM(100), px2MM(28),'Start Date',align='C')
        
        #//*--Col 4
        pdf.rect(px2MM(630), px2MM(256), px2MM(110), px2MM(124),'FD')
        pdf.set_xy(px2MM(650),px2MM(290)) 
        pdf.multi_cell(px2MM(70), px2MM(28),'Policy Tenure',align='R')
        
        #//*--Col 5
        pdf.rect(px2MM(740), px2MM(256), px2MM(130), px2MM(124),'FD')
        pdf.set_xy(px2MM(760),px2MM(290)) 
        pdf.multi_cell(px2MM(95), px2MM(28),'Annual Premium* ',align='R')
        
        #//*--Col 6
        pdf.rect(px2MM(870), px2MM(256), px2MM(160), px2MM(124),'FD')
        pdf.set_xy(px2MM(890),px2MM(304)) 
        pdf.cell(px2MM(100), px2MM(28),'Life Cover',align='R')
        
        #//*--Col 7
        pdf.rect(px2MM(1030), px2MM(256), px2MM(149), px2MM(124),'FD')
        pdf.set_xy(px2MM(1050),px2MM(276)) 
        pdf.multi_cell(px2MM(120), px2MM(28),'Premium paid till date amount',align='R')
        
        #//*--Col 8
        pdf.rect(px2MM(1179), px2MM(256), px2MM(150), px2MM(124),'FD')
        pdf.set_xy(px2MM(1199),px2MM(290)) 
        pdf.multi_cell(px2MM(110), px2MM(28),'Premium Payable',align='R')
        
        #//*--Col 9
        pdf.rect(px2MM(1329), px2MM(256), px2MM(266), px2MM(124),'FD')
        pdf.set_xy(px2MM(1349),px2MM(304)) 
        pdf.cell(px2MM(226), px2MM(28),'Suggested Action',align='L')

        #//*--Col 10
        pdf.rect(px2MM(1595), px2MM(256), px2MM(205), px2MM(124),'FD')
        pdf.set_xy(px2MM(1615),px2MM(304)) 
        pdf.cell(px2MM(165), px2MM(28),'Surrender Value**',align='R')

        
        
    insur_page_create(pdf)    
    rect_y = mm2PX(pdf.get_y())+63
    text_y = rect_y+15
    col = '#F3F6F9'
    
    for i in range(len(insurance_policy)):

        if 1080-rect_y < 182:
            insur_page_create(pdf) 
            rect_y = mm2PX(pdf.get_y())+63
            text_y = rect_y+15
            col = '#F3F6F9'
            
        h_rect = 58
        h_text = w1 = w2 = 28
        gp = 15
        
        pdf.set_text_color(*hex2RGB('#1A1A1D'))
        pdf.set_font('LeagueSpartan-Regular', size=px2pts(20))
    
        if mm2PX(pdf.get_string_width(insurance_policy["policy_name"].iloc[i])) > 190 or mm2PX(pdf.get_string_width(insurance_policy["suggested_action"].iloc[i])) > 226 :
            h_rect = 94
            h_text = 56
            h_text = w1 = w2 = 56
            
        if mm2PX(pdf.get_string_width(insurance_policy["policy_name"].iloc[i])) > 190:
            w1 = 28
        if mm2PX(pdf.get_string_width(insurance_policy["suggested_action"].iloc[i])) > 226:
            w2 = 28
            
        if i == (len(insurance_policy)-1):
            pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(20))
            pdf.set_fill_color(*hex2RGB('#B9BABE'))
            pdf.set_draw_color(*hex2RGB('#B9BABE'))
            pdf.set_line_width(px2MM(1))
            pdf.rect(px2MM(120), px2MM(rect_y), px2MM(1680), px2MM(1),'FD')
            pdf.set_fill_color(*hex2RGB('#FFFFFF'))
            pdf.set_draw_color(*hex2RGB('#E9EAEE'))
            pdf.set_line_width(px2MM(0.2))
            pdf.rect(px2MM(120), px2MM(rect_y+1), px2MM(1680), px2MM(h_rect),'FD')
        else: 
            pdf.set_fill_color(*hex2RGB(col))
            pdf.rect(px2MM(120), px2MM(rect_y), px2MM(210), px2MM(h_rect),'FD')
            pdf.rect(px2MM(330), px2MM(rect_y), px2MM(170), px2MM(h_rect),'FD')
            pdf.rect(px2MM(500), px2MM(rect_y), px2MM(130), px2MM(h_rect),'FD')
            pdf.rect(px2MM(630), px2MM(rect_y), px2MM(110), px2MM(h_rect),'FD')
            pdf.rect(px2MM(740), px2MM(rect_y), px2MM(130), px2MM(h_rect),'FD')
            pdf.rect(px2MM(870), px2MM(rect_y), px2MM(160), px2MM(h_rect),'FD')
            pdf.rect(px2MM(1030), px2MM(rect_y), px2MM(149), px2MM(h_rect),'FD')
            pdf.rect(px2MM(1179), px2MM(rect_y), px2MM(150), px2MM(h_rect),'FD')
            pdf.rect(px2MM(1329), px2MM(rect_y), px2MM(266), px2MM(h_rect),'FD')
            pdf.rect(px2MM(1595), px2MM(rect_y), px2MM(205), px2MM(h_rect),'FD')
        
        #//*--Col 1
        pdf.set_xy(px2MM(140),px2MM(text_y)) 
        pdf.multi_cell(px2MM(190), px2MM(w1),insurance_policy["policy_name"].iloc[i],align='L')
        
        #//*--Col 2
        pdf.set_xy(px2MM(350),px2MM(text_y)) 
        pdf.multi_cell(px2MM(150), px2MM(h_text),insurance_policy["plan_type"].iloc[i],align='L')
        
        #//*--Col 3
        pdf.set_xy(px2MM(520),px2MM(text_y)) 
        pdf.multi_cell(px2MM(100), px2MM(h_text),insurance_policy["start_date"].iloc[i],align='C')
        
        #//*--Col 4
        pdf.set_xy(px2MM(650),px2MM(text_y)) 
        pdf.multi_cell(px2MM(70), px2MM(h_text),insurance_policy["policy_tenure"].iloc[i],align='R')
        
        #//*--Col 5
        pdf.set_xy(px2MM(760),px2MM(text_y)) 
        if insurance_policy["annual_premium"].iloc[i] == "":
            pdf.multi_cell(px2MM(90), px2MM(h_text),'₹ 0.0K',align='R')
        else:   
            val1 = '₹ '+str(format_cash2(float(insurance_policy["annual_premium"].iloc[i])))
            pdf.multi_cell(px2MM(90), px2MM(h_text),val1,align='R')
        
        #//*--Col 6    
        pdf.set_xy(px2MM(890),px2MM(text_y)) 
        if insurance_policy["life_cover"].iloc[i] == 0:
            pdf.multi_cell(px2MM(120), px2MM(h_text),'₹ 0.0K',align='R')
        else:
            val1 = '₹ '+str(format_cash2(float(insurance_policy["life_cover"].iloc[i])))
            pdf.multi_cell(px2MM(120), px2MM(h_text),val1,align='R')
            
        #//*--Col 7
        pdf.set_xy(px2MM(1050),px2MM(text_y)) 
        if insurance_policy["premium_paid_till_date"].iloc[i]=="":
            pdf.multi_cell(px2MM(120), px2MM(h_text),"₹ 0.0K",align='R')
        else:
            val1 = '₹ '+str(format_cash2(float(insurance_policy["premium_paid_till_date"].iloc[i])))
            pdf.multi_cell(px2MM(120), px2MM(h_text),val1,align='R')
                   
        #//*--Col 8
        pdf.set_xy(px2MM(1199),px2MM(text_y)) 
        if insurance_policy["premium_payable"].iloc[i]=="":
            pdf.multi_cell(px2MM(110), px2MM(h_text),"₹ 0.0K",align='R')
        else:
            val = '₹ '+str(format_cash3(float(insurance_policy["premium_payable"].iloc[i])))
            pdf.multi_cell(px2MM(110), px2MM(h_text),val,align='R')
            
            
        #//*--Col 9
        pdf.set_xy(px2MM(1349),px2MM(text_y)) 
        pdf.multi_cell(px2MM(246), px2MM(w2),insurance_policy["suggested_action"].iloc[i],align='L')
        
        #//*--Col 10
        pdf.set_xy(px2MM(1605),px2MM(text_y)) 
        if insurance_policy["surrender_value"].iloc[i]== "":
            pdf.multi_cell(px2MM(185), px2MM(h_text),'-',align='R')    
        elif insurance_policy["surrender_value"].iloc[i].isdigit() or (insurance_policy["surrender_value"].iloc[i].count('.') == 1 and insurance_policy["surrender_value"].iloc[i].replace('.', '').isdigit()):
            val = '₹ '+str(format_cash3(float(insurance_policy["surrender_value"].iloc[i])))
            pdf.multi_cell(px2MM(185), px2MM(h_text),val,align='R')  
        else:
            pdf.multi_cell(px2MM(185), px2MM(h_text),insurance_policy["surrender_value"].iloc[i],align='R')
            


        if col == '#F3F6F9':
            col = '#FFFFFF'
        else:
            col = '#F3F6F9'
            
        rect_y=mm2PX(pdf.get_y())+gp
        text_y=rect_y+15
        
        
        #//*-----Index Text of Page--**////
        rem = mm2PX(pdf.get_y())+57
    
        pdf.set_xy(px2MM(1870), px2MM(1018))  
        pdf.set_font('LeagueSpartan-Light', size=px2pts(30))
        pdf.set_text_color(*hex2RGB('#1A1A1D'))
        pdf.cell(px2MM(20), px2MM(42),str(pdf.page_no()),align='R')
        
    #//*----Add page for comments----*//
        
    def ins_page_add(pdf):
        pdf.add_page()
        pdf.set_fill_color(*hex2RGB('#FCF8ED'))
        pdf.rect(0, 0, px2MM(1920), px2MM(1080),'F')

        #//*----Featured List of Financial Products----*//
        pdf.set_xy(px2MM(120),px2MM(80)) 
        pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(60))
        pdf.set_text_color(*hex2RGB('#1A1A1D'))
        pdf.cell(px2MM(698), px2MM(84),'Insurance Policy Evaluation',align='L')
        
        #//*---Top Black box
        pdf.set_fill_color(*hex2RGB('#000000'))
        pdf.rect(px2MM(0), px2MM(80), px2MM(15), px2MM(84),'F') 
        
        #//*-----Index Text of Page--**////
        rem = mm2PX(pdf.get_y())+120
        index_text(pdf,'#1A1A1D') 
        return rem

    suggested_action = list(set(insurance_policy["suggested_action"].tolist()))
    suggested_action = [x.lower() for x in suggested_action]
    pt_no = 1
 

    #//*---1st Comment----*//
    tab_comments = ['* All premium amounts are converted to annual figures based on the frequency of premium payments (quarterly, semi-annual, etc.)','** Surrender value is an estimate derived from the general surrender value factor applied to insurance policies in case of surrender. The total surrender value excludes ULIPs and Annuities.']
    if not (1080-rem) > 100 :
        rem = ins_page_add(pdf)
        
    for i in range(2):
        pdf.set_font('LeagueSpartan-Light', size=px2pts(18))
        pdf.set_text_color(*hex2RGB('#000000'))
        pdf.set_xy(px2MM(120),px2MM(rem)) 
        pdf.cell(px2MM(1680), px2MM(25),tab_comments[i],align='L')
        rem = mm2PX(pdf.get_y())+35
        
    #//*---4 points Comments----*//
            
    if 'surrender' in suggested_action:
        if not (1080-rem) > 330:
            rem = ins_page_add(pdf)
        else:
            rem = mm2PX(pdf.get_y())+75
            
        pdf.set_font('LeagueSpartan-Regular', size=px2pts(40))
        pdf.set_text_color(*hex2RGB('#000000'))
        pdf.set_xy(px2MM(120),px2MM(rem)) 
        pdf.cell(px2MM(170), px2MM(56),'Comments',align='L')
        
        rem = mm2PX(pdf.get_y())+86
        
        #//*--1st point
        pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(30))
        pdf.set_text_color(*hex2RGB('#1A1A1D'))
        pdf.set_xy(px2MM(120),px2MM(rem)) 
        pdf.cell(px2MM(20), px2MM(42),str(pt_no)+'.',align='L')
        
        pdf.set_font('LeagueSpartan-Regular', size=px2pts(30))
        pdf.set_xy(px2MM(150),px2MM(rem)) 
        pdf.cell(px2MM(200), px2MM(42),'For policies where,',align='L')
        
        pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(30))
        pdf.set_xy(px2MM(385),px2MM(rem)) 
        pdf.cell(px2MM(100), px2MM(42),'"Surrender"',align='L')
        
        pdf.set_font('LeagueSpartan-Regular', size=px2pts(30))
        pdf.set_xy(px2MM(545),px2MM(rem)) 
        pdf.cell(px2MM(250), px2MM(42),"is suggested:",align='L')
        
        rem = mm2PX(pdf.get_y())+58
        pdf.set_fill_color(*hex2RGB('#000000'))
        pdf.rect(px2MM(150), px2MM(rem+20), px2MM(10), px2MM(10),'F')
        
        pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(30))
        pdf.set_text_color(*hex2RGB('#1A1A1D'))
        pdf.set_xy(px2MM(180),px2MM(rem)) 
        pdf.cell(px2MM(160), px2MM(42),'Endowment:',align='L')
        
        pdf.set_font('LeagueSpartan-Regular', size=px2pts(30))
        pdf.set_xy(px2MM(347),px2MM(rem)) 
        pdf.cell(px2MM(1450), px2MM(42),'Unless the policy is linked to a financial goal, limit your losses by surrendering as the yield is generally low at 3-4% p.a.',align='L')
        
        
        if not (1080-rem) > 220:
            rem = ins_page_add(pdf)
        else:
            rem = mm2PX(pdf.get_y())+58
            
        pdf.rect(px2MM(150), px2MM(rem+20), px2MM(10), px2MM(10),'F')
        pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(30))
        pdf.set_text_color(*hex2RGB('#1A1A1D'))
        pdf.set_xy(px2MM(180),px2MM(rem)) 
        pdf.cell(px2MM(200), px2MM(42),'ULIP Policies:',align='L')
        
        pdf.set_font('LeagueSpartan-Regular', size=px2pts(30))
        pdf.set_xy(px2MM(360),px2MM(rem)) 
        pdf.cell(px2MM(1450), px2MM(42),'Unless the policy is linked to a financial goal, surrendering is better because charges related to administration, fund',align='L')
        
        pdf.set_xy(px2MM(180),px2MM(rem+42)) 
        pdf.cell(px2MM(1650), px2MM(42),'management, premium allocation, mortality, etc. greatly reduce the investible amount.',align='L')
        
        if not (1080-rem) > 220:
            rem = ins_page_add(pdf)
        else:
            rem = mm2PX(pdf.get_y())+58
            
        pdf.rect(px2MM(150),px2MM(rem+20), px2MM(10), px2MM(10),'F')
        pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(30))
        pdf.set_text_color(*hex2RGB('#1A1A1D'))
        pdf.set_xy(px2MM(180),px2MM(rem)) 
        pdf.cell(px2MM(200), px2MM(42),'Annuity:',align='L')
        
        pdf.set_font('LeagueSpartan-Regular', size=px2pts(30))
        pdf.set_xy(px2MM(300),px2MM(rem)) 
        pdf.cell(px2MM(1450), px2MM(42),'Avoid incurring surrender charges when the policy has less time to maturity.',align='L')
        
        if not (1080-rem) > 226:
            rem = ins_page_add(pdf)
        else:
            rem = mm2PX(pdf.get_y())+58
            
        pdf.rect(px2MM(150), px2MM(rem+20), px2MM(10), px2MM(10),'F')
        pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(30))
        pdf.set_text_color(*hex2RGB('#1A1A1D'))
        pdf.set_xy(px2MM(180),px2MM(rem)) 
        pdf.cell(px2MM(200), px2MM(42),'Whole Life Policies:',align='L')
        
        pdf.set_font('LeagueSpartan-Regular', size=px2pts(30))
        pdf.set_xy(px2MM(430),px2MM(rem)) 
        pdf.cell(px2MM(1450), px2MM(42),'Choosing an annuity plan with an insurer is comparatively costly due to charges related to administration, fund',align='L')
        
        pdf.set_xy(px2MM(180),px2MM(rem+42)) 
        pdf.multi_cell(px2MM(1650), px2MM(42),'management, premium allocation, mortality (if applicable), etc. Consider standalone pension solutions like NPS, which have low fees, more investment options, and better long-term returns.',align='L')
        pt_no+=1
        
    if 'stop premium payment' in suggested_action:
    
        if not (1080-rem) > 220:
            rem = ins_page_add(pdf)
        else:
            rem = mm2PX(pdf.get_y())+40
        
        #//*--2nd point
        pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(30))
        pdf.set_text_color(*hex2RGB('#1A1A1D'))
        pdf.set_xy(px2MM(120),px2MM(rem)) 
        pdf.cell(px2MM(20), px2MM(42),str(pt_no)+'.',align='L')
        
        pdf.set_font('LeagueSpartan-Regular', size=px2pts(30))
        pdf.set_xy(px2MM(150),px2MM(rem)) 
        pdf.cell(px2MM(300), px2MM(42),'For policies where,',align='L')
        
        pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(30))
        pdf.set_xy(px2MM(380),px2MM(rem)) 
        pdf.cell(px2MM(300), px2MM(42),'"Stop premium payment"',align='L')
        
        pdf.set_font('LeagueSpartan-Regular', size=px2pts(30))
        pdf.set_xy(px2MM(715),px2MM(rem)) 
        pdf.cell(px2MM(250), px2MM(42),"is suggested:",align='L')
        
        rem = mm2PX(pdf.get_y())+58
        pdf.set_fill_color(*hex2RGB('#000000'))
        pdf.rect(px2MM(150), px2MM(rem+20), px2MM(10), px2MM(10),'F')
        
        pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(30))
        pdf.set_text_color(*hex2RGB('#1A1A1D'))
        pdf.set_xy(px2MM(180),px2MM(rem)) 
        pdf.cell(px2MM(160), px2MM(42),'Endowment:',align='L')
        
        pdf.set_font('LeagueSpartan-Regular', size=px2pts(30))
        pdf.set_xy(px2MM(347),px2MM(rem)) 
        pdf.cell(px2MM(1450), px2MM(42),'Limit your loss early on as surrendering after 3 years leads to significant value erosion. Also, the yield on such traditional ',align='L')
        
        pdf.set_font('LeagueSpartan-Regular', size=px2pts(30))
        pdf.set_xy(px2MM(180),px2MM(rem+42)) 
        pdf.cell(px2MM(1450), px2MM(42),'products is generally low at 3-4% p.a.',align='L')
        
        pt_no+=1
        
    if 'continue till maturity' in suggested_action:
    
        if not (1080-rem) > 300:
            rem = ins_page_add(pdf)
        else:
            rem = mm2PX(pdf.get_y())+80
            
        #//*--3rd point
        pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(30))
        pdf.set_text_color(*hex2RGB('#1A1A1D'))
        pdf.set_xy(px2MM(120),px2MM(rem)) 
        pdf.cell(px2MM(20), px2MM(42),str(pt_no)+'.',align='L')
        
        pdf.set_font('LeagueSpartan-Regular', size=px2pts(30))
        pdf.set_xy(px2MM(150),px2MM(rem)) 
        pdf.cell(px2MM(300), px2MM(42),'For policies where,',align='L')
        
        pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(30))
        pdf.set_xy(px2MM(380),px2MM(rem)) 
        pdf.cell(px2MM(300), px2MM(42),'"Continue till Maturity"',align='L')
        
        pdf.set_font('LeagueSpartan-Regular', size=px2pts(30))
        pdf.set_xy(px2MM(685),px2MM(rem)) 
        pdf.cell(px2MM(250), px2MM(42),"is suggested:",align='L')
        
        rem = mm2PX(pdf.get_y())+58
        pdf.set_fill_color(*hex2RGB('#000000'))
        pdf.rect(px2MM(150), px2MM(rem+20), px2MM(10), px2MM(10),'F')
        
        pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(30))
        pdf.set_text_color(*hex2RGB('#1A1A1D'))
        pdf.set_xy(px2MM(180),px2MM(rem)) 
        pdf.cell(px2MM(300), px2MM(42),'Whole Life Plan:',align='L')
        
        pdf.set_font('LeagueSpartan-Regular', size=px2pts(30))
        pdf.set_xy(px2MM(400),px2MM(rem)) 
        pdf.cell(px2MM(1450), px2MM(42),'Avoid incurring surrender charges when the policy has less time to maturity.',align='L')
        
        pt_no+=1
    
    if 'continue till lock-in period' in suggested_action:

        if not (1080-rem) > 320:
            rem = ins_page_add(pdf)
        else:
            rem = mm2PX(pdf.get_y())+80
        
        #//*--2nd point
        pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(30))
        pdf.set_text_color(*hex2RGB('#1A1A1D'))
        pdf.set_xy(px2MM(120),px2MM(rem)) 
        pdf.cell(px2MM(20), px2MM(42),str(pt_no)+'.',align='L')
        
        pdf.set_font('LeagueSpartan-Regular', size=px2pts(30))
        pdf.set_xy(px2MM(150),px2MM(rem)) 
        pdf.cell(px2MM(300), px2MM(42),'For policies where,',align='L')
        
        pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(30))
        pdf.set_xy(px2MM(380),px2MM(rem)) 
        pdf.cell(px2MM(500), px2MM(42),'"Continue till lock-in period"',align='L')
        
        pdf.set_font('LeagueSpartan-Regular', size=px2pts(30))
        pdf.set_xy(px2MM(750),px2MM(rem)) 
        pdf.cell(px2MM(250), px2MM(42),"is suggested:",align='L')
        
        rem = mm2PX(pdf.get_y())+58
        pdf.set_fill_color(*hex2RGB('#000000'))
        pdf.rect(px2MM(150), px2MM(rem+20), px2MM(10), px2MM(10),'F')
        
        pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(30))
        pdf.set_text_color(*hex2RGB('#1A1A1D'))
        pdf.set_xy(px2MM(180),px2MM(rem)) 
        pdf.cell(px2MM(160), px2MM(42),'ULIPS:',align='L')
        
        pdf.set_font('LeagueSpartan-Regular', size=px2pts(30))
        pdf.set_xy(px2MM(270),px2MM(rem)) 
        pdf.cell(px2MM(1450), px2MM(42),'Continue the policy until the end of the lock-in period, as early surrender will move your funds to the discontinuance fund, the',align='L')
        
        pdf.set_font('LeagueSpartan-Regular', size=px2pts(30))
        pdf.set_xy(px2MM(180),px2MM(rem+42)) 
        pdf.multi_cell(px2MM(1650), px2MM(42),"life cover will cease and the fund management charges will still be levied. Post-lock-in period, you will receive the policy's fund value upon surrendering.",align='L')
        
        


#//*----------------------------------**------------------------------------------------//
#//*-----Insurance Policy Evaluation----*//
def insurance_policy_recommendation_summary(pdf,json_data,c_MoneyS,money_signData):
    try:
        tab_val2 = json_data["insurance_policy_evaluation"]['recommendation_table']
        if tab_val2==[]:
            return None
    except:
        return None 
    
    insurance_policy_recommendation = pd.DataFrame.from_dict(tab_val2)
    if insurance_policy_recommendation.empty:
        return None
    
    plan = insurance_policy_recommendation['plan'].tolist()
    cover = insurance_policy_recommendation['cover'].tolist()
    annual_premium = insurance_policy_recommendation['annual_premium'].tolist()
    
    pdf.add_page()
    pdf.set_fill_color(*hex2RGB('#FCF8ED'))
    pdf.rect(0, 0, px2MM(1920), px2MM(1080),'F')

    #//*----Featured List of Financial Products----*//
    pdf.set_xy(px2MM(120),px2MM(80)) 
    pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(60))
    pdf.set_text_color(*hex2RGB('#1A1A1D'))
    pdf.cell(px2MM(602), px2MM(84),'Insurance Policy Evaluation',align='L')
    
    #//*---Top Black box
    pdf.set_fill_color(*hex2RGB('#000000'))
    pdf.rect(px2MM(0), px2MM(80), px2MM(15), px2MM(84),'F') 
    
    #//*----Heading Statements----*//
    statement1 = "By separating your insurance and investment needs, you can increase your life coverage significantly (with term insurance) and earn better returns on your investments (with instruments like mutual funds)."
    statement2 = "Refer to our “Financial Products Featured List” section for high-quality term insurance and mutual fund options."
    
    pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(30))
    pdf.set_text_color(*hex2RGB('#000000'))
    pdf.set_xy(px2MM(120),px2MM(204)) 
    pdf.multi_cell(px2MM(1680), px2MM(42),statement1,align='L')
    
    pdf.set_xy(px2MM(120),px2MM(308)) 
    pdf.multi_cell(px2MM(1680), px2MM(42),statement2,align='L')
    
    pdf.rect(px2MM(120), px2MM(410), px2MM(315), px2MM(42),'F') 
    pdf.set_fill_color(*hex2RGB('#313236'))
    pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(24))
    pdf.set_text_color(*hex2RGB('#FFFFFF'))
    pdf.set_xy(px2MM(135),px2MM(415)) 
    pdf.multi_cell(px2MM(300), px2MM(32),'Recommendation Summary',align='L')
    
    pdf.set_draw_color(*hex2RGB('#E9EAEE'))
    pdf.set_line_width(px2MM(0.02))
    
    #//*---Table Rectangles---*//
    for i in range(3):
        if i == 1:
            pdf.set_fill_color(*hex2RGB('#F3F6F9'))
        else:
            pdf.set_fill_color(*hex2RGB('#FFFFFF'))
        pdf.rect(px2MM(120), px2MM(452+(i*52)), px2MM(646), px2MM(52),'FD') 
        pdf.rect(px2MM(766), px2MM(452+(i*52)), px2MM(517), px2MM(52),'FD') 
        pdf.rect(px2MM(1283), px2MM(452+(i*52)), px2MM(517), px2MM(52),'FD')
        
    #//*---For last Row (Full White)    
    pdf.set_fill_color(*hex2RGB('#B9BABE'))
    pdf.set_draw_color(*hex2RGB('#B9BABE'))
    pdf.set_line_width(px2MM(1))
    # pdf.rect(px2MM(126), px2MM(mm2PX(tot_height)+43), px2MM(1674), px2MM(1),'F') 
    # pdf.set_fill_color(*hex2RGB('#E9EAEE'))    
    pdf.rect(px2MM(120), px2MM(608), px2MM(1680), px2MM(1),'FD')
    pdf.set_fill_color(*hex2RGB('#FFFFFF'))   
    pdf.set_draw_color(*hex2RGB('#E9EAEE'))
    pdf.set_line_width(px2MM(0.2)) 
    pdf.rect(px2MM(120), px2MM(609), px2MM(1680), px2MM(52),'FD')
        
    #//*---Column Names---------*//
    #//*---Col 1
    pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(24))
    pdf.set_text_color(*hex2RGB('#1A1A1D'))
    pdf.set_xy(px2MM(140),px2MM(462)) 
    pdf.multi_cell(px2MM(350), px2MM(32),'',align='L')
    
    #//*---Col 2
    pdf.set_xy(px2MM(786),px2MM(462)) 
    pdf.multi_cell(px2MM(477), px2MM(32),'Cover',align='C')
    
    #//*--col 3
    pdf.set_xy(px2MM(1303),px2MM(462)) 
    pdf.multi_cell(px2MM(477), px2MM(32),'Annual Premium',align='C')
    

    for i in range(3):
        #//*---Field Values---------*//
        #//*---Col 1
        pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(24))
        pdf.set_text_color(*hex2RGB('#1A1A1D'))
        pdf.set_xy(px2MM(140),px2MM(514+(i*52))) 
        pdf.multi_cell(px2MM(350), px2MM(32),plan[i],align='L')
        
        #//*---Col 2
        pdf.set_font('LeagueSpartan-Regular', size=px2pts(24))  
        pdf.set_xy(px2MM(786),px2MM(514+(i*52))) 
        if cover[i]== "":
            pdf.multi_cell(px2MM(477), px2MM(32),"",align='C')
        else:
            val1 = "₹ "+str(format_cash2(float(cover[i])))
            pdf.multi_cell(px2MM(477), px2MM(32),val1,border='0',align='C')
            
        if i == 1:
            deg_one = mm2PX(pdf.get_x())+3
            deg_one = 1024+(mm2PX(pdf.get_string_width(str(val1)))/2)
            pdf.set_font('LeagueSpartan-Regular', size=px2pts(9))
            pdf.set_xy(px2MM(deg_one),px2MM(514+(i*52)+4)) 
            pdf.multi_cell(px2MM(15), px2MM(12),"1",border='0',align='L')
        
        #//*--col 3
        pdf.set_xy(px2MM(1303),px2MM(514+(i*52))) 
        if i == 2:
            pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(24))
            pdf.set_xy(px2MM(1303),px2MM(514+(i*52)+1)) 
        else:
            pdf.set_font('LeagueSpartan-Regular', size=px2pts(24))
        if annual_premium[i] == "" :   
            pdf.multi_cell(px2MM(477), px2MM(32),' ',align='C')
        else:
            val2 = "₹ "+str(format_cash2(float(annual_premium[i])))
            pdf.multi_cell(px2MM(477), px2MM(32),val2,align='C')
            
        deg_two = 1541+(mm2PX(pdf.get_string_width(str(val2)))/2)    
        pdf.set_font('LeagueSpartan-Regular', size=px2pts(9))
        if i == 1:
            pdf.set_xy(px2MM(deg_two),px2MM(514+(i*52)+4)) 
            pdf.multi_cell(px2MM(15), px2MM(12),"2",border='0',align='L')
        elif i ==2:
            pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(9))
            pdf.set_xy(px2MM(deg_two),px2MM(514+(i*52)+4)) 
            pdf.multi_cell(px2MM(15), px2MM(12),"3",border='0',align='L')
            
    #//*---Comments----*//
    comments = ["Estimated based on your need-based analysis, considering the identified mortality protection gap.","Estimated using your age, gender, the above cover, and coverage until the age of 65 years, for an affordable policy. The exact premium may vary depending on other factors like policy tenure, cover amount, life insurer, etc.","Net savings in premiums can be reinvested in high-quality instruments."]
    comm_num = [1,2,3]
    for i in range(3) :
        pdf.set_font('LeagueSpartan-Light', size=px2pts(8))
        pdf.set_xy(px2MM(120),px2MM(697+(i*25)+10)) 
        pdf.multi_cell(px2MM(15), px2MM(25),str(comm_num[i]),align='L')
        
        pdf.set_font('LeagueSpartan-Light', size=px2pts(18))
        pdf.set_xy(px2MM(127),px2MM(700+(i*25)+10)) 
        pdf.multi_cell(px2MM(1680), px2MM(25),comments[i],align='L')
        
    #//*-----Index Text of Page--**////
    index_text(pdf,'#1A1A1D')
        
              
#//*-----Planning Your Estate and Will
def planning_your_esate_will(pdf,json_data,c_MoneyS,money_signData):
    pdf.add_page()
    pdf.set_fill_color(*hex2RGB('#FCF8ED'))
    pdf.rect(0,0, px2MM(1920),px2MM(1080),'F')
    pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(60))
    pdf.set_text_color(*hex2RGB('#1A1A1D'))
    pdf.set_xy(px2MM(120),px2MM(80))  
    pdf.cell(px2MM(8018),px2MM(84),"Planning Your Estate and Will")

    pdf.set_fill_color(*hex2RGB('#000000'))
    pdf.rect(0,px2MM(80), px2MM(15),px2MM(84),'F')


    pdf.set_font('LeagueSpartan-Regular', size=px2pts(30))
    pdf.set_xy(px2MM(120), px2MM(224))
    pdf.multi_cell(px2MM(1500), px2MM(42), "Creating a will is essential for individuals who wish to distribute their assets as per their preference after their demise. In India, legal heirs differ based on religious identity, and it is crucial to understand the legalities involved in estate and will planning. Here are some key points to keep in mind:")

    pdf.set_fill_color(*hex2RGB('#000000'))

    # Rectangle Bullets in PDF.
    pdf.rect(px2MM(120),px2MM(408),px2MM(10), px2MM(10), 'F')
    pdf.rect(px2MM(120),px2MM(470),px2MM(10), px2MM(10), 'F')
    pdf.rect(px2MM(120),px2MM(574),px2MM(10), px2MM(10), 'F')
    pdf.rect(px2MM(120),px2MM(636),px2MM(10), px2MM(10), 'F')
    pdf.rect(px2MM(120),px2MM(740),px2MM(10), px2MM(10), 'F')
    
    
    
    pdf.set_xy(px2MM(150), px2MM(390))
    pdf.multi_cell(px2MM(1470),px2MM(42),"Any adult over the age of 18 with sound mind can create a will that outlines the distribution of assets.",align = "L")

    pdf.set_xy(px2MM(150), px2MM(452))
    pdf.multi_cell(px2MM(1470),px2MM(42), "Legal heirs differ based on religious identity, with Hindus following the Hindu Succession Act of 1956, Muslims following the Muslim Personal Law, and other Indians following the Indian Succession Act of 1925.",align = "L")
    
    pdf.set_xy(px2MM(150) , px2MM(556))
    pdf.multi_cell(px2MM(1470), px2MM(42), "Nominees are caretakers and not owners of the assets, and the assets will later be distributed to legal heirs.",align = "L")

    pdf.set_xy(px2MM(150), px2MM(618))
    pdf.multi_cell(px2MM(1480),px2MM(42), "Joint accounts are considered as equal ownership between individuals, so the survivor does not become the owner. Half the wealth is distributed to legal heirs or according to the will.",align = "L")

    pdf.set_xy(px2MM(150), px2MM(722))
    pdf.multi_cell(px2MM(1470),px2MM(42),"It is not mandatory for a will to be stamped, typed, or registered, and it only requires the individual's signature along with two other witnesses.",align = "L")
    
    #//*-----Index Text of Page--**////
    index_text(pdf,'#1A1A1D')


#//*-----Building a Strong Credit Profile
def building_strong_credit_profile(pdf,json_data,c_MoneyS,money_signData):
    pdf.add_page()
    pdf.set_fill_color(*hex2RGB('#FCF8ED'))
    pdf.rect(0,0, px2MM(1920),px2MM(1080),'F')
    pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(60))
    pdf.set_text_color(*hex2RGB('#1A1A1D'))
    pdf.set_xy(px2MM(120),px2MM(80))  
    pdf.cell(px2MM(8018),px2MM(84),"Building a Strong Credit Profile")

    pdf.set_fill_color(*hex2RGB('#000000'))
    pdf.rect(0,px2MM(80), px2MM(15),px2MM(84),'F')

    pdf.set_font('LeagueSpartan-Regular', size=px2pts(30))
    pdf.set_xy(px2MM(120), px2MM(224))
    pdf.multi_cell(px2MM(1500), px2MM(42), "A strong credit profile is critical for financial stability and securing credit. Cultivating good financial habits can help you achieve a healthy credit profile. Here are some valuable tips to consider:")

    pdf.set_fill_color(*hex2RGB('#000000'))

    # Rectangle Bullets in PDF.
    for i in range(6):
        pdf.rect(px2MM(120),px2MM(366 + (i * 104)),px2MM(10), px2MM(10), 'F')


    pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(30))
    pdf.set_xy(px2MM(150), px2MM(348))
    pdf.multi_cell(px2MM(500),px2MM(42), "Make timely payments - ",align = "L")

    pdf.set_font('LeagueSpartan-Regular', size=px2pts(30))
    pdf.set_xy(px2MM(475), px2MM(348))
    pdf.cell(px2MM(1470 -200),px2MM(42), "Consistently paying your bills on time demonstrates your creditworthiness and protects you")

    pdf.set_xy(px2MM(150), px2MM(348+42))
    pdf.multi_cell(px2MM(1470),px2MM(42), "from unnecessary fees.")

    
    pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(30))
    pdf.set_xy(px2MM(150), px2MM(452))
    pdf.multi_cell(px2MM(1470),px2MM(42), "Keep a credit card in use - ",align = "L")

    pdf.set_font('LeagueSpartan-Regular', size=px2pts(30))
    pdf.set_xy(px2MM(498), px2MM(452))
    pdf.multi_cell(px2MM(1470),px2MM(42), "Using a credit card responsibly can positively impact your credit score, but it is important to")

    pdf.set_xy(px2MM(150), px2MM(452+42))
    pdf.multi_cell(px2MM(1470),px2MM(42), "pay off the entire balance on time to avoid accruing debt.")
    
# _____________________________________________________________________________________________________________________________________________________

    pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(30))
    pdf.set_xy(px2MM(150), px2MM(556))
    pdf.multi_cell(px2MM(1470),px2MM(42), "Apply for credit mindfully - ",align = "L")

    pdf.set_font('LeagueSpartan-Regular', size=px2pts(30))
    pdf.set_xy(px2MM(512), px2MM(556))
    pdf.multi_cell(px2MM(1470),px2MM(42), "While having a healthy mix of credit lines can boost your creditworthiness, avoid applying")

    pdf.set_xy(px2MM(150), px2MM(556+42))
    pdf.multi_cell(px2MM(1470),px2MM(42), "for credit unnecessarily or having multiple rejections, which can negatively impact your credit profile.")

# ____________________________________________________________________________________________________________________________________________________

# ____________________________________________________________________________________________________________________________________________________

    pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(30))
    pdf.set_xy(px2MM(150) , px2MM(660))
    pdf.multi_cell(px2MM(1470), px2MM(42), "Always close accounts - ",align = "L")

    pdf.set_font('LeagueSpartan-Regular', size=px2pts(30))
    pdf.set_xy(px2MM(465), px2MM(660))
    pdf.multi_cell(px2MM(1470),px2MM(42), "Make timely repayments, request the return of any hypothecated documents, and verify that ")

    pdf.set_xy(px2MM(150), px2MM(660 + 42))
    pdf.multi_cell(px2MM(1470),px2MM(42), "closure letters are updated with credit bureaus.")
# ____________________________________________________________________________________________________________________________________________________

# ____________________________________________________________________________________________________________________________________________________

    pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(30))
    pdf.set_xy(px2MM(150), px2MM(660 + 104))
    pdf.multi_cell(px2MM(1470),px2MM(42), "Maintain aged accounts - ",align = "L")

    pdf.set_font('LeagueSpartan-Regular', size=px2pts(30))
    pdf.set_xy(px2MM(490), px2MM(660 + 104))
    pdf.multi_cell(px2MM(1470),px2MM(42), "Keep your oldest accounts open, as they can help build a longer credit history and improve")
    
    pdf.set_xy(px2MM(150), px2MM(660 + 104 + 42))
    pdf.multi_cell(px2MM(1470),px2MM(42), "your credit profile. Use your available credit lines wisely and pay them off in a timely manner.")
# ____________________________________________________________________________________________________________________________________________________

# ____________________________________________________________________________________________________________________________________________________

    pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(30))
    pdf.set_xy(px2MM(150), px2MM(764 + 104))
    pdf.multi_cell(px2MM(1470),px2MM(42),"Communicate with your lender - ",align = "L")

    pdf.set_font('LeagueSpartan-Regular', size=px2pts(30))
    pdf.set_xy(px2MM(578), px2MM(764 + 104))
    pdf.multi_cell(px2MM(1470),px2MM(42), "If you face any financial difficulties, be open and honest with your lender, and seek")

    pdf.set_xy(px2MM(150), px2MM(764 + 104 + 42))
    pdf.multi_cell(px2MM(1470),px2MM(42), "help when needed. Most lenders are willing to listen and provide assistance if you can demonstrate the legitimacy of your situation.")
    #//*-----Index Text of Page--**////
    index_text(pdf,'#1A1A1D')
    global best_practices_idx
    best_practices_idx = pdf.page_no()
# ____________________________________________________________________________________________________________________________________________________

def planning_your_taxes(pdf,json_data,c_MoneyS,money_signData):
    pdf.add_page()
    pdf.set_fill_color(*hex2RGB('#FCF8ED'))
    pdf.rect(0,0, px2MM(1920),px2MM(1080),'F')
    pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(60))
    pdf.set_text_color(*hex2RGB('#1A1A1D'))
    pdf.set_xy(px2MM(120),px2MM(80))  
    pdf.cell(px2MM(740),px2MM(84),"Planning Your Income Taxes",align='L')

    pdf.set_fill_color(*hex2RGB('#000000'))
    pdf.rect(0,px2MM(80), px2MM(15),px2MM(84),'F')


    pdf.set_font('LeagueSpartan-Regular', size=px2pts(30))
    pdf.set_xy(px2MM(120), px2MM(224))
    pdf.multi_cell(px2MM(1500), px2MM(42), "Tax planning is a crucial aspect of personal finance that cannot be overlooked. It can help you maximize your returns and minimize your tax liability. Here are some best practices that you should consider:")

    pdf.set_fill_color(*hex2RGB('#000000'))

    # Rectangle Bullets in PDF.
       # Rectangle Bullets in PDF.
    pdf.rect(px2MM(120),px2MM(366),px2MM(10), px2MM(10), 'F')
    pdf.rect(px2MM(120),px2MM(428),px2MM(10), px2MM(10), 'F')
    pdf.rect(px2MM(120),px2MM(532),px2MM(10), px2MM(10), 'F')
    pdf.rect(px2MM(120),px2MM(636),px2MM(10), px2MM(10), 'F')
    pdf.rect(px2MM(120),px2MM(740),px2MM(10), px2MM(10), 'F')

    pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(30))
    pdf.set_xy(px2MM(150), px2MM(348))
    pdf.cell(px2MM(340),px2MM(42), "Start tax planning early - ",align = "L")
    pdf.set_font('LeagueSpartan-Regular', size=px2pts(30))
    pdf.set_xy(px2MM(490), px2MM(348))
    pdf.multi_cell(px2MM(1130),px2MM(42), "Start at the beginning of the financial year instead of waiting till the last minute.",align = "L")

    pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(30))
    pdf.set_xy(px2MM(150), px2MM(410))
    pdf.cell(px2MM(420),px2MM(42), "Utilize tax-saving investments - ",align = "L")
    pdf.set_font('LeagueSpartan-Regular', size=px2pts(30))
    pdf.set_xy(px2MM(570), px2MM(410))
    pdf.cell(px2MM(1050),px2MM(42), "Invest in tax-saving instruments to reduce your tax liability. Refer to the 'Available ",align = "L")
    pdf.set_xy(px2MM(150), px2MM(452))
    pdf.multi_cell(px2MM(1470),px2MM(42), "Tax Deductions' table on the following pages.",align = "L")
    
    pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(30))
    pdf.set_xy(px2MM(150), px2MM(514))
    pdf.cell(px2MM(470),px2MM(42), "Claim all available tax deductions - ",align = "L")
    pdf.set_font('LeagueSpartan-Regular', size=px2pts(30))
    pdf.set_xy(px2MM(620), px2MM(514))
    pdf.cell(px2MM(1000),px2MM(42), "Make sure to claim all possible deductions under different sections of the Income ",align = "L")
    pdf.set_xy(px2MM(150), px2MM(556))
    pdf.multi_cell(px2MM(1470),px2MM(42), "Tax Act to reduce your tax liability. Refer to the 'Available Tax Deductions' table on the following pages.",align = "L")

    pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(30))
    pdf.set_xy(px2MM(150), px2MM(618))
    pdf.cell(px2MM(400),px2MM(42), "Review your salary structure -",align = "L")
    pdf.set_font('LeagueSpartan-Regular', size=px2pts(30))
    pdf.set_xy(px2MM(550), px2MM(618))
    pdf.cell(px2MM(1070),px2MM(42), "Optimize your salary structure to reduce your tax liability by including components such",align = "L")
    pdf.set_xy(px2MM(150), px2MM(660))
    pdf.multi_cell(px2MM(1470),px2MM(42), "as House Rent Allowance (HRA), Leave Travel Allowance (LTA), and medical reimbursements.",align = "L")
    
    pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(30))
    pdf.set_xy(px2MM(150), px2MM(722))
    pdf.cell(px2MM(320),px2MM(42), "File your taxes on time -",align = "L")
    pdf.set_font('LeagueSpartan-Regular', size=px2pts(30))
    pdf.set_xy(px2MM(470), px2MM(722))
    pdf.cell(px2MM(1150),px2MM(42), "Ensure that you file your tax returns on time to avoid penalties and interest charges.",align = "L")
    
    #//*-----Index Text of Page--**////
    index_text(pdf,'#1A1A1D')
          
#//*----Page Written By Gurutirth 

#//*----Your Financial Profile---*//  
def fin_profile(pdf, json_data,c_MoneyS,money_signData):
     # //*---User Values---*//

    money_sign_desc = {
        "Eagle":"Far-Sighted Eagle", "Horse":"Persistent Horse",
        "Tiger":"Tactical Tiger", "Lion":"Opportunistic Lion",
        "Elephant":"Virtuous Elephant", "Turtle":"Vigilant Turtle",
        "Whale":"Enlightened Whale", "Shark":"Stealthy Shark"
    }
    

    #  generation
    try:
        # df = pd.DataFrame.from_dict(json_data["Generation Profile"])
        fin_score=json_data['oneview']['fbs']
        if fin_score==None:
            fin_score = 0
    except:
        return None
    gen_profile = json_data['gen_profile']["gen_profile"] 

        
    # card 4 data
    age_range = json_data['gen_profile']['life_stage_age_range']
    phase = json_data['gen_profile']['life_stage']
    generation = json_data['gen_profile']['gen_profile']
    generation_desc = json_data['gen_profile']['gen_profile_desc']

    
    life_stage_pts =  json_data['gen_profile']['life_stage_desc']   
    age_range_color = money_signData[c_MoneyS]['fin_profile'][0] 
    meter_stick_xpos_dict = {20:(0, 63), 40:(70, 137), 60:(144, 211), 80:(218, 285), 100:(292, 359)}
    meter_img_dict = {20:'meter_1_20.png', 40:'meter_20_40.png', 
    60:'meter_40_60.png', 80:'meter_60_80.png', 100:'meter_80_100.png'}
    for val in meter_img_dict:
        if fin_score <= val:
            meter_img = meter_img_dict[val]
            meter_stick_xpos = (meter_stick_xpos_dict[val][1] - meter_stick_xpos_dict[val][0])/20*(fin_score - (val-20))
            meter_stick_xpos = (201 + meter_stick_xpos_dict[val][0] + meter_stick_xpos)
            if fin_score <= 20:
                score_box_xpos = 190
            elif fin_score >=81:
                score_box_xpos = 416
            else:
                score_box_xpos = meter_stick_xpos - 74
            break
    
    your_money_sign = c_MoneyS.capitalize()

   
   
    #//*---Page setup----*//
    pdf.add_page()
    pdf.set_fill_color(*hex2RGB('#000000'))
    pdf.rect(0, 0, px2MM(1920), px2MM(1080),'F')

    vl_color = money_signData[c_MoneyS]['content'][3]
    # purple rectangle
    pdf.set_fill_color(*hex2RGB(vl_color))
    pdf.rect(px2MM(0), px2MM(80), px2MM(15), px2MM(84), 'F')
    

    # cards
    ms_tm = 'MoneySign<sup>TM</sup>'
    card_titles = ['Financial Behaviour Score', 'MoneySign', 'Generation Profile']
    for card_num in range(3):
        # card background
        pdf.set_fill_color(*hex2RGB('#FCF8ED'))
        if card_num == 2:
            card_height = 309
        else:
            card_height = 786
        pdf.rect(px2MM(120+(card_num*577)), px2MM(214), px2MM(527), px2MM(card_height), 'F')

        if card_num == 1:
        # card 2 background
            pdf.image(join(cwd,'assets','images','MoneySign',f'{your_money_sign}_text.png'),px2MM(697), px2MM(216), px2MM(527), px2MM(784))
            pdf.image(join(cwd,'assets','images','MoneySign','cream_bg_mask.png'),px2MM(697), px2MM(216), px2MM(527), px2MM(784))
            pdf.image(join(cwd,'assets','images','MoneySign',f'{your_money_sign}.png'),px2MM(810), px2MM(422), px2MM(300), px2MM(300))
            # black boxes to hide your_money_sign_bg.png vertical overflow
            pdf.set_fill_color(*hex2RGB('#000000'))
            pdf.rect(px2MM(697), px2MM(0), px2MM(527), px2MM(216), 'F')
            pdf.rect(px2MM(647), px2MM(0), px2MM(50), px2MM(1080), 'F')
            pdf.rect(px2MM(697), px2MM(1000), px2MM(527), px2MM(80), 'F')

        # card titles
        pdf.set_xy(px2MM(168+(card_num*577)), px2MM(254))  
        pdf.set_font('LeagueSpartan-Medium', size=px2pts(40))
        pdf.set_text_color(*hex2RGB('#000000'))


        pdf.cell(px2MM(431), px2MM(56), card_titles[card_num], align='C')
        
        # pdf.set_xy(px2MM(1048), px2MM((266)))
        # pdf.set_font('LeagueSpartan-Medium', size=16)
        # pdf.set_text_color(*hex2RGB('#000000'))
        # pdf.cell(px2MM(16), px2MM(8), 'TM')
        
        #//*--To print superscritp R 
        pdf.set_xy(px2MM(1046), px2MM(252))
        pdf.set_font('Inter-ExtraLight', size=26)
        pdf.set_text_color(*hex2RGB('#000000'))
        pdf.cell(px2MM(16), px2MM(56), '®') 

    # ------------- meter ----------------
    
    # fin score stick
    pdf.set_fill_color(*hex2RGB("#FFFFFF"))
    pdf.rect(px2MM(meter_stick_xpos), px2MM(492), px2MM(6), px2MM(95), 'F')
    # fin score box
    pdf.rect(px2MM(score_box_xpos), px2MM(380), px2MM(160), px2MM(148), 'F')

    # fin score 
    pdf.set_xy(px2MM(score_box_xpos), px2MM(416))  
    pdf.set_font('Prata', size=px2pts(64))
    pdf.set_text_color(*hex2RGB('#000000'))
    pdf.cell(px2MM(160), px2MM(87), str(fin_score), align='C')

    # Actual meter image
    pdf.image(join(cwd,'assets','images','BehaviourMeter', meter_img),
    px2MM(190), px2MM(575), px2MM(386), px2MM(74))

    # -------------meter labels------------

    # 0 label
    pdf.set_xy(px2MM(190), px2MM(669))  
    pdf.set_font('LeagueSpartan-semiBold', size=px2pts(24))
    pdf.set_text_color(*hex2RGB('#898B90'))
    pdf.multi_cell(px2MM(40), px2MM(32), '0', align='L')

    # 100 label
    pdf.set_xy(px2MM(540), px2MM(669))  
    pdf.set_font('LeagueSpartan-semiBold', size=px2pts(24))
    pdf.set_text_color(*hex2RGB('#898B90'))
    pdf.multi_cell(px2MM(50), px2MM(32), '100', align='L')

    # card 1 footer
    # card footer range
    pdf.set_font('LeagueSpartan-Medium', size=px2pts(24))
    pdf.set_text_color(*hex2RGB('#1A1A1D'))
    pdf.set_xy(px2MM(160), px2MM(761))
    pdf.cell(px2MM(63), px2MM(32), '0-50 : ', align='L')
    pdf.set_xy(px2MM(160), px2MM(808))
    pdf.cell(px2MM(74), px2MM(32), '50-75 : ', align='L')
    pdf.set_xy(px2MM(160), px2MM(855))
    pdf.cell(px2MM(83), px2MM(32), '75-100 : ', align='L')
    # card footer range descriptions
    pdf.set_font('LeagueSpartan-Regular', size=px2pts(24))
    pdf.set_text_color(*hex2RGB('#1A1A1D'))
    pdf.set_xy(px2MM(233), px2MM(761))
    pdf.cell(px2MM(213), px2MM(32), 'Financially vulnerable', align='L')
    pdf.set_xy(px2MM(244), px2MM(808))
    pdf.cell(px2MM(175), px2MM(32), 'Financially coping', align='L')
    pdf.set_xy(px2MM(253), px2MM(855))
    pdf.cell(px2MM(131), px2MM(32), 'Financially fit', align='L')

    # card 2 footer
    pdf.set_xy(px2MM(766), px2MM(822))
    pdf.set_font('Prata', size=px2pts(42))
    pdf.set_text_color(*hex2RGB('#000000'))
    # pdf.cell(px2MM(400), px2MM(66),['moneySign'], align='C')
    pdf.cell(px2MM(400), px2MM(66), json_data['money_sign']['money_sign'], align='C')

    # card 3 content
    # --Titles
    pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(30))
    pdf.set_text_color(*hex2RGB('#1A1A1D'))
    # generation 
    pdf.set_xy(px2MM(1313), px2MM(330))
    pdf.cell(px2MM(447), px2MM(42), generation)

    # content
    pdf.set_font('LeagueSpartan-Regular', size=px2pts(24))
    pdf.set_text_color(*hex2RGB('#313236'))
    pdf.set_xy(px2MM(1313), px2MM(387))
    pdf.multi_cell(px2MM(447), px2MM(32),generation_desc, align='L')

    # -----card 4
    # background
    pdf.set_fill_color(*hex2RGB('#FCF8ED'))
    pdf.rect(px2MM(1273), px2MM(563), px2MM(527), px2MM(437), 'F')

    # title
    pdf.set_xy(px2MM(1454), px2MM(603))  
    pdf.set_font('LeagueSpartan-Medium', size=px2pts(40))
    pdf.set_text_color(*hex2RGB('#000000'))
    pdf.cell(px2MM(164), px2MM(56), 'Life stage', align='C')

    # subtitle
    pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(30))
    pdf.set_text_color(*hex2RGB('#1A1A1D')) 
    pdf.set_xy(px2MM(1313), px2MM(679))
    pdf.cell(px2MM(250), px2MM(42), phase, align='L')

    # label
    pdf.set_fill_color(*hex2RGB(age_range_color))
    pdf.rect(px2MM(1605), px2MM(682.5), px2MM(166), px2MM(35), 'F')
    # label text
    pdf.set_font('LeagueSpartan-Regular', size=px2pts(18))
    pdf.set_text_color(*hex2RGB('#000000')) 
    pdf.set_xy(px2MM(1610), px2MM(687.5))
    pdf.cell(px2MM(154), px2MM(25), f'Age Range: {age_range}', align='C')

    y_h = pdf.get_y()+13
    # bullet points
    for idx, point in enumerate(life_stage_pts):
        pdf.set_fill_color(*hex2RGB('#313236'))
        # pdf.rect(px2MM(1295), px2MM(mm2PX(y_h)+15), px2MM(5), px2MM(5), 'F')
        pdf.circle(x=px2MM(1333), y=px2MM(mm2PX(y_h)+14), r=px2MM(5), style='F')
        # text
        pdf.set_font('LeagueSpartan-Regular', size=px2pts(24))
        pdf.set_text_color(*hex2RGB('#313236')) 
        pdf.set_xy(px2MM(1353), px2MM(mm2PX(y_h)))
        pdf.multi_cell(px2MM(427), px2MM(32), point,align='L')
        y_h = pdf.get_y()
        # pdf.cell(px2MM(1334), px2MM(pdf.get_y()+32), point, align='L')

    # page tile 
    pdf.set_xy(px2MM(120), px2MM(80))  
    pdf.set_font('LeagueSpartan-Medium', size=px2pts(60))
    pdf.set_text_color(*hex2RGB('#FFFFFF'))
    pdf.cell(px2MM(792), px2MM(84), 'Your Financial Profile')
 
    #//*-----Index Text of Page--**////
    index_text(pdf,'#FFFFFF')
    global your_fin_prof_idx
    your_fin_prof_idx = pdf.page_no()
#//*----Our Assumptions------*//
def assumptions(pdf,json_data,c_MoneyS,money_signData):
    try:
        df = pd.DataFrame.from_dict(json_data["our_assumption"]['assets'])
        df2 = pd.DataFrame.from_dict(json_data["our_assumption"]['yoy_growth_to_income'])
        df3 = pd.DataFrame.from_dict(json_data["our_assumption"]['liabilities_interest_ratio'])
        
    except:
        return None
   
    #//*---Page setup----*//
    pdf.add_page()

    # pg background color
    pdf.set_fill_color(*hex2RGB('#FCF8ED'))
    pdf.rect(0, 0, px2MM(1920), px2MM(1080), 'F')
    pdf.image(join(cwd,'assets','images','backgrounds','doubleLine.png'),px2MM(1449),px2MM(0),px2MM(471),px2MM(1080))
    # black rectangle besides title
    pdf.set_fill_color(*hex2RGB('#000000'))
    pdf.rect(px2MM(0), px2MM(80), px2MM(15), px2MM(80), 'F')

    # Page title
    pdf.set_xy(px2MM(120), px2MM(80))  
    pdf.set_font('LeagueSpartan-Medium', size=px2pts(60))
    pdf.set_text_color(*hex2RGB('#000000'))
    pdf.multi_cell(px2MM(441), px2MM(84), 'Our Assumptions', align='L')

    # ------cards--------------
    pdf.set_fill_color(*hex2RGB('#FFFFFF'))
    pdf.set_draw_color(*hex2RGB('#E9EAEE'))
    pdf.set_line_width(px2MM(0.2))
    # card 1: Asset class risk level table  card
    pdf.rect(px2MM(120), px2MM(232), px2MM(820), px2MM(752), 'FD')
    # card 2: income/expense table card
    pdf.rect(px2MM(980), px2MM(184), px2MM(820), px2MM(462), 'FD')
    # card 3: Interest rate table
    pdf.rect(px2MM(980), px2MM(676), px2MM(540), px2MM(357), 'FD')

    # card 2 title
    pdf.set_xy(px2MM(1020), px2MM(224))  
    pdf.set_font('LeagueSpartan-Medium', size=px2pts(24))
    pdf.set_text_color(*hex2RGB('#000000'))
    pdf.multi_cell(px2MM(400), px2MM(32), 'Income/Expense YoY Growth', align='L')

    # card 2 title
    pdf.set_xy(px2MM(1020), px2MM(716))  
    pdf.set_font('LeagueSpartan-Medium', size=px2pts(24))
    pdf.set_text_color(*hex2RGB('#000000'))
    pdf.multi_cell(px2MM(400), px2MM(32), 'Interest Rates on Liabilities', align='L')

    # ---------tables-------------------------
    # --card 1 table--
    # asset class table title row
    pdf.set_fill_color(*hex2RGB('#F3F6F9'))
    pdf.set_draw_color(*hex2RGB('#E9EAEE'))
    pdf.set_line_width(px2MM(0.2))
    pdf.rect(px2MM(160), px2MM(325), px2MM(177), px2MM(45), 'DF')
    pdf.rect(px2MM(337), px2MM(325), px2MM(248), px2MM(45), 'DF')
    pdf.rect(px2MM(585), px2MM(325), px2MM(119), px2MM(45), 'DF')
    pdf.rect(px2MM(704), px2MM(325), px2MM(196), px2MM(45), 'DF')  
    pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(18))
    pdf.set_text_color(*hex2RGB('#000000'))
    pdf.set_xy(px2MM(180),px2MM(335))
    pdf.cell(px2MM(137), px2MM(25),'Asset Classes',border=0,align='L')
    pdf.set_xy(px2MM(357),px2MM(335))
    pdf.cell(px2MM(208), px2MM(25),'Examples',border=0,align='L')
    pdf.set_xy(px2MM(605),px2MM(335))
    pdf.cell(px2MM(79), px2MM(25),'Returns %',border=0,align='L')
    pdf.set_xy(px2MM(724),px2MM(335))
    pdf.cell(px2MM(156), px2MM(25),'Risk Level',border=0,align='C')

    instrument = []
    risk_images = []
    
    pdf.set_text_color(*hex2RGB('#000000'))
    pdf.set_font('LeagueSpartan-Regular', size=px2pts(24))
    pdf.set_xy(px2MM(160),px2MM(273))
    pdf.cell(px2MM(357), px2MM(32),'Risk/Return Profile of Asset Classes',border=0,align='L')
    
    for i in range(len(df)):
        # txt = df["Instrument 1"][i]+'\n'+df["Instrument 2"][i]+'\n'+df["Instrument 3"][i]
        txt = "\n".join(df['examples'].iloc[i])
        instrument.append(txt)
        
        if df["risk_level"][i]=="Moderate to High":
            risk_images.append('Riskmeter_m2h.png')
        elif df["risk_level"][i]=="Low to High":
            risk_images.append('Riskmeter_l2h.png')
        elif df["risk_level"][i]=="Very Low to Moderate":
            risk_images.append('Riskmeter_vl2m.png')
        elif df["risk_level"][i]=="Low to Very High":
            risk_images.append('Riskmeter_l2vh.png')
              
    table1_col_vals = [list(df['asset_class']),instrument,list(df["return_percentage"]),list(df['risk_level'])]
    risk_images = ['Riskmeter_m2h.png', 'Riskmeter_l2h.png', 'Riskmeter_l2h.png', 'Riskmeter_vl2m.png', 'Riskmeter_l2vh.png']

    pdf.set_fill_color(*hex2RGB('#F3F6F9'))
    pdf.set_draw_color(*hex2RGB('#E9EAEE'))
    col_x_pos = (160, 337, 585, 704)
    col_text_y_pos = (415, 389, 415, 440)
    col_widths = (177, 248, 119, 196)
    col_text_widths = (137, 218, 100, 166)
    for row in range(len(df)):
        for column in range(len(table1_col_vals)):
            # backgrounds
            pdf.set_draw_color(*hex2RGB('#E9EAEE'))
            if row%2 == 0:
                pdf.set_fill_color(*hex2RGB('#ffffff')) 
                pdf.rect(px2MM(col_x_pos[column]), px2MM(369+(row*115)), px2MM(col_widths[column]), px2MM(115), 'DF')
            else:
                pdf.set_fill_color(*hex2RGB('#F3F6F9'))
                pdf.rect(px2MM(col_x_pos[column]), px2MM(369+(row*115)), px2MM(col_widths[column]), px2MM(115), 'DF')
            
            # text weigth
            if column == 0 or column == 2:
                pdf.set_font('LeagueSpartan-Regular', size=px2pts(18))
            else:
                pdf.set_font('LeagueSpartan-Light', size=px2pts(18))

            # text color
            pdf.set_text_color(*hex2RGB('#000000'))

            # text positions
            pdf.set_xy(px2MM((col_x_pos[column])+20),px2MM(col_text_y_pos[column]+row*115))
            
            # text cells
            if column == 3:
                pdf.multi_cell(px2MM(col_text_widths[column]), px2MM(25),table1_col_vals[column][row],border=0,align='C')
            else:
                pdf.multi_cell(px2MM(col_text_widths[column]), px2MM(25),table1_col_vals[column][row],border=0,align='L')
            # Risk Images
            pdf.image(join(cwd,'assets','images','RiskMeters',risk_images[row]), px2MM(763), px2MM(389+row*115), px2MM(78), px2MM(40))
                
    # --card2 table--
    
    col_x_pos = (1020, 1175, 1299)
    col_widths = (155, 124, 156)
    col_text_widths = (115, 84, 116)
    col_align = ('L', 'C', 'R')

    l1 = ['Lifestage']+list(df2["life_stage"])
    l2 = ['Age Range']+list(df2["age_range"])
    # l3 = ['Income Growth']+list(str(x*100)+'%' for x in df2["Percentage"])
    l3 = ['Income Growth']+list(df2["income_growth"])


    
    for i in range(len(l1)):
        if i%2==0:
            pdf.set_fill_color(*hex2RGB('#F3F6F9'))
        else:
            pdf.set_fill_color(*hex2RGB('#FFFFFF'))
        if i==0:
            pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(18))
        else:
            pdf.set_font('LeagueSpartan-Regular', size=px2pts(18))
            
 
        pdf.set_draw_color(*hex2RGB('#E9EAEE'))
        pdf.set_text_color(*hex2RGB('#000000'))
        pdf.set_line_width(px2MM(0.2))
        pdf.rect(px2MM(1020),px2MM(276+(i*45)),px2MM(290),px2MM(45),'FD')
        pdf.set_xy(px2MM(1040),px2MM(286+(i*45)))
        pdf.cell(px2MM(250),px2MM(25),l1[i],align='L')
        
        pdf.rect(px2MM(1310),px2MM(276+(i*45)),px2MM(258),px2MM(45),'FD')
        pdf.set_xy(px2MM(1330),px2MM(286+(i*45)))
        pdf.cell(px2MM(218),px2MM(25),str(l2[i]),align='C')
        
        pdf.rect(px2MM(1568),px2MM(276+(i*45)),px2MM(192),px2MM(45),'FD')
        pdf.set_xy(px2MM(1588),px2MM(286+(i*45)))
        pdf.cell(px2MM(152),px2MM(25),str(l3[i]),align='R')
        
    pdf.set_font('LeagueSpartan-Regular', size=px2pts(18))  
    pdf.set_xy(px2MM(1310),px2MM(511))
    pdf.cell(px2MM(170),px2MM(25),'Expense Growth: '+ json_data["our_assumption"]['yoy_growth_expense'],align='L')  
    
    pdf.set_font('LeagueSpartan-Light', size=px2pts(18))  
    pdf.set_xy(px2MM(1020),px2MM(556))
    pdf.multi_cell(px2MM(740),px2MM(25),'The timing of life stages varies based on profession, industry trends, career goals, and other factors, making it unique to each individual.',align='L') 
    

    # --card 3 table--
    col_x_pos = (1020, 1260)
    col_widths = (240, 220)
    col_text_widths = (200, 180)
    col_align = ('L', 'R')

    l1=['Liabilities']+list(df3["liabilities"])
    l2=['Interest Rates']+list(df3["interest_rate_range"])
    
    table3_col_vals = [
      l1,
      l2
    ]

    for row in range(len(l1)):
        for column in range(2):
            # cel backgrounds
            pdf.set_draw_color(*hex2RGB('#E9EAEE'))
            pdf.set_line_width(px2MM(0.2))
            pdf.set_font('LeagueSpartan-Regular', size=px2pts(18))
            if row == 0:
                pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(18))
                pdf.set_fill_color(*hex2RGB('#F3F6F9'))
            elif row%2 == 0:
                pdf.set_fill_color(*hex2RGB('#F3F6F9'))
            else:
                pdf.set_fill_color(*hex2RGB('#ffffff'))
            pdf.rect(px2MM(col_x_pos[column]), px2MM(768+(row*45)), px2MM(col_widths[column]), px2MM(45), 'DF')
            # col text
            pdf.set_xy(px2MM(col_x_pos[column]+20), px2MM(778+row*45))
            pdf.set_text_color(*hex2RGB('#000000'))
            pdf.cell(px2MM(col_text_widths[column]), px2MM(25),table3_col_vals[column][row],border=0,align=col_align[column])
    #//*-----Index Text of Page--**////
    index_text(pdf,'#1A1A1D')

#//*----Your Financial wellness plan------*//

def fin_wellness_plan(pdf,json_data,c_MoneyS,money_signData):
    try:
        fwp = pd.DataFrame.from_dict(json_data['fwp'])
        if fwp.empty:
            return None
        
        
        df_exp_lib_manage = fwp['desc'].iloc[0]
        df_asset = fwp['desc'].iloc[1]
        df_expense = fwp['desc'].iloc[2]
        
        if df_exp_lib_manage==[] and df_asset==[] and df_expense==[]:
            return None
    
    except:
        return None
    
    #//*---Page setup----*//
    pdf.add_page()
    pdf.set_fill_color(*hex2RGB('#FCF8ED'))
    pdf.rect(0, 0, px2MM(1920), px2MM(1080),'F') 

    # black background of page
    pdf.set_fill_color(*hex2RGB('#000000'))
    pdf.rect(0, 0, 1920, 1080, 'F')

    # white rectangular backgrount at bottom
    pdf.set_fill_color(*hex2RGB('#FCF8ED'))
    pdf.rect(px2MM(0), px2MM(736), px2MM(1920), px2MM(344), 'F')
    
    #//*--Purple vertical line
    # pdf.set_xy(px2MM(125),px2MM(78))
    pdf.set_fill_color(*hex2RGB('#ffffff'))
    pdf.rect(px2MM(0), px2MM(80), px2MM(15), px2MM(84), 'F') 
    
    #//*---heading statement
    pdf.set_xy(px2MM(120),px2MM(80))  
    pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(60))
    pdf.set_text_color(*hex2RGB('#ffffff'))
    pdf.cell(px2MM(724), px2MM(84),'Your Financial Wellness Plan')
        
    # subtitle
    pdf.set_xy(px2MM(120), px2MM(244))  
    pdf.set_font('LeagueSpartan-Regular', size=px2pts(40))
    pdf.set_text_color(*hex2RGB('#ffffff'))
    pdf.cell(px2MM(234), px2MM(56), 'Key Takeaways')
    # Cards

    card_title_list = fwp['title'].tolist()
    image_list = []
    for i in card_title_list:
        if i.lower() == "emergency planning":
           image_list.append('Shield.png') 
        elif i.lower() == "expense and liability management":
            image_list.append('Expense.png') 
        elif i.lower() == "asset allocation":
            image_list.append('Assets.png') 
        else:
            image_list.append('Assets.png') 
            
    #//*---Image name: (Shield = Emergency Planning)
    # image_list = ['Shield.png','Expense.png', 'Assets.png']
   
    
    len_p = []
    for card_num in range(len(card_title_list)):
        # Card Boxes
        pdf.set_fill_color(*hex2RGB('#FFFFFF'))
        pdf.set_draw_color(*hex2RGB('#E9EAEE'))
        pdf.set_line_width(px2MM(0.2))
        
        pdf.rect(px2MM(120+(card_num*577)), px2MM(340), px2MM(527), px2MM(654), 'FD')
            
        # logo 
        logo = join(cwd,'assets','images','icons', image_list[card_num])
        pdf.image(logo, px2MM(160+(card_num*577)), px2MM(382), px2MM(80), px2MM(80))

        # Card titles  
        pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(30))
        pdf.set_text_color(*hex2RGB('#000000'))
        if len(card_title_list[card_num]) >= 20:
            pdf.set_xy(px2MM(260+(card_num*579)), px2MM(380))
        else:
            pdf.set_xy(px2MM(260+(card_num*579)), px2MM(399))
        pdf.multi_cell(px2MM(347), px2MM(42), card_title_list[card_num], align="L")
        
        pi_gap_rect=522
        pi_gap_text = 504  
        for i in range(len(fwp['desc'].iloc[card_num])):
            if fwp['desc'].iloc[card_num][i] == "":
                continue
            pdf.set_fill_color(*hex2RGB('#000000'))
            pdf.rect(px2MM(160+(card_num*576)), px2MM(pi_gap_rect+(i*48)), px2MM(10), px2MM(10), 'F')

            pdf.set_xy(px2MM(195+(card_num*577)), px2MM(pi_gap_text+(i*30)))  
            pdf.set_font('LeagueSpartan-Regular', size=px2pts(30))
            pdf.set_text_color(*hex2RGB('#000000'))
            pdf.multi_cell(px2MM(417), px2MM(42), fwp['desc'].iloc[card_num][i], align='L')
            
            pi_gap_rect=pi_gap_text = mm2PX(pdf.get_y())

    #//*-----Index Text of Page--**////
    index_text(pdf,'#1A1A1D')
    global your_fw_plan_idx
    your_fw_plan_idx = pdf.page_no()
    
    

#//*------cashflow_plan
def cashflow_plan(pdf,json_data,c_MoneyS,money_signData):
    try:
        df_cash_flow = pd.DataFrame.from_dict(json_data['next_three_months_action_plan']["table"])
    except:
        return None
    
    lcol_val_list = ["Next 3 Months Cashflows"]+list(df_cash_flow["particular"])

    rcol_val_list = ["Amount"]+list("₹ "+str(format_cash2(float(x))) for x in df_cash_flow["amount"])
  
    #//*---Page setup----*//
    pdf.add_page()
    pdf.set_fill_color(*hex2RGB('#FCF8ED'))
    pdf.rect(0, 0, px2MM(1920), px2MM(1080),'F')
    pdf.image(join(cwd,'assets','images','backgrounds','doubleLine.png'),px2MM(1449),px2MM(0),px2MM(471),px2MM(1080))
    
    #//*-----Index Text of Page--**////
    index_text(pdf,'#1A1A1D')
    global your_fw_plan_idx
    if your_fw_plan_idx == 0:
        your_fw_plan_idx = pdf.page_no()

    # black rectangle
    pdf.set_fill_color(*hex2RGB('#000000'))
    pdf.rect(px2MM(0), px2MM(80), px2MM(15), px2MM(84), 'F')

    # page tile 
    pdf.set_xy(px2MM(120), px2MM(80))  
    pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(60))
    pdf.set_text_color(*hex2RGB('#1A1A1D'))
    pdf.cell(px2MM(792), px2MM(84), "Next 3 Months Action Plan")

    pdf.set_draw_color(*hex2RGB('#E9EAEE'))
    pdf.set_line_width(px2MM(0.05))
    # pdf.set_draw_color(*hex2RGB('#D3D3D3'))
    # pdf.rect(px2MM(120), px2MM(224), px2MM(516), px2MM(432), 'D')


    pdf.set_font('LeagueSpartan-Regular', size=px2pts(24))
    pdf.set_text_color(*hex2RGB('#000000'))
    for row in range(len(lcol_val_list)):
        if row%2 == 0:
            col = '#ffffff'
        else:
            col = '#F3F6F9'
        pdf.set_fill_color(*hex2RGB(col))
            
        if row == 0 or row==len(lcol_val_list)-1: 
            pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(24))
        else:
            pdf.set_font('LeagueSpartan-Regular', size=px2pts(24))
            
            

        pdf.rect(px2MM(120), px2MM(204+(row*65)), px2MM(360), px2MM(65), 'FD')
        pdf.rect(px2MM(480), px2MM(204+(row*65)), px2MM(156), px2MM(65), 'FD')
        if row ==len(lcol_val_list)-1:
            pdf.set_draw_color(*hex2RGB('#B9BABE'))
            pdf.set_fill_color(*hex2RGB('#B9BABE'))
            pdf.rect(px2MM(120), px2MM(204+(row*65)), px2MM(516), px2MM(1), 'FD') 
            pdf.set_fill_color(*hex2RGB(col))
            pdf.set_draw_color(*hex2RGB('#E9EAEE'))
            pdf.set_line_width(px2MM(0.2))
            pdf.rect(px2MM(120), px2MM(205+(row*65)), px2MM(516), px2MM(65), 'FD') 
            
            
        # col1 text
        pdf.set_xy(px2MM(140), px2MM(224+(row*65)))  
        pdf.cell(px2MM(320), px2MM(32), lcol_val_list[row], align='L')
        # col2 text
        pdf.set_xy(px2MM(500), px2MM(224+(row*65)))
        if rcol_val_list[row]=='₹0.0' or rcol_val_list[row]=='₹0':
            pdf.cell(px2MM(116), px2MM(32),' ', align='R')
        else:
            pdf.cell(px2MM(116), px2MM(32), rcol_val_list[row], align='R')

    featured_list = pd.DataFrame.from_dict(json_data['next_three_months_action_plan']['comments'])

    flg = "False"
    for i in range(len(featured_list['desc'])):
        if featured_list['desc'].iloc[i]:
            flg = 'True'
            
    if flg =='False':
        return None

    df_0 = pd.DataFrame(featured_list['desc'].iloc[0],columns=[featured_list['title'].iloc[0]])
    df_1 = pd.DataFrame(featured_list['desc'].iloc[1],columns=[featured_list['title'].iloc[1]])
    df_2 = pd.DataFrame(featured_list['desc'].iloc[2],columns=[featured_list['title'].iloc[2]])


    df_list = [df_0,df_1,df_2]

    y_hight = 0

    top_mrgn = 204   
    for i in range(3):

        if not df_list[i].empty:
            
            pdf.set_xy(px2MM(696), px2MM(mm2PX(y_hight)+top_mrgn))
            pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(32))
            pdf.set_text_color(*hex2RGB('#000000'))
            pdf.cell(px2MM(1077), px2MM(42),df_list[i].columns[0], align='L')
            
            cl_name = df_list[i].columns[0]
            h_end_high = pdf.get_y()
            pdf.set_font('LeagueSpartan-Regular', size=px2pts(30))
            pdf.set_text_color(*hex2RGB('#1A1A1D'))
            pdf.set_fill_color(*hex2RGB('#000000'))
            
            for k in range(len(df_list[i])):
                j = 10
                if k==0:
                    j = 50
                pdf.rect(px2MM(696), px2MM(mm2PX(h_end_high)+j+20), px2MM(10), px2MM(10), 'F')
                pdf.set_xy(px2MM(726), px2MM(mm2PX(h_end_high)+j+0))
                pdf.multi_cell(px2MM(1077), px2MM(42),df_list[i][cl_name][k] , align='L')
                
                h_end_high = pdf.get_y()
            top_mrgn = 30
        else:
            continue
                     
        y_hight = pdf.get_y()
        
            
#//*-----disclaimer----*//
def disclaimer(pdf,json_data,c_MoneyS,money_signData):
    pdf.add_page()
    pdf.set_fill_color(*hex2RGB('#FCF8ED'))
    pdf.rect(0,0,px2MM(1920),px2MM(1080),'F')

    pdf.set_xy(px2MM(140),px2MM(78))  
    pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(60))
    pdf.set_text_color(*hex2RGB('#1A1A1D'))
    pdf.cell(px2MM(263), px2MM(84),"Disclaimer",border=0)
    
    
    pdf.set_xy(px2MM(140),px2MM(202))  
    pdf.set_font('LeagueSpartan-Medium', size=px2pts(36))
    pdf.set_text_color(*hex2RGB('#1A1A1D'))
    pdf.cell(px2MM(1143), px2MM(45),"The Disclaimer page should be read in conjunction with this report.",border=0)
    
    pdf.set_xy(px2MM(140),px2MM(287))  
    pdf.set_font('LeagueSpartan-Light', size=px2pts(24))
    pdf.set_text_color(*hex2RGB('#000000'))
    pdf.cell(px2MM(1760), px2MM(32),"This report is based on the data and presumptions supplied by you (client/ user/ member).",border=0)
    
    pdf.set_xy(px2MM(140),px2MM(343))  
    pdf.multi_cell(px2MM(1640), px2MM(32),"This report is designed to assess your present financial condition and recommend planning ideas and concepts that may be beneficial. This report aims to demonstrate how well-established financial planning principles can enhance your existing financial situation. This report does not imply a recommendation of any specific method, but rather offers broad, general advice on the benefits of a few financial planning principles.",border=0)

    pdf.set_xy(px2MM(140),px2MM(463))  
    text1="""The reports give estimates based on multiple hypotheses; thus they are purely speculative and do not represent assurances of investment returns. Before making any financial decisions or adopting any transactions or plans, you should speak with your tax and/or legal counsel and solely decide on the execution and implementation. """
    pdf.multi_cell(px2MM(1640), px2MM(32),text1,border=0,align="L")
    
    
    pdf.set_xy(px2MM(140),px2MM(527))  
    txt1 = """1 Finance Private Limited or any of its representatives will not be liable or responsible for any losses or damages incurred by the client/user/member as a result of this report."""
    pdf.multi_cell(px2MM(1640), px2MM(32),txt1,border=0,align="L")
    
    pdf.set_xy(px2MM(140),px2MM(615))  
    txt3 = """Prices mentioned in this report may have come from sources we believe to be dependable, but they are not guaranteed. It’s crucial to understand that past performance does not guarantee future outcomes and that actual results may vary from the forecasts in this report."""
    pdf.multi_cell(px2MM(1640), px2MM(32),txt3,border=0,align="L")
    
    pdf.set_xy(px2MM(140),px2MM(703))  
    txt4 = """Unless changes to your financial or personal situation necessitate a more frequent review, we advise that you evaluate your plan once a quarter. Please be aware that some discrepancies could occur due to different calculation methods."""
    pdf.multi_cell(px2MM(1650), px2MM(32),txt4,border=0,align="L")

    pdf.set_fill_color(*hex2RGB('#FFFFFF'))
    pdf.rect(px2MM(140),px2MM(807),px2MM(517),px2MM(42),'F')

    pdf.set_font('LeagueSpartan-Regular', size=px2pts(18))
    pdf.set_text_color(*hex2RGB('#313236'))
    pdf.set_xy(px2MM(150),px2MM(815))  
    pdf.cell(px2MM(150), px2MM(22),"For any questions or queries, send an email to ",border=0,align="L")

    pdf.set_xy(px2MM(493),px2MM(815))  
    pdf.set_font('LeagueSpartan-Medium','U', size=px2pts(18))
    pdf.multi_cell(px2MM(491), px2MM(22),"care@1finance.co.in",border=0,align="L")

    
    pdf.set_xy(px2MM(1595),px2MM(814))  
    pdf.set_font('LeagueSpartan-Regular', size=px2pts(20))
    pdf.set_text_color(*hex2RGB('#000000'))
    pdf.cell(px2MM(185), px2MM(28),str("https://1finance.co.in/"),border=0)
    
    pdf.set_xy(px2MM(140),px2MM(944))  
    pdf.set_font('LeagueSpartan-Regular', size=px2pts(24))
    pdf.set_text_color(*hex2RGB('#000000'))
    pdf.multi_cell(px2MM(1640), px2MM(32),"Investment in securities market are subject to market risks. Read all the related documents carefully before investing. \nRegistration granted by SEBI, membership of BASL and certification from National Institute of Securities Markets (NISM) in no way guarantee performance of the \nintermediary or provide any assurance of returns to investors.",border=0,align="C")
    
    #//*-----Index Text of Page--**////
    index_text(pdf,'#1A1A1D')
    
    
#//*-----Def last Page
def lastpage(pdf,json_data,c_MoneyS,money_signData):
    pdf.add_page()
    pdf.set_fill_color(*hex2RGB('#000000'))
    pdf.rect(0,0,px2MM(1920),px2MM(1080),'F')

    pdf.image(logo,px2MM(904),px2MM(394),px2MM(104),px2MM(119.88))

    pdf.set_xy(px2MM(518),px2MM(579.27))  
    pdf.set_font('LeagueSpartan-Light', size=px2pts(40))
    pdf.set_text_color(*hex2RGB('#FFFFFF'))
    text3="""Unit No. 1101 & 1102, 11th Floor, B - Wing, \nLotus Corporate Park, Goregaon (E), Mumbai-400063,"""
    pdf.multi_cell(px2MM(887),px2MM(56),text3,border=0,align="C")

    pdf.image(join(cwd,'assets','images','icons','gmail.svg'),px2MM(110),px2MM(854),px2MM(32),px2MM(32))
    pdf.set_xy(px2MM(158),px2MM(852))  
    pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(25.33))
    pdf.set_text_color(*hex2RGB('#FFFFFF'))
    pdf.cell(px2MM(241),px2MM(32),"care@1finance.co.in",border=0,align="L")

    pdf.image(join(cwd,'assets','images','icons','globe.svg'),px2MM(110),px2MM(900),px2MM(32),px2MM(32))
    pdf.set_xy(px2MM(158),px2MM(900))  
    pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(24))
    pdf.set_text_color(*hex2RGB('#FFFFFF'))
    pdf.cell(px2MM(243),px2MM(32),"https://1finance.co.in",border=0,align="L")


    pdf.image(join(cwd,'assets','images','icons','call.svg'),px2MM(110),px2MM(948),px2MM(32),px2MM(32))
    pdf.set_xy(px2MM(158),px2MM(948))  
    pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(25.33))
    pdf.set_text_color(*hex2RGB('#FFFFFF'))
    pdf.cell(px2MM(158),px2MM(32),"022 - 6912 0000",border=0,align="L")


    pdf.set_xy(px2MM(1485),px2MM(896))  
    pdf.set_font('LeagueSpartan-Light', size=px2pts(24))
    pdf.set_text_color(*hex2RGB('#FFFFFF'))
    pdf.cell(px2MM(325),px2MM(32),"Prepared by",border=0,align="R")

    pdf.set_xy(px2MM(1485),px2MM(938))  
    pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(30))
    pdf.set_text_color(*hex2RGB('#FFFFFF'))
    pdf.cell(px2MM(325),px2MM(42),"1 Finance Private Limited",border=0,align="R")  

    pdf.line(110,791,1700,0)
    pdf.image(join(cwd,'assets','images','icons','Line 3.png'),px2MM(110),px2MM(791),px2MM(1700),px2MM(0.02))
    
    #//*-----Index Text of Page--**////
    index_text(pdf,'#FFFFFF') 

#//*-----Def Your 1 view
def your_1_view_detail(pdf,json_data,c_MoneyS,money_signData):
    pdf.add_page()
    pdf.set_fill_color(*hex2RGB('#000000'))
    pdf.rect(0,0,px2MM(1920),px2MM(1080),'F')
    
    # df_asset = pd.DataFrame.from_dict(json_data['Asset'])

    pdf.set_fill_color(*hex2RGB('#FCF8ED'))
    pdf.rect(px2MM(0),px2MM(80),px2MM(15),px2MM(84),'F')

    pdf.set_xy(px2MM(120),px2MM(80))  
    pdf.set_font('LeagueSpartan-Bold', size=px2pts(60))
    pdf.set_text_color(*hex2RGB('#ffffff'))
    pdf.cell(px2MM(293),px2MM(84),"Your 1 View",border=0,align="L")

    # card 1
    pdf.set_fill_color(*hex2RGB('#E6E0FF'))
    pdf.rect(px2MM(120),px2MM(204),px2MM(527),px2MM(520),'F')
    # pdf.image(join(cwd,'assets','images','1_view_table','table_bg1.png'),px2MM(120),px2MM(204),px2MM(527),px2MM(592))

    pdf.image(join(cwd,'assets','images','icons','Assets.png'),px2MM(160),px2MM(244),px2MM(60),px2MM(60))
    pdf.set_xy(px2MM(240),px2MM(246))  
    pdf.set_font('LeagueSpartan-Regular', size=px2pts(40))
    pdf.set_text_color(*hex2RGB('#000000'))
    pdf.cell(px2MM(105),px2MM(56),"Assets",border=0,align="L")

    asset_table = pd.DataFrame(json_data['oneview']['assets'])
    assets = list(asset_table['title'])
    
    pdf.set_xy(px2MM(500),px2MM(253))  
    pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(30))
    pdf.set_text_color(*hex2RGB('#000000'))
    pdf.cell(px2MM(105),px2MM(42),'₹ '+ format_cash2(float(json_data['oneview']['total']['assets'])),border=0,align="L")
    
    pdf.set_text_color(*hex2RGB('#000000'))
    for rows in range(len(assets)):
        pdf.set_line_width(px2MM(0.1))
        pdf.set_draw_color(*hex2RGB('#E9EAEE'))
        pdf.set_fill_color(*hex2RGB('#FFFFFF'))
        pdf.rect(px2MM(160), px2MM(324+(rows*72)), px2MM(290), px2MM(72), 'DF')
        pdf.rect(px2MM(450), px2MM(324+(rows*72)), px2MM(157), px2MM(72), 'DF')
        
        pdf.set_xy(px2MM(180), px2MM(344+(rows*72)))
        pdf.set_font('LeagueSpartan-Regular', size=px2pts(24))
        pdf.cell(px2MM(250),px2MM(32),assets[rows],border=0,align="L")

            #cal2 text
        pdf.set_draw_color(*hex2RGB('#E9EAEE'))
        pdf.set_xy(px2MM(470), px2MM(344+(rows*72)))
        pdf.set_font('LeagueSpartan-Regular', size=px2pts(24))
        if asset_table['value'].iloc[rows] ==' ' or asset_table['value'].iloc[rows] =='':
            pdf.cell(px2MM(117),px2MM(25),'-',border=0,align="R")
        elif asset_table['value'].iloc[rows] ==' ' or asset_table['value'].iloc[rows] =='':
            pdf.cell(px2MM(117),px2MM(25),'-',border=0,align="R")
        else:
            val = format_cash2(float(asset_table['value'].iloc[rows]))
            # pdf.cell(px2MM(117),px2MM(25),f"₹ {asset_table['value'].iloc[rows]}",border=0,align="R")
            pdf.cell(px2MM(117),px2MM(25),f"₹ {val}",border=0,align="R")
  
     # card 2
    pdf.set_fill_color(*hex2RGB('#DEEDFF'))
    pdf.rect(px2MM(697),px2MM(558),px2MM(527),px2MM(304),'F')

    pdf.image(join(cwd,'assets','images','icons','Income.png'),px2MM(737),px2MM(598),px2MM(60),px2MM(60))
    pdf.set_xy(px2MM(817),px2MM(600))  
    pdf.set_font('LeagueSpartan-Regular', size=px2pts(40))
    pdf.set_text_color(*hex2RGB('#000000'))
    pdf.cell(px2MM(116),px2MM(56),"Income",border=0,align="L")
    
    income_table = pd.DataFrame(json_data['oneview']['income'])
    income = list(income_table['title'])

    
    pdf.set_xy(px2MM(1050),px2MM(607))  
    pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(30))
    pdf.set_text_color(*hex2RGB('#000000'))
    pdf.cell(px2MM(130),px2MM(42),'₹ '+format_cash2(float(json_data['oneview']['total']['income'])),border=0,align="R")
    
    pdf.set_text_color(*hex2RGB('#000000'))
    for rows in range(len(income_table)):
        pdf.set_draw_color(*hex2RGB('#E9EAEE'))
        pdf.set_fill_color(*hex2RGB('#FFFFFF'))
        pdf.rect(px2MM(737), px2MM(678+(rows*72)), px2MM(447), px2MM(72), 'DF')
        pdf.rect(px2MM(1027), px2MM(678+(rows*72)), px2MM(157), px2MM(72), 'DF')

        # col1 text
        pdf.set_xy(px2MM(757), px2MM(698+(rows*72)))
        pdf.set_font('LeagueSpartan-Regular', size=px2pts(24))
        pdf.cell(px2MM(250),px2MM(32),income[rows],border=0,align="L")

            #cal2 text
        pdf.set_xy(px2MM(1047), px2MM(698+(rows*72)))
        pdf.set_font('LeagueSpartan-Regular', size=px2pts(24))
        if income_table['value'].iloc[rows] == '':
            pdf.cell(px2MM(117),px2MM(32),'-',border=0,align="R")
        else:
            val = format_cash2(float(income_table['value'].iloc[rows]))
            pdf.cell(px2MM(117),px2MM(32),f"₹ {val}",border=0,align="R")

    pdf.set_fill_color(*hex2RGB('#FFDDDA'))
    pdf.rect(px2MM(1273),px2MM(558),px2MM(527),px2MM(304),'F')

    pdf.image(join(cwd,'assets','images','icons','Expense.png'),px2MM(1313),px2MM(598),px2MM(60),px2MM(60))
    pdf.set_xy(px2MM(1393),px2MM(600))  
    pdf.set_font('LeagueSpartan-Regular', size=px2pts(40))
    pdf.set_text_color(*hex2RGB('#000000'))
    pdf.cell(px2MM(154),px2MM(56),"Expenses",border=0,align="L")

    
    expense_table = pd.DataFrame(json_data['oneview']['expense'])
    expense_keys = list(expense_table['title'])

    
    pdf.set_xy(px2MM(1630),px2MM(607))  
    pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(30))
    pdf.set_text_color(*hex2RGB('#000000'))
    pdf.cell(px2MM(130),px2MM(42),'₹ '+format_cash2(float(json_data['oneview']['total']['expense'])),border=0,align="R")
    
    pdf.set_text_color(*hex2RGB('#000000'))
    for rows in range(len(expense_keys)):
        pdf.set_draw_color(*hex2RGB('#E9EAEE'))
        pdf.set_fill_color(*hex2RGB('#FFFFFF'))
        pdf.rect(px2MM(1313), px2MM(678+(rows*72)), px2MM(290), px2MM(72), 'DF')
        pdf.rect(px2MM(1603), px2MM(678+(rows*72)), px2MM(157), px2MM(72), 'DF')
        
        pdf.set_xy(px2MM(1333), px2MM(698+(rows*72)))
        pdf.set_font('LeagueSpartan-Regular', size=px2pts(24))
        pdf.cell(px2MM(250),px2MM(32),expense_keys[rows],border=0,align="L")

            #cal2 text
        pdf.set_xy(px2MM(1623), px2MM(698+(rows*72)))
        pdf.set_font('LeagueSpartan-Regular', size=px2pts(24))
        if expense_table['value'].iloc[rows] == '':
            pdf.cell(px2MM(117),px2MM(32),"-",border=0,align="R")
        else:
            val = format_cash2(float(expense_table['value'].iloc[rows]))
            pdf.cell(px2MM(117),px2MM(32),f"₹ {val}",border=0,align="R")
    
    Insurance_table = pd.DataFrame(json_data['oneview']['insurance'])
    Insurance_keys = list(Insurance_table['title'])

    
    pdf.set_fill_color(*hex2RGB('#FFE7CC'))
    pdf.rect(px2MM(1273),px2MM(204),px2MM(527),px2MM(304),'F')

    pdf.image(join(cwd,'assets','images','icons','Insurance.png'),px2MM(1313),px2MM(244),px2MM(60),px2MM(60))
    pdf.set_xy(px2MM(1393),px2MM(246))  
    pdf.set_font('LeagueSpartan-Regular', size=px2pts(40))
    pdf.set_text_color(*hex2RGB('#000000'))
    pdf.cell(px2MM(158),px2MM(56),"Insurance",border=0,align="L")
    
    
    pdf.set_text_color(*hex2RGB('#000000'))
    for row in range(len(Insurance_table)):
        pdf.set_draw_color(*hex2RGB('#E9EAEE'))
        pdf.set_fill_color(*hex2RGB('#FFFFFF'))
        pdf.rect(px2MM(1313), px2MM(324+(row*72)), px2MM(290), px2MM(72), 'DF')
        pdf.rect(px2MM(1603), px2MM(324+(row*72)), px2MM(157), px2MM(72), 'DF')
        
        # col1 text
        pdf.set_xy(px2MM(1333), px2MM(344+(row*72)))
        pdf.set_font('LeagueSpartan-Regular', size=px2pts(24))
        pdf.cell(px2MM(250),px2MM(32),Insurance_keys[row],border=0,align="L")

            #cal2 text
        pdf.set_xy(px2MM(1623), px2MM(344+(row*72)))
        pdf.set_font('LeagueSpartan-Regular', size=px2pts(24))
        if Insurance_table['value'].iloc[row]== '':  
            pdf.cell(px2MM(117),px2MM(32),'-',border=0,align="R")
        else:   
            pdf.cell(px2MM(117),px2MM(32),f"₹ {format_cash2(float(Insurance_table['value'].iloc[row]))}",border=0,align="R")


    #  # card 5
    
    try:
        val = json_data['oneview']['total']['liabilities']
    except:
        val = 'N/A'
    pdf.set_fill_color(*hex2RGB('#FFF3DB'))
    pdf.rect(px2MM(696),px2MM(204),px2MM(527),px2MM(304),'F')

    pdf.image(join(cwd,'assets','images','icons','Liabilities.png'),px2MM(736),px2MM(244),px2MM(60),px2MM(60))
    pdf.set_xy(px2MM(816),px2MM(246))  
    pdf.set_font('LeagueSpartan-Regular', size=px2pts(40))
    pdf.set_text_color(*hex2RGB('#000000'))
    pdf.cell(px2MM(155),px2MM(56),"Liabilities",border=0,align="L")
    
    pdf.set_xy(px2MM(1050),px2MM(253))  
    pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(30))
    pdf.set_text_color(*hex2RGB('#000000'))
    if val == 'N/A':
        pdf.cell(px2MM(130),px2MM(42),'0',border=0,align="R")
    else:
        
        pdf.cell(px2MM(130),px2MM(42),'₹ '+str(format_cash2(float(val))),border=0,align="R")
    
    
    
    liabilities_table = pd.DataFrame(json_data['oneview']['liabilities'])
    liabilities_keys = list(liabilities_table['title'])
    # liabilities_keys = list(liabilities.keys())
    
    pdf.set_text_color(*hex2RGB('#000000'))
    for row in range(len(liabilities_keys)):
        pdf.set_draw_color(*hex2RGB('#E9EAEE'))
        pdf.set_fill_color(*hex2RGB('#FFFFFF'))
        pdf.rect(px2MM(736), px2MM(324+(row*72)), px2MM(290), px2MM(72), 'DF')
        pdf.rect(px2MM(1026), px2MM(324+(row*72)), px2MM(157), px2MM(72), 'DF')

        #     # col1 text
        pdf.set_xy(px2MM(756), px2MM(344+(row*72)))
        pdf.set_font('LeagueSpartan-Regular', size=px2pts(24))
        pdf.cell(px2MM(250),px2MM(32),liabilities_keys[row],border=0,align="L")

            #cal2 text
        pdf.set_xy(px2MM(1046), px2MM(344+(row*72)))
        pdf.set_font('LeagueSpartan-Regular', size=px2pts(24))
        if liabilities_table['value'].iloc[row] =='':
            pdf.cell(px2MM(117),px2MM(25),'-',border=0,align="R")
        else:
            pdf.cell(px2MM(117),px2MM(25),f'₹ {format_cash2(float(liabilities_table["value"].iloc[row]))}',border=0,align="R")

    desc_text = '''Disclaimer: The accuracy and comprehensiveness of this information is dependent on the details provided to us. The more accurate the information, the better our financial suggestions will be.'''
    pdf.set_xy(px2MM(405), px2MM(976))
    pdf.set_font('LeagueSpartan-Regular', size=px2pts(24))
    pdf.set_text_color(*hex2RGB('#ffffff'))
    pdf.multi_cell(px2MM(1110),px2MM(32),desc_text,border=0,align="C")
    
    #//*-----Index Text of Page--**////
    index_text(pdf,'#FFFFFF')  
    global your_1_view_idx
    your_1_view_idx = pdf.page_no()
     
        

#//*----Calling of main functiong by taking sys.argv----*//

